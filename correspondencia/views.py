"""
Vistas del módulo de correspondencia.

Este módulo contiene todas las vistas necesarias para gestionar el flujo completo
de correspondencia entrante, incluyendo:

- Radicación manual y automática de correspondencia
- Gestión de bandejas de trabajo (personal, ventanilla, clasificados)
- Procesamiento de correos electrónicos entrantes
- Gestión de correspondencia saliente
- APIs para funcionalidades dinámicas (SLA, carga de subseries)
- Control de acceso basado en roles y permisos

Las vistas implementan:
- Autenticación y autorización requerida
- Validación de formularios
- Manejo de transacciones de base de datos
- Paginación y filtrado
- Respuestas JSON para funcionalidades AJAX
- Integración con sistema de mensajes de Django

Autor: Sistema de Gestión Documental
Fecha: 2025
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404, HttpResponseForbidden
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.contrib import messages
from .forms import (
    CorrespondenciaForm, ContactoForm, CompartirCorrespondenciaForm, CompartirOtrasOficinasForm, ManualRadicacionCorreoForm,
    EntidadExternaForm, RespuestaCorrespondenciaForm, AprobarRechazarRespuestaForm, HistorialFilterForm,
    GrupoAgendaForm, ComunicacionMasivaForm, ComunicacionInternaForm,
    RadicacionRapidaEntranteForm, RadicacionRapidaSalienteForm
)
from .models import (
    Correspondencia, HistorialCorrespondencia, Contacto, CorreoEntrante, CorreoProblematico, OficinaProductora, SerieDocumental,
    SubserieDocumental, AdjuntoCorreoEntrante, AdjuntoCorreo, AdjuntoCorrespondenciaRapida,
    DistribucionInternaUsuario, EntidadExterna,
    CorrespondenciaSalida, AdjuntoSalida, HistorialSalida, SalidaDestinatario,
    GrupoAgenda, ComunicacionMasiva, ComunicacionDestinatario, Notificacion, AccesoCorrespondenciaOficina,
    ComunicacionInterna, MOTIVO_PAPELERA_CHOICES, EstadoSincronizacionCorreos, MEDIO_RECIBIDO_CHOICES,
    MEDIO_RECEPCION_CHOICES, ORIGEN_RADICACION_CHOICES, ESTADO_RESPUESTA_RAPIDA_CHOICES,
    TIPO_TRAMITE_CHOICES, DIAS_RESPUESTA_POR_TIPO_TRAMITE, calcular_dias_habiles, TipoTramite,
    AuditoriaContacto, extraer_dominios_candidatos, normalizar_dominio_correo,
)
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, View, CreateView, TemplateView, RedirectView
from django.utils import timezone
from django.db import transaction
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from documentos.models import PerfilUsuario, OficinaProductora # Necesario para obtener perfil
from correspondencia.permisos import usuario_puede_gestion_operativa
from django.contrib.auth.models import User
from django.db.models import Q, Count, Case, When, BooleanField, Subquery, OuterRef, F, Value, Exists # Para consultas OR y anotaciones
from datetime import timedelta
from django.core.paginator import Paginator
from django.core.management import call_command
from .tasks import procesar_emails_imap_manual, procesar_emails_periodico
from django.core.files.base import ContentFile
import traceback
from django.db.utils import IntegrityError
import os
from django.core.exceptions import ValidationError
from django.http import JsonResponse
from django.http import HttpResponse
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from django.db.models import Max
from math import ceil
# justo debajo de los imports de modelos
# views.py (encabezado)
from documentos.models import SubserieDocumental as SubserieDoc
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from django.core.mail import EmailMessage, EmailMultiAlternatives
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.views import PasswordResetView
from django.urls import reverse_lazy
from django.db.models.functions import Coalesce, Concat
from django.db.models import F # <-- Importar F
from django.db.models import Q # Importar Q para búsquedas complejas
from django.db import models # <-- IMPORTAR BASE MODELS
from django.http import JsonResponse
from .modelos_minimos_sla import SubserieTramite
from .trd_interna import obtener_clasificacion_comunicacion_interna
from .utils_sla import get_cutoff_time, aplicar_corte, sumar_habiles
from .utils.asunto_salida import normalizar_asunto_salida, asunto_respuesta_desde_entrada
from .utils.dashboard_kpis import get_ventanilla_dashboard_kpis
from .utils.sla_queries import excluir_entrantes_con_respuesta, ids_entrantes_con_respuesta
from .utils.historial_queries import fetch_historial_combinado, hydrate_historial_urls
from django.utils import timezone
import datetime
from math import ceil
from django.core.cache import cache
from io import BytesIO
import logging
import re
import smtplib
import socket
import requests
import json
try:
    import dns.resolver  # type: ignore
except Exception:  # pragma: no cover
    dns = None

DOMINIOS_GENERICOS = [
    'hotmail.com', 'yahoo.com', 'outlook.com', 'icloud.com'
]

logger = logging.getLogger(__name__)

PERMISO_RESPUESTA_DISCRECIONAL = 'correspondencia.responder_correspondencia_discrecional'


class PasswordResetHTMLView(PasswordResetView):
    form_class = PasswordResetForm
    template_name = 'registration/password_reset_form.html'
    email_template_name = 'registration/password_reset_email.html'
    subject_template_name = 'registration/password_reset_subject.txt'
    success_url = reverse_lazy('password_reset_done')

    def form_valid(self, form):
        opts = {
            'use_https': self.request.is_secure(),
            'token_generator': self.token_generator,
            'from_email': self.from_email,
            'email_template_name': self.email_template_name,
            'subject_template_name': self.subject_template_name,
            'request': self.request,
            'html_email_template_name': self.email_template_name,
        }
        form.save(**opts)
        return redirect(self.get_success_url())


def usuario_tiene_permiso_respuesta_discrecional(user):
    return bool(
        user and user.is_authenticated and (
            user.is_superuser or user.has_perm(PERMISO_RESPUESTA_DISCRECIONAL)
        )
    )


def usuario_puede_responder_correspondencia(correspondencia, user):
    perfil_usuario = getattr(user, 'perfil', None)
    es_ventanilla_o_admin = user.groups.filter(name__in=['Ventanilla', 'Admin']).exists() or user.is_superuser
    es_de_oficina_destino = perfil_usuario and perfil_usuario.oficina == correspondencia.oficina_destino
    es_asignado_inicial = correspondencia.usuario_destino_inicial == user
    fue_compartido_con_usuario = DistribucionInternaUsuario.objects.filter(
        correspondencia=correspondencia,
        usuario_asignado=user,
    ).exists()

    tiene_acceso_interoficina_respuesta = False
    if perfil_usuario and getattr(perfil_usuario, 'oficina_id', None):
        acceso_interoficina = correspondencia.accesos_oficinas.filter(
            oficina_id=perfil_usuario.oficina_id
        ).first()
        if acceso_interoficina and acceso_interoficina.puede_responder:
            if not acceso_interoficina.solo_lider or user.groups.filter(name='Lider de Oficina').exists():
                tiene_acceso_interoficina_respuesta = True

    return bool(
        es_ventanilla_o_admin or
        es_de_oficina_destino or
        es_asignado_inicial or
        fue_compartido_con_usuario or
        tiene_acceso_interoficina_respuesta
    )


def _get_public_domain_whitelist():
    dominios = []
    for valor in getattr(settings, 'EMAIL_DOMAINS_WHITELIST', []):
        dominio = normalizar_dominio_correo(valor)
        if dominio and dominio not in dominios:
            dominios.append(dominio)
    return dominios


def _build_domain_policy(dominio, entidad_id=None):
    dominio_normalizado = normalizar_dominio_correo(dominio)
    entidad_seleccionada = None
    entidad_asociada = EntidadExterna.buscar_por_dominio(dominio_normalizado)
    whitelist_publica = _get_public_domain_whitelist()

    if entidad_id and str(entidad_id).isdigit():
        entidad_seleccionada = EntidadExterna.objects.filter(pk=int(entidad_id)).first()

    known_public = dominio_normalizado in whitelist_publica
    known_for_entity = entidad_seleccionada.tiene_dominio_autorizado(dominio_normalizado) if entidad_seleccionada else False

    return {
        'selected_entity_id': entidad_seleccionada.pk if entidad_seleccionada else None,
        'selected_entity_name': entidad_seleccionada.nombre if entidad_seleccionada else '',
        'matching_entity_id': entidad_asociada.pk if entidad_asociada else None,
        'matching_entity_name': entidad_asociada.nombre if entidad_asociada else '',
        'known_in_system': bool(entidad_asociada or known_public),
        'known_for_entity': known_for_entity,
        'known_in_public_whitelist': known_public,
        'can_register_for_entity': bool(entidad_seleccionada and dominio_normalizado and not known_for_entity),
    }


def _resolver_clasificacion_comunicacion_interna(oficina_remitente, serie_id=None, subserie_id=None):
    serie, subserie = obtener_clasificacion_comunicacion_interna()

    if not serie_id or not subserie_id:
        trd = ComunicacionInterna.construir_trd_desde_estructura(oficina_remitente, serie=serie, subserie=subserie)
        if not trd:
            aviso_trd = getattr(
                oficina_remitente,
                'codigo_trd_comunicacion_interna_display',
                'sin trd por falta de mapeo',
            )
            raise ValidationError(
                f"No fue posible construir la TRD. La oficina remitente está {aviso_trd}."
            )
        return serie, subserie, trd

    if str(serie.pk) != str(serie_id) or str(subserie.pk) != str(subserie_id):
        raise ValidationError(
            'Para comunicaciones internas solo está permitida la serie COMUNICACIONES OFICIALES '
            'con la subserie Comunicaciones Internas.'
        )

    trd = ComunicacionInterna.construir_trd_desde_estructura(oficina_remitente, serie=serie, subserie=subserie)
    if not trd:
        aviso_trd = getattr(
            oficina_remitente,
            'codigo_trd_comunicacion_interna_display',
            'sin trd por falta de mapeo',
        )
        raise ValidationError(
            f"No fue posible construir la TRD. La oficina remitente está {aviso_trd}."
        )

    return serie, subserie, trd



def _base_bandeja_personal_queryset(usuario, oficina):
    """Query base: correspondencia asignada o compartida directamente con el usuario."""
    return (
        Correspondencia.objects.filter(
            oficina_destino=oficina,
        ).filter(
            Q(usuario_destino_inicial=usuario) |
            Q(distribuciones_internas__usuario_asignado=usuario),
        ).distinct()
    )


def _base_bandeja_oficina_queryset(oficina):
    """Query base compartido entre dashboard y bandeja de oficina."""
    return (
        Correspondencia.objects.filter(
            tipo_radicado='ENTRANTE',
            oficina_destino=oficina,
        ).annotate(
            num_distribuciones=Count('distribuciones_internas', distinct=True),
            total_destinatarios=Count('distribuciones_internas', distinct=True),
            total_leidos=Count(
                'distribuciones_internas',
                filter=Q(distribuciones_internas__leido=True),
                distinct=True,
            ),
            acceso_solo_lectura=Value(False, output_field=BooleanField()),
        ).filter(
            num_distribuciones__gt=1
        )
    )


def _correspondencias_compartidas_ids(usuario):
    """IDs de correspondencia asignadas o compartidas con el usuario (materializado)."""
    return list(
        Correspondencia.objects.filter(
            Q(usuario_destino_inicial=usuario) |
            Q(distribuciones_internas__usuario_asignado=usuario)
        ).values_list('pk', flat=True).distinct()
    )


def _base_bandeja_salientes_queryset(usuario, *, alcance='mias', oficina=None):
    """Query base liviano para bandeja de respuestas salientes (sin anotaciones pesadas)."""
    if alcance == 'oficina' and oficina is not None:
        return CorrespondenciaSalida.objects.filter(oficina_emisora=oficina)

    compartida_ids = _correspondencias_compartidas_ids(usuario)
    return CorrespondenciaSalida.objects.filter(
        Q(usuario_redactor=usuario) |
        (
            Q(respuesta_a_id__in=compartida_ids) &
            ~Q(usuario_redactor=usuario)
        )
    )


def _usuario_puede_ver_salida(usuario, respuesta):
    """Mismos criterios de acceso que la bandeja salientes: ventanilla/admin, redactor, entrada compartida u oficina."""
    if usuario.groups.filter(name__in=['Ventanilla', 'Admin']).exists() or usuario.is_superuser:
        return True
    if respuesta.usuario_redactor_id == usuario.pk:
        return True
    if respuesta.respuesta_a_id and respuesta.respuesta_a_id in _correspondencias_compartidas_ids(usuario):
        return True
    try:
        oficina_usuario_id = usuario.perfil.oficina_id
    except AttributeError:
        oficina_usuario_id = None
    if (
        oficina_usuario_id
        and respuesta.oficina_emisora_id == oficina_usuario_id
    ):
        return True
    return False


def _annotate_bandeja_salientes_page(queryset, usuario):
    """Anotaciones de destinatarios solo para la página visible (evita COUNT sobre GROUP BY)."""
    return queryset.defer(
        'cuerpo',
        'motivo_respuesta_discrecional',
        'envio_detalle_snapshot',
    ).annotate(
        total_destinatarios=Count('destinatarios', distinct=True),
        destinatarios_enviados=Count(
            'destinatarios',
            filter=Q(destinatarios__estado='ENVIADO'),
            distinct=True,
        ),
        destinatarios_fallidos=Count(
            'destinatarios',
            filter=Q(destinatarios__estado='FALLO'),
            distinct=True,
        ),
        es_respuesta_compartida=Case(
            When(usuario_redactor=usuario, then=False),
            default=True,
            output_field=models.BooleanField(),
        ),
    ).select_related(
        'destinatario_contacto',
        'destinatario_contacto__entidad_externa',
        'usuario_redactor',
        'usuario_aprobador',
        'envio_grupo',
    ).prefetch_related('adjuntos', 'destinatarios')


def _build_respuesta_valida_subquery(outer_ref_field: str):
    return (
        CorrespondenciaSalida.objects
        .filter(
            respuesta_a_id=OuterRef(outer_ref_field),
            estado__in=['APROBADA', 'ENVIADA']
        )
        .annotate(
            fecha_referencia=Coalesce('fecha_envio', 'fecha_aprobacion', 'fecha_creacion')
        )
        .order_by('-fecha_referencia', '-fecha_creacion')
    )


def _marcar_lectura_usuario_en_pagina(correspondencias, usuario):
    """Marca leido_por_usuario_actual sin Exists correlacionado en el queryset principal."""
    ids = [item.id for item in correspondencias]
    if not ids:
        return
    leidos_ids = set(
        DistribucionInternaUsuario.objects.filter(
            correspondencia_id__in=ids,
            usuario_asignado=usuario,
            leido=True,
        ).values_list('correspondencia_id', flat=True)
    )
    for item in correspondencias:
        item.leido_por_usuario_actual = item.id in leidos_ids


def _asignar_estado_respuesta_bandeja(correspondencia, fecha_respuesta=None):
    fecha_limite = correspondencia.fecha_limite_respuesta
    correspondencia.respuesta_registrada_en = fecha_respuesta
    correspondencia.fue_respondida = fecha_respuesta is not None
    correspondencia.respondida_a_tiempo = bool(
        correspondencia.fue_respondida and fecha_limite and fecha_respuesta <= fecha_limite
    )

    if not correspondencia.requiere_respuesta:
        correspondencia.estado_respuesta_bandeja = 'no_requiere'
    elif not correspondencia.fue_respondida:
        correspondencia.estado_respuesta_bandeja = 'pendiente'
    elif not fecha_limite:
        correspondencia.estado_respuesta_bandeja = 'sin_plazo'
    elif correspondencia.respondida_a_tiempo:
        correspondencia.estado_respuesta_bandeja = 'a_tiempo'
    else:
        correspondencia.estado_respuesta_bandeja = 'fuera_tiempo'


def _guardar_adjuntos_radicacion_fisica(correspondencia, request):
    """Guarda adjuntos manuales cuando el usuario los envía (opcionales para cualquier medio)."""

    def _obtener_archivos_adjuntos(req):
        posibles_campos = (
            'adjuntos_entrada',
            'radicar-adjuntos_archivos',
            'adjuntos_archivos',
        )
        for nombre_campo in posibles_campos:
            archivos = req.FILES.getlist(nombre_campo)
            if archivos:
                return archivos
        return []

    adjuntos_guardados = 0
    total_size = 0
    max_files = 10
    max_total_size = 15 * 1024 * 1024
    max_file_size = 5 * 1024 * 1024
    allowed_extensions_fisico = ['.pdf', '.jpg', '.jpeg', '.png', '.tiff', '.bmp']
    allowed_extensions_general = ['.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png', '.tiff', '.bmp', '.xls', '.xlsx', '.zip']
    adjuntos_files = _obtener_archivos_adjuntos(request)

    if not adjuntos_files:
        return 0

    if len(adjuntos_files) > max_files:
        raise ValidationError(f"Máximo {max_files} archivos permitidos por correspondencia.")

    allowed_extensions = (
        allowed_extensions_fisico
        if correspondencia.medio_recepcion == 'FISICO'
        else allowed_extensions_general
    )

    for adjunto_file in adjuntos_files:
        file_extension = os.path.splitext(adjunto_file.name)[1].lower()
        if file_extension not in allowed_extensions:
            raise ValidationError(
                f"Formato no permitido: {adjunto_file.name}. Solo se permiten: {', '.join(allowed_extensions)}"
            )

        if adjunto_file.size > max_file_size:
            raise ValidationError(f"Archivo {adjunto_file.name} excede el tamaño máximo de 5MB.")

        total_size += adjunto_file.size
        if total_size > max_total_size:
            raise ValidationError(
                f"El tamaño total de archivos ({total_size / (1024*1024):.1f}MB) excede el límite de 15MB."
            )

    for adjunto_file in adjuntos_files:
        try:
            adjunto = AdjuntoCorreo(
                correspondencia=correspondencia,
                nombre_original=adjunto_file.name,
                tipo_mime=adjunto_file.content_type or 'application/octet-stream'
            )
            adjunto.archivo.save(adjunto_file.name, adjunto_file, save=True)
            adjuntos_guardados += 1
        except Exception as exc:
            messages.warning(request, f"No se pudo guardar el adjunto {adjunto_file.name}: {exc}")

    return adjuntos_guardados



# Create your views here.

# --- Vista para la Bandeja de Correos Clasificados ---
class EsVentanillaMixin(UserPassesTestMixin):
    """Mixin para verificar si el usuario pertenece al grupo 'Ventanilla'."""
@login_required
def grupos_agenda_index(request):
    """Lista de grupos del usuario según su oficina."""
    oficina = getattr(request.user, 'perfil', None).oficina if hasattr(request.user, 'perfil') else None
    if not oficina:
        messages.error(request, 'Tu usuario no tiene oficina asignada. Contacta al administrador.')
        return redirect('correspondencia:dashboard_usuario')
    grupos = GrupoAgenda.objects.filter(oficina_propietaria=oficina).prefetch_related('contactos').annotate(
        total_contactos=Count('contactos', distinct=True)
    )

    grupos_total = grupos.count()
    grupos_activos_total = grupos.filter(activo=True).count()
    contactos_total = sum(grupo.total_contactos for grupo in grupos)
    promedio_contactos = round(contactos_total / grupos_total) if grupos_total else 0

    return render(request, 'correspondencia/agenda/grupos_index.html', {
        'titulo_pagina': 'Categorización por agenda',
        'grupos': grupos,
        'oficina_usuario': oficina,
        'grupos_total': grupos_total,
        'grupos_activos_total': grupos_activos_total,
        'contactos_total': contactos_total,
        'promedio_contactos': promedio_contactos,
    })


@login_required
def grupo_agenda_crear(request):
    oficina = getattr(request.user, 'perfil', None).oficina if hasattr(request.user, 'perfil') else None
    if not oficina:
        messages.error(request, 'Tu usuario no tiene oficina asignada.')
        return redirect('correspondencia:dashboard_usuario')
    if request.method == 'POST':
        form = GrupoAgendaForm(request.POST, oficina_propietaria=oficina)
        if form.is_valid():
            grupo = form.save(commit=False)
            grupo.oficina_propietaria = oficina
            grupo.creado_por = request.user
            grupo.save()
            form.save_m2m()
            messages.success(request, 'Grupo creado correctamente.')
            return redirect('correspondencia:grupos_agenda_index')
    else:
        form = GrupoAgendaForm(oficina_propietaria=oficina)
    return render(request, 'correspondencia/agenda/grupo_form.html', {
        'titulo_pagina': 'Nuevo grupo de agenda',
        'form': form,
        'accion': 'crear'
    })


@login_required
def grupo_agenda_editar(request, pk: int):
    oficina = getattr(request.user, 'perfil', None).oficina if hasattr(request.user, 'perfil') else None
    grupo = get_object_or_404(GrupoAgenda, pk=pk)
    if not oficina or grupo.oficina_propietaria_id != oficina.id:
        raise Http404()
    if request.method == 'POST':
        form = GrupoAgendaForm(request.POST, instance=grupo, oficina_propietaria=oficina)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grupo actualizado.')
            return redirect('correspondencia:grupos_agenda_index')
    else:
        form = GrupoAgendaForm(instance=grupo, oficina_propietaria=oficina)
    return render(request, 'correspondencia/agenda/grupo_form.html', {
        'titulo_pagina': f'Editar grupo: {grupo.nombre}',
        'form': form,
        'accion': 'editar'
    })


@login_required
def grupo_agenda_eliminar(request, pk: int):
    oficina = getattr(request.user, 'perfil', None).oficina if hasattr(request.user, 'perfil') else None
    grupo = get_object_or_404(GrupoAgenda, pk=pk)
    if not oficina or grupo.oficina_propietaria_id != oficina.id:
        raise Http404()
    if request.method == 'POST':
        grupo.delete()
        messages.success(request, 'Grupo eliminado.')
        return redirect('correspondencia:grupos_agenda_index')
    return render(request, 'correspondencia/agenda/grupo_confirm_delete.html', {
        'titulo_pagina': f'Eliminar grupo: {grupo.nombre}',
        'grupo': grupo
    })


@login_required
def contactos_agenda_paginado_ajax(request):
    """Devuelve contactos paginados de la oficina del usuario, con filtros por entidad y búsqueda.
    Responde JSON: {success, results: [...], page, page_size, total, total_pages}
    """
    if request.method != 'GET' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'})

    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.oficina:
        return JsonResponse({'success': False, 'error': 'Usuario sin oficina asignada.'})

    try:
        entidad_id = request.GET.get('entidad_id')
        query = (request.GET.get('q') or '').strip()
        page = max(int(request.GET.get('page') or 1), 1)
        page_size = min(max(int(request.GET.get('page_size') or 25), 5), 100)

        qs = Contacto.objects.select_related('entidad_externa')\
            .order_by('entidad_externa__nombre', 'apellidos', 'nombres')

        if entidad_id and entidad_id.isdigit():
            qs = qs.filter(entidad_externa_id=int(entidad_id))
        if query:
            from django.db.models import Q
            qs = qs.filter(Q(nombres__icontains=query) | Q(apellidos__icontains=query) | Q(correo_electronico__icontains=query))

        total = qs.count()
        total_pages = max(ceil(total / page_size), 1)
        if page > total_pages:
            page = total_pages
        offset = (page - 1) * page_size
        items = list(qs[offset:offset+page_size])

        def serialize_contacto(c: Contacto):
            tiene_email = bool(c.correo_electronico)
            return {
                'id': c.id,
                'nombre': c.nombre_completo,
                'entidad': c.entidad_externa.nombre if c.entidad_externa else 'Sin entidad',
                'entidad_id': c.entidad_externa.id if c.entidad_externa else None,
                'email': c.correo_electronico,
                'tiene_email': tiene_email,
            }

        return JsonResponse({
            'success': True,
            'results': [serialize_contacto(c) for c in items],
            'page': page,
            'page_size': page_size,
            'total': total,
            'total_pages': total_pages,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al listar contactos: {str(e)}'})


@login_required
def validar_nombre_grupo_ajax(request):
    """Valida unicidad del nombre de grupo dentro de la oficina del usuario."""
    if request.method != 'GET' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'})

    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.oficina:
        return JsonResponse({'ok': False, 'error': 'Usuario sin oficina asignada.'})

    nombre = (request.GET.get('nombre') or '').strip()
    exclude_id = request.GET.get('exclude_id')
    if not nombre:
        return JsonResponse({'ok': False, 'exists': False})
    qs = GrupoAgenda.objects.filter(oficina_propietaria=perfil.oficina, nombre__iexact=nombre)
    if exclude_id and exclude_id.isdigit():
        qs = qs.exclude(pk=int(exclude_id))
    return JsonResponse({'ok': True, 'exists': qs.exists()})


@login_required
def grupo_agenda_detalle_ajax(request, pk: int):
    """Devuelve datos del grupo (para edición en modal) y sus contactos seleccionados.
    JSON: {success, grupo: {id, nombre, activo, descripcion}, contactos: [{id,nombre,entidad,email}]}
    """
    if request.method != 'GET' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'})

    perfil = getattr(request.user, 'perfil', None)
    if not perfil or not perfil.oficina:
        return JsonResponse({'success': False, 'error': 'Usuario sin oficina asignada.'})

    grupo = get_object_or_404(GrupoAgenda, pk=pk)
    if grupo.oficina_propietaria_id != perfil.oficina.id:
        return JsonResponse({'success': False, 'error': 'No autorizado.'}, status=403)

    contactos = grupo.contactos.select_related('entidad_externa').all()
    data_contactos = [{
        'id': c.id,
        'nombre': c.nombre_completo,
        'entidad': c.entidad_externa.nombre if c.entidad_externa else 'Sin entidad',
        'email': c.correo_electronico,
    } for c in contactos]

    return JsonResponse({
        'success': True,
        'grupo': {
            'id': grupo.id,
            'nombre': grupo.nombre,
            'activo': grupo.activo,
            'descripcion': grupo.descripcion or ''
        },
        'contactos': data_contactos
    })

@login_required
def destinatarios_salida_ajax(request, pk):
    """Devuelve los destinatarios de una correspondencia saliente en formato JSON."""
    print(f"DEBUG: destinatarios_salida_ajax llamado con pk={pk}")
    print(f"DEBUG: Método: {request.method}")
    print(f"DEBUG: Headers: {dict(request.headers)}")
    
    if request.method != 'GET' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        print("DEBUG: Método no permitido o no es AJAX")
        return JsonResponse({'success': False, 'error': 'Método no permitido.'})

    perfil = getattr(request.user, 'perfil', None)
    print(f"DEBUG: Usuario: {request.user}, Perfil: {perfil}")
    if not perfil or not perfil.oficina:
        print("DEBUG: Usuario sin oficina asignada")
        return JsonResponse({'success': False, 'error': 'Usuario sin oficina asignada.'})

    try:
        salida = get_object_or_404(CorrespondenciaSalida, pk=pk)
        print(f"DEBUG: Salida encontrada: {salida}")
        print(f"DEBUG: Tipo de envío: {salida.envio_tipo}")
        print(f"DEBUG: Oficina emisora: {salida.oficina_emisora}")
        print(f"DEBUG: Oficina del usuario: {perfil.oficina}")
        
        # Verificar permisos - el usuario debe tener acceso a esta salida
        puede_ver = (
            salida.oficina_emisora == perfil.oficina or
            request.user.groups.filter(name__in=['Admin', 'Ventanilla']).exists() or
            request.user.is_superuser
        )
        print(f"DEBUG: Puede ver: {puede_ver}")
        
        if not puede_ver:
            print("DEBUG: No autorizado")
            return JsonResponse({'success': False, 'error': 'No autorizado.'}, status=403)

        # Obtener destinatarios según el tipo de envío
        data_destinatarios = []
        
        if salida.envio_tipo == 'MULTIPLE_SELECTIVO':
            print("DEBUG: Procesando envío múltiple selectivo")
            # Para envíos múltiples, buscar en SalidaDestinatario
            destinatarios = salida.destinatarios.select_related('contacto__entidad_externa').all()
            print(f"DEBUG: Total destinatarios encontrados: {destinatarios.count()}")
            
            for dest in destinatarios:
                print(f"DEBUG: Procesando destinatario: {dest}")
                data_destinatarios.append({
                    'id': dest.id,
                    'nombre': dest.nombre_snapshot or (dest.contacto.nombre_completo if dest.contacto else 'Sin nombre'),
                    'email': dest.email_snapshot,
                    'entidad': dest.contacto.entidad_externa.nombre if dest.contacto and dest.contacto.entidad_externa else 'Sin entidad',
                    'estado': dest.get_estado_display(),
                    'fecha_envio': dest.fecha_envio.strftime('%Y-%m-%d %H:%M') if dest.fecha_envio else None,
                    'estado_class': 'success' if dest.estado == 'ENVIADO' else 'danger' if dest.estado == 'FALLO' else 'warning'
                })
        elif salida.envio_tipo == 'GRUPO' and salida.envio_grupo:
            print("DEBUG: Procesando envío a grupo")
            # Para envíos a grupo, obtener contactos del grupo
            contactos_grupo = salida.envio_grupo.contactos.select_related('entidad_externa').all()
            print(f"DEBUG: Total contactos en grupo: {contactos_grupo.count()}")
            
            for contacto in contactos_grupo:
                print(f"DEBUG: Procesando contacto del grupo: {contacto}")
                data_destinatarios.append({
                    'id': contacto.id,
                    'nombre': contacto.nombre_completo,
                    'email': contacto.correo_electronico or 'Sin email',
                    'entidad': contacto.entidad_externa.nombre if contacto.entidad_externa else 'Sin entidad',
                    'estado': 'Enviado',  # Asumir que se envió al grupo
                    'fecha_envio': salida.fecha_envio.strftime('%Y-%m-%d %H:%M') if salida.fecha_envio else None,
                    'estado_class': 'success'
                })
        else:
            print("DEBUG: Procesando envío individual")
            # Para envíos individuales, usar el destinatario principal
            if salida.destinatario_contacto:
                print(f"DEBUG: Destinatario individual: {salida.destinatario_contacto}")
                data_destinatarios.append({
                    'id': salida.destinatario_contacto.id,
                    'nombre': salida.destinatario_contacto.nombre_completo,
                    'email': salida.destinatario_email or salida.destinatario_contacto.correo_electronico or 'Sin email',
                    'entidad': salida.destinatario_contacto.entidad_externa.nombre if salida.destinatario_contacto.entidad_externa else 'Sin entidad',
                    'estado': salida.get_estado_display(),
                    'fecha_envio': salida.fecha_envio.strftime('%Y-%m-%d %H:%M') if salida.fecha_envio else None,
                    'estado_class': 'success' if salida.estado == 'ENVIADA' else 'danger' if salida.estado == 'ERROR_ENVIO' else 'warning'
                })

        print(f"DEBUG: Total destinatarios procesados: {len(data_destinatarios)}")
        print(f"DEBUG: Datos finales: {data_destinatarios}")

        response_data = {
            'success': True,
            'salida': {
                'id': salida.id,
                'numero_radicado': salida.numero_radicado_salida,
                'asunto': salida.asunto,
                'total_destinatarios': len(data_destinatarios)
            },
            'destinatarios': data_destinatarios
        }
        
        print(f"DEBUG: Respuesta final: {response_data}")
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"DEBUG: Error en destinatarios_salida_ajax: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Error al obtener destinatarios: {str(e)}'})

@login_required
def comunicaciones_list(request):
    oficina = getattr(request.user, 'perfil', None).oficina if hasattr(request.user, 'perfil') else None
    if not oficina:
        messages.error(request, 'Tu usuario no tiene oficina asignada.')
        return redirect('correspondencia:dashboard_usuario')
    comunicaciones = ComunicacionMasiva.objects.filter(oficina_emisora=oficina).order_by('-fecha_creacion')
    return render(request, 'correspondencia/agenda/comunicaciones_list.html', {
        'titulo_pagina': 'Comunicaciones masivas',
        'comunicaciones': comunicaciones
    })


@login_required
def comunicacion_masiva_crear(request):
    oficina = getattr(request.user, 'perfil', None).oficina if hasattr(request.user, 'perfil') else None
    if not oficina:
        messages.error(request, 'Tu usuario no tiene oficina asignada.')
        return redirect('correspondencia:dashboard_usuario')
    if request.method == 'POST':
        form = ComunicacionMasivaForm(request.POST, oficina_emisora=oficina)
        if form.is_valid():
            comunicacion = form.save(commit=False)
            comunicacion.oficina_emisora = oficina
            comunicacion.usuario_creador = request.user
            comunicacion.save()
            # Crear destinatarios desde grupos seleccionados
            grupos = form.cleaned_data.get('grupos')
            contactos_ids = set()
            if grupos:
                for g in grupos:
                    for cid in g.contactos.values_list('id', flat=True):
                        contactos_ids.add(cid)
            # Validar y crear destinatarios
            contactos = Contacto.objects.filter(id__in=list(contactos_ids)).exclude(correo_electronico__isnull=True).exclude(correo_electronico='')
            for contacto in contactos:
                try:
                    ComunicacionDestinatario.objects.create(
                        comunicacion=comunicacion,
                        contacto=contacto,
                        email_snapshot=contacto.correo_electronico,
                        nombre_snapshot=contacto.nombre_completo
                    )
                except IntegrityError:
                    pass
            messages.success(request, 'Borrador de comunicación creado. Puedes enviarlo cuando quieras.')
            return redirect('correspondencia:comunicacion_masiva_detalle', pk=comunicacion.pk)
    else:
        form = ComunicacionMasivaForm(oficina_emisora=oficina)
    return render(request, 'correspondencia/agenda/comunicacion_form.html', {
        'titulo_pagina': 'Nueva comunicación masiva',
        'form': form
    })


@login_required
def comunicacion_masiva_detalle(request, pk: int):
    oficina = getattr(request.user, 'perfil', None).oficina if hasattr(request.user, 'perfil') else None
    comunicacion = get_object_or_404(ComunicacionMasiva, pk=pk)
    if not oficina or comunicacion.oficina_emisora_id != oficina.id:
        raise Http404()
    return render(request, 'correspondencia/agenda/comunicacion_detalle.html', {
        'titulo_pagina': 'Detalle de comunicación',
        'comunicacion': comunicacion,
        'destinatarios': comunicacion.destinatarios.select_related('contacto')
    })


@login_required
def comunicacion_masiva_enviar(request, pk: int):
    if request.method != 'POST':
        raise Http404()
    oficina = getattr(request.user, 'perfil', None).oficina if hasattr(request.user, 'perfil') else None
    comunicacion = get_object_or_404(ComunicacionMasiva, pk=pk)
    if not oficina or comunicacion.oficina_emisora_id != oficina.id:
        raise Http404()

    destinatarios = list(comunicacion.destinatarios.select_related('contacto'))
    if not destinatarios:
        messages.error(request, 'No hay destinatarios para enviar.')
        return redirect('correspondencia:comunicacion_masiva_detalle', pk=pk)

    html_message = comunicacion.cuerpo
    plain_message = strip_tags(html_message)
    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', getattr(settings, 'EMAIL_HOST_USER', None))
    
    enviados = 0
    for d in destinatarios:
        if d.estado == 'ENVIADO':
            continue
        try:
            email = EmailMessage(
                subject=comunicacion.asunto,
                body=plain_message,
                from_email=from_email,
                to=[d.email_snapshot]
            )
            email.content_subtype = 'html'
            email.body = html_message
            email.send(fail_silently=False)
            d.estado = 'ENVIADO'
            d.fecha_envio = timezone.now()
            d.save(update_fields=['estado', 'fecha_envio'])
            enviados += 1
        except Exception as e_send:
            d.estado = 'FALLO'
            d.detalle_error = str(e_send)
            d.save(update_fields=['estado', 'detalle_error'])
    
    # Actualizar estado de la comunicación
    total_destinatarios = len(destinatarios)
    if enviados == total_destinatarios:
        comunicacion.estado = 'ENVIADA'
    elif enviados > 0:
        comunicacion.estado = 'PARCIAL'
    else:
        comunicacion.estado = 'ERROR'
    
    if enviados > 0:
        comunicacion.fecha_envio = timezone.now()
    
    comunicacion.save(update_fields=['estado', 'fecha_envio'])
    
    if enviados == total_destinatarios:
        messages.success(request, f'Comunicación enviada exitosamente a {enviados} destinatarios.')
    elif enviados > 0:
        messages.warning(request, f'Comunicación enviada parcialmente: {enviados} de {total_destinatarios} destinatarios.')
    else:
        messages.error(request, 'No se pudo enviar la comunicación a ningún destinatario.')
    return redirect('correspondencia:comunicacion_masiva_detalle', pk=pk)


@login_required
# @permission_required('correspondencia.view_contacto', raise_exception=True) # Permiso opcional
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def editar_contacto(request, pk):
    """Edita un contacto existente. Solo accesible por Ventanilla.
    Incluye: auditoría de cambios y alerta si se cambia email con correspondencia activa."""
    contacto = get_object_or_404(Contacto, pk=pk)
    
    # Capturar valores originales ANTES de la edición para auditoría
    campos_originales = {
        'nombres': contacto.nombres,
        'apellidos': contacto.apellidos or '',
        'cargo': contacto.cargo or '',
        'correo_electronico': contacto.correo_electronico or '',
        'telefono_contacto': contacto.telefono_contacto or '',
        'numero_documento': contacto.numero_documento or '',
        'entidad_externa': contacto.entidad_externa_id,
        'entidad_externa_nombre': contacto.entidad_externa.nombre if contacto.entidad_externa else '',
    }
    email_original = contacto.correo_electronico
    
    if request.method == 'POST':
        form = ContactoForm(request.POST, instance=contacto, user=request.user)
        if form.is_valid():
            try:
                nuevo_email = (form.cleaned_data.get('correo_electronico') or '').strip().lower()
                
                # --- Alerta: cambio de email con correspondencia activa ---
                if email_original and nuevo_email != email_original:
                    correspondencia_activa_entrante = Correspondencia.objects.filter(
                        remitente=contacto
                    ).exclude(
                        estado__in=['RESPONDIDA']
                    ).count()
                    
                    correspondencia_activa_salida = SalidaDestinatario.objects.filter(
                        contacto=contacto,
                        estado='PENDIENTE'
                    ).count()
                    
                    total_activas = correspondencia_activa_entrante + correspondencia_activa_salida
                    
                    if total_activas > 0:
                        messages.warning(
                            request,
                            f"⚠️ Atención: Se cambió el email de '{email_original}' a '{nuevo_email}'. "
                            f"Este contacto tiene {total_activas} correspondencia(s) activa(s). "
                            f"Los envíos ya realizados NO se ven afectados (tienen snapshot del email anterior), "
                            f"pero los envíos futuros usarán el nuevo email."
                        )
                
                # --- Registrar cambios en auditoría ---
                campos_modificados = {}
                campos_nuevos = {
                    'nombres': form.cleaned_data.get('nombres', ''),
                    'apellidos': form.cleaned_data.get('apellidos', '') or '',
                    'cargo': form.cleaned_data.get('cargo', '') or '',
                    'correo_electronico': nuevo_email,
                    'telefono_contacto': form.cleaned_data.get('telefono_contacto', '') or '',
                    'numero_documento': form.cleaned_data.get('numero_documento', '') or '',
                    'entidad_externa': form.cleaned_data.get('entidad_externa').id if form.cleaned_data.get('entidad_externa') else None,
                    'entidad_externa_nombre': form.cleaned_data.get('entidad_externa').nombre if form.cleaned_data.get('entidad_externa') else '',
                }
                
                for campo, valor_nuevo in campos_nuevos.items():
                    valor_original = campos_originales.get(campo, '')
                    if str(valor_original) != str(valor_nuevo):
                        campos_modificados[campo] = {
                            'antes': str(valor_original),
                            'despues': str(valor_nuevo)
                        }
                
                contacto_editado = form.save()
                
                # Solo registrar auditoría si hubo cambios reales
                if campos_modificados:
                    AuditoriaContacto.registrar_cambio(
                        contacto=contacto_editado,
                        usuario=request.user,
                        tipo_cambio='EDICION',
                        campos_modificados=campos_modificados,
                        request=request
                    )
                
                messages.success(
                    request,
                    f"Contacto '{contacto_editado.nombre_completo}' de la entidad '{contacto_editado.entidad_externa.nombre}' actualizado exitosamente."
                )
                return redirect('correspondencia:listar_contactos')
            except IntegrityError as e:
                correo = (form.cleaned_data.get('correo_electronico') or '').strip().lower()
                contacto_existente = Contacto.objects.filter(
                    correo_electronico__iexact=correo
                ).exclude(pk=contacto.pk).first()
                if contacto_existente:
                    messages.error(
                        request,
                        f"El correo '{correo}' ya está registrado en el contacto '{contacto_existente.nombre_completo}' "
                        f"de la entidad '{contacto_existente.entidad_externa.nombre}'."
                    )
                else:
                    messages.error(
                        request,
                        f"Error al actualizar el contacto."
                    )
            except Exception as e:
                 messages.error(request, f"Error al actualizar el contacto: {e}")
        else:
             messages.error(request, "Por favor corrija los errores en el formulario.")
    else:
        form = ContactoForm(instance=contacto, user=request.user)
    
    # Obtener historial de auditoría del contacto
    historial_auditoria = AuditoriaContacto.objects.filter(
        contacto=contacto
    ).select_related('usuario').order_by('-fecha_cambio')[:20]
    
    context = {
        'form': form,
        'contacto': contacto,
        'titulo_pagina': f'Editar Contacto: {contacto.nombre_completo}',
        'is_edit': True,
        'historial_auditoria': historial_auditoria,
    }
    return render(request, 'correspondencia/admin/contacto_form.html', context)

@login_required
@require_GET
def auditoria_contacto_ajax(request, pk):
    """Retorna el historial de auditoría de un contacto en formato JSON."""
    try:
        contacto = get_object_or_404(Contacto, pk=pk)
        registros = AuditoriaContacto.objects.filter(
            contacto=contacto
        ).select_related('usuario').order_by('-fecha_cambio')[:20]
        
        historial = []
        for reg in registros:
            historial.append({
                'tipo_cambio': reg.tipo_cambio,
                'tipo_cambio_display': reg.get_tipo_cambio_display(),
                'fecha': reg.fecha_cambio.strftime('%d/%m/%Y %H:%M'),
                'usuario': reg.usuario.get_full_name() or reg.usuario.username if reg.usuario else 'Sistema',
                'campos_modificados': reg.campos_modificados or {},
            })
        
        return JsonResponse({'success': True, 'historial': historial})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_GET
def verificar_contacto_activo_ajax(request, pk):
    """Verifica si un contacto tiene correspondencia activa."""
    try:
        contacto = get_object_or_404(Contacto, pk=pk)
        
        entrante_activa = Correspondencia.objects.filter(
            remitente=contacto
        ).exclude(
            estado__in=['RESPONDIDA']
        ).count()
        
        salida_pendiente = SalidaDestinatario.objects.filter(
            contacto=contacto,
            estado='PENDIENTE'
        ).count()
        
        total = entrante_activa + salida_pendiente
        detalle_parts = []
        if entrante_activa > 0:
            detalle_parts.append(f"{entrante_activa} entrante(s)")
        if salida_pendiente > 0:
            detalle_parts.append(f"{salida_pendiente} salida(s) pendiente(s)")
        
        return JsonResponse({
            'success': True,
            'total_activas': total,
            'entrante_activa': entrante_activa,
            'salida_pendiente': salida_pendiente,
            'detalle': ', '.join(detalle_parts) if detalle_parts else 'ninguna',
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def compartir_correspondencia(request, pk):
    """Comparte automáticamente una correspondencia con todos los usuarios de la misma oficina (excepto el actual)."""
    correspondencia = get_object_or_404(Correspondencia, pk=pk)
    usuario_actual = request.user
    perfil_usuario_actual = getattr(usuario_actual, 'perfil', None)

    # Verificación de permisos
    puede_compartir = (
        correspondencia.usuario_destino_inicial == usuario_actual or
        DistribucionInternaUsuario.objects.filter(correspondencia=correspondencia, usuario_asignado=usuario_actual).exists()
    )

    if not puede_compartir:
        messages.error(request, "No tienes permiso para compartir esta correspondencia.")
        return redirect('correspondencia:detalle_correspondencia', pk=pk)

    if not perfil_usuario_actual or not perfil_usuario_actual.oficina:
        messages.error(request, "No tienes una oficina asignada para compartir.")
        return redirect('correspondencia:detalle_correspondencia', pk=pk)

    oficina_usuario = perfil_usuario_actual.oficina

    if request.method == 'POST':
        form = CompartirCorrespondenciaForm(
            request.POST,
            oficina=oficina_usuario,
            usuario_actual=usuario_actual,
            correspondencia=correspondencia
        )
        if form.is_valid():
            observaciones = form.cleaned_data['observaciones']

            # Asegurar que el usuario actual tenga una distribución (marcada como leída)
            # Esto es importante para el conteo X/Y correcto
            if not DistribucionInternaUsuario.objects.filter(correspondencia=correspondencia, usuario_asignado=usuario_actual).exists():
                DistribucionInternaUsuario.objects.create(
                    correspondencia=correspondencia,
                    usuario_asignado=usuario_actual,
                    asignado_por=usuario_actual,
                    leido=True,  # Marcar como leída porque el usuario actual está compartiendo
                    observaciones="Auto-creada al compartir"
                )

            usuarios_destino = form.get_usuarios_destino()
            compartidos = 0

            with transaction.atomic():
                for usuario in usuarios_destino:
                    dist, creado = DistribucionInternaUsuario.objects.get_or_create(
                        correspondencia=correspondencia,
                        usuario_asignado=usuario,
                        defaults={
                            'asignado_por': usuario_actual,
                            'observaciones': observaciones
                        }
                    )
                    if creado:
                        HistorialCorrespondencia.objects.create(
                            correspondencia=correspondencia,
                            evento='REDISTRIBUIDA_INTERNA',
                            usuario=usuario_actual,
                            descripcion=f"Compartida internamente con {usuario.get_full_name() or usuario.username}. Obs: {observaciones[:100]}"
                        )
                        crear_notificacion_compartir_oficina(correspondencia, usuario, usuario_actual, observaciones)
                        compartidos += 1

            if compartidos:
                messages.success(
                    request,
                    f"Correspondencia compartida con {compartidos} usuario(s) de tu oficina."
                )
            else:
                messages.info(request, "Todos los usuarios de tu oficina ya tenían acceso a esta correspondencia.")

            return redirect('correspondencia:detalle_correspondencia', pk=pk)
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = CompartirCorrespondenciaForm(
            oficina=oficina_usuario,
            usuario_actual=usuario_actual,
            correspondencia=correspondencia
        )

    context = {
        'titulo_pagina': f'Compartir Radicado: {correspondencia.numero_radicado}',
        'correspondencia': correspondencia,
        'form': form,
        'usuarios_destino': form.get_usuarios_destino()
    }
    return render(request, 'correspondencia/usuario/compartir_form.html', context)

@login_required
def lider_compartir_con_oficina(request, pk):
    """
    Permite que un líder de oficina externa comparta una correspondencia compartida
    con toda su oficina (cambia solo_lider=False).
    """
    correspondencia = get_object_or_404(Correspondencia, pk=pk)
    usuario = request.user
    
    # Verificar si es líder
    if not usuario.groups.filter(name='Lider de Oficina').exists():
        messages.error(request, "Solo los líderes pueden realizar esta acción.")
        return redirect('correspondencia:detalle_correspondencia', pk=pk)
    
    # Verificar que tenga acceso interoficina
    if not hasattr(usuario, 'perfil') or not usuario.perfil.oficina:
        messages.error(request, "No tienes una oficina asignada.")
        return redirect('correspondencia:detalle_correspondencia', pk=pk)
    
    # Buscar el acceso de SU oficina
    acceso = AccesoCorrespondenciaOficina.objects.filter(
        correspondencia=correspondencia,
        oficina=usuario.perfil.oficina,
        solo_lider=True  # Solo puede hacerlo si actualmente está restringido
    ).first()
    
    if not acceso:
        messages.warning(request, "No se encontró un acceso restringido para tu oficina o ya está compartido con todos.")
        return redirect('correspondencia:detalle_correspondencia', pk=pk)
    
    # Cambiar el flag
    acceso.solo_lider = False
    acceso.save()
    
    # Registrar en historial
    HistorialCorrespondencia.objects.create(
        correspondencia=correspondencia,
        evento='COMPARTIDA_OFICINA',
        usuario=usuario,
        descripcion=f"Líder {usuario.get_full_name() or usuario.username} compartió esta correspondencia con toda la oficina {usuario.perfil.oficina.nombre}."
    )
    
    # Crear notificaciones para el resto de la oficina (excluyendo al líder)
    from django.contrib.auth.models import User
    usuarios_oficina = User.objects.filter(
        perfil__oficina=usuario.perfil.oficina,
        is_active=True
    ).exclude(pk=usuario.pk).exclude(groups__name='Lider de Oficina')  # Excluir líderes que ya lo vieron
    
    notificaciones_creadas = 0
    from django.urls import reverse
    url = reverse('correspondencia:detalle_correspondencia', kwargs={'pk': correspondencia.pk})
    
    for u in usuarios_oficina:
        try:
            Notificacion.objects.create(
                usuario=u,
                tipo='acceso_oficina',
                titulo="Correspondencia compartida por tu líder",
                mensaje=(
                    f"Tu líder {usuario.get_full_name() or usuario.username} compartió el radicado {correspondencia.numero_radicado} "
                    f"con toda la oficina. Asunto: {correspondencia.asunto[:100]}"
                ),
                correspondencia=correspondencia,
                url=url
            )
            notificaciones_creadas += 1
        except Exception:
            continue
    
    messages.success(
        request, 
        f"La correspondencia ahora es visible para toda tu oficina. Se notificó a {notificaciones_creadas} usuario(s)."
    )
    
    return redirect('correspondencia:detalle_correspondencia', pk=pk)

@login_required
def redistribuir_oficinas(request, pk):
    """Comparte la correspondencia con otras oficinas otorgando acceso de solo lectura."""
    correspondencia = get_object_or_404(Correspondencia, pk=pk)
    usuario_actual = request.user
    perfil_usuario_actual = getattr(usuario_actual, 'perfil', None)

    puede_compartir = (
        correspondencia.usuario_destino_inicial == usuario_actual or
        DistribucionInternaUsuario.objects.filter(correspondencia=correspondencia, usuario_asignado=usuario_actual).exists()
    )

    if not puede_compartir:
        messages.error(request, "No tienes permiso para redistribuir esta correspondencia a otras oficinas.")
        return redirect('correspondencia:detalle_correspondencia', pk=pk)

    oficina_origen = getattr(perfil_usuario_actual, 'oficina', None)

    if request.method == 'POST':
        form = CompartirOtrasOficinasForm(
            request.POST,
            correspondencia=correspondencia,
            oficina_origen=oficina_origen
        )
        if form.is_valid():
            oficinas_seleccionadas = form.cleaned_data['oficinas']
            observaciones = form.cleaned_data['observaciones']
            oficinas_agregadas = 0

            with transaction.atomic():
                for oficina in oficinas_seleccionadas:
                    # Verificar si se marcó explícitamente "Compartir con toda la oficina"
                    share_all_key = f'share_all_{oficina.id}'
                    compartir_con_todos = request.POST.get(share_all_key) == 'on'
                    es_solo_lider = not compartir_con_todos  # Si NO está marcado "compartir con todos", entonces solo líderes
                    
                    # Verificar si se marcó "Permitir responder"
                    puede_responder_key = f'puede_responder_{oficina.id}'
                    permiso_respuesta = request.POST.get(puede_responder_key) == 'on'
                    
                    acceso, creado = AccesoCorrespondenciaOficina.objects.get_or_create(
                        correspondencia=correspondencia,
                        oficina=oficina,
                        defaults={
                            'compartido_por': usuario_actual,
                            'observaciones': observaciones,
                            'solo_lider': es_solo_lider,
                            'puede_responder': permiso_respuesta
                        }
                    )
                    # Si ya existía, actualizar solo_lider y puede_responder si es necesario
                    if not creado:
                        acceso.solo_lider = es_solo_lider
                        acceso.puede_responder = permiso_respuesta
                        acceso.save(update_fields=['solo_lider', 'puede_responder'])
                        # Registrar actualización en historial
                        visibilidad_msg = " (Solo Líderes)" if es_solo_lider else " (Toda la oficina)"
                        respuesta_msg = " - Puede responder" if permiso_respuesta else ""
                        HistorialCorrespondencia.objects.create(
                            correspondencia=correspondencia,
                            evento='COMPARTIDA_OFICINA',
                            usuario=usuario_actual,
                            descripcion=f"Configuración actualizada para oficina {oficina.nombre}{visibilidad_msg}{respuesta_msg}. Obs: {observaciones[:100]}"
                        )
                    
                    if creado:
                        visibilidad_msg = " (Solo Líderes)" if es_solo_lider else " (Toda la oficina)"
                        respuesta_msg = " - Puede responder" if permiso_respuesta else ""
                        HistorialCorrespondencia.objects.create(
                            correspondencia=correspondencia,
                            evento='COMPARTIDA_OFICINA',
                            usuario=usuario_actual,
                            descripcion=f"Acceso de solo lectura otorgado a la oficina {oficina.nombre}{visibilidad_msg}{respuesta_msg}. Obs: {observaciones[:100]}"
                        )
                        crear_notificaciones_acceso_oficina(correspondencia, oficina, usuario_actual, observaciones)
                        oficinas_agregadas += 1

            if oficinas_agregadas:
                messages.success(
                    request,
                    f"Correspondencia compartida con {oficinas_agregadas} oficina(s) adicionales."
                )
            else:
                messages.info(request, "Las oficinas seleccionadas ya contaban con acceso a esta correspondencia.")

            return redirect('correspondencia:detalle_correspondencia', pk=pk)
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = CompartirOtrasOficinasForm(
            correspondencia=correspondencia,
            oficina_origen=oficina_origen
        )

    context = {
        'titulo_pagina': f'Redistribuir Radicado: {correspondencia.numero_radicado}',
        'correspondencia': correspondencia,
        'form': form,
        'accesos_existentes': correspondencia.accesos_oficinas.select_related('oficina', 'compartido_por').all()
    }
    return render(request, 'correspondencia/usuario/redistribuir_oficinas.html', context)

@login_required
@user_passes_test(lambda u: not u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:pendientes_distribuir')
def bandeja_personal(request):
    """Muestra la bandeja personal del usuario (asignada + compartida) con estado de lectura X/Y y plazo."""
    usuario_actual = request.user
    perfil_usuario = getattr(usuario_actual, 'perfil', None)
    correspondencias_list = [] 
    load_success = False

    # Parámetros de filtro
    search_term = request.GET.get('search_term', '').strip()
    fecha_radicacion_desde = request.GET.get('fecha_radicacion_desde', '').strip()
    fecha_radicacion_hasta = request.GET.get('fecha_radicacion_hasta', '').strip()
    remitente_term = request.GET.get('remitente', '').strip()
    entidad_term = request.GET.get('entidad', '').strip()
    estado_plazo = request.GET.get('estado_plazo', '').strip()  # ok|proximo|urgente|critico|vencido
    estado_lectura = request.GET.get('estado_lectura', '').strip()  # leidos|no_leidos
    fecha_limite_desde = request.GET.get('fecha_limite_desde', '').strip()
    fecha_limite_hasta = request.GET.get('fecha_limite_hasta', '').strip()

    if perfil_usuario and perfil_usuario.oficina:
        try:
            correspondencias_qs = _base_bandeja_personal_queryset(
                usuario_actual,
                perfil_usuario.oficina,
            )
            
            # Aplicar filtros de búsqueda de texto
            if search_term:
                correspondencias_qs = correspondencias_qs.filter(
                    Q(numero_radicado__icontains=search_term) |
                    Q(asunto__icontains=search_term) |
                    Q(remitente__nombres__icontains=search_term) |
                    Q(remitente__apellidos__icontains=search_term) |
                    Q(remitente__entidad_externa__nombre__icontains=search_term)
                )

            # Filtro adicional por remitente exacto/parcial
            if remitente_term:
                correspondencias_qs = correspondencias_qs.filter(
                    Q(remitente__nombres__icontains=remitente_term) |
                    Q(remitente__apellidos__icontains=remitente_term) |
                    Q(remitente__correo_electronico__icontains=remitente_term)
                )

            # Filtro adicional por entidad
            if entidad_term:
                correspondencias_qs = correspondencias_qs.filter(
                    remitente__entidad_externa__nombre__icontains=entidad_term
                )
            
            # Aplicar filtros de fecha de radicación
            if fecha_radicacion_desde:
                try:
                    correspondencias_qs = correspondencias_qs.filter(fecha_radicacion__date__gte=fecha_radicacion_desde)
                except Exception:
                    pass
            
            if fecha_radicacion_hasta:
                try:
                    correspondencias_qs = correspondencias_qs.filter(fecha_radicacion__date__lte=fecha_radicacion_hasta)
                except Exception:
                    pass
            
            # Filtro por fecha límite de respuesta (para el botón de "por vencer" - próximas 48 horas)
            if fecha_limite_desde and fecha_limite_hasta:
                try:
                    # Si ambas fechas son iguales, significa que queremos filtrar por las próximas 48 horas
                    # (igual que el KPI kpi_sla_por_vencer_personal)
                    ahora = timezone.now()
                    correspondencias_qs = correspondencias_qs.filter(
                        requiere_respuesta=True,
                        fecha_limite_respuesta_persist__gte=ahora,
                        fecha_limite_respuesta_persist__lte=ahora + timedelta(hours=48)
                    )
                except Exception:
                    pass
            
            # Filtro por estado de plazo usando días restantes (calcular en memoria pero de forma eficiente)
            if estado_plazo:
                mapped = 'ok' if estado_plazo == 'en_tramite' else estado_plazo
                ahora = timezone.now()
                hoy = ahora.date()
                
                # Para vencido, usar filtro directo en base de datos (más eficiente)
                if mapped == 'vencido':
                    correspondencias_qs = excluir_entrantes_con_respuesta(
                        correspondencias_qs.filter(
                            requiere_respuesta=True,
                            fecha_limite_respuesta_persist__isnull=False,
                            fecha_limite_respuesta_persist__lt=ahora,
                        )
                    )
                else:
                    # Para otros estados, calcular días restantes en memoria pero optimizado
                    try:
                        # Solo cargar los campos necesarios
                        candidatos = list(correspondencias_qs.select_related('subserie').only(
                            'id', 'requiere_respuesta', 'fecha_limite_respuesta_persist'
                        ))
                        
                        ids_match = []
                        for c in candidatos:
                            if not c.requiere_respuesta or not c.fecha_limite_respuesta_persist:
                                continue
                            
                            # Calcular días restantes de forma simple (igual que el modelo cuando está vencido)
                            fecha_limite_date = c.fecha_limite_respuesta_persist.date()
                            delta = fecha_limite_date - hoy
                            dias_restantes = delta.days
                            
                            # Si no está vencido, usar la lógica del modelo (días hábiles)
                            if dias_restantes >= 0:
                                # Usar la propiedad del modelo para días hábiles
                                try:
                                    dias_restantes = c.dias_restantes
                                    if dias_restantes is None:
                                        continue
                                except:
                                    continue
                            
                            # Determinar estado basado en días restantes
                            if mapped == 'critico' and 0 <= dias_restantes <= 1:
                                ids_match.append(c.id)
                            elif mapped == 'urgente' and 1 < dias_restantes <= 4:
                                ids_match.append(c.id)
                            elif mapped == 'proximo' and 4 < dias_restantes <= 10:
                                ids_match.append(c.id)
                            elif mapped == 'ok' and dias_restantes > 10:
                                ids_match.append(c.id)
                        
                        if ids_match:
                            correspondencias_qs = correspondencias_qs.filter(id__in=ids_match)
                        else:
                            correspondencias_qs = correspondencias_qs.none()
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Error filtrando por estado_plazo '{mapped}': {e}")
                        pass

            respuesta_valida_subquery = (
                CorrespondenciaSalida.objects
                .filter(
                    respuesta_a=OuterRef('pk'),
                    estado__in=['APROBADA', 'ENVIADA']
                )
                .annotate(
                    fecha_referencia=Coalesce('fecha_envio', 'fecha_aprobacion', 'fecha_creacion')
                )
                .order_by('-fecha_referencia', '-fecha_creacion')
            )

            # Anotar el queryset para obtener conteos de lectura (para mostrar X/Y)
            correspondencias_qs = correspondencias_qs.annotate(
                # Contar total de distribuciones (Y)
                total_destinatarios=Count('distribuciones_internas', distinct=True),
                # Contar distribuciones marcadas como leídas (X)
                total_leidos=Count('distribuciones_internas', filter=Q(distribuciones_internas__leido=True), distinct=True),
                respuesta_registrada_en=Subquery(
                    respuesta_valida_subquery.values('fecha_referencia')[:1]
                ),
            )

            # Filtro por estado de lectura basado en X/Y
            if estado_lectura:
                if estado_lectura == 'leidos':
                    correspondencias_qs = correspondencias_qs.filter(total_leidos=F('total_destinatarios'))
                elif estado_lectura == 'no_leidos':
                    correspondencias_qs = correspondencias_qs.filter(total_leidos__lt=F('total_destinatarios'))

            # Optimizar relaciones relacionadas después de anotar
            correspondencias_qs = correspondencias_qs.select_related(
                'remitente', 'oficina_destino', 'usuario_destino_inicial'
            ).prefetch_related(
                'adjuntos_correo'
            ).order_by('-fecha_radicacion')

            # Paginación para eficiencia
            paginator = Paginator(correspondencias_qs, 25)  # 25 registros por página
            page_number = request.GET.get('page', 1)
            try:
                correspondencias_page = paginator.page(page_number)
            except (EmptyPage, PageNotAnInteger):
                correspondencias_page = paginator.page(1)
            
            correspondencias_list = correspondencias_page.object_list
            _marcar_lectura_usuario_en_pagina(correspondencias_list, usuario_actual)
            for item in correspondencias_list:
                fecha_respuesta = getattr(item, 'respuesta_registrada_en', None)
                fecha_limite = item.fecha_limite_respuesta

                item.fue_respondida = fecha_respuesta is not None
                item.respondida_a_tiempo = bool(
                    item.fue_respondida and fecha_limite and fecha_respuesta <= fecha_limite
                )

                if not item.requiere_respuesta:
                    item.estado_respuesta_bandeja = 'no_requiere'
                elif not item.fue_respondida:
                    item.estado_respuesta_bandeja = 'pendiente'
                elif not fecha_limite:
                    item.estado_respuesta_bandeja = 'sin_plazo'
                elif item.respondida_a_tiempo:
                    item.estado_respuesta_bandeja = 'a_tiempo'
                else:
                    item.estado_respuesta_bandeja = 'fuera_tiempo'

            load_success = True
            
        except Exception as e:
            print(f"Error crítico cargando bandeja personal {usuario_actual.username}: {e}")
            messages.error(request, f"Ocurrió un error inesperado al cargar tu bandeja: {e}")
            load_success = False
    else:
        messages.warning(request, "No tienes un perfil o una oficina asignada para ver tu bandeja.")
        load_success = False

    context = {
        'page_title': 'Bandeja Personal',
        'correspondencias': correspondencias_list,
        'correspondencias_page': correspondencias_page if 'correspondencias_page' in locals() else None,
        'load_success': load_success,
        'mostrar_boton_compartir': True,
        'tipo_tabla': 'personal',
        # Valores de filtros para rehidratación en la UI
        'filtro_search_term': search_term,
        'filtro_fecha_radicacion_desde': fecha_radicacion_desde,
        'filtro_fecha_radicacion_hasta': fecha_radicacion_hasta,
        'filtro_remitente': remitente_term,
        'filtro_entidad': entidad_term,
        'filtro_estado_plazo': estado_plazo,
        'filtro_estado_lectura': estado_lectura,
    }
    return render(request, 'correspondencia/usuario/bandeja_personal.html', context)

@login_required
@require_GET
def lectura_detalle_ajax(request, correspondencia_id: int):
    """Devuelve en JSON los usuarios de la oficina y si han leído la correspondencia indicada."""
    usuario_actual = request.user
    perfil = getattr(usuario_actual, 'perfil', None)
    if not perfil or not perfil.oficina:
        return JsonResponse({'ok': False, 'error': 'Sin oficina asignada'}, status=400)

    try:
        correspondencia = Correspondencia.objects.get(pk=correspondencia_id, oficina_destino=perfil.oficina)
    except Correspondencia.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Correspondencia no encontrada'}, status=404)

    distros = (DistribucionInternaUsuario.objects
               .filter(correspondencia=correspondencia)
               .select_related('usuario_asignado')
               .order_by('usuario_asignado__first_name', 'usuario_asignado__last_name'))

    leidos = []
    no_leidos = []
    for d in distros:
        nombre = d.usuario_asignado.get_full_name() or d.usuario_asignado.username
        item = {
            'id': d.usuario_asignado_id, 
            'nombre': nombre, 
            'leido': d.leido,
            'fecha_lectura': d.fecha_lectura.isoformat() if d.fecha_lectura else None
        }
        (leidos if d.leido else no_leidos).append(item)

    return JsonResponse({
        'ok': True,
        'correspondencia': correspondencia.id,
        'leidos': leidos,
        'no_leidos': no_leidos,
        'totales': {
            'destinatarios': len(leidos) + len(no_leidos),
            'leidos': len(leidos),
        }
    })

@login_required
@require_GET
def lectura_detalle_interoficina_ajax(request, correspondencia_id: int):
    """Devuelve el estado de lectura por oficina para accesos interoficina."""
    usuario_actual = request.user
    perfil = getattr(usuario_actual, 'perfil', None)
    if not perfil or not perfil.oficina:
        return JsonResponse({'ok': False, 'error': 'Sin oficina asignada'}, status=400)

    try:
        correspondencia = Correspondencia.objects.get(pk=correspondencia_id)
    except Correspondencia.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Correspondencia no encontrada'}, status=404)

    accesos = (
        AccesoCorrespondenciaOficina.objects
        .filter(correspondencia=correspondencia)
        .select_related('oficina', 'compartido_por', 'compartido_por__perfil')
        .order_by('oficina__nombre')
    )

    oficinas_map: dict[int, dict] = {}
    for acceso in accesos:
        entry = oficinas_map.setdefault(acceso.oficina_id, {
            'nombre_oficina': acceso.oficina.nombre,
            'usuarios': []
        })

        # Estado general de la oficina como lector colectivo
        entry['usuarios'].append({
            'nombre': acceso.oficina.nombre,
            'rol': 'Oficina',
            'leido': acceso.leido,
            'fecha_lectura': acceso.fecha_lectura.isoformat() if acceso.fecha_lectura else None
        })

        # Usuario que compartió el acceso (referencia)
        if acceso.compartido_por:
            entry['usuarios'].append({
                'nombre': acceso.compartido_por.get_full_name() or acceso.compartido_por.username,
                'rol': 'Compartido por',
                'leido': acceso.leido,
                'fecha_lectura': acceso.fecha_lectura.isoformat() if acceso.fecha_lectura else None
            })

    return JsonResponse({
        'ok': True,
        'oficinas': list(oficinas_map.values())
    })

@login_required
@user_passes_test(lambda u: not u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:pendientes_distribuir') # Corregido previamente
def bandeja_oficina(request):
    """
    Muestra la correspondencia compartida dentro de la oficina y los accesos interoficina de solo lectura.

    IMPORTANTE: Esta bandeja NO muestra correspondencia personal (asignada inicialmente a un usuario específico).
    Incluye:
    - Correspondencia compartida internamente (2+ distribuciones internas)
    - Correspondencia de otras oficinas con acceso solo lectura (AccesoCorrespondenciaOficina)
    """
    usuario_actual = request.user
    perfil_usuario = getattr(usuario_actual, 'perfil', None)
    correspondencias_list = []
    load_success = False
    nombre_oficina = "Desconocida"

    # Inicializar variables de filtro ANTES del if para evitar UnboundLocalError
    search_term = request.GET.get('search_term', '').strip()
    fecha_radicacion_desde = request.GET.get('fecha_radicacion_desde', '').strip()
    fecha_radicacion_hasta = request.GET.get('fecha_radicacion_hasta', '').strip()
    remitente_term = request.GET.get('remitente', '').strip()
    entidad_term = request.GET.get('entidad', '').strip()
    estado_plazo = request.GET.get('estado_plazo', '').strip()
    estado_lectura = request.GET.get('estado_lectura', '').strip()

    if perfil_usuario and perfil_usuario.oficina:
        nombre_oficina = perfil_usuario.oficina.nombre
        try:
            # Parámetros de filtro (ya inicializados arriba)
            
            # Correspondencia compartida internamente (multiusuario) o con acceso interoficina (solo lectura)
            correspondencias_qs = _base_bandeja_oficina_queryset(perfil_usuario.oficina)
            
            if search_term:
                correspondencias_qs = correspondencias_qs.filter(
                    Q(asunto__icontains=search_term) |
                    Q(numero_radicado__icontains=search_term) |
                    Q(remitente__nombres__icontains=search_term) |
                    Q(remitente__apellidos__icontains=search_term) |
                    Q(remitente__entidad_externa__nombre__icontains=search_term)
                )

            # Filtro adicional por remitente exacto/parcial
            if remitente_term:
                correspondencias_qs = correspondencias_qs.filter(
                    Q(remitente__nombres__icontains=remitente_term) |
                    Q(remitente__apellidos__icontains=remitente_term) |
                    Q(remitente__correo_electronico__icontains=remitente_term)
                )

            # Filtro adicional por entidad
            if entidad_term:
                correspondencias_qs = correspondencias_qs.filter(
                    remitente__entidad_externa__nombre__icontains=entidad_term
                )

            # Filtros de fecha de radicación
            if fecha_radicacion_desde:
                try:
                    correspondencias_qs = correspondencias_qs.filter(fecha_radicacion__date__gte=fecha_radicacion_desde)
                except Exception:
                    pass
            if fecha_radicacion_hasta:
                try:
                    correspondencias_qs = correspondencias_qs.filter(fecha_radicacion__date__lte=fecha_radicacion_hasta)
                except Exception:
                    pass
            
            # Subquery para obtener el primer usuario que compartió (excluye la distribución inicial)
            primer_compartidor_qs = DistribucionInternaUsuario.objects.filter(
                correspondencia=OuterRef('pk'),
                asignado_por__isnull=False
            ).exclude(
                usuario_asignado=F('correspondencia__usuario_destino_inicial')
            ).order_by('fecha_asignacion')

            respuesta_valida_subquery = _build_respuesta_valida_subquery('pk')

            # Anotar el queryset para obtener conteos de lectura y nombre del compartidor
            correspondencias_qs = correspondencias_qs.annotate(
                compartido_por_username=Subquery(primer_compartidor_qs.values('asignado_por__username')[:1]),
                compartido_por_first_name=Subquery(primer_compartidor_qs.values('asignado_por__first_name')[:1]),
                compartido_por_last_name=Subquery(primer_compartidor_qs.values('asignado_por__last_name')[:1]),
                respuesta_registrada_en=Subquery(
                    respuesta_valida_subquery.values('fecha_referencia')[:1]
                )
            )

            # Filtro por estado de lectura
            if estado_lectura:
                if estado_lectura == 'leidos':
                    correspondencias_qs = correspondencias_qs.filter(
                        total_destinatarios__gt=0,
                        total_leidos=F('total_destinatarios')
                    )
                elif estado_lectura == 'no_leidos':
                    correspondencias_qs = correspondencias_qs.filter(
                        total_destinatarios__gt=0,
                        total_leidos__lt=F('total_destinatarios')
                    )

            # Filtro por estado de plazo usando propiedad del modelo
            if estado_plazo:
                mapped = 'ok' if estado_plazo == 'en_tramite' else estado_plazo
                try:
                    candidatos = list(correspondencias_qs.select_related('subserie'))
                    ids_match = [c.id for c in candidatos if getattr(c, 'estado_plazo', None) == mapped]
                    correspondencias_qs = correspondencias_qs.filter(id__in=ids_match)
                except Exception:
                    pass
            
            # Optimizar relaciones relacionadas después de anotar
            correspondencias_qs = correspondencias_qs.select_related(
                'remitente', 'oficina_destino', 'usuario_destino_inicial'
            ).prefetch_related(
                'adjuntos_correo',
                'accesos_oficinas__oficina'
            ).order_by('-fecha_radicacion')
            
            # Paginación para eficiencia
            paginator = Paginator(correspondencias_qs, 25)
            page_number = request.GET.get('page', 1)
            try:
                correspondencias_page = paginator.page(page_number)
            except (EmptyPage, PageNotAnInteger):
                correspondencias_page = paginator.page(1)

            correspondencias_list = correspondencias_page.object_list
            for item in correspondencias_list:
                _asignar_estado_respuesta_bandeja(
                    item,
                    getattr(item, 'respuesta_registrada_en', None)
                )
                
            load_success = True
        except Exception as e:
            print(f"Error crítico cargando bandeja oficina {nombre_oficina}: {e}")
            messages.error(request, f"Ocurrió un error inesperado al cargar la bandeja de la oficina: {e}")
            load_success = False
    else:
        messages.warning(request, "No tienes un perfil o una oficina asignada para ver la bandeja de oficina.")
        load_success = False

    context = {
        'page_title': f'Bandeja Oficina y Accesos: {nombre_oficina}',
        'correspondencias': correspondencias_list, # Ya incluye los counts
        'correspondencias_page': correspondencias_page if 'correspondencias_page' in locals() else None,
        'load_success': load_success,
        'mostrar_boton_compartir': False, 
        'tipo_tabla': 'oficina',
        # Valores de filtros para rehidratación en la UI
        'filtro_search_term': search_term,
        'filtro_fecha_radicacion_desde': fecha_radicacion_desde,
        'filtro_fecha_radicacion_hasta': fecha_radicacion_hasta,
        'filtro_remitente': remitente_term,
        'filtro_entidad': entidad_term,
        'filtro_estado_plazo': estado_plazo,
        'filtro_estado_lectura': estado_lectura,
    }
    return render(request, 'correspondencia/usuario/bandeja_oficina.html', context)

@login_required
@user_passes_test(lambda u: not u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:pendientes_distribuir')
def bandeja_interoficina(request):
    """
    Muestra las correspondencias compartidas desde otras oficinas hacia la oficina
    actual mediante accesos de solo lectura.
    """
    usuario_actual = request.user
    perfil_usuario = getattr(usuario_actual, 'perfil', None)
    nombre_oficina = "Desconocida"
    load_success = False
    accesos_list = []
    accesos_page = None

    search_term = request.GET.get('search_term', '').strip()
    remitente_term = request.GET.get('remitente', '').strip()
    entidad_term = request.GET.get('entidad', '').strip()
    estado_lectura = request.GET.get('estado_lectura', '').strip()
    estado_plazo = request.GET.get('estado_plazo', '').strip()
    fecha_compartido_desde = request.GET.get('fecha_compartido_desde', '').strip()
    fecha_compartido_hasta = request.GET.get('fecha_compartido_hasta', '').strip()

    total_accesos = 0
    total_pendientes = 0

    if perfil_usuario and perfil_usuario.oficina:
        nombre_oficina = perfil_usuario.oficina.nombre
        try:
            accesos_qs = AccesoCorrespondenciaOficina.objects.filter(
                oficina=perfil_usuario.oficina
            ).select_related(
                'correspondencia',
                'correspondencia__oficina_destino',
                'correspondencia__remitente',
                'correspondencia__usuario_destino_inicial',
                'compartido_por',
                'compartido_por__perfil',
                'oficina'
            ).order_by('-fecha_compartido')

            # --- FILTRO DE LÍDERES ---
            # Si el usuario NO es del grupo 'Lider de Oficina', solo ve lo compartido con toda la oficina (solo_lider=False)
            if not usuario_actual.groups.filter(name='Lider de Oficina').exists():
                accesos_qs = accesos_qs.filter(solo_lider=False)
            # -------------------------

            if search_term:
                accesos_qs = accesos_qs.filter(
                    Q(correspondencia__numero_radicado__icontains=search_term) |
                    Q(correspondencia__asunto__icontains=search_term) |
                    Q(correspondencia__remitente__nombres__icontains=search_term) |
                    Q(correspondencia__remitente__apellidos__icontains=search_term) |
                    Q(correspondencia__remitente__entidad_externa__nombre__icontains=search_term)
                )

            if remitente_term:
                accesos_qs = accesos_qs.filter(
                    Q(correspondencia__remitente__nombres__icontains=remitente_term) |
                    Q(correspondencia__remitente__apellidos__icontains=remitente_term) |
                    Q(correspondencia__remitente__correo_electronico__icontains=remitente_term)
                )

            if entidad_term:
                accesos_qs = accesos_qs.filter(
                    Q(correspondencia__remitente__entidad_externa__nombre__icontains=entidad_term)
                )

            if fecha_compartido_desde:
                try:
                    accesos_qs = accesos_qs.filter(fecha_compartido__date__gte=fecha_compartido_desde)
                except Exception:
                    pass

            if fecha_compartido_hasta:
                try:
                    accesos_qs = accesos_qs.filter(fecha_compartido__date__lte=fecha_compartido_hasta)
                except Exception:
                    pass

            if estado_lectura == 'leidos':
                accesos_qs = accesos_qs.filter(leido=True)
            elif estado_lectura == 'no_leidos':
                accesos_qs = accesos_qs.filter(leido=False)

            if estado_plazo:
                mapped = 'ok' if estado_plazo == 'en_tramite' else estado_plazo
                try:
                    candidatos = list(accesos_qs.select_related('correspondencia__subserie'))
                    ids_match = [a.id for a in candidatos if getattr(a.correspondencia, 'estado_plazo', None) == mapped]
                    accesos_qs = accesos_qs.filter(id__in=ids_match)
                except Exception:
                    pass

            respuesta_valida_subquery = _build_respuesta_valida_subquery('correspondencia_id')
            accesos_qs = accesos_qs.annotate(
                respuesta_registrada_en=Subquery(
                    respuesta_valida_subquery.values('fecha_referencia')[:1]
                )
            )

            total_accesos = accesos_qs.count()
            total_pendientes = accesos_qs.filter(leido=False).count()

            paginator = Paginator(accesos_qs, 25)
            page_number = request.GET.get('page', 1)
            try:
                accesos_page = paginator.page(page_number)
            except (EmptyPage, PageNotAnInteger):
                accesos_page = paginator.page(1)

            accesos_list = accesos_page.object_list
            for acceso in accesos_list:
                _asignar_estado_respuesta_bandeja(
                    acceso.correspondencia,
                    getattr(acceso, 'respuesta_registrada_en', None)
                )
            load_success = True
        except Exception as e:
            print(f"Error cargando bandeja interoficina {nombre_oficina}: {e}")
            messages.error(request, f"Ocurrió un error al cargar la bandeja interoficina: {e}")
    else:
        messages.warning(request, "No tienes un perfil o una oficina asignada para ver la bandeja interoficina.")

    context = {
        'page_title': f'Bandeja Interoficina: {nombre_oficina}',
        'accesos': accesos_list,
        'accesos_page': accesos_page,
        'load_success': load_success,
        'total_accesos': total_accesos,
        'total_pendientes': total_pendientes,
        'filtro_search_term': search_term,
        'filtro_remitente': remitente_term,
        'filtro_entidad': entidad_term,
        'filtro_estado_plazo': estado_plazo,
        'filtro_estado_lectura': estado_lectura,
        'filtro_fecha_compartido_desde': fecha_compartido_desde,
        'filtro_fecha_compartido_hasta': fecha_compartido_hasta,
    }
    return render(request, 'correspondencia/usuario/bandeja_interoficina.html', context)

# === VISTAS PARA VENTANILLA - RADICACIÓN MANUAL DE CORREOS ===

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def marcar_correo_papelera_view(request, correo_id):
    """Marca un correo entrante como enviado a papelera (excluir del flujo sin borrar). POST: motivo_papelera."""
    correo = get_object_or_404(CorreoEntrante, pk=correo_id)
    if request.method != 'POST':
        return redirect('correspondencia:detalle_correo_entrante', correo_id=correo.id)
    motivo = (request.POST.get('motivo_papelera') or '').strip()
    if motivo not in dict(MOTIVO_PAPELERA_CHOICES):
        messages.error(request, 'Debe seleccionar un motivo válido para enviar a papelera.')
        return redirect('correspondencia:detalle_correo_entrante', correo_id=correo.id)
    correo.en_papelera = True
    correo.motivo_papelera = motivo
    correo.fecha_papelera = timezone.now()
    correo.usuario_papelera = request.user
    correo.save(update_fields=['en_papelera', 'motivo_papelera', 'fecha_papelera', 'usuario_papelera'])
    messages.success(request, 'Correo enviado a papelera. Ya no aparecerá en la bandeja activa.')
    return redirect('correspondencia:bandeja_correos_pendientes')


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def restaurar_correo_papelera_view(request, correo_id):
    """Restaura un correo de la papelera a la bandeja activa."""
    correo = get_object_or_404(CorreoEntrante, pk=correo_id)
    if request.method != 'POST':
        return redirect('correspondencia:detalle_correo_entrante', correo_id=correo.id)
    correo.en_papelera = False
    correo.motivo_papelera = ''
    correo.fecha_papelera = None
    correo.usuario_papelera = None
    correo.save(update_fields=['en_papelera', 'motivo_papelera', 'fecha_papelera', 'usuario_papelera'])
    messages.success(request, 'Correo restaurado a la bandeja de correos entrantes.')
    return redirect('correspondencia:bandeja_correos_pendientes')


# === DASHBOARD VENTANILLA (MVP) ===

def _extract_int_pks_from_post(post, keys):
    """PKs enviados en POST (select simple o múltiple)."""
    pks = set()
    for key in keys:
        for raw in post.getlist(key):
            try:
                pks.add(int(raw))
            except (TypeError, ValueError):
                continue
    return pks


def _configure_dashboard_ventanilla_form_querysets(*forms, post=None, user=None):
    """
    Evita renderizar cientos de <option> en el HTML inicial (~196 KB).
    En GET: querysets vacíos; Select2 carga vía AJAX al abrir el modal.
    En POST: solo las opciones enviadas (invocar ANTES de is_valid()).
    """
    contacto_keys = (
        'radicar-remitente',
        'rapida_ent-remitente',
        'rapida_sal-destinatario_contacto',
    )
    oficina_keys = (
        'radicar-oficina_destino',
        'radicar-oficina_selector',
        'rapida_ent-oficina_destino',
        'rapida_sal-oficina_emisora',
    )
    entidad_keys = ('radicar-entidad_selector',)
    otras_oficinas_keys = ('radicar-otras_oficinas',)
    usuario_rapido_keys = ('radicar-usuario_destino_rapido',)

    contacto_pks = _extract_int_pks_from_post(post, contacto_keys) if post else set()
    oficina_pks = _extract_int_pks_from_post(post, oficina_keys) if post else set()
    usuario_rapido_pks = _extract_int_pks_from_post(post, usuario_rapido_keys) if post else set()
    entidad_pks = _extract_int_pks_from_post(post, entidad_keys) if post else set()
    otras_oficinas_pks = _extract_int_pks_from_post(post, otras_oficinas_keys) if post else set()
    if not post and user is not None:
        perfil = getattr(user, 'perfil', None)
        if perfil and perfil.oficina_id:
            oficina_pks.add(perfil.oficina_id)

    if contacto_pks:
        qs_contacto = Contacto.objects.filter(pk__in=contacto_pks).select_related(
            'entidad_externa',
        ).only(
            'id', 'nombres', 'apellidos', 'correo_electronico',
            'entidad_externa_id', 'entidad_externa__nombre',
        )
    else:
        qs_contacto = Contacto.objects.none()

    if oficina_pks:
        qs_oficina = OficinaProductora.objects.filter(pk__in=oficina_pks).only('id', 'nombre')
    else:
        qs_oficina = OficinaProductora.objects.none()

    if entidad_pks:
        qs_entidad = EntidadExterna.objects.filter(pk__in=entidad_pks).only('id', 'nombre')
    else:
        qs_entidad = EntidadExterna.objects.none()

    if otras_oficinas_pks:
        qs_otras_oficinas = OficinaProductora.objects.filter(
            pk__in=otras_oficinas_pks,
        ).only('id', 'nombre')
    else:
        qs_otras_oficinas = OficinaProductora.objects.none()

    for form in forms:
        if form is None:
            continue
        for field_name in ('remitente', 'destinatario_contacto'):
            if field_name in form.fields:
                form.fields[field_name].queryset = qs_contacto
        for field_name in ('oficina_destino', 'oficina_emisora', 'oficina_selector'):
            if field_name in form.fields:
                form.fields[field_name].queryset = qs_oficina
        if 'entidad_selector' in form.fields:
            form.fields['entidad_selector'].queryset = qs_entidad
        if 'otras_oficinas' in form.fields:
            form.fields['otras_oficinas'].queryset = qs_otras_oficinas
        if 'usuario_destino_rapido' in form.fields and usuario_rapido_pks:
            qs_usuario = User.objects.filter(pk__in=usuario_rapido_pks, is_active=True).select_related('perfil')
            actual = form.fields['usuario_destino_rapido'].queryset
            if actual.exists():
                qs_usuario = qs_usuario | actual
            form.fields['usuario_destino_rapido'].queryset = qs_usuario.distinct()


def _dashboard_radicacion_form(*args, **kwargs):
    kwargs.setdefault('defer_select_options', True)
    return ManualRadicacionCorreoForm(*args, **kwargs)


def _dashboard_rapida_ent_form(*args, **kwargs):
    kwargs.setdefault('defer_select_options', True)
    return RadicacionRapidaEntranteForm(*args, **kwargs)


def _dashboard_rapida_sal_form(user, *args, **kwargs):
    kwargs.setdefault('defer_select_options', True)
    return RadicacionRapidaSalienteForm(*args, user=user, **kwargs)


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def dashboard_ventanilla(request):
    """Dashboard Ventanilla: KPIs cacheados y modales de radicación (sin IMAP en GET)."""
    from datetime import timedelta

    # --- Procesamiento de formularios enviados desde modales ---
    open_contacto_modal = False
    open_entidad_modal = False
    open_radicacion_modal = False
    open_rapida_entrante_modal = False
    open_rapida_saliente_modal = False

    if request.method == 'POST':
        form_prefix = request.POST.get('form_prefix', '')
        if form_prefix == 'contacto':
            contacto_form = ContactoForm(request.POST)
            entidad_form = EntidadExternaForm()
            radicacion_form = _dashboard_radicacion_form(prefix='radicar')
            rapida_entrante_form = _dashboard_rapida_ent_form(prefix='rapida_ent')
            rapida_saliente_form = _dashboard_rapida_sal_form(request.user, prefix='rapida_sal')
            if contacto_form.is_valid():
                contacto_form.save()
                messages.success(request, 'Contacto creado exitosamente.')
                return redirect('correspondencia:dashboard_ventanilla')
            else:
                open_contacto_modal = True
        elif form_prefix == 'entidad':
            entidad_form = EntidadExternaForm(request.POST)
            contacto_form = ContactoForm()
            radicacion_form = _dashboard_radicacion_form(prefix='radicar')
            rapida_entrante_form = _dashboard_rapida_ent_form(prefix='rapida_ent')
            rapida_saliente_form = _dashboard_rapida_sal_form(request.user, prefix='rapida_sal')
            if entidad_form.is_valid():
                entidad_form.save()
                messages.success(request, 'Entidad creada exitosamente.')
                return redirect('correspondencia:dashboard_ventanilla')
            else:
                open_entidad_modal = True
        elif form_prefix == 'radicar':
            contacto_form = ContactoForm()
            entidad_form = EntidadExternaForm()
            radicacion_form = _dashboard_radicacion_form(request.POST, request.FILES, prefix='radicar')
            rapida_entrante_form = _dashboard_rapida_ent_form(prefix='rapida_ent')
            rapida_saliente_form = _dashboard_rapida_sal_form(request.user, prefix='rapida_sal')
            _configure_dashboard_ventanilla_form_querysets(
                radicacion_form, post=request.POST, user=request.user,
            )
            if radicacion_form.is_valid():
                try:
                    with transaction.atomic():
                        correspondencia = radicacion_form.save(commit=False)
                        correspondencia.tipo_radicado = 'ENTRANTE'
                        correspondencia.usuario_radicador = request.user
                        correspondencia.save()
                        adjuntos_guardados = _guardar_adjuntos_radicacion_fisica(correspondencia, request)

                        HistorialCorrespondencia.objects.create(
                            correspondencia=correspondencia,
                            evento='RADICADA',
                            usuario=request.user,
                            descripcion=(
                                f"Radicada manualmente por {request.user.username} desde dashboard de ventanilla"
                                + (f" con {adjuntos_guardados} adjuntos" if adjuntos_guardados > 0 else "")
                            )
                        )

                        suffix = _aplicar_distribucion_rapida_desde_form(
                            correspondencia,
                            request,
                            radicacion_form.cleaned_data,
                        )

                    messages.success(request, f'Correspondencia {correspondencia.numero_radicado} radicada exitosamente{suffix}.')
                    return redirect('correspondencia:dashboard_ventanilla')
                except Exception as e:
                    open_radicacion_modal = True
                    messages.error(request, f'Error al radicar la correspondencia: {e}')
            else:
                open_radicacion_modal = True
        elif form_prefix == 'rapida_entrante':
            # Radicación rápida ENTRANTE
            contacto_form = ContactoForm()
            entidad_form = EntidadExternaForm()
            radicacion_form = _dashboard_radicacion_form(prefix='radicar')
            rapida_entrante_form = _dashboard_rapida_ent_form(request.POST, request.FILES, prefix='rapida_ent')
            rapida_saliente_form = _dashboard_rapida_sal_form(request.user, prefix='rapida_sal')
            _configure_dashboard_ventanilla_form_querysets(
                rapida_entrante_form, post=request.POST, user=request.user,
            )
            if rapida_entrante_form.is_valid():
                correspondencia = rapida_entrante_form.save(commit=False)
                correspondencia.tipo_radicado = 'ENTRANTE'
                correspondencia.origen_radicacion = 'RAPIDA'  # Marcar como radicación rápida
                correspondencia.usuario_radicador = request.user
                
                # Calcular fecha límite de respuesta automáticamente según tipo de trámite
                tipo_tramite_codigo = correspondencia.tipo_tramite
                if tipo_tramite_codigo:
                    try:
                        tipo_tramite_obj = TipoTramite.objects.get(codigo=tipo_tramite_codigo, activo=True)
                        if tipo_tramite_obj.dias_respuesta is not None:
                            # Usar fecha de recepción del documento si está disponible, sino fecha de radicación
                            fecha_base = correspondencia.fecha_recepcion_documento or correspondencia.fecha_radicacion or timezone.now()
                            # Calcular fecha límite desde la fecha base
                            fecha_limite = calcular_dias_habiles(fecha_base, tipo_tramite_obj.dias_respuesta)
                            correspondencia.fecha_limite_respuesta_manual = fecha_limite
                    except TipoTramite.DoesNotExist:
                        # Si no existe el tipo, continuar sin calcular fecha límite
                        pass
                
                correspondencia.save()
                
                # Procesar archivos adjuntos si existen
                archivos_subidos = request.FILES.getlist('rapida_ent-adjuntos_archivos')
                if archivos_subidos:
                    from correspondencia.models import AdjuntoCorrespondenciaRapida
                    import mimetypes
                    for archivo in archivos_subidos:
                        # Detectar tipo MIME
                        tipo_mime = mimetypes.guess_type(archivo.name)[0] or 'application/octet-stream'
                        # Crear adjunto
                        AdjuntoCorrespondenciaRapida.objects.create(
                            correspondencia=correspondencia,
                            archivo=archivo,
                            nombre_original=archivo.name,
                            tipo_mime=tipo_mime
                        )
                
                # === ENVÍO DE CORREO AL FUNCIONARIO RESPONSABLE (radicación rápida entrante) ===
                email_funcionario = correspondencia.email_funcionario_responsable
                if email_funcionario:
                    try:
                        from correspondencia.utils.radicacion_rapida_email import (
                            adjuntos_desde_queryset,
                            enviar_notificacion_radicacion_rapida_entrante,
                        )

                        contexto_email = {
                            'nombre_funcionario': correspondencia.funcionario_responsable_tramite or 'Funcionario',
                            'numero_radicado': correspondencia.numero_radicado,
                            'fecha_radicacion': timezone.localtime(
                                correspondencia.fecha_radicacion or timezone.now()
                            ).strftime('%d/%m/%Y %H:%M'),
                            'remitente': (
                                correspondencia.entidad_persona_remitente
                                or 'No especificado'
                            ),
                            'direccion_correo_remitente': correspondencia.direccion_correo_remitente or '',
                            'oficina_destino': str(correspondencia.oficina_destino) if correspondencia.oficina_destino else 'No especificada',
                            'medio_recepcion': correspondencia.get_medio_recepcion_display() if correspondencia.medio_recepcion else '',
                            'tipo_tramite': correspondencia.tipo_tramite or '',
                            'fecha_limite': correspondencia.fecha_limite_respuesta_manual.strftime('%d/%m/%Y') if correspondencia.fecha_limite_respuesta_manual else '',
                            'asunto': correspondencia.asunto,
                            'cuerpo_correo': '',  # No hay correo original en radicación física
                            'usuario_radicador': request.user.get_full_name() or request.user.username,
                        }

                        enviar_notificacion_radicacion_rapida_entrante(
                            email_funcionario=email_funcionario,
                            contexto_email=contexto_email,
                            asunto=f"Correspondencia asignada - {correspondencia.numero_radicado}",
                            adjuntos=adjuntos_desde_queryset(correspondencia.adjuntos_rapidos.all()),
                        )

                        HistorialCorrespondencia.objects.create(
                            correspondencia=correspondencia,
                            evento='NOTIFICACION',
                            usuario=request.user,
                            descripcion=(
                                f"Notificación enviada a {correspondencia.funcionario_responsable_tramite} "
                                f"({email_funcionario}) — radicación rápida entrante (física)"
                            )
                        )
                        messages.info(request, f"📧 Notificación enviada a {email_funcionario}")

                    except Exception as e_email:
                        HistorialCorrespondencia.objects.create(
                            correspondencia=correspondencia,
                            evento='ERROR',
                            usuario=request.user,
                            descripcion=f"Error al enviar notificación a {email_funcionario}: {e_email}"
                        )
                        messages.warning(
                            request,
                            f"⚠️ Radicación exitosa, pero falló el envío de notificación a {email_funcionario}: {e_email}"
                        )

                messages.success(request, f'Entrante {correspondencia.numero_radicado} radicada exitosamente (rápida).')
                return redirect('correspondencia:dashboard_ventanilla')
            else:
                open_rapida_entrante_modal = True
        elif form_prefix == 'rapida_saliente':
            # Radicación rápida SALIENTE
            contacto_form = ContactoForm()
            entidad_form = EntidadExternaForm()
            radicacion_form = _dashboard_radicacion_form(prefix='radicar')
            rapida_entrante_form = _dashboard_rapida_ent_form(prefix='rapida_ent')
            rapida_saliente_form = _dashboard_rapida_sal_form(request.user, request.POST, prefix='rapida_sal')
            _configure_dashboard_ventanilla_form_querysets(
                rapida_saliente_form, post=request.POST, user=request.user,
            )
            if rapida_saliente_form.is_valid():
                salida = rapida_saliente_form.save(commit=False)
                salida.usuario_redactor = request.user
                # Asignar oficina emisora desde el formulario
                salida.oficina_emisora = rapida_saliente_form.cleaned_data.get('oficina_emisora')
                # Asignar destinatario si se seleccionó contacto
                destinatario_contacto = rapida_saliente_form.cleaned_data.get('destinatario_contacto')
                if destinatario_contacto:
                    salida.destinatario_contacto = destinatario_contacto
                    salida.destinatario_email = destinatario_contacto.correo_electronico or ''
                else:
                    # Crear un snapshot del destinatario texto en el cuerpo si no hay contacto
                    destinatario_texto = rapida_saliente_form.cleaned_data.get('destinatario_texto', '')
                    if destinatario_texto:
                        salida.cuerpo = f"[Destinatario: {destinatario_texto}]\n\n{salida.cuerpo}"
                # Estado: radicada directamente como ENVIADA (registro de constancia)
                salida.estado = 'ENVIADA'
                salida.fecha_envio = timezone.now()
                salida.save()
                messages.success(request, f'Saliente {salida.numero_radicado_salida} radicada exitosamente (rápida).')
                return redirect('correspondencia:dashboard_ventanilla')
            else:
                open_rapida_saliente_modal = True
        else:
            contacto_form = ContactoForm()
            entidad_form = EntidadExternaForm()
            radicacion_form = _dashboard_radicacion_form(prefix='radicar')
            rapida_entrante_form = _dashboard_rapida_ent_form(prefix='rapida_ent')
            rapida_saliente_form = _dashboard_rapida_sal_form(request.user, prefix='rapida_sal')
    else:
        contacto_form = ContactoForm()
        entidad_form = EntidadExternaForm()
        radicacion_form = _dashboard_radicacion_form(prefix='radicar')
        rapida_entrante_form = _dashboard_rapida_ent_form(prefix='rapida_ent')
        rapida_saliente_form = _dashboard_rapida_sal_form(request.user, prefix='rapida_sal')

    post_data = request.POST if request.method == 'POST' else None
    _configure_dashboard_ventanilla_form_querysets(
        radicacion_form,
        rapida_entrante_form,
        rapida_saliente_form,
        post=post_data,
        user=request.user,
    )

    # --- KPIs (cache Redis 90s; SLA vencido/por_vencer en un solo aggregate) ---
    kpis = get_ventanilla_dashboard_kpis()

    context = {
        'titulo_pagina': 'Dashboard Ventanilla',
        'defer_select_options': True,
        **kpis,
        # Formularios para modales
        'form_contacto': contacto_form,
        'form_entidad': entidad_form,
        'form_radicacion': radicacion_form,
        # Formularios de radicación rápida
        'form_rapida_entrante': rapida_entrante_form,
        'form_rapida_saliente': rapida_saliente_form,
        # Flags para reabrir modales tras POST inválido
        'open_contacto_modal': open_contacto_modal,
        'open_radicacion_modal': open_radicacion_modal,
        'open_entidad_modal': open_entidad_modal,
        'open_rapida_entrante_modal': open_rapida_entrante_modal,
        'open_rapida_saliente_modal': open_rapida_saliente_modal,
    }

    return render(request, 'correspondencia/admin/dashboard_ventanilla.html', context)

# ------------- ENDPOINT PARA CONTACTOS EN FORMULARIO DE RADICACIÓN MANUAL -------------

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def listar_entidades(request):
    """Muestra una lista paginada de todas las Entidades Externas."""
    
    # Manejar POST del modal de crear entidad
    if request.method == 'POST':
        form_prefix = request.POST.get('form_prefix')
        if form_prefix == 'entidad':
            form = EntidadExternaForm(request.POST)
            if form.is_valid():
                try:
                    entidad = form.save()
                    messages.success(request, f"Entidad '{entidad.nombre}' creada exitosamente.")
                    return redirect('correspondencia:listar_entidades')
                except Exception as e:
                    messages.error(request, f"Error al crear la entidad: {e}")
            else:
                messages.error(request, "Por favor corrija los errores en el formulario.")
    
    entidades_list = EntidadExterna.objects.all()
    
    # Búsqueda por nombre, NIT, teléfono o dirección
    query = request.GET.get('q')
    if query:
        entidades_list = entidades_list.filter(
            Q(nombre__icontains=query) |
            Q(nit__icontains=query) |
            Q(telefono__icontains=query) |
            Q(direccion__icontains=query)
        )
    
    paginator = Paginator(entidades_list, 25) # 25 entidades por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'entidades': page_obj,
        'titulo_pagina': 'Gestionar Entidades Externas',
        'search_query': query or "" # Pasar query a la plantilla para mostrarlo en el input
    }
    return render(request, 'correspondencia/admin/lista_entidades.html', context)

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def crear_entidad(request):
    """Crea una nueva Entidad Externa."""
    if request.method == 'POST':
        form = EntidadExternaForm(request.POST)
        if form.is_valid():
            try:
                entidad = form.save()
                messages.success(request, f"Entidad '{entidad.nombre}' creada exitosamente.")
                return redirect('correspondencia:listar_entidades') # Redirigir a la lista
            except Exception as e:
                 messages.error(request, f"Error al crear la entidad: {e}")
        else:
             messages.error(request, "Por favor corrija los errores en el formulario.")
    else:
        form = EntidadExternaForm()
    
    context = {
        'form': form,
        'titulo_pagina': 'Crear Nueva Entidad Externa'
    }
    return render(request, 'correspondencia/admin/entidad_form.html', context)

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def editar_entidad(request, pk):
    """Edita una entidad externa existente. Solo accesible por Ventanilla."""
    entidad = get_object_or_404(EntidadExterna, pk=pk)
    
    if request.method == 'POST':
        form = EntidadExternaForm(request.POST, instance=entidad)
        if form.is_valid():
            try:
                entidad_editada = form.save()
                messages.success(
                    request,
                    f"Entidad '{entidad_editada.nombre}' actualizada exitosamente."
                )
                return redirect('correspondencia:listar_entidades')
            except IntegrityError:
                nombre = form.cleaned_data.get('nombre', '').strip()
                messages.error(
                    request,
                    f"Ya existe una entidad con el nombre '{nombre}' o el dominio indicado."
                )
            except Exception as e:
                messages.error(request, f"Error al actualizar la entidad: {e}")
        else:
            messages.error(request, "Por favor corrija los errores en el formulario.")
    else:
        form = EntidadExternaForm(instance=entidad)
    
    context = {
        'form': form,
        'entidad': entidad,
        'titulo_pagina': f'Editar Entidad: {entidad.nombre}',
        'is_edit': True
    }
    return render(request, 'correspondencia/admin/entidad_form.html', context)

# === FIN CRUD Entidades Externas ===

# === CRUD Contactos (Ajustado para EntidadExterna) ===
@login_required
@require_POST
def editar_radicacion_rapida_saliente(request, pk):
    """
    Vista AJAX para editar una radicación rápida saliente.
    Recibe POST con prefix='rapida_sal', retorna JSON.
    """
    import re

    salida = get_object_or_404(CorrespondenciaSalida, pk=pk)

    # Permitir edición de salidas independientes y vinculadas a radicación rápida entrante
    if salida.respuesta_a is not None and salida.respuesta_a.origen_radicacion != 'RAPIDA':
        return JsonResponse({'success': False, 'error': 'Solo se pueden editar radicaciones rápidas salientes.'}, status=400)

    form = RadicacionRapidaSalienteForm(
        request.POST,
        request.FILES,
        instance=salida,
        user=request.user,
        prefix='rapida_sal'
    )
    if form.is_valid():
        obj = form.save(commit=False)
        obj.oficina_emisora = form.cleaned_data.get('oficina_emisora')

        destinatario_contacto = form.cleaned_data.get('destinatario_contacto')
        if destinatario_contacto:
            obj.destinatario_contacto = destinatario_contacto
            obj.destinatario_email = destinatario_contacto.correo_electronico or ''
        else:
            obj.destinatario_contacto = None
            obj.destinatario_email = ''
            dest_texto = form.cleaned_data.get('destinatario_texto', '').strip()
            if dest_texto:
                cuerpo_actual = obj.cuerpo or ''
                if cuerpo_actual.startswith('[Destinatario:'):
                    match_old = re.match(r'^\[Destinatario:.*?\]\s*\n*', cuerpo_actual)
                    if match_old:
                        cuerpo_actual = cuerpo_actual[match_old.end():].strip()
                obj.cuerpo = f"[Destinatario: {dest_texto}]\n\n{cuerpo_actual}"

        obj.save()
        return JsonResponse({
            'success': True,
            'message': f'Radicación {salida.numero_radicado_salida} actualizada exitosamente.'
        })
    else:
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = [str(e) for e in error_list]
        return JsonResponse({'success': False, 'errors': errors}, status=400)


@login_required
@require_POST
def editar_radicacion_rapida_entrante(request, pk):
    """
    Vista AJAX para editar una radicación rápida entrante.
    Recibe POST con prefix='rapida_ent', retorna JSON.
    """
    corresp = get_object_or_404(Correspondencia, pk=pk)

    if corresp.origen_radicacion != 'RAPIDA':
        return JsonResponse({'success': False, 'error': 'Solo se pueden editar radicaciones rápidas entrantes.'}, status=400)

    form = RadicacionRapidaEntranteForm(request.POST, instance=corresp, prefix='rapida_ent')
    if form.is_valid():
        obj = form.save(commit=False)
        obj.tipo_radicado = 'ENTRANTE'
        obj.origen_radicacion = 'RAPIDA'
        
        # Calcular fecha límite de respuesta automáticamente según tipo de trámite
        tipo_tramite_codigo = obj.tipo_tramite
        if tipo_tramite_codigo:
            try:
                tipo_tramite_obj = TipoTramite.objects.get(codigo=tipo_tramite_codigo, activo=True)
                if tipo_tramite_obj.dias_respuesta is not None:
                    # Usar fecha de recepción del documento si está disponible, sino fecha de radicación
                    fecha_base = obj.fecha_recepcion_documento or obj.fecha_radicacion or timezone.now()
                    # Calcular fecha límite desde la fecha base
                    fecha_limite = calcular_dias_habiles(fecha_base, tipo_tramite_obj.dias_respuesta)
                    obj.fecha_limite_respuesta_manual = fecha_limite
            except TipoTramite.DoesNotExist:
                # Si no existe el tipo, continuar sin calcular fecha límite
                pass
        
        obj.save()
        return JsonResponse({
            'success': True,
            'message': f'Radicación {corresp.numero_radicado} actualizada exitosamente.'
        })
    else:
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = [str(e) for e in error_list]
        return JsonResponse({'success': False, 'errors': errors}, status=400)


@login_required
@require_GET
def api_radicacion_rapida_entrante_data(request, pk):
    """API para obtener datos de una radicación rápida entrante (para poblar modal de edición)."""
    corresp = get_object_or_404(Correspondencia, pk=pk)
    if corresp.origen_radicacion != 'RAPIDA':
        return JsonResponse({'success': False, 'error': 'No es radicación rápida.'}, status=400)

    data = {
        'success': True,
        'numero_radicado': corresp.numero_radicado,
        'fields': {
            'asunto': corresp.asunto or '',
            'remitente': corresp.remitente_id or '',
            'remitente_texto': corresp.entidad_persona_remitente or '',
            'oficina_destino': corresp.oficina_destino_id or '',
            'medio_recepcion': corresp.medio_recepcion or '',
            'tipo_tramite': corresp.tipo_tramite or '',
            'entidad_persona_remitente': corresp.entidad_persona_remitente or '',
            'funcionario_responsable_tramite': corresp.funcionario_responsable_tramite or '',
            'clasificacion_comunicacion': corresp.clasificacion_comunicacion or '',
            'numero_folios': corresp.numero_folios if corresp.numero_folios is not None else '',
            'anexos': corresp.anexos or '',
            'medio_recibido': corresp.medio_recibido or '',
            'direccion_correo_remitente': corresp.direccion_correo_remitente or '',
            'empresa_transportadora': corresp.empresa_transportadora or '',
            'numero_guia': corresp.numero_guia or '',
            'fecha_limite_respuesta_manual': corresp.fecha_limite_respuesta_manual.isoformat() if corresp.fecha_limite_respuesta_manual else '',
            'fecha_primer_seguimiento': corresp.fecha_primer_seguimiento.isoformat() if corresp.fecha_primer_seguimiento else '',
            'fecha_segundo_seguimiento': corresp.fecha_segundo_seguimiento.isoformat() if corresp.fecha_segundo_seguimiento else '',
            'fecha_notificacion_vencimiento': corresp.fecha_notificacion_vencimiento.isoformat() if corresp.fecha_notificacion_vencimiento else '',
            'fecha_respuesta': corresp.fecha_respuesta.isoformat() if corresp.fecha_respuesta else '',
            'estado_respuesta': corresp.estado_respuesta if corresp.estado_respuesta in ('PENDIENTE', 'RESPONDIDA', 'VENCIDA') else '',
            'radicado_enviado_respuesta': corresp.radicado_enviado_respuesta or '',
        }
    }
    return JsonResponse(data)


@login_required
@require_GET
def api_tipos_tramite(request):
    """
    API para obtener tipos de trámite activos con sus días de respuesta.
    Usado por JavaScript para calcular fechas límite dinámicamente.
    """
    tipos = TipoTramite.objects.filter(activo=True).order_by('orden', 'codigo')
    data = {}
    for tipo in tipos:
        data[tipo.codigo] = {
            'nombre': tipo.nombre,
            'dias_respuesta': tipo.dias_respuesta,
            'descripcion': tipo.descripcion
        }
    return JsonResponse(data)


@login_required
@require_GET
def api_radicacion_rapida_saliente_data(request, pk):
    """API para obtener datos de una radicación rápida saliente (para poblar modal de edición)."""
    import re
    salida = get_object_or_404(CorrespondenciaSalida, pk=pk)
    # Permitir acceso para salidas independientes y vinculadas a radicación rápida entrante
    if salida.respuesta_a is not None and salida.respuesta_a.origen_radicacion != 'RAPIDA':
        return JsonResponse({'success': False, 'error': 'No es radicación rápida.'}, status=400)

    # Extraer destinatario_texto del cuerpo
    cuerpo_original = salida.cuerpo or ''
    destinatario_texto = ''
    cuerpo_limpio = cuerpo_original
    if cuerpo_original.startswith('[Destinatario:'):
        match = re.match(r'^\[Destinatario:\s*(.*?)\]\s*\n*', cuerpo_original)
        if match:
            destinatario_texto = match.group(1).strip()
            cuerpo_limpio = cuerpo_original[match.end():].strip()

    evidencia_url = ''
    evidencia_nombre = ''
    if salida.evidencia_respuesta:
        try:
            evidencia_url = salida.evidencia_respuesta.url
            evidencia_nombre = salida.evidencia_respuesta.name.split('/')[-1]
        except Exception:
            evidencia_url = ''
            evidencia_nombre = ''

    data = {
        'success': True,
        'numero_radicado': salida.numero_radicado_salida,
        'radicado_entrada': salida.respuesta_a.numero_radicado if salida.respuesta_a else '',
        'fields': {
            'asunto': salida.asunto or '',
            'cuerpo': cuerpo_limpio,
            'destinatario_contacto': salida.destinatario_contacto_id or '',
            'destinatario_texto': destinatario_texto,
            'oficina_emisora': salida.oficina_emisora_id or '',
            'funcionario_envia': salida.funcionario_envia or '',
            'fue_respondida': bool(salida.fue_respondida),
            'evidencia_url': evidencia_url,
            'evidencia_nombre': evidencia_nombre,
        }
    }
    return JsonResponse(data)


def buscar_entidades(request):
    """
    API endpoint para buscar entidades externas.
    
    FUNCIONALIDAD:
    - Busca entidades por nombre (filtro opcional)
    - Retorna máximo 10 resultados ordenados por nombre
    - Formato de respuesta compatible con Select2
    
    PARÁMETROS:
    - q: Término de búsqueda (opcional)
    
    RESPUESTA:
    {
        "results": [
            {"id": 1, "text": "Nombre Entidad"},
            ...
        ]
    }
    """
    q = request.GET.get('q', '').strip()
    from .models import EntidadExterna
    qs = EntidadExterna.objects.all()
    if q:
        qs = qs.filter(nombre__icontains=q)
    entidades = qs.order_by('nombre')[:10]
    data = {
        'results': [
            {
                'id': e.id,
                'text': e.nombre
            }
            for e in entidades
        ]
    }
    return JsonResponse(data)

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def buscar_oficinas(request):
    """
    API endpoint para buscar oficinas productoras.
    
    FUNCIONALIDAD:
    - Busca oficinas por nombre (filtro opcional)
    - Retorna máximo 10 resultados ordenados por nombre
    - Formato de respuesta compatible con Select2
    
    PARÁMETROS:
    - q: Término de búsqueda (opcional)
    
    RESPUESTA:
    {
        "results": [
            {"id": 1, "text": "Nombre Oficina"},
            ...
        ]
    }
    """
    q = request.GET.get('q', '').strip()
    default_limit = 50 if not q else 20
    try:
        limit = int(request.GET.get('limit') or default_limit)
    except (TypeError, ValueError):
        limit = default_limit
    limit = max(1, min(limit, 100))

    from .models import OficinaProductora
    qs = OficinaProductora.objects.all()
    if q:
        qs = qs.filter(nombre__icontains=q)
    oficinas = qs.order_by('nombre')[:limit]
    data = {
        'results': [
            {
                'id': o.id,
                'text': o.nombre
            }
            for o in oficinas
        ]
    }
    return JsonResponse(data)


def _aplicar_distribucion_rapida_desde_form(correspondencia, request, cleaned_data):
    """Aplica la asignación inicial y opcionalmente comparte con toda la oficina destino."""
    if not cleaned_data.get('distribuir_rapido'):
        return ''

    usuario_destino = cleaned_data.get('usuario_destino_rapido')
    if not usuario_destino:
        raise ValidationError('Debe seleccionar un responsable principal para la distribución inmediata.')

    if not hasattr(usuario_destino, 'perfil') or usuario_destino.perfil.oficina_id != correspondencia.oficina_destino_id:
        raise ValidationError('El responsable principal debe pertenecer a la oficina destino seleccionada.')

    observaciones = (cleaned_data.get('observaciones_distribucion') or '').strip() or 'Asignación inicial desde Ventanilla'
    otras_oficinas = cleaned_data.get('otras_oficinas')
    compartir_con_toda_oficina = bool(cleaned_data.get('compartir_con_toda_oficina'))

    correspondencia.usuario_destino_inicial = usuario_destino
    correspondencia.estado = 'ASIGNADA_USUARIO'
    correspondencia.save(update_fields=['usuario_destino_inicial', 'estado'])

    usuarios_oficina = User.objects.filter(
        perfil__oficina=correspondencia.oficina_destino,
        is_active=True,
    ).select_related('perfil')

    DistribucionInternaUsuario.objects.update_or_create(
        correspondencia=correspondencia,
        usuario_asignado=usuario_destino,
        defaults={
            'asignado_por': request.user,
            'fecha_asignacion': timezone.now(),
            'observaciones': observaciones,
            'leido': False,
        }
    )

    if compartir_con_toda_oficina:
        for usuario in usuarios_oficina:
            DistribucionInternaUsuario.objects.update_or_create(
                correspondencia=correspondencia,
                usuario_asignado=usuario,
                defaults={
                    'asignado_por': request.user,
                    'fecha_asignacion': timezone.now(),
                    'observaciones': f"Compartido por ventanilla. {observaciones}".strip(),
                    'leido': False,
                }
            )

        HistorialCorrespondencia.objects.create(
            correspondencia=correspondencia,
            evento='REDISTRIBUIDA_INTERNA',
            usuario=request.user,
            descripcion=f"Compartida con toda la oficina {correspondencia.oficina_destino.nombre} desde el modal de radicación."
        )

    oficinas_agregadas = 0
    for oficina in (otras_oficinas or []):
        if oficina.pk == correspondencia.oficina_destino_id:
            continue
        acceso, creado = AccesoCorrespondenciaOficina.objects.update_or_create(
            correspondencia=correspondencia,
            oficina=oficina,
            defaults={
                'compartido_por': request.user,
                'observaciones': f"Pre-redistribución desde modal de radicación. {observaciones}".strip(),
                'solo_lider': False,
                'puede_responder': False,
            }
        )
        if creado:
            oficinas_agregadas += 1
            HistorialCorrespondencia.objects.create(
                correspondencia=correspondencia,
                evento='COMPARTIDA_OFICINA',
                usuario=request.user,
                descripcion=f"Acceso de solo lectura otorgado a la oficina {oficina.nombre} desde el modal de radicación."
            )
            crear_notificaciones_acceso_oficina(correspondencia, oficina, request.user, observaciones)

    HistorialCorrespondencia.objects.create(
        correspondencia=correspondencia,
        evento='ASIGNADA_USUARIO',
        usuario=request.user,
        descripcion=f"Asignada a {usuario_destino.get_full_name() or usuario_destino.username} desde el modal de radicación"
    )

    extras = []
    if compartir_con_toda_oficina:
        extras.append('compartida con toda la oficina')
    else:
        extras.append('solo al responsable principal')
    if oficinas_agregadas:
        extras.append(f'compartida con {oficinas_agregadas} oficina(s) adicionales')
    return f" y distribuida ({', '.join(extras)})" if extras else ' y distribuida'


def dashboard_usuario(request):
    """
    Dashboard para usuarios regulares que muestra:
    - KPIs de correspondencia personal y de oficina
    - Correspondencia reciente
    - Estado de SLA
    """
    from datetime import timedelta
    
    # Verificar si el usuario tiene perfil y oficina asignada
    try:
        perfil_usuario = request.user.perfil
        oficina_usuario = perfil_usuario.oficina if hasattr(perfil_usuario, 'oficina') else None
    except:
        perfil_usuario = None
        oficina_usuario = None
    
    # --- KPIs ---
    ahora = timezone.now()
    hoy_inicio = ahora.astimezone(timezone.get_current_timezone()).replace(hour=0, minute=0, second=0, microsecond=0)
    hoy_fin = hoy_inicio + datetime.timedelta(days=1)
    
    # KPIs de correspondencia personal usando la misma base que la bandeja personal.
    if oficina_usuario:
        correspondencias_personales_qs = _base_bandeja_personal_queryset(
            request.user,
            oficina_usuario,
        )

        correspondencias_pendientes = correspondencias_personales_qs.annotate(
            num_distribuciones=Count('distribuciones_internas', distinct=True),
            total_destinatarios=Count('distribuciones_internas', distinct=True),
            total_leidos=Count('distribuciones_internas', filter=Q(distribuciones_internas__leido=True), distinct=True)
        ).filter(
            total_leidos__lt=F('total_destinatarios')  # X < Y
        ).distinct()
        kpi_recibidos_pendientes = correspondencias_pendientes.count()

        correspondencias_hoy = correspondencias_personales_qs.filter(
            fecha_radicacion__gte=hoy_inicio,
            fecha_radicacion__lt=hoy_fin,
        )
        kpi_recibidos_hoy = correspondencias_hoy.count()
    else:
        kpi_recibidos_pendientes = 0
        kpi_recibidos_hoy = 0
    
    # KPIs de oficina usando la misma base que la bandeja de oficina.
    kpi_oficina_pendientes = 0
    kpi_oficina_hoy = 0
    if oficina_usuario:
        correspondencias_oficina_qs = _base_bandeja_oficina_queryset(oficina_usuario)

        kpi_oficina_pendientes = correspondencias_oficina_qs.filter(
            total_destinatarios__gt=0,
            total_leidos__lt=F('total_destinatarios'),
        ).count()

        kpi_oficina_hoy = correspondencias_oficina_qs.filter(
            fecha_radicacion__gte=hoy_inicio,
            fecha_radicacion__lt=hoy_fin,
        ).count()
    
    # SLA vencido personal (correspondencia asignada al usuario)
    correspondencias_usuario = DistribucionInternaUsuario.objects.filter(
        usuario_asignado=request.user
    ).values_list('correspondencia_id', flat=True)
    
    kpi_sla_vencido_personal = excluir_entrantes_con_respuesta(
        Correspondencia.objects.filter(
            id__in=correspondencias_usuario,
            tipo_radicado='ENTRANTE',
            requiere_respuesta=True,
            fecha_limite_respuesta_persist__lt=ahora,
        )
    ).count()
    
    # SLA por vencer personal (48h)
    kpi_sla_por_vencer_personal = excluir_entrantes_con_respuesta(
        Correspondencia.objects.filter(
            id__in=correspondencias_usuario,
            tipo_radicado='ENTRANTE',
            requiere_respuesta=True,
            fecha_limite_respuesta_persist__gte=ahora,
            fecha_limite_respuesta_persist__lte=ahora + timedelta(hours=48),
        )
    ).count()
    
    # --- Correspondencia reciente personal ---
    distribuciones_recientes = DistribucionInternaUsuario.objects.filter(
        usuario_asignado=request.user
    ).select_related('correspondencia__remitente', 'correspondencia__oficina_destino').order_by('-fecha_asignacion')[:8]
    
    # --- Correspondencia reciente de oficina ---
    correspondencia_reciente_oficina = []
    if oficina_usuario:
        correspondencia_reciente_oficina = Correspondencia.objects.filter(
            tipo_radicado='ENTRANTE',
            oficina_destino=oficina_usuario
        ).select_related('remitente', 'usuario_destino_inicial').order_by('-fecha_radicacion')[:8]
    
    # --- Correspondencia compartida recientemente ---
    # Buscar correspondencia que el usuario haya compartido (esto requeriría un campo adicional)
    # Por ahora, mostraremos correspondencia asignada inicialmente al usuario
    correspondencia_compartida = []
    
    # --- DATOS PARA GRÁFICO DE BARRAS POR ENTIDAD (últimos 30 días) ---
    fecha_hace_30_dias = hoy_inicio - timedelta(days=30)
    
    # Obtener correspondencia recibida agrupada por entidad remitente
    if oficina_usuario:
        correspondencia_por_entidad = correspondencias_personales_qs.filter(
            tipo_radicado='ENTRANTE',
            fecha_radicacion__gte=fecha_hace_30_dias,
            remitente__isnull=False,
            remitente__entidad_externa__isnull=False
        ).values('remitente__entidad_externa__nombre').annotate(
            total=Count('id')
        ).order_by('-total')[:10]  # Top 10 entidades
        
        entidades_labels = []
        entidades_valores = []
        
        for item in correspondencia_por_entidad:
            nombre_entidad = item['remitente__entidad_externa__nombre']
            # Truncar nombres largos
            if len(nombre_entidad) > 25:
                nombre_entidad = nombre_entidad[:22] + '...'
            entidades_labels.append(nombre_entidad)
            entidades_valores.append(item['total'])
        
        # Si no hay datos, mostrar mensaje
        if not entidades_labels:
            entidades_labels = ['Sin datos']
            entidades_valores = [0]
    else:
        entidades_labels = ['Sin oficina asignada']
        entidades_valores = [0]
    
    # --- TIMELINE DE ACTIVIDAD RECIENTE ---
    timeline_actividades = []
    
    # 1. Últimas correspondencias enviadas
    salidas_recientes = CorrespondenciaSalida.objects.filter(
        usuario_redactor=request.user
    ).order_by('-fecha_envio')[:5]
    
    for salida in salidas_recientes:
        timeline_actividades.append({
            'tipo': 'envio',
            'icono': 'bi-send-check',
            'color': 'success',
            'titulo': 'Correspondencia enviada',
            'descripcion': f"{salida.numero_radicado_salida}",
            'detalle': salida.asunto[:50] if salida.asunto else 'Sin asunto',
            'fecha': salida.fecha_envio,
            'url': reverse('correspondencia:detalle_respuesta_salida', args=[salida.id])
        })
    
    # 2. Últimas correspondencias asignadas
    distribuciones_asignadas = DistribucionInternaUsuario.objects.filter(
        usuario_asignado=request.user
    ).select_related('correspondencia__remitente__entidad_externa', 'asignado_por').order_by('-fecha_asignacion')[:5]
    
    for dist in distribuciones_asignadas:
        asignado_por = dist.asignado_por.get_full_name() if dist.asignado_por else 'Sistema'
        timeline_actividades.append({
            'tipo': 'asignacion',
            'icono': 'bi-person-check',
            'color': 'info',
            'titulo': 'Correspondencia entrante',
            'descripcion': f"{dist.correspondencia.numero_radicado}",
            'detalle': f"Asignado por: {asignado_por}",
            'fecha': dist.fecha_asignacion,
            'url': reverse('correspondencia:detalle_correspondencia', args=[dist.correspondencia.id])
        })
    
    # Ordenar timeline por fecha descendente y limitar a 10
    # Usar timezone.now() como fecha de referencia para None
    timeline_actividades.sort(key=lambda x: x['fecha'] if x['fecha'] is not None else timezone.now(), reverse=True)
    timeline_actividades = timeline_actividades[:10]

    # KPI: rebotes de envío del usuario
    kpi_rebotes = SalidaDestinatario.objects.filter(
        estado='REBOTE',
        correspondencia_salida__usuario_redactor=request.user,
    ).count()

    context = {
        'titulo_pagina': 'Dashboard Usuario',
        'perfil_usuario': perfil_usuario,
        'oficina_usuario': oficina_usuario,
        # KPIs personales
        'kpi_recibidos_pendientes': kpi_recibidos_pendientes,
        'kpi_recibidos_hoy': kpi_recibidos_hoy,
        'kpi_sla_vencido_personal': kpi_sla_vencido_personal,
        'kpi_sla_por_vencer_personal': kpi_sla_por_vencer_personal,
        # KPIs de oficina
        'kpi_oficina_pendientes': kpi_oficina_pendientes,
        'kpi_oficina_hoy': kpi_oficina_hoy,
        # Correspondencia reciente
        'distribuciones_recientes': distribuciones_recientes,
        'correspondencia_reciente_oficina': correspondencia_reciente_oficina,
        'correspondencia_compartida': correspondencia_compartida,
        # Fecha de hoy para filtros
        'hoy': hoy_inicio,
        # Datos para gráfico de barras por entidad (serializados como JSON)
        'entidades_labels': json.dumps(entidades_labels),
        'entidades_valores': json.dumps(entidades_valores),
        # Timeline de actividad
        'timeline_actividades': timeline_actividades,
        # Rebotes
        'kpi_rebotes': kpi_rebotes,
    }
    
    return render(request, 'correspondencia/usuario/dashboard_usuario.html', context)

@login_required
# @user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def metricas_destinatarios(request):
    """Vista para mostrar métricas de envío por destinatario."""
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    
    # Filtros
    oficina_id = request.GET.get('oficina')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    # Base queryset
    qs = SalidaDestinatario.objects.select_related(
        'contacto__entidad_externa',
        'correspondencia_salida__oficina_emisora'
    ).filter(
        correspondencia_salida__oficina_emisora__isnull=False
    )
    
    # Aplicar filtros
    if oficina_id:
        qs = qs.filter(correspondencia_salida__oficina_emisora_id=oficina_id)
    
    if fecha_desde:
        try:
            qs = qs.filter(correspondencia_salida__fecha_creacion__date__gte=fecha_desde)
        except:
            pass
    
    if fecha_hasta:
        try:
            qs = qs.filter(correspondencia_salida__fecha_creacion__date__lte=fecha_hasta)
        except:
            pass
    
    # Métricas por destinatario
    metricas_destinatario = qs.values(
        'contacto__nombres', 'contacto__apellidos', 'contacto__entidad_externa__nombre',
        'email_snapshot'
    ).annotate(
        total_envios=Count('id'),
        exitosos=Count('id', filter=Q(estado='ENVIADO')),
        fallidos=Count('id', filter=Q(estado='FALLO')),
        rebotes=Count('id', filter=Q(estado='REBOTE')),
        pendientes=Count('id', filter=Q(estado='PENDIENTE'))
    ).order_by('-total_envios')
    
    # Métricas por oficina
    metricas_oficina = qs.values(
        'correspondencia_salida__oficina_emisora__nombre'
    ).annotate(
        total_envios=Count('id'),
        exitosos=Count('id', filter=Q(estado='ENVIADO')),
        fallidos=Count('id', filter=Q(estado='FALLO')),
        rebotes=Count('id', filter=Q(estado='REBOTE')),
        pendientes=Count('id', filter=Q(estado='PENDIENTE'))
    ).order_by('-total_envios')
    
    # Destinatarios con fallos o rebotes recientes (últimos 30 días)
    fecha_limite = datetime.now() - timedelta(days=30)
    fallos_recientes = qs.filter(
        estado__in=['FALLO', 'REBOTE'],
        correspondencia_salida__fecha_creacion__gte=fecha_limite
    ).select_related('contacto', 'correspondencia_salida').order_by('-correspondencia_salida__fecha_creacion')
    
    context = {
        'metricas_destinatario': metricas_destinatario,
        'metricas_oficina': metricas_oficina,
        'fallos_recientes': fallos_recientes,
        'oficinas': OficinaProductora.objects.all(),
        'titulo_pagina': 'Métricas de Envío por Destinatario'
    }
    
    return render(request, 'correspondencia/admin/metricas_destinatarios.html', context)
# ===================== LISTADO DE CONTACTOS GLOBAL PARA USUARIOS =====================
@login_required
def lista_contactos_usuario(request):
    """Listado con filtros para que los usuarios consulten la agenda global."""
    contactos_qs = Contacto.objects.select_related('entidad_externa').all()

    search = (request.GET.get('q') or '').strip()
    entidad_param = request.GET.get('entidad')
    correo_electronico = (request.GET.get('correo_electronico') or '').strip()
    tiene_email = request.GET.get('tiene_email')
    orden = request.GET.get('orden') or 'entidad'

    if search:
        contactos_qs = contactos_qs.filter(
            Q(nombres__icontains=search) |
            Q(apellidos__icontains=search) |
            Q(cargo__icontains=search) |
            Q(correo_electronico__icontains=search) |
            Q(entidad_externa__nombre__icontains=search)
        )

    if entidad_param and entidad_param.isdigit():
        contactos_qs = contactos_qs.filter(entidad_externa_id=int(entidad_param))

    if correo_electronico:
        contactos_qs = contactos_qs.filter(correo_electronico__icontains=correo_electronico)

    if tiene_email == 'con':
        contactos_qs = contactos_qs.exclude(Q(correo_electronico__isnull=True) | Q(correo_electronico=''))
    elif tiene_email == 'sin':
        contactos_qs = contactos_qs.filter(Q(correo_electronico__isnull=True) | Q(correo_electronico=''))

    orden_map = {
        'entidad': ('entidad_externa__nombre', 'apellidos', 'nombres'),
        'nombre': ('apellidos', 'nombres'),
        'recientes': ('-id',),
    }
    contactos_qs = contactos_qs.order_by(*orden_map.get(orden, ('entidad_externa__nombre', 'apellidos', 'nombres')))

    contactos_filtrados_total = contactos_qs.count()
    contactos_con_email_total = contactos_qs.exclude(
        Q(correo_electronico__isnull=True) | Q(correo_electronico='')
    ).count()
    contactos_sin_email_total = contactos_filtrados_total - contactos_con_email_total
    entidades_total = contactos_qs.exclude(entidad_externa__isnull=True).values('entidad_externa_id').distinct().count()

    paginator = Paginator(contactos_qs, 30)
    page_number = request.GET.get('page')
    contactos_page = paginator.get_page(page_number)

    entidades = EntidadExterna.objects.order_by('nombre')

    params = request.GET.copy()
    params.pop('page', None)
    querystring = params.urlencode()

    context = {
        'contactos': contactos_page,
        'entidades': entidades,
        'filtros': {
            'q': search,
            'entidad': entidad_param or '',
            'tiene_email': tiene_email or '',
            'orden': orden,
        },
        'titulo_pagina': 'Contactos Globales',
        'querystring': querystring,
        'contactos_filtrados_total': contactos_filtrados_total,
        'contactos_con_email_total': contactos_con_email_total,
        'contactos_sin_email_total': contactos_sin_email_total,
        'entidades_total': entidades_total,
    }
    return render(request, 'correspondencia/usuario/contactos_globales.html', context)


@login_required
def contactos_agenda_modal_ajax(request):
    """Endpoint AJAX para el selector ampliado de contactos."""
    if request.method != 'GET' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Solicitud no válida.'}, status=400)

    search = (request.GET.get('q') or '').strip()
    entidad_param = request.GET.get('entidad')
    tiene_email = request.GET.get('tiene_email')
    orden = request.GET.get('orden') or 'entidad'
    try:
        page_number = max(int(request.GET.get('page') or 1), 1)
    except (TypeError, ValueError):
        page_number = 1
    try:
        per_page = int(request.GET.get('per_page') or 10)
    except (TypeError, ValueError):
        per_page = 10
    per_page = max(5, min(per_page, 50))

    contactos_qs = Contacto.objects.select_related('entidad_externa').all()

    if search:
        contactos_qs = contactos_qs.filter(
            Q(nombres__icontains=search) |
            Q(apellidos__icontains=search) |
            Q(cargo__icontains=search) |
            Q(correo_electronico__icontains=search) |
            Q(entidad_externa__nombre__icontains=search)
        )

    if entidad_param and entidad_param.isdigit():
        contactos_qs = contactos_qs.filter(entidad_externa_id=int(entidad_param))

    if tiene_email == 'con':
        contactos_qs = contactos_qs.exclude(Q(correo_electronico__isnull=True) | Q(correo_electronico=''))
    elif tiene_email == 'sin':
        contactos_qs = contactos_qs.filter(Q(correo_electronico__isnull=True) | Q(correo_electronico=''))

    orden_map = {
        'entidad': ('entidad_externa__nombre', 'apellidos', 'nombres'),
        'nombre': ('apellidos', 'nombres'),
        'recientes': ('-id',),
    }
    contactos_qs = contactos_qs.order_by(*orden_map.get(orden, ('entidad_externa__nombre', 'apellidos', 'nombres')))

    paginator = Paginator(contactos_qs, per_page)
    page_obj = paginator.get_page(page_number)

    contactos_data = [{
        'id': contacto.id,
        'nombre': contacto.nombre_completo,
        'entidad': contacto.entidad_externa.nombre if contacto.entidad_externa else 'Sin entidad',
        'cargo': contacto.cargo or '',
        'email': contacto.correo_electronico or '',
        'telefono': contacto.telefono_contacto or '',
    } for contacto in page_obj.object_list]

    return JsonResponse({
        'success': True,
        'contactos': contactos_data,
        'meta': {
            'total': paginator.count,
            'pages': paginator.num_pages or 1,
            'page': page_obj.number,
        }
    })

# ===================== VISTAS AJAX PARA MODALES =====================

@login_required
def crear_contacto_ajax(request):
    """Vista AJAX para crear contactos desde el modal."""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            # Usar el formulario para validaciones
            form = ContactoForm(request.POST)
            
            if not form.is_valid():
                # Extraer errores del formulario de manera amigable
                errores = []
                correo = form.data.get('correo_electronico', '').strip().lower()
                
                for field, errors in form.errors.items():
                    # Obtener el nombre del campo en español
                    field_label = form.fields[field].label if field in form.fields else field.replace('_', ' ').title()
                    
                    for error in errors:
                        error_str = str(error).lower()
                        
                        # Mensajes específicos para correo electrónico duplicado
                        if 'correo' in field.lower() and ('ya existe' in error_str or 'unique' in error_str or 'duplicate' in error_str):
                            if correo:
                                contacto_existente = Contacto.objects.filter(correo_electronico__iexact=correo).first()
                                if contacto_existente:
                                    errores.append(
                                        f"El correo electrónico <strong>'{correo}'</strong> ya está registrado para "
                                        f"<strong>'{contacto_existente.nombre_completo}'</strong> de la entidad "
                                        f"<strong>'{contacto_existente.entidad_externa.nombre}'</strong>."
                                    )
                                else:
                                    errores.append(f"El correo electrónico '{correo}' ya está en uso.")
                            else:
                                errores.append("El correo electrónico ingresado ya está registrado.")
                        # Otros errores de validación
                        elif 'required' in error_str or 'obligatorio' in error_str:
                            errores.append(f"El campo <strong>{field_label}</strong> es obligatorio.")
                        elif 'invalid' in error_str:
                            errores.append(f"El valor ingresado en <strong>{field_label}</strong> no es válido.")
                        else:
                            # Mensaje genérico pero amigable
                            mensaje = str(error)
                            # Limpiar mensajes de Django que contienen diccionarios
                            if '{' in mensaje and '}' in mensaje:
                                # Si el error contiene un diccionario, extraer solo el mensaje útil
                                if 'correo' in field.lower():
                                    if correo:
                                        contacto_existente = Contacto.objects.filter(correo_electronico__iexact=correo).first()
                                        if contacto_existente:
                                            mensaje = (
                                                f"El correo electrónico '{correo}' ya está registrado para "
                                                f"'{contacto_existente.nombre_completo}' de la entidad "
                                                f"'{contacto_existente.entidad_externa.nombre}'."
                                            )
                                        else:
                                            mensaje = f"El correo electrónico '{correo}' ya está en uso."
                                    else:
                                        mensaje = "El correo electrónico ingresado ya está registrado."
                                else:
                                    mensaje = f"Error en el campo {field_label}."
                            errores.append(mensaje)
                
                mensaje_error = '<br>'.join(errores) if errores else 'Por favor corrija los errores en el formulario.'
                
                return JsonResponse({
                    'success': False,
                    'error': mensaje_error
                })
            
            # Si el formulario es válido, guardar
            contacto = form.save()
            warning = ''
            registrar_dominio = str(request.POST.get('registrar_dominio_entidad', '')).lower() in {'1', 'true', 'si', 'sí', 'yes', 'on'}
            dominio_contacto = normalizar_dominio_correo(contacto.correo_electronico)

            if registrar_dominio and contacto.entidad_externa and dominio_contacto and not contacto.entidad_externa.tiene_dominio_autorizado(dominio_contacto):
                resultado_dominios = contacto.entidad_externa.registrar_dominios([dominio_contacto])
                if resultado_dominios['registrados']:
                    warning = f"Se agregó el dominio '{dominio_contacto}' a la entidad '{contacto.entidad_externa.nombre}'."
                elif resultado_dominios['conflictos']:
                    conflicto = resultado_dominios['conflictos'][0]
                    warning = (
                        f"El contacto se creó, pero el dominio '{conflicto['dominio']}' ya pertenece a "
                        f"'{conflicto['entidad_nombre']}' y no se reasignó."
                    )
            
            return JsonResponse({
                'success': True,
                'contacto': {
                    'id': contacto.id,
                    'nombre_completo': contacto.nombre_completo,
                    'entidad': contacto.entidad_externa.nombre,
                    'email': contacto.correo_electronico,
                },
                'warning': warning,
                'message': 'Contacto creado exitosamente.'
            })
            
        except IntegrityError as exc:
            # Capturar errores de integridad (duplicados, etc.)
            correo = request.POST.get('correo_electronico', '').strip().lower()
            contacto_existente = Contacto.objects.filter(correo_electronico__iexact=correo).first()
            
            if contacto_existente:
                mensaje = (
                    f"El correo electrónico '{correo}' ya está registrado para "
                    f"'{contacto_existente.nombre_completo}' de la entidad "
                    f"'{contacto_existente.entidad_externa.nombre}'."
                )
            else:
                mensaje = f"Ya existe un contacto con el correo electrónico '{correo}'."
            
            logger.warning("Error de integridad al crear contacto (AJAX): %s", exc)
            return JsonResponse({
                'success': False,
                'error': mensaje
            })
        except Exception as e:
            logger.exception("Error inesperado al crear contacto (AJAX)")
            # Mensaje genérico para errores inesperados
            return JsonResponse({
                'success': False,
                'error': 'Ocurrió un error inesperado al crear el contacto. Por favor, intente nuevamente.'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido.'
    })

@login_required
def crear_entidad_ajax(request):
    """Vista AJAX para crear entidades desde el modal."""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            # Obtener datos del formulario
            nombre = request.POST.get('nombre')
            nit = request.POST.get('nit', '')
            dominio = request.POST.get('dominio', '')
            
            # Validaciones básicas
            if not nombre:
                return JsonResponse({
                    'success': False,
                    'error': 'El nombre de la entidad es obligatorio.'
                })
            
            # Verificar si ya existe una entidad con ese nombre
            if EntidadExterna.objects.filter(nombre=nombre).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Ya existe una entidad con ese nombre.'
                })
            
            dominios_extraidos = extraer_dominios_candidatos(dominio)
            dominio_principal = dominios_extraidos[0] if dominios_extraidos else None

            if dominio_principal:
                entidad_existente_por_dominio = EntidadExterna.buscar_por_dominio(dominio_principal)
                if entidad_existente_por_dominio:
                    return JsonResponse({
                        'success': False,
                        'error': f"El dominio '{dominio_principal}' ya está asociado a la entidad '{entidad_existente_por_dominio.nombre}'."
                    })

            # Crear la entidad
            entidad = EntidadExterna.objects.create(
                nombre=nombre,
                nit=nit if nit else None,
                dominio=dominio_principal
            )

            registro_dominios = entidad.registrar_dominios(dominios_extraidos[1:]) if len(dominios_extraidos) > 1 else {'registrados': [], 'conflictos': []}
            
            return JsonResponse({
                'success': True,
                'entidad': {
                    'id': entidad.id,
                    'nombre': entidad.nombre,
                    'nit': entidad.nit,
                    'dominio': entidad.dominio,
                    'dominios': entidad.dominios_configurados(),
                },
                'warning': registro_dominios['conflictos'],
                'message': 'Entidad creada exitosamente.'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al crear la entidad: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido.'
    })

@login_required
def responder_correspondencia_ajax(request):
    """Vista AJAX para responder correspondencia desde el modal."""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            # Obtener datos del formulario
            correspondencia_entrada_id = request.POST.get('correspondencia_entrada_id')
            asunto = normalizar_asunto_salida(request.POST.get('asunto') or '')
            cuerpo = request.POST.get('cuerpo')
            destinatarios_contacto_ids = request.POST.getlist('destinatarios_contacto')
            destinatarios_email = [e.strip() for e in request.POST.getlist('destinatarios_email') if e.strip()]
            categoria_contacto_id = (request.POST.get('categoria_contacto_id') or '').strip()
            enviar_inmediatamente = request.POST.get('enviar_inmediatamente') == 'on'
            marcar_como_leido = request.POST.get('marcar_como_leido') == 'on'
            
            # Límite de carga de adjuntos (25 MB negocio; Postmark limita a 10 MB al enviar).
            from correspondencia.utils.postmark_outbound import (
                mensaje_error_carga_adjuntos,
                salida_adjuntos_upload_limit_bytes,
            )
            adjuntos_files = request.FILES.getlist('adjuntos')
            total_adjuntos_bytes = sum(getattr(f, 'size', 0) for f in adjuntos_files)
            upload_limit_bytes = salida_adjuntos_upload_limit_bytes()
            if total_adjuntos_bytes > upload_limit_bytes:
                return JsonResponse({
                    'success': False,
                    'error': mensaje_error_carga_adjuntos(total_adjuntos_bytes),
                    'error_code': 'ADJUNTOS_SUPERAN_LIMITE_CARGA',
                })

            # Validaciones básicas
            total_destinatarios = len(destinatarios_contacto_ids) + len(destinatarios_email)
            print(f"DEBUG: contactos={destinatarios_contacto_ids}, emails={destinatarios_email}, categoria={categoria_contacto_id}, total={total_destinatarios}")
            
            # Validar que no se mezclen categoría con destinatarios manuales
            if categoria_contacto_id and (destinatarios_contacto_ids or destinatarios_email):
                return JsonResponse({
                    'success': False,
                    'error': 'No puede seleccionar una categoría y destinatarios manuales al mismo tiempo.'
                })
            
            # Validar que haya al menos un tipo de destinatario
            if not categoria_contacto_id and total_destinatarios == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe seleccionar al menos un destinatario o una categoría.'
                })
            
            if not all([correspondencia_entrada_id, asunto, cuerpo]):
                return JsonResponse({
                    'success': False,
                    'error': 'Los campos correspondencia, asunto y cuerpo son obligatorios.'
                })
            
            # Obtener la correspondencia entrante
            try:
                correspondencia_entrada = Correspondencia.objects.get(id=correspondencia_entrada_id)
            except Correspondencia.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'La correspondencia no existe.'
                })

            if not usuario_puede_responder_correspondencia(correspondencia_entrada, request.user):
                return JsonResponse({
                    'success': False,
                    'error': 'No tienes permiso para responder a esta correspondencia.'
                })

            es_respuesta_discrecional = False
            motivo_respuesta_discrecional = ''
            if not correspondencia_entrada.requiere_respuesta:
                if not usuario_tiene_permiso_respuesta_discrecional(request.user):
                    return JsonResponse({
                        'success': False,
                        'error': 'Esta correspondencia no requiere respuesta y no tiene permiso para responderla discrecionalmente.'
                    })
                es_respuesta_discrecional = True
                motivo_respuesta_discrecional = (request.POST.get('motivo_respuesta_discrecional') or '').strip()
                if not motivo_respuesta_discrecional:
                    return JsonResponse({
                        'success': False,
                        'error': 'Debe indicar el motivo de la respuesta discrecional.'
                    })
            
            # Obtener perfil y oficina del usuario
            perfil_usuario = getattr(request.user, 'perfil', None)
            oficina_usuario = getattr(perfil_usuario, 'oficina', None)
            
            # Procesar destinatarios según el tipo seleccionado
            contactos = []
            if categoria_contacto_id:
                # Obtener contactos de la categoría (grupo de agenda)
                try:
                    grupo = GrupoAgenda.objects.get(
                        id=categoria_contacto_id,
                        oficina_propietaria=oficina_usuario,
                        activo=True
                    )
                    contactos = list(grupo.contactos.filter(
                        correo_electronico__isnull=False
                    ).exclude(correo_electronico=''))
                        
                except GrupoAgenda.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'La categoría seleccionada no existe o no pertenece a su oficina.'
                    })
            else:
                # Procesar destinatarios manuales
                contactos = list(Contacto.objects.filter(id__in=destinatarios_contacto_ids))

            # NUEVO: normalización + deduplicación por email (backend, silencioso)
            contactos_by_email = { (c.correo_electronico or '').strip().lower(): c for c in contactos if c.correo_electronico }
            manual_validos = []
            for em in destinatarios_email:
                em_norm = em.strip().lower()
                manual_validos.append(em_norm)

            # Unir y deduplicar
            emails_unicos = set(list(contactos_by_email.keys()) + manual_validos)

            # Límite 50 tras dedup (silencioso si reduce)
            if len(emails_unicos) > 50:
                return JsonResponse({'success': False, 'error': f'No se permiten más de 50 destinatarios (tras normalización).'})

            # Validar que haya al menos un destinatario válido
            if not emails_unicos:
                return JsonResponse({
                        'success': False,
                        'error': 'Debe seleccionar al menos un destinatario válido.'
                    })

            from correspondencia.utils.blocked_recipients import validar_emails_destinatario_permitidos
            ok_destinatarios, error_destinatarios = validar_emails_destinatario_permitidos(emails_unicos)
            if not ok_destinatarios:
                return JsonResponse({
                    'success': False,
                    'error': error_destinatarios,
                    'error_code': 'DESTINATARIO_INSTITUCIONAL_BLOQUEADO',
                })
            
            # Crear la correspondencia de salida
            try:
                with transaction.atomic():
                    # Crear la respuesta
                    respuesta = CorrespondenciaSalida.objects.create(
                        respuesta_a=correspondencia_entrada,
                        usuario_redactor=request.user,
                        asunto=asunto,
                        cuerpo=cuerpo,
                        tipo_respuesta='DISCRECIONAL' if es_respuesta_discrecional else 'OBLIGATORIA',
                        motivo_respuesta_discrecional=motivo_respuesta_discrecional,
                        destinatario_contacto=contactos[0] if contactos else None,
                        estado='BORRADOR' if not enviar_inmediatamente else 'PENDIENTE_APROBACION'
                    )

                    HistorialSalida.objects.create(
                        correspondencia_salida=respuesta,
                        tipo_evento='RESPUESTA_DISCRECIONAL' if es_respuesta_discrecional else 'CREACION',
                        usuario=request.user,
                        descripcion=(
                            f'Respuesta discrecional creada desde modal. Motivo: {motivo_respuesta_discrecional}'
                            if es_respuesta_discrecional else
                            'Respuesta creada desde modal.'
                        )
                    )

                    # Trazabilidad de modalidad de envío
                    try:
                        total_dest = len(emails_unicos)
                        if categoria_contacto_id:
                            respuesta.envio_tipo = 'GRUPO'
                            try:
                                respuesta.envio_grupo = GrupoAgenda.objects.get(id=categoria_contacto_id)
                            except GrupoAgenda.DoesNotExist:
                                respuesta.envio_grupo = None
                            respuesta.envio_total_destinatarios = total_dest
                            respuesta.envio_detalle_snapshot = f"{total_dest} contactos (deduplicados)"
                        else:
                            if total_dest <= 1:
                                respuesta.envio_tipo = 'INDIVIDUAL'
                            else:
                                respuesta.envio_tipo = 'MULTIPLE_SELECTIVO'
                            respuesta.envio_total_destinatarios = total_dest
                            respuesta.envio_detalle_snapshot = f"{total_dest} destinatarios (deduplicados)"
                        respuesta.save(update_fields=['envio_tipo','envio_grupo','envio_total_destinatarios','envio_detalle_snapshot'])
                    except Exception:
                        pass

                    # Crear destinatarios desde emails_unicos
                    destinatarios_a_crear = []
                    for em in emails_unicos:
                        c = contactos_by_email.get(em)
                        destinatarios_a_crear.append(SalidaDestinatario(
                            correspondencia_salida=respuesta,
                            contacto=c,
                            email_snapshot=em,
                            nombre_snapshot=(c.nombre_completo if c else em),
                            estado='PENDIENTE'
                        ))
                    
                    # Crear todos los destinatarios de una vez
                    if destinatarios_a_crear:
                        SalidaDestinatario.objects.bulk_create(destinatarios_a_crear)
                    
                    # Procesar archivos adjuntos
                    archivos = request.FILES.getlist('adjuntos')
                    adjuntos_a_crear = []
                    for archivo in archivos:
                        adjuntos_a_crear.append(AdjuntoSalida(
                            correspondencia_salida=respuesta,
                            archivo=archivo,
                            nombre_original=archivo.name
                        ))
                    
                    if adjuntos_a_crear:
                        AdjuntoSalida.objects.bulk_create(adjuntos_a_crear)
                    
                    # Marcar como leído si se solicita
                    if marcar_como_leido:
                        correspondencia_entrada.leido_por_oficina = True
                        correspondencia_entrada.save()
                    
                    # Si se debe enviar inmediatamente, cambiar estado
                    if enviar_inmediatamente:
                        respuesta.estado = 'PENDIENTE_APROBACION'
                        respuesta.save()
                        
                        # Crear historial
                        HistorialSalida.objects.create(
                            correspondencia_salida=respuesta,
                            tipo_evento='ENVIO_APROBACION',
                            usuario=request.user,
                            descripcion=(
                                f'Respuesta discrecional enviada para aprobación desde modal. Motivo: {motivo_respuesta_discrecional}'
                                if es_respuesta_discrecional else
                                'Respuesta enviada para aprobación desde modal.'
                            )
                        )
            except Exception as e:
                # Log del error para debugging
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error en transacción de respuesta: {str(e)}")
                raise
            
            # Preparar mensaje de respuesta
            if categoria_contacto_id:
                grupo = GrupoAgenda.objects.get(id=categoria_contacto_id)
                prefijo = 'Respuesta discrecional' if es_respuesta_discrecional else 'Respuesta'
                message = f'{prefijo} creada exitosamente con {len(contactos)} destinatarios de la categoría "{grupo.nombre}".'
            else:
                total_dest = len(contactos) + len(destinatarios_email)
                prefijo = 'Respuesta discrecional' if es_respuesta_discrecional else 'Respuesta'
                message = f'{prefijo} creada exitosamente con {total_dest} destinatarios.'
            
            return JsonResponse({
                'success': True,
                'respuesta_id': respuesta.id,
                'numero_radicado_salida': respuesta.numero_radicado_salida,
                'message': message
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al crear la respuesta: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido.'
    })
def crear_correspondencia_salida_ajax(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            asunto = normalizar_asunto_salida(request.POST.get('asunto') or '')
            cuerpo = request.POST.get('cuerpo')
            destinatarios_contacto_ids = request.POST.getlist('destinatarios_contacto')
            destinatarios_email = [e.strip() for e in request.POST.getlist('destinatarios_email') if e.strip()]
            categoria_contacto_id = (request.POST.get('categoria_contacto_id') or '').strip()
            enviar_inmediatamente = request.POST.get('enviar_inmediatamente') == 'on'

            if not all([asunto, cuerpo]):
                return JsonResponse({'success': False, 'error': 'Asunto y cuerpo son obligatorios.'})

            perfil_usuario = getattr(request.user, 'perfil', None)
            oficina_usuario = getattr(perfil_usuario, 'oficina', None)
            if not oficina_usuario:
                return JsonResponse({'success': False, 'error': 'Usuario sin oficina asignada.'})

            # Límite de carga de adjuntos (25 MB negocio; Postmark limita a 10 MB al enviar).
            from correspondencia.utils.postmark_outbound import (
                mensaje_error_carga_adjuntos,
                salida_adjuntos_upload_limit_bytes,
            )
            adjuntos_files = request.FILES.getlist('adjuntos')
            total_adjuntos_bytes = sum(getattr(f, 'size', 0) for f in adjuntos_files)
            upload_limit_bytes = salida_adjuntos_upload_limit_bytes()
            if total_adjuntos_bytes > upload_limit_bytes:
                return JsonResponse({
                    'success': False,
                    'error': mensaje_error_carga_adjuntos(total_adjuntos_bytes),
                    'error_code': 'ADJUNTOS_SUPERAN_LIMITE_CARGA',
                })

            # No mezclar categoría con selección manual
            if categoria_contacto_id and (destinatarios_contacto_ids or destinatarios_email):
                return JsonResponse({'success': False, 'error': 'No puede seleccionar una categoría y destinatarios manuales al mismo tiempo.'})

            contactos = []
            grupo = None
            if categoria_contacto_id:
                try:
                    grupo = GrupoAgenda.objects.get(id=categoria_contacto_id, oficina_propietaria=oficina_usuario, activo=True)
                    contactos = list(grupo.contactos.filter(correo_electronico__isnull=False).exclude(correo_electronico=''))
                except GrupoAgenda.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'La categoría no existe o no pertenece a su oficina.'})
            else:
                contactos_qs = Contacto.objects.filter(id__in=destinatarios_contacto_ids)
                contactos = list(contactos_qs)

            # NUEVO: normalización + deduplicación por email (backend, silencioso)
            contactos_by_email = { (c.correo_electronico or '').strip().lower(): c for c in contactos if c.correo_electronico }
            manual_validos = []
            for em in destinatarios_email:
                em_norm = em.strip().lower()
                manual_validos.append(em_norm)

            emails_unicos = set(list(contactos_by_email.keys()) + manual_validos)

            # Límite 50 tras dedup
            if len(emails_unicos) > 50:
                return JsonResponse({'success': False, 'error': f'No se permiten más de 50 destinatarios (tras normalización).'})

            # Validar que haya al menos un destinatario válido
            if not emails_unicos:
                return JsonResponse({'success': False, 'error': 'Debe seleccionar al menos un destinatario válido.'})

            from correspondencia.utils.blocked_recipients import validar_emails_destinatario_permitidos
            ok_destinatarios, error_destinatarios = validar_emails_destinatario_permitidos(emails_unicos)
            if not ok_destinatarios:
                return JsonResponse({
                    'success': False,
                    'error': error_destinatarios,
                    'error_code': 'DESTINATARIO_INSTITUCIONAL_BLOQUEADO',
                })

            with transaction.atomic():
                salida = CorrespondenciaSalida.objects.create(
                    respuesta_a=None,
                    usuario_redactor=request.user,
                    oficina_emisora=oficina_usuario,
                    asunto=asunto,
                    cuerpo=cuerpo,
                    destinatario_contacto=contactos[0] if contactos else None,
                    estado='PENDIENTE_APROBACION' if enviar_inmediatamente else 'BORRADOR'
                )

                # Trazabilidad de modalidad de envío
                total_dest = len(emails_unicos)
                if categoria_contacto_id and grupo:
                    salida.envio_tipo = 'GRUPO'
                    salida.envio_grupo = grupo
                    salida.envio_total_destinatarios = total_dest
                    salida.envio_detalle_snapshot = f"{total_dest} contactos (deduplicados)"
                else:
                    salida.envio_tipo = 'INDIVIDUAL' if total_dest <= 1 else 'MULTIPLE_SELECTIVO'
                    salida.envio_total_destinatarios = total_dest
                    salida.envio_detalle_snapshot = f"{total_dest} destinatarios (deduplicados)"
                salida.save(update_fields=['envio_tipo','envio_grupo','envio_total_destinatarios','envio_detalle_snapshot'])

                # Crear destinatarios desde emails_unicos
                dest_objs = []
                for em in emails_unicos:
                    c = contactos_by_email.get(em)
                    dest_objs.append(SalidaDestinatario(
                        correspondencia_salida=salida,
                        contacto=c,
                        email_snapshot=em,
                        nombre_snapshot=(c.nombre_completo if c else em)
                    ))
                
                if dest_objs:
                    SalidaDestinatario.objects.bulk_create(dest_objs)

                archivos = request.FILES.getlist('adjuntos')
                adj_objs = [AdjuntoSalida(correspondencia_salida=salida, archivo=a, nombre_original=a.name) for a in archivos]
                if adj_objs:
                    AdjuntoSalida.objects.bulk_create(adj_objs)

                if enviar_inmediatamente:
                    HistorialSalida.objects.create(
                        correspondencia_salida=salida,
                        tipo_evento='ENVIO_APROBACION',
                        usuario=request.user,
                        descripcion='Salida creada y enviada a aprobación desde modal.'
                    )
                else:
                    HistorialSalida.objects.create(
                        correspondencia_salida=salida,
                        tipo_evento='CREACION',
                        usuario=request.user,
                        descripcion='Borrador de salida creado desde modal.'
                    )

            message = (f'Salida creada con {len(contactos)} destinatarios de la categoría "{grupo.nombre}".' if categoria_contacto_id and grupo 
                       else f'Salida creada con {len(contactos) + len(destinatarios_email)} destinatarios.')

            return JsonResponse({'success': True, 'respuesta_id': salida.id, 'numero_radicado_salida': salida.numero_radicado_salida, 'message': message})

        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al crear correspondencia de salida: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': f'Ocurrió un error inesperado: {str(e)}'})

    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
def buscar_contactos_ajax(request):
    """Vista AJAX para buscar contactos de la agenda global."""
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            entidad_id = request.GET.get('entidad_id')
            query = (request.GET.get('q') or '').strip()
            limit = int(request.GET.get('limit') or 20)
            limit = max(1, min(limit, 50))

            if query and len(query) < 2:
                return JsonResponse({'success': True, 'contactos': []})

            contactos_qs = Contacto.objects.select_related('entidad_externa').order_by('entidad_externa__nombre', 'apellidos', 'nombres')

            if entidad_id and entidad_id.isdigit():
                contactos_qs = contactos_qs.filter(entidad_externa_id=int(entidad_id))

            if query:
                contactos_qs = contactos_qs.filter(
                    Q(nombres__icontains=query) |
                    Q(apellidos__icontains=query) |
                    Q(correo_electronico__icontains=query) |
                    Q(entidad_externa__nombre__icontains=query)
                )
            
            contactos_qs = contactos_qs[:limit]
            contactos_data = [{
                    'id': contacto.id,
                    'nombre_completo': contacto.nombre_completo,
                    'entidad': contacto.entidad_externa.nombre if contacto.entidad_externa else 'Sin entidad',
                    'entidad_id': contacto.entidad_externa.id if contacto.entidad_externa else None,
                    'email': contacto.correo_electronico,
                    'cargo': contacto.cargo
            } for contacto in contactos_qs]
            
            return JsonResponse({'success': True, 'contactos': contactos_data})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error al buscar contactos: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'})


@login_required
def buscar_grupos_ajax(request):
    """Vista AJAX para buscar grupos de agenda por nombre (de la oficina del usuario).
    Retorna: {success, grupos: [{id, nombre, total}]}
    """
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            perfil_usuario = getattr(request.user, 'perfil', None)
            if not perfil_usuario or not perfil_usuario.oficina:
                return JsonResponse({'success': False, 'error': 'Usuario sin oficina asignada.'})

            q = (request.GET.get('q') or '').strip()
            limit = int(request.GET.get('limit') or 10)
            limit = max(1, min(limit, 25))

            if q and len(q) < 2:
                return JsonResponse({'success': True, 'grupos': []})

            qs = GrupoAgenda.objects.filter(oficina_propietaria=perfil_usuario.oficina, activo=True)
            if q:
                qs = qs.filter(nombre__icontains=q)
            qs = qs.order_by('nombre')[:limit]

            grupos = [{'id': g.id, 'nombre': g.nombre, 'total': g.contactos.count()} for g in qs]
            return JsonResponse({'success': True, 'grupos': grupos})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error al buscar grupos: {str(e)}'})

    return JsonResponse({'success': False, 'error': 'Método no permitido.'})

@login_required
def buscar_entidades_ajax(request):
    """Vista AJAX para buscar entidades externas."""
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            # Buscar todas las entidades
            entidades = EntidadExterna.objects.all().order_by('nombre')
            
            entidades_data = []
            for entidad in entidades:
                entidades_data.append({
                    'id': entidad.id,
                    'nombre': entidad.nombre,
                    'nit': entidad.nit,
                    'dominio': entidad.dominio,
                    'dominios': entidad.dominios_configurados(),
                })
            
            return JsonResponse({
                'success': True,
                'entidades': entidades_data
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al buscar entidades: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido.'
    })

def crear_contacto_automatico_si_no_existe(correspondencia, oficina_destino=None):
    """
    Crea automáticamente un contacto en la agenda de la oficina si no existe.
    Solo se ejecuta cuando se distribuye correspondencia a un usuario.
    
    Validaciones inteligentes:
    - Solo si el remitente tiene email válido
    - Solo si no existe ya en la agenda de la oficina
    - Validar que la entidad sea razonable (no spam)
    - Crear entidad si no existe
    """
    if not correspondencia.remitente:
        return None
    
    remitente = correspondencia.remitente
    
    # 1. Validar que tenga email (esencial para respuestas)
    if not remitente.correo_electronico or '@' not in remitente.correo_electronico:
        return None
    
    # 2. Verificar si ya existe un contacto global con este correo
    contacto_existente = Contacto.objects.filter(
        correo_electronico__iexact=remitente.correo_electronico
    ).first()
    
    if contacto_existente:
        return contacto_existente
    
    # 3. Validar que la entidad no sea spam (nombres muy cortos o sospechosos)
    entidad_nombre = remitente.entidad_externa.nombre if remitente.entidad_externa else "Sin entidad"
    if len(entidad_nombre) < 3 or entidad_nombre.lower() in ['test', 'spam', 'fake', 'prueba']:
        return None
    
    # 4. Buscar entidad existente por nombre similar (evitar duplicados)
    entidad_existente = None
    if remitente.entidad_externa:
        # Buscar por nombre exacto o similar
        entidad_existente = EntidadExterna.objects.filter(
            nombre__iexact=remitente.entidad_externa.nombre
        ).first()
        
        if not entidad_existente:
            # Buscar por similitud (nombres que contengan palabras clave)
            palabras_clave = remitente.entidad_externa.nombre.lower().split()
            for palabra in palabras_clave:
                if len(palabra) > 3:  # Ignorar palabras muy cortas
                    entidad_similar = EntidadExterna.objects.filter(
                        nombre__icontains=palabra
                    ).first()
                    if entidad_similar:
                        entidad_existente = entidad_similar
                        break
    
    # 5. Crear entidad si no existe
    if not entidad_existente and remitente.entidad_externa:
        try:
            entidad_existente = EntidadExterna.objects.create(
                nombre=remitente.entidad_externa.nombre,
                dominio=remitente.entidad_externa.dominio
            )
        except Exception:
            # Si falla la creación, usar la entidad por defecto
            entidad_existente = EntidadExterna.get_entidad_por_defecto()
    elif not entidad_existente:
        entidad_existente = EntidadExterna.get_entidad_por_defecto()
    
    # 6. Crear el contacto global
    try:
        nuevo_contacto = Contacto.objects.create(
            entidad_externa=entidad_existente,
            nombres=remitente.nombres,
            apellidos=remitente.apellidos,
            cargo=remitente.cargo,
            correo_electronico=remitente.correo_electronico,
            telefono_contacto=remitente.telefono_contacto
        )
        
        # Log para auditoría
        print(f"Contacto global creado automáticamente: {nuevo_contacto.nombre_completo} ({nuevo_contacto.correo_electronico})")
        
        return nuevo_contacto
        
    except Exception as e:
        # Si falla la creación, no interrumpir el flujo principal
        print(f"Error al crear contacto automático: {str(e)}")
        return None

@login_required
def contactos_automaticos(request):
    """Muestra los contactos que fueron creados automáticamente al distribuir correspondencia."""
    # Obtener la oficina del usuario
    perfil_usuario = getattr(request.user, 'perfil', None)
    if not perfil_usuario or not perfil_usuario.oficina:
        messages.error(request, "No tienes una oficina asignada.")
        return redirect('correspondencia:welcome')
    
    # Mostrar todos los contactos globales recientes (creados automáticamente o por cualquier usuario)
    contactos_automaticos = Contacto.objects.select_related('entidad_externa').order_by('-id')
    
    # Búsqueda
    query = request.GET.get('q')
    if query:
        contactos_automaticos = contactos_automaticos.filter(
            Q(nombres__icontains=query) | 
            Q(apellidos__icontains=query) | 
            Q(correo_electronico__icontains=query) |
            Q(entidad_externa__nombre__icontains=query)
        )
    
    # Paginación
    paginator = Paginator(contactos_automaticos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'contactos': page_obj,
        'titulo_pagina': 'Contactos de Agenda',
        'subtitulo': 'Agenda global de contactos',
        'total_contactos': contactos_automaticos.count(),
    }
    
    return render(request, 'correspondencia/admin/contactos_automaticos.html', context)

 
"""
Vistas del módulo de correspondencia.

Este módulo contiene todas las vistas necesarias para gestionar el flujo completo
de correspondencia entrante, incluyendo:

- Radicación manual y automática de correspondencia
- Gestión de bandejas de trabajo (personal, ventanilla, clasificados)
- Procesamiento de correos electrónicos entrantes
- Gestión de correspondencia saliente
- APIs para funcionalidades dinámicas (SLA, carga de subseries)
- Control de acceso basado en roles y permisos
Las vistas implementan:
- Autenticación y autorización requerida
- Validación de formularios
- Manejo de transacciones de base de datos
- Paginación y filtrado
- Respuestas JSON para funcionalidades AJAX
- Integración con sistema de mensajes de Django

Autor: Sistema de Gestión Documental
Fecha: 2025
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.contrib import messages
from .forms import CorrespondenciaForm, ContactoForm, CompartirCorrespondenciaForm, ManualRadicacionCorreoForm, EntidadExternaForm, RespuestaCorrespondenciaForm, AprobarRechazarRespuestaForm, HistorialFilterForm
from .models import Correspondencia, HistorialCorrespondencia, Contacto, CorreoEntrante, OficinaProductora, SerieDocumental, SubserieDocumental, AdjuntoCorreoEntrante, AdjuntoCorreo, DistribucionInternaUsuario, EntidadExterna, CorrespondenciaSalida, AdjuntoSalida, HistorialSalida
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, View
from django.utils import timezone
from django.db import transaction
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from documentos.models import PerfilUsuario # Necesario para obtener perfil
from django.contrib.auth.models import User
from .models import DistribucionInternaUsuario
from django.db.models import Q, Count, Case, When, BooleanField # Para consultas OR y anotaciones
from django.core.management import call_command
from django.core.files.base import ContentFile
import traceback
from django.db.utils import IntegrityError
import os
from django.core.exceptions import ValidationError
from django.http import JsonResponse
from django.http import HttpResponse
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from django.db.models import Max
# justo debajo de los imports de modelos
# views.py (encabezado)
from documentos.models import SubserieDocumental as SubserieDoc
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from django.core.mail import EmailMessage # <--- Añadir esta importación
from django.urls import reverse_lazy
from django.db.models.functions import Coalesce
from django.db.models import F # <-- Importar F
from django.db.models import Q # Importar Q para búsquedas complejas
from django.db import models # <-- IMPORTAR BASE MODELS
from django.http import JsonResponse
from .modelos_minimos_sla import SubserieTramite
from .utils_sla import get_cutoff_time, aplicar_corte, sumar_habiles
from django.utils import timezone
import datetime
from io import BytesIO

DOMINIOS_GENERICOS = [
    'hotmail.com', 'yahoo.com', 'outlook.com', 'icloud.com'
]



# Create your views here.

# --- Vista para la Bandeja de Correos Clasificados ---
class EsVentanillaMixin(UserPassesTestMixin):
    """Mixin para verificar si el usuario pertenece al grupo 'Ventanilla'."""
    def test_func(self):
        # Asume que tienes un grupo llamado 'Ventanilla'. Ajusta si es necesario.
        return self.request.user.groups.filter(name='Ventanilla').exists()

    def handle_no_permission(self):
        messages.error(self.request, "No tienes permiso para acceder a esta sección.")
        # Redirigir a una página segura, por ejemplo, la home o welcome
        return redirect('correspondencia:welcome') # Ajusta el nombre de la URL si es diferente

class BandejaClasificadosView(LoginRequiredMixin, EsVentanillaMixin, ListView):
    model = CorreoEntrante
    template_name = 'correspondencia/admin/bandeja_clasificados.html'
    context_object_name = 'correos_clasificados'
    # paginate_by = 15 # Quitar paginación del servidor

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def radicar_correspondencia(request, correo_id=None):
    """
    Vista para radicar correspondencia manualmente.
    
    Esta vista permite a usuarios del grupo 'Ventanilla' radicar correspondencia
    tanto manualmente como desde correos electrónicos pre-clasificados.
    
    Funcionalidades:
    - Radicación manual de correspondencia física/electrónica
    - Pre-llenado automático desde correos clasificados por IA
    - Validación de formularios con manejo de errores
    - Creación automática de historial de radicación
    - Cálculo automático de SLA (plazos de respuesta)
    - Vinculación con correos origen cuando aplica
    
    Args:
        request: Objeto HttpRequest de Django
        correo_id (int, optional): ID del correo entrante para pre-llenar datos
        
    Returns:
        HttpResponse: Renderiza formulario de radicación o redirige tras éxito
        
    Raises:
        404: Si el correo_id no existe o ya fue radicado
        PermissionError: Si el usuario no pertenece al grupo Ventanilla
    """
    correo_origen = None
    initial_data = {}

    if correo_id:
        correo_origen = get_object_or_404(CorreoEntrante, pk=correo_id, procesado=True, radicado_asociado__isnull=True, urgencia_asociada__isnull=True)
        # Pre-llenar datos iniciales del formulario
        initial_data = {
            'asunto': correo_origen.asunto,
            'medio_recepcion': 'ELECTRONICO', # Asumir electrónico si viene de correo
            # Intentar encontrar o sugerir contacto basado en remitente
            # 'remitente': encontrar_o_crear_contacto(correo_origen.remitente),
            'oficina_destino': correo_origen.oficina_clasificada,
            'serie': correo_origen.serie_clasificada,
            'subserie': correo_origen.subserie_clasificada,
            # Otros campos que puedas pre-llenar
        }
        # Buscar/Crear contacto (ejemplo simple)
        try:
            # Asumiendo que Contacto tiene un campo 'correo_electronico' único o prioritario
            contacto, created = Contacto.objects.get_or_create(
                correo_electronico=correo_origen.remitente,
                defaults={'entidad': f'Entidad de {correo_origen.remitente}'} # Valor por defecto simple
            )
            initial_data['remitente'] = contacto
        except Exception as e:
            messages.warning(request, f"No se pudo encontrar o crear automáticamente el contacto para {correo_origen.remitente}. Por favor, selecciónelo manualmente. Error: {e}")

    # Pasar initial_data al formulario
    form = CorrespondenciaForm(request.POST or None, initial=initial_data)
    # form_contacto = ContactoForm(request.POST or None, prefix="contacto") # Si tienes form aparte para contacto

    if request.method == 'POST':
        # Validar y guardar como antes...
        if form.is_valid(): # and form_contacto.is_valid():
            try:
                with transaction.atomic():
                    # Crear o usar contacto existente
                    # contacto = form_contacto.save() # O obtenerlo del form principal si está integrado
                    # correspondencia.remitente = contacto
                    correspondencia = form.save(commit=False)
                    correspondencia.usuario_radicador = request.user
                    # Asegúrate que los campos obligatorios estén
                    if not correspondencia.oficina_destino or not correspondencia.serie:
                        messages.error(request, "La Oficina Destino y la Serie son obligatorias.")
                        # Re-renderizar formulario con errores
                        context = {'form': form, 'titulo_pagina': 'Radicar Nueva Correspondencia'}
                        return render(request, 'correspondencia/admin/radicar_form.html', context)

                    correspondencia.save() # Guardar para obtener ID

                    adjuntos_guardados = _guardar_adjuntos_radicacion_fisica(correspondencia, request)
                    
                    # Crear historial inicial
                    HistorialCorrespondencia.objects.create(
                        correspondencia=correspondencia,
                        evento='RADICADA',
                        usuario=request.user,
                        descripcion=f"Radicada por {request.user.username} desde correo ID {correo_origen.id}" if correo_origen else f"Radicada manualmente por {request.user.username}" + (f" con {adjuntos_guardados} adjuntos" if adjuntos_guardados > 0 else "")
                    )
                    
                    # Marcar el correo origen como radicado
                    if correo_origen:
                        correo_origen.radicado_asociado = correspondencia
                        correo_origen.save(update_fields=['radicado_asociado'])

                        # Copiar adjuntos del CorreoEntrante a AdjuntoCorreo
                        for adj_origen in correo_origen.adjuntos.all():
                            adj_destino = AdjuntoCorreo(
                                correspondencia=correspondencia,
                                nombre_original=adj_origen.nombre_original,
                                tipo_mime=adj_origen.tipo_mime
                            )
                            if adj_origen.archivo:
                                try:
                                    file_content = ContentFile(adj_origen.archivo.read())
                                    file_name = adj_origen.nombre_original or os.path.basename(adj_origen.archivo.name)
                                    adj_destino.archivo.save(file_name, file_content, save=False)
                                except Exception as e:
                                    print(f"Error copiando archivo adjunto {adj_origen.id}: {e}")
                                    messages.warning(request, f"No se pudo copiar el adjunto: {adj_origen.nombre_original}")
                            adj_destino.save()

                    messages.success(request, f"Correspondencia {correspondencia.numero_radicado} radicada exitosamente.")
                    # Redirigir a la lista de pendientes, que es la acción siguiente para Ventanilla
                    return redirect('correspondencia:pendientes_distribuir') 

            except Exception as e:
                messages.error(request, f"Error al radicar la correspondencia: {e}")
                # Considera loggear el error completo: logger.error(f"Error radicando: {e}", exc_info=True)

        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")

    # Renderizar formulario GET o si hubo error POST
    context = {
        'form': form,
        # 'form_contacto': form_contacto,
        'titulo_pagina': 'Radicar Nueva Correspondencia' + (f" (Desde Correo ID: {correo_id})" if correo_id else ""),
        'correo_origen': correo_origen,
        'sla_enabled': True,  # Habilitar funcionalidad SLA
    }
    return render(request, 'correspondencia/admin/radicar_form.html', context)
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def lista_pendientes_distribuir(request):
    """Muestra la lista de correspondencia radicada pendiente de asignar a usuario (OPTIMIZADA CON PAGINACIÓN)."""
    
    # Base queryset: únicamente RADICADA (optimizado para reducir carga)
    qs = (
        Correspondencia.objects
        .filter(estado='RADICADA')
        .select_related(
            'oficina_destino',
            'remitente',
            'remitente__entidad_externa',
        )
        .order_by('-fecha_radicacion')
    )

    # --- Filtros GET (según estándares) ---
    search_term = (request.GET.get('search_term') or '').strip()
    oficina = (request.GET.get('oficina') or '').strip()
    serie = (request.GET.get('serie') or '').strip()
    subserie = (request.GET.get('subserie') or '').strip()
    serie_id = request.GET.get('serie_id') or ''
    subserie_id = request.GET.get('subserie_id') or ''
    fecha_inicio = request.GET.get('fecha_inicio') or ''
    fecha_fin = request.GET.get('fecha_fin') or ''
    medio = (request.GET.get('medio_recepcion') or '').strip()

    if search_term:
        qs = qs.filter(
            Q(numero_radicado__icontains=search_term)
            | Q(asunto__icontains=search_term)
            | Q(remitente__nombres__icontains=search_term)
            | Q(remitente__apellidos__icontains=search_term)
            | Q(remitente__entidad_externa__nombre__icontains=search_term)
        )

    if oficina:
        qs = qs.filter(oficina_destino__nombre__icontains=oficina)

    if serie_id.isdigit():
        qs = qs.filter(serie_id=int(serie_id))
    elif serie:
        qs = qs.filter(serie__nombre__icontains=serie)

    if subserie_id.isdigit():
        qs = qs.filter(subserie_id=int(subserie_id))
    elif subserie:
        qs = qs.filter(subserie__nombre__icontains=subserie)

    # Fechas (YYYY-MM-DD)
    if fecha_inicio:
        try:
            qs = qs.filter(fecha_radicacion__date__gte=fecha_inicio)
        except Exception:
            pass
    if fecha_fin:
        try:
            qs = qs.filter(fecha_radicacion__date__lte=fecha_fin)
        except Exception:
            pass

    if medio:
        qs = qs.filter(medio_recepcion__iexact=medio)

    # --- Paginación ---
    paginator = Paginator(qs, 25)  # 25 items por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'correspondencias': page_obj,  # retrocompatibilidad en plantilla
        'page_obj': page_obj,
        'titulo_pagina': 'Correspondencia Pendiente de Asignar',
        # Valores de filtros para rehidratación en la UI
        'filtro_search_term': search_term,
        'filtro_oficina': oficina,
        'filtro_serie': serie,
        'filtro_subserie': subserie,
        'filtro_serie_id': serie_id,
        'filtro_subserie_id': subserie_id,
        'filtro_fecha_inicio': fecha_inicio,
        'filtro_fecha_fin': fecha_fin,
        'filtro_medio': medio,
    }
    return render(request, 'correspondencia/admin/lista_pendientes.html', context)

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def distribuir_correspondencia(request, pk):
    """Distribuye una correspondencia a un usuario específico y crea el registro inicial de distribución."""
    correspondencia = get_object_or_404(Correspondencia, pk=pk)
    
    # Obtener usuarios de la oficina destino (con perfil asociado)
    usuarios_oficina = User.objects.filter(
        perfil__oficina=correspondencia.oficina_destino
    ).select_related('perfil') # Incluir perfil para asegurar que existe
    # Otras oficinas disponibles para pre-redistribuir (solo lectura)
    otras_oficinas_opciones = (
        OficinaProductora.objects
        .order_by('nombre')
        .exclude(pk=correspondencia.oficina_destino_id)
        .select_related('unidad_administrativa')
    )
    
    if request.method == 'POST':
        usuario_id = request.POST.get('usuario')
        observaciones = request.POST.get('observaciones', 'Asignación inicial desde Ventanilla') # Observación por defecto
        compartir_oficina = request.POST.get('compartir_oficina') == 'on'
        otras_oficinas_ids = request.POST.getlist('otras_oficinas')
        
        if not usuario_id:
             messages.error(request, "Debe seleccionar un usuario.")
             # Re-renderizar con error
             # Convertir IDs a strings para comparación en template
             selected_oficinas = [str(id) for id in otras_oficinas_ids]
             context = {
                 'correspondencia': correspondencia,
                 'usuarios_oficina': usuarios_oficina,
                 'titulo_pagina': 'Asignar Correspondencia a Usuario',
                 'otras_oficinas_opciones': otras_oficinas_opciones,
                 'selected_oficinas': selected_oficinas,
                 'selected_usuario': usuario_id,
                 'observaciones_text': observaciones,
                 'compartir_oficina_checked': compartir_oficina
             }
             return render(request, 'correspondencia/admin/distribuir_correspondencia.html', context)

        try:
            usuario_destino = User.objects.get(id=usuario_id)
            # Verificar que el usuario pertenece a la oficina destino (doble chequeo)
            if not hasattr(usuario_destino, 'perfil') or usuario_destino.perfil.oficina != correspondencia.oficina_destino:
                messages.error(request, "El usuario seleccionado no pertenece a la oficina destino o no tiene perfil.")
                return redirect('correspondencia:distribuir_correspondencia', pk=pk)
            
            with transaction.atomic():
                # 1. Actualizar la correspondencia
                correspondencia.usuario_destino_inicial = usuario_destino
                correspondencia.estado = 'ASIGNADA_USUARIO'
                correspondencia.save(update_fields=['usuario_destino_inicial', 'estado'])
                
                # 2. Crear el registro de DistribucionInternaUsuario para el asignado inicial
                # Esto centraliza el seguimiento de quién debe verla y si la ha leído.
                distribucion_inicial, created = DistribucionInternaUsuario.objects.update_or_create(
                    correspondencia=correspondencia,
                    usuario_asignado=usuario_destino,
                    defaults={
                        'asignado_por': request.user, # Quién hizo la asignación (ventanilla)
                        'fecha_asignacion': timezone.now(),
                        'observaciones': observaciones,
                        'leido': False # Inicialmente no leído
                    }
                )
                
                # 3. Compartir con toda la oficina (opcional)
                if compartir_oficina:
                    for u in usuarios_oficina:
                        DistribucionInternaUsuario.objects.update_or_create(
                            correspondencia=correspondencia,
                            usuario_asignado=u,
                            defaults={
                                'asignado_por': request.user,
                                'fecha_asignacion': timezone.now(),
                                'observaciones': f"Compartido por ventanilla. {observaciones}".strip(),
                                'leido': False
                            }
                        )
                    # Historial
                    HistorialCorrespondencia.objects.create(
                        correspondencia=correspondencia,
                        evento='REDISTRIBUIDA_INTERNA',
                        usuario=request.user,
                        descripcion=f"Compartida con toda la oficina {correspondencia.oficina_destino.nombre} desde Ventanilla."
                    )

                # 4. Pre-redistribuir a otras oficinas (acceso solo lectura)
                oficinas_agregadas = 0
                if otras_oficinas_ids:
                    oficinas_seleccionadas = OficinaProductora.objects.filter(pk__in=otras_oficinas_ids)
                    for oficina in oficinas_seleccionadas:
                        # Verificar si se marcó explícitamente "Compartir con toda la oficina"
                        share_all_key = f'share_all_{oficina.id}'
                        compartir_con_todos = request.POST.get(share_all_key) == 'on'
                        es_solo_lider = not compartir_con_todos
                        
                        # Verificar si se marcó "Permitir responder"
                        puede_responder_key = f'puede_responder_{oficina.id}'
                        permiso_respuesta = request.POST.get(puede_responder_key) == 'on'

                        acceso, creado = AccesoCorrespondenciaOficina.objects.update_or_create(
                            correspondencia=correspondencia,
                            oficina=oficina,
                            defaults={
                                'compartido_por': request.user,
                                'observaciones': f"Pre-redistribución desde Ventanilla. {observaciones}".strip(),
                                'solo_lider': es_solo_lider,
                                'puede_responder': permiso_respuesta
                            }
                        )
                        if creado:
                            oficinas_agregadas += 1
                            
                            visibilidad_msg = " (Solo Líderes)" if es_solo_lider else " (Toda la oficina)"
                            
                            HistorialCorrespondencia.objects.create(
                                correspondencia=correspondencia,
                                evento='COMPARTIDA_OFICINA',
                                usuario=request.user,
                                descripcion=f"Acceso de solo lectura otorgado a la oficina {oficina.nombre}{visibilidad_msg}."
                            )
                            crear_notificaciones_acceso_oficina(correspondencia, oficina, request.user, observaciones)
                
                # 5. Registrar en el historial general de asignación
                HistorialCorrespondencia.objects.create(
                    correspondencia=correspondencia,
                    evento='ASIGNADA_USUARIO',
                    usuario=request.user,
                    descripcion=f"Asignada a {usuario_destino.get_full_name() or usuario_destino.username}"
                )
            
            extra = []
            if compartir_oficina:
                extra.append("compartida con la oficina")
            if otras_oficinas_ids:
                extra.append(f"pre-redistribuida a {len(otras_oficinas_ids)} oficina(s)")
            suffix = f" ({', '.join(extra)})" if extra else ""
            messages.success(request, f'Correspondencia {correspondencia.numero_radicado} asignada exitosamente a {usuario_destino.get_full_name() or usuario_destino.username}{suffix}')
            return redirect('correspondencia:pendientes_distribuir')
            
        except User.DoesNotExist:
            messages.error(request, "Usuario seleccionado no válido.")
            # Re-renderizar con valores previos
            selected_oficinas = [str(id) for id in request.POST.getlist('otras_oficinas')]
            context = {
                'correspondencia': correspondencia,
                'usuarios_oficina': usuarios_oficina,
                'titulo_pagina': 'Asignar Correspondencia a Usuario',
                'otras_oficinas_opciones': otras_oficinas_opciones,
                'selected_oficinas': selected_oficinas,
                'selected_usuario': request.POST.get('usuario'),
                'observaciones_text': request.POST.get('observaciones', ''),
                'compartir_oficina_checked': request.POST.get('compartir_oficina') == 'on'
            }
            return render(request, 'correspondencia/admin/distribuir_correspondencia.html', context)
        except Exception as e:
            messages.error(request, f"Error al asignar la correspondencia: {str(e)}")
            # Re-renderizar con valores previos
            selected_oficinas = [str(id) for id in request.POST.getlist('otras_oficinas')]
            context = {
                'correspondencia': correspondencia,
                'usuarios_oficina': usuarios_oficina,
                'titulo_pagina': 'Asignar Correspondencia a Usuario',
                'otras_oficinas_opciones': otras_oficinas_opciones,
                'selected_oficinas': selected_oficinas,
                'selected_usuario': request.POST.get('usuario'),
                'observaciones_text': request.POST.get('observaciones', ''),
                'compartir_oficina_checked': request.POST.get('compartir_oficina') == 'on'
            }
            return render(request, 'correspondencia/admin/distribuir_correspondencia.html', context)
    
    # Contexto para el método GET
    context = {
        'correspondencia': correspondencia,
        'usuarios_oficina': usuarios_oficina,
        'titulo_pagina': 'Asignar Correspondencia a Usuario',
        'otras_oficinas_opciones': otras_oficinas_opciones,
        'selected_oficinas': [],  # Lista vacía para GET
        'selected_usuario': None,
        'observaciones_text': '',
        'compartir_oficina_checked': False
    }
    return render(request, 'correspondencia/admin/distribuir_correspondencia.html', context)

# --- Vistas para Contactos --- 

@login_required
# @permission_required('correspondencia.view_contacto', raise_exception=True) # Permiso opcional
@login_required
def home_view(request):
    return render(request, 'correspondencia/admin/lista_pendientes.html')

@login_required
def ver_perfil(request):
    return render(request, 'correspondencia/admin/lista_pendientes.html')

@login_required
def bandeja_entrada(request):
    """Muestra la correspondencia asignada directamente o compartida con el usuario."""
    usuario_actual = request.user
    perfil_usuario = getattr(usuario_actual, 'perfil', None)
    nombre_oficina = "Oficina no asignada"
    correspondencias_list = Correspondencia.objects.none() # Queryset vacío por defecto
    load_success = False

    if perfil_usuario and perfil_usuario.oficina:
        nombre_oficina = perfil_usuario.oficina.nombre
        try:
            # 1. Correspondencia asignada directamente al usuario
            directamente_asignada = Q(usuario_destino_inicial=usuario_actual)
            
            # 2. Correspondencia compartida con el usuario a través de DistribucionInternaUsuario
            compartida_con_usuario = Q(distribuciones_internas__usuario_asignado=usuario_actual)
            
            # Combinar ambos filtros con OR
            correspondencias_list = Correspondencia.objects.filter(
                directamente_asignada | compartida_con_usuario,
                oficina_destino=perfil_usuario.oficina # Asegurar que sea de su oficina
            ).distinct().select_related(
                'remitente', 
                'oficina_destino',
                'usuario_destino_inicial' # Puede ser útil mostrar quién fue el asignado inicial
            ).prefetch_related(
                'adjuntos_correo', # Para mostrar adjuntos si es necesario
                'distribuciones_internas' # Para posible lógica adicional
            ).order_by('-fecha_radicacion') # O por fecha de asignación/compartido?

            load_success = True
            print(f"Usuario: {usuario_actual}, Oficina: {nombre_oficina}, Correspondencias encontradas: {correspondencias_list.count()}")

        except AttributeError as e:
             print(f"Error de atributo en bandeja_entrada (posiblemente perfil): {e}")
             messages.error(request, "Error al cargar datos de usuario/oficina.")
             # load_success permanece False
        except Exception as e:
            print(f"Error capturado en bandeja_entrada: {e}")
            messages.error(request, f"Ocurrió un error inesperado al cargar la bandeja: {e}")
            # load_success permanece False
    else:
        messages.warning(request, "No tienes un perfil o una oficina asignada para ver tu bandeja de entrada.")
        # load_success permanece False

    # Configurar el contexto final
    context = {
        'titulo_pagina': f"Bandeja de Entrada - {nombre_oficina}",
        'correspondencias': correspondencias_list, 
        'nombre_oficina': nombre_oficina,
        'load_success': load_success
    }

    print("Contexto antes de render:", context)
    return render(request, 'correspondencia/admin/bandeja_entrada.html', context)
@login_required
def detalle_correspondencia(request, pk):
    correspondencia = Correspondencia.objects.select_related(
        'remitente', 'oficina_destino', 'serie', 'subserie', 'usuario_radicador'
    ).prefetch_related(
        'adjuntos_correo',
        'correo_origen__adjuntos',
        'historial__usuario',
        'distribuciones_internas__usuario_asignado',
        'accesos_oficinas__oficina',
        'accesos_oficinas__compartido_por'
    ).filter(pk=pk).first()

    salida = None
    if correspondencia:
        salida = correspondencia.respuestas_salientes.first()
    else:
        salida = CorrespondenciaSalida.objects.select_related(
            'respuesta_a__oficina_destino', 'respuesta_a__remitente',
            'usuario_redactor', 'usuario_aprobador', 'destinatario_contacto'
        ).prefetch_related('adjuntos', 'historial__usuario').filter(pk=pk).first()

    if not correspondencia and not salida:
        raise Http404("Correspondencia no encontrada.")

    # Si tenemos correspondencia entrante:
    historial = []
    adjuntos = []
    correo_origen_obj = None
    cuerpo_html_renderizado = ''
    mostrar_cuerpo_correo_origen = False
    puede_compartir = False
    puede_responder = False
    usuarios_que_leyeron = []

    if correspondencia:
        usuario_actual = request.user
        perfil_usuario = getattr(usuario_actual, 'perfil', None)
        is_ventanilla_or_admin = usuario_actual.groups.filter(name__in=['Ventanilla', 'Admin']).exists() or usuario_actual.is_superuser
        es_de_oficina_destino = perfil_usuario and perfil_usuario.oficina == correspondencia.oficina_destino
        es_asignado_inicial = correspondencia.usuario_destino_inicial == usuario_actual
        fue_compartido_con_usuario = DistribucionInternaUsuario.objects.filter(correspondencia=correspondencia, usuario_asignado=usuario_actual).exists()
        acceso_interoficina = None
        tiene_acceso_interoficina = False
        
        # --- BARRERA DE SEGURIDAD ABSOLUTA ---
        # Si la oficina del usuario tiene un acceso restringido (solo_lider=True),
        # bloqueamos inmediatamente a los no-líderes, a menos que tengan permisos especiales.
        if perfil_usuario and perfil_usuario.oficina:
            acceso_restrictivo = correspondencia.accesos_oficinas.filter(
                oficina=perfil_usuario.oficina,
                solo_lider=True
            ).exists()
            
            if acceso_restrictivo:
                soy_lider = usuario_actual.groups.filter(name='Lider de Oficina').exists()
                if not soy_lider:
                    # Excepciones: Ventanilla, Admin, Radicador, o Asignación Directa
                    es_privilegiado = (
                        correspondencia.usuario_radicador == usuario_actual or
                        is_ventanilla_or_admin or
                        DistribucionInternaUsuario.objects.filter(
                            correspondencia=correspondencia, usuario_asignado=usuario_actual
                        ).exists()
                    )
                    
                    if not es_privilegiado:
                        messages.error(request, "ACCESO DENEGADO: Documento restringido a líderes.")
                        redirect_url = 'correspondencia:bandeja_personal' if perfil_usuario else 'correspondencia:welcome'
                        return redirect(redirect_url)

        # Continuar con lógica normal de detección de acceso para mostrar info correcta
        if perfil_usuario and perfil_usuario.oficina:
            for acceso in correspondencia.accesos_oficinas.all():
                if acceso.oficina_id == perfil_usuario.oficina_id:
                    # Aquí ya sabemos que si era restrictivo y pasó la barrera, es porque puede verlo
                    acceso_interoficina = acceso
                    tiene_acceso_interoficina = True
                    break

        puede_ver = (
            es_de_oficina_destino or correspondencia.usuario_radicador == usuario_actual or
            is_ventanilla_or_admin or fue_compartido_con_usuario or tiene_acceso_interoficina
        )

        if not puede_ver:
            messages.error(request, "No tienes permiso para ver esta correspondencia.")
            redirect_url = 'correspondencia:bandeja_personal' if perfil_usuario else 'correspondencia:welcome'
            return redirect(redirect_url)

        if es_de_oficina_destino and correspondencia.estado != 'RADICADA':
            try:
                distribucion = DistribucionInternaUsuario.objects.get(correspondencia=correspondencia, usuario_asignado=usuario_actual)
                if not distribucion.leido:
                    distribucion.leido = True
                    distribucion.fecha_lectura = timezone.now()
                    distribucion.save(update_fields=['leido', 'fecha_lectura'])

                    leidos_count = DistribucionInternaUsuario.objects.filter(correspondencia=correspondencia, leido=True).count()
                    if leidos_count == 1 and correspondencia.estado != 'LEIDA':
                        with transaction.atomic():
                            correspondencia.estado = 'LEIDA'
                            correspondencia.leido_por_oficina = True
                            correspondencia.save(update_fields=['estado', 'leido_por_oficina'])
                            HistorialCorrespondencia.objects.create(
                                correspondencia=correspondencia,
                                evento='LEIDA',
                                usuario=usuario_actual,
                                descripcion=f"Leída por primera vez por {usuario_actual.username}."
                            )
            except:
                pass

        if is_ventanilla_or_admin:
            puede_compartir = True
            puede_responder = True

        if es_asignado_inicial or fue_compartido_con_usuario:
            puede_compartir = True
            puede_responder = True

        if es_de_oficina_destino and not puede_compartir:
            puede_compartir = True
            # Si es de la oficina destino, puede responder por defecto
            if not puede_responder:
                puede_responder = True

        # NUEVA LÓGICA: Si tiene acceso interoficina con permiso de respuesta
        if tiene_acceso_interoficina and acceso_interoficina:
            # Marcar lectura por oficina externa con acceso de solo lectura
            acceso_interoficina.marcar_leido()
            # Si tiene permiso explícito de respuesta, otorgarlo
            if acceso_interoficina.puede_responder:
                puede_responder = True
                puede_compartir = True  # Si puede responder, también puede compartir

        form_redistribuir = None
        if puede_compartir:
            if perfil_usuario and perfil_usuario.oficina:
                oficina_origen_form = perfil_usuario.oficina
            else:
                oficina_origen_form = correspondencia.oficina_destino

            form_redistribuir = CompartirOtrasOficinasForm(
                correspondencia=correspondencia,
                oficina_origen=oficina_origen_form
            )

        historial = correspondencia.historial.all()
        adjuntos = correspondencia.adjuntos_correo.all()
        correo_origen_obj = correspondencia.correo_origen.first()
        if correo_origen_obj:
            cuerpo_html_renderizado = correo_origen_obj.obtener_cuerpo_html_renderizado()
            mostrar_cuerpo_correo_origen = bool(cuerpo_html_renderizado or correo_origen_obj.cuerpo_texto)
        usuarios_que_leyeron = User.objects.filter(
            correspondencia_asignada__correspondencia=correspondencia,
            correspondencia_asignada__leido=True
        ).distinct()

        # Detectar si la respuesta de salida fue enviada por una categoría (GrupoAgenda)
        categoria_envio = None
        destinatarios_salientes = None
        try:
            if salida:
                destinatarios_salientes = salida.destinatarios.select_related('contacto')
                destinatarios_contact_ids = list(destinatarios_salientes.values_list('contacto_id', flat=True))
                total_destinatarios = len(destinatarios_contact_ids)
                if total_destinatarios > 0 and salida.oficina_emisora_id:
                    grupos_posibles = (
                        GrupoAgenda.objects
                        .filter(
                            oficina_propietaria_id=salida.oficina_emisora_id,
                            activo=True,
                            contactos__in=destinatarios_contact_ids
                        )
                        .annotate(
                            match_count=Count('contactos', filter=Q(contactos__in=destinatarios_contact_ids), distinct=True)
                        )
                        .filter(match_count=total_destinatarios)
                        .distinct()
                    )
                    categoria_envio = grupos_posibles.first() if grupos_posibles.exists() else None
        except Exception:
            categoria_envio = None

    puede_respuesta_discrecional = bool(
        correspondencia and
        not correspondencia.requiere_respuesta and
        puede_responder and
        usuario_tiene_permiso_respuesta_discrecional(request.user)
    )

    context = {
        'titulo_pagina': f"Detalle Radicado: {(correspondencia.numero_radicado if correspondencia else salida.numero_radicado_salida)}",
        'correspondencia': correspondencia,
        'salida': salida,
        'historial': historial,
        'adjuntos': adjuntos,
        'correo_origen_obj': correo_origen_obj,
        'cuerpo_html_renderizado': cuerpo_html_renderizado,
        'mostrar_cuerpo_correo_origen': mostrar_cuerpo_correo_origen,
        'usuarios_que_leyeron': usuarios_que_leyeron,
        'puede_compartir': puede_compartir,
        'puede_responder': puede_responder,
        'puede_respuesta_discrecional': puede_respuesta_discrecional,
        'es_respuesta_discrecional': puede_respuesta_discrecional,
        'puede_redistribuir': puede_compartir,
        'accesos_oficinas': correspondencia.accesos_oficinas.all() if correspondencia else [],
        'acceso_interoficina': acceso_interoficina,  # Para verificar si el usuario tiene acceso interoficina
        'form_redistribuir': form_redistribuir,
        # Extras para destinatario de salida
        'categoria_envio': categoria_envio,
        'destinatarios_salientes': destinatarios_salientes,
        'tiene_rebote': destinatarios_salientes.filter(estado='REBOTE').exists() if destinatarios_salientes else False,
    }
    return render(request, 'correspondencia/usuario/detalle_correspondencia.html', context)


@login_required
def imprimir_sello_correspondencia(request, pk):
    """Genera un PDF 3x2 pulgadas con sello de radicado y QR.

    - El botón es visible para todos, pero solo Ventanilla puede usarlo.
    - Si el usuario no es Ventanilla, muestra mensaje "Sin permisos" y redirige.
    """
    correspondencia = get_object_or_404(Correspondencia, pk=pk)

    # Verificación de permiso dinámico (no oculta el botón, solo valida en acción)
    es_ventanilla = request.user.groups.filter(name='Ventanilla').exists() or request.user.is_superuser
    if not es_ventanilla:
        messages.error(request, "Sin permisos")
        return redirect('correspondencia:detalle_correspondencia', pk=pk)

    # Validación previa: debe estar radicada (tener número de radicado)
    if not correspondencia.numero_radicado:
        messages.error(request, "La correspondencia debe estar radicada antes de imprimir el sello.")
        return redirect('correspondencia:detalle_correspondencia', pk=pk)

    # URL absoluta al detalle para el QR
    try:
        detalle_url = request.build_absolute_uri(
            reverse('correspondencia:detalle_correspondencia', args=[pk])
        )
    except Exception:
        detalle_url = ''

    # Generación del PDF con ReportLab y QRCode
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
        from reportlab.lib.utils import ImageReader
        import qrcode
        from io import BytesIO as _BytesIO

        buf = _BytesIO()
        page_size = (3 * inch, 2 * inch)
        c = canvas.Canvas(buf, pagesize=page_size)
        width, height = page_size

        # Encabezado
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width / 2, height - 0.25 * inch, "HOSPITAL DEL SARARE")

        # Radicado y fecha
        c.setFont("Helvetica-Bold", 10)
        c.drawString(0.2 * inch, height - 0.6 * inch, f"RADICADO: {correspondencia.numero_radicado}")
        c.setFont("Helvetica", 8)
        try:
            fecha_str = timezone.localtime(correspondencia.fecha_radicacion).strftime('%d/%m/%Y %H:%M')
        except Exception:
            fecha_str = ''
        c.drawString(0.2 * inch, height - 0.8 * inch, f"FECHA: {fecha_str}")

        # Oficina destino
        oficina = getattr(correspondencia.oficina_destino, 'nombre', None) or 'Sin oficina'
        c.setFont("Helvetica", 7)
        c.drawString(0.2 * inch, height - 1.0 * inch, f"OFICINA: {oficina[:28]}")

        # QR
        try:
            qr = qrcode.QRCode(version=1, box_size=6, border=1)
            qr.add_data(detalle_url or correspondencia.numero_radicado)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            qr_buf = _BytesIO()
            img.save(qr_buf, format='PNG')
            qr_buf.seek(0)
            qr_img = ImageReader(qr_buf)
            qr_size = 0.95 * inch
            c.drawImage(qr_img, width - (qr_size + 0.2 * inch), 0.2 * inch, qr_size, qr_size, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

        # Pie
        #c.setFont("Helvetica-Oblique", 6)
        #c.drawString(0.2 * inch, 0.2 * inch, "Sello de radicado generado por plataforma de Correspondencia")

        c.showPage()
        c.save()
        pdf_bytes = buf.getvalue()
        buf.close()
    except Exception as e:
        messages.error(request, f"No se pudo generar el sello: {e}")
        return redirect('correspondencia:detalle_correspondencia', pk=pk)

    # Marcar como sellado y registrar historial
    try:
        correspondencia.marcar_sellado()
        try:
            HistorialCorrespondencia.objects.create(
                correspondencia=correspondencia,
                evento='SELLO_IMPRESO',
                usuario=request.user,
                descripcion='Sello de radicado generado.'
            )
        except Exception:
            pass
    except Exception:
        pass

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="sello_{correspondencia.numero_radicado}.pdf"'
    return response
# @login_required
# def marcar_como_leido(request, pk):
#     # ... (código anterior) ...
#     pass # Mantener comentado o eliminar

@login_required
def redistribuir_interna(request, pk):
    """Permite redistribuir una correspondencia a usuarios dentro de la misma oficina."""
    correspondencia = get_object_or_404(Correspondencia, pk=pk)
    
    # Verificar que el usuario pertenece a la oficina destino
    if not request.user.perfil.oficina == correspondencia.oficina_destino:
        messages.error(request, "No tienes permiso para redistribuir esta correspondencia.")
        return redirect('bandeja_entrada')
    
    # Obtener usuarios de la misma oficina
    usuarios_oficina = User.objects.filter(
        perfil__oficina=correspondencia.oficina_destino
    ).exclude(
        id=request.user.id  # Excluir al usuario actual
    )
    
    if request.method == 'POST':
        usuario_id = request.POST.get('usuario')
        observaciones = request.POST.get('observaciones', '')
        
        try:
            usuario = User.objects.get(id=usuario_id)
            # Crear la distribución interna
            DistribucionInternaUsuario.objects.create(
                correspondencia=correspondencia,
                usuario_asignado=usuario,
                asignado_por=request.user,
                observaciones=observaciones
            )
            
            # Registrar en el historial
            HistorialCorrespondencia.objects.create(
                correspondencia=correspondencia,
                evento='REDISTRIBUIDA_INTERNA',
                usuario=request.user,
                descripcion=f"Redistribuida internamente a {usuario.get_full_name() or usuario.username}"
            )
            
            messages.success(request, f'Correspondencia redistribuida exitosamente a {usuario.get_full_name() or usuario.username}')
            return redirect('bandeja_entrada')
            
        except User.DoesNotExist:
            messages.error(request, "Usuario seleccionado no válido.")
        except Exception as e:
            messages.error(request, f"Error al redistribuir: {str(e)}")
    
    context = {
        'correspondencia': correspondencia,
        'usuarios_oficina': usuarios_oficina,
        'titulo_pagina': 'Redistribuir Correspondencia'
    }
    return render(request, 'correspondencia/admin/redistribuir_interna.html', context)

# === VISTAS PARA VENTANILLA - RADICACIÓN MANUAL DE CORREOS ===

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def bandeja_correos_pendientes_view(request):
    """Muestra la lista de correos entrantes con filtros (búsqueda, fechas, estado, papelera y adjuntos)."""
    from django.utils import timezone as tz
    import datetime

    qs = CorreoEntrante.objects.select_related('radicado_asociado', 'usuario_papelera').prefetch_related('adjuntos').order_by('-fecha_lectura_imap')

    estado = request.GET.get('estado')
    papelera = request.GET.get('papelera')
    q = request.GET.get('q', '').strip()
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    con_adjuntos = request.GET.get('con_adjuntos')
    fecha_base = request.GET.get('fecha_base', 'recepcion_gmail')
    orden = request.GET.get('orden', 'recepcion_gmail_desc')

    if papelera == '1':
        qs = qs.filter(en_papelera=True)
    else:
        qs = qs.filter(en_papelera=False)

    if estado == 'pendiente' and papelera != '1':
        qs = qs.filter(radicado_asociado__isnull=True, urgencia_asociada__isnull=True)
    elif estado == 'radicado' and papelera != '1':
        qs = qs.filter(models.Q(radicado_asociado__isnull=False) | models.Q(urgencia_asociada__isnull=False))

    if q:
        qs = qs.filter(models.Q(remitente__icontains=q) | models.Q(asunto__icontains=q))

    # Filtro de fecha: si el usuario especificó un rango manual se respeta;
    # si NO especificó ningún filtro de fecha, se muestran por defecto solo los
    # correos desde medianoche del día actual (hora local del proyecto).
    campos_fecha = {
        'recepcion_gmail': 'fecha_recibida_gmail',
        'fecha_correo': 'fecha_recepcion_original',
        'lectura_sistema': 'fecha_lectura_imap',
    }
    if fecha_base not in campos_fecha:
        fecha_base = 'recepcion_gmail'
    campo_fecha = campos_fecha[fecha_base]

    if desde:
        try:
            qs = qs.filter(**{f'{campo_fecha}__date__gte': desde})
        except Exception:
            pass
    elif not hasta and papelera != '1':
        # ─ Filtro por defecto: hoy desde las 00:00 hora local ─
        inicio_hoy = tz.localtime(tz.now()).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        qs = qs.filter(**{f'{campo_fecha}__gte': inicio_hoy})

    if hasta:
        try:
            qs = qs.filter(**{f'{campo_fecha}__date__lte': hasta})
        except Exception:
            pass

    if con_adjuntos == '1':
        qs = qs.filter(adjuntos__isnull=False).distinct()

    ordenamientos = {
        'recepcion_gmail_desc': ['-fecha_recibida_gmail', '-fecha_lectura_imap', '-id'],
        'recepcion_gmail_asc': ['fecha_recibida_gmail', 'fecha_lectura_imap', 'id'],
        'fecha_correo_desc': ['-fecha_recepcion_original', '-fecha_recibida_gmail', '-id'],
        'fecha_correo_asc': ['fecha_recepcion_original', 'fecha_recibida_gmail', 'id'],
    }
    if orden not in ordenamientos:
        orden = 'recepcion_gmail_desc'
    qs = qs.order_by(*ordenamientos[orden])

    query_params = request.GET.copy()
    query_params.pop('page', None)
    orden_query_base = query_params.copy()
    orden_query_base.pop('orden', None)

    paginator = Paginator(qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Última vez que se guardó un correo (útil para saber si Celery está trayendo correos)
    ultimo_fetch = None
    sync_estado = None
    if papelera != '1':
        ultimo_fetch = CorreoEntrante.objects.aggregate(ultima=Max('fecha_lectura_imap'))['ultima']
        sync_estado = EstadoSincronizacionCorreos.objects.filter(fuente='GMAIL_IMAP').first()

    # Indicar al template si está activo el filtro por defecto (para mostrar aviso)
    filtro_desde_default = not desde and not hasta and papelera != '1'

    context = {
        'correos': page_obj,
        'titulo_pagina': 'Gestión de Correos Entrantes',
        'filtro_estado': estado or '',
        'filtro_papelera': papelera or '',
        'filtro_q': q,
        'filtro_desde': desde or '',
        'filtro_hasta': hasta or '',
        'filtro_con_adjuntos': con_adjuntos or '',
        'filtro_fecha_base': fecha_base,
        'orden_actual': orden,
        'orden_query_base': orden_query_base.urlencode(),
        'ultimo_fetch': ultimo_fetch,
        'sync_estado': sync_estado,
        'filtro_desde_default': filtro_desde_default,
        'correos_problematicos_pendientes': CorreoProblematico.objects.filter(resuelto=False).count(),
    }
    return render(request, 'correspondencia/admin/bandeja_correos_pendientes.html', context)


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def bandeja_correos_problematicos_view(request):
    """Muestra correos detectados por IMAP pero excluidos del flujo normal."""
    qs = CorreoProblematico.objects.select_related('correo_entrante_asociado').order_by('resuelto', '-fecha_recibida_gmail', '-fecha_lectura_imap')

    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '').strip()
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')

    if estado == 'pendiente':
        qs = qs.filter(resuelto=False)
    elif estado == 'resuelto':
        qs = qs.filter(resuelto=True)

    if q:
        qs = qs.filter(
            models.Q(remitente__icontains=q)
            | models.Q(asunto__icontains=q)
            | models.Q(detalle_problema__icontains=q)
            | models.Q(message_id__icontains=q)
        )

    if desde:
        try:
            qs = qs.filter(fecha_lectura_imap__date__gte=desde)
        except Exception:
            pass
    if hasta:
        try:
            qs = qs.filter(fecha_lectura_imap__date__lte=hasta)
        except Exception:
            pass

    paginator = Paginator(qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'correos': page_obj,
        'titulo_pagina': 'Bandeja de Correos Problemáticos',
        'filtro_q': q,
        'filtro_estado': estado,
        'filtro_desde': desde or '',
        'filtro_hasta': hasta or '',
        'pendientes_count': CorreoProblematico.objects.filter(resuelto=False).count(),
        'resueltos_count': CorreoProblematico.objects.filter(resuelto=True).count(),
    }
    return render(request, 'correspondencia/admin/bandeja_correos_problematicos.html', context)


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def detalle_correo_problematico_view(request, problem_id):
    correo = get_object_or_404(CorreoProblematico, pk=problem_id)
    context = {
        'correo': correo,
        'titulo_pagina': f'Detalle Correo Problemático: {correo.asunto[:50]}',
    }
    return render(request, 'correspondencia/admin/detalle_correo_problematico.html', context)


@login_required
@require_POST
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def forzar_ingreso_correo_problematico_view(request, problem_id):
    """Fuerza el ingreso de un correo problemático a la bandeja principal."""
    from correspondencia.utils.email_ingestion import forzar_ingreso_correo_problematico

    resultado = forzar_ingreso_correo_problematico(problem_id, usuario=request.user)
    if resultado['ok']:
        messages.success(request, resultado['detail'])
        correo_entrante = resultado.get('correo_entrante')
        if correo_entrante:
            return redirect('correspondencia:detalle_correo_entrante', correo_id=correo_entrante.pk)
    else:
        messages.error(request, resultado['detail'])

    return redirect('correspondencia:detalle_correo_problematico', problem_id=problem_id)


# === DASHBOARD VENTANILLA (MVP) - DUPLICADO - ELIMINAR EN REFACTOR ===
# NOTA: Esta función está duplicada. La versión principal está más arriba.
# Por ahora se mantiene sincronizada pero debería eliminarse en un refactor.
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def dashboard_ventanilla_legacy(request):
    """Dashboard inicial para Ventanilla: KPIs, recientes y estado IMAP."""
    from datetime import timedelta

    # --- Procesamiento de formularios enviados desde modales ---
    open_contacto_modal = False
    open_entidad_modal = False
    open_radicacion_modal = False
    open_rapida_entrante_modal = False
    open_rapida_saliente_modal = False

    if request.method == 'POST':
        form_prefix = request.POST.get('form_prefix', '')
        if form_prefix == 'contacto':
            contacto_form = ContactoForm(request.POST)
            entidad_form = EntidadExternaForm()
            radicacion_form = ManualRadicacionCorreoForm(prefix='radicar')
            rapida_entrante_form = RadicacionRapidaEntranteForm(prefix='rapida_ent')
            rapida_saliente_form = RadicacionRapidaSalienteForm(prefix='rapida_sal', user=request.user)
            if contacto_form.is_valid():
                contacto_form.save()
                messages.success(request, 'Contacto creado exitosamente.')
                return redirect('correspondencia:dashboard_ventanilla')
            else:
                open_contacto_modal = True
        elif form_prefix == 'entidad':
            entidad_form = EntidadExternaForm(request.POST)
            contacto_form = ContactoForm()
            radicacion_form = ManualRadicacionCorreoForm(prefix='radicar')
            rapida_entrante_form = RadicacionRapidaEntranteForm(prefix='rapida_ent')
            rapida_saliente_form = RadicacionRapidaSalienteForm(prefix='rapida_sal', user=request.user)
            if entidad_form.is_valid():
                entidad_form.save()
                messages.success(request, 'Entidad creada exitosamente.')
                return redirect('correspondencia:dashboard_ventanilla')
            else:
                open_entidad_modal = True
        elif form_prefix == 'radicar':
            contacto_form = ContactoForm()
            entidad_form = EntidadExternaForm()
            radicacion_form = ManualRadicacionCorreoForm(request.POST, prefix='radicar')
            rapida_entrante_form = RadicacionRapidaEntranteForm(prefix='rapida_ent')
            rapida_saliente_form = RadicacionRapidaSalienteForm(prefix='rapida_sal', user=request.user)
            if radicacion_form.is_valid():
                correspondencia = radicacion_form.save(commit=False)
                correspondencia.tipo_radicado = 'ENTRANTE'
                correspondencia.usuario_radicador = request.user
                correspondencia.save()
                messages.success(request, f'Correspondencia {correspondencia.numero_radicado} radicada exitosamente.')
                return redirect('correspondencia:dashboard_ventanilla')
            else:
                open_radicacion_modal = True
        elif form_prefix == 'rapida_entrante':
            contacto_form = ContactoForm()
            entidad_form = EntidadExternaForm()
            radicacion_form = ManualRadicacionCorreoForm(prefix='radicar')
            rapida_entrante_form = RadicacionRapidaEntranteForm(request.POST, prefix='rapida_ent')
            rapida_saliente_form = RadicacionRapidaSalienteForm(prefix='rapida_sal', user=request.user)
            if rapida_entrante_form.is_valid():
                correspondencia = rapida_entrante_form.save(commit=False)
                correspondencia.tipo_radicado = 'ENTRANTE'
                correspondencia.origen_radicacion = 'RAPIDA'  # Marcar como radicación rápida
                correspondencia.usuario_radicador = request.user
                correspondencia.save()
                messages.success(request, f'Entrante {correspondencia.numero_radicado} radicada exitosamente (rápida).')
                return redirect('correspondencia:dashboard_ventanilla')
            else:
                open_rapida_entrante_modal = True
        elif form_prefix == 'rapida_saliente':
            contacto_form = ContactoForm()
            entidad_form = EntidadExternaForm()
            radicacion_form = ManualRadicacionCorreoForm(prefix='radicar')
            rapida_entrante_form = RadicacionRapidaEntranteForm(prefix='rapida_ent')
            rapida_saliente_form = RadicacionRapidaSalienteForm(request.POST, prefix='rapida_sal', user=request.user)
            if rapida_saliente_form.is_valid():
                salida = rapida_saliente_form.save(commit=False)
                salida.usuario_redactor = request.user
                salida.oficina_emisora = rapida_saliente_form.cleaned_data.get('oficina_emisora')
                destinatario_contacto = rapida_saliente_form.cleaned_data.get('destinatario_contacto')
                if destinatario_contacto:
                    salida.destinatario_contacto = destinatario_contacto
                    salida.destinatario_email = destinatario_contacto.correo_electronico or ''
                else:
                    destinatario_texto = rapida_saliente_form.cleaned_data.get('destinatario_texto', '')
                    if destinatario_texto:
                        salida.cuerpo = f"[Destinatario: {destinatario_texto}]\n\n{salida.cuerpo}"
                salida.estado = 'ENVIADA'
                salida.fecha_envio = timezone.now()
                salida.save()
                messages.success(request, f'Saliente {salida.numero_radicado_salida} radicada exitosamente (rápida).')
                return redirect('correspondencia:dashboard_ventanilla')
            else:
                open_rapida_saliente_modal = True
        else:
            contacto_form = ContactoForm()
            entidad_form = EntidadExternaForm()
            radicacion_form = ManualRadicacionCorreoForm(prefix='radicar')
            rapida_entrante_form = RadicacionRapidaEntranteForm(prefix='rapida_ent')
            rapida_saliente_form = RadicacionRapidaSalienteForm(prefix='rapida_sal', user=request.user)
    else:
        contacto_form = ContactoForm()
        entidad_form = EntidadExternaForm()
        radicacion_form = ManualRadicacionCorreoForm(prefix='radicar')
        rapida_entrante_form = RadicacionRapidaEntranteForm(prefix='rapida_ent')
        rapida_saliente_form = RadicacionRapidaSalienteForm(prefix='rapida_sal', user=request.user)

    # --- KPIs ---
    ahora = timezone.now()
    hoy_inicio = ahora.astimezone(timezone.get_current_timezone()).replace(hour=0, minute=0, second=0, microsecond=0)
    hoy_fin = hoy_inicio + datetime.timedelta(days=1)

    kpi_pendientes_radicacion = CorreoEntrante.objects.filter(radicado_asociado__isnull=True, urgencia_asociada__isnull=True).count()
    kpi_radicados_hoy = Correspondencia.objects.filter(
        tipo_radicado='ENTRANTE',
        fecha_radicacion__gte=hoy_inicio,
        fecha_radicacion__lt=hoy_fin
    ).count()
    qs_sla_legacy = excluir_entrantes_con_respuesta(
        Correspondencia.objects.filter(
            tipo_radicado='ENTRANTE',
            requiere_respuesta=True,
        )
    )
    kpi_sla_vencido = qs_sla_legacy.filter(
        fecha_limite_respuesta_persist__lt=ahora,
    ).count()
    kpi_sla_por_vencer = qs_sla_legacy.filter(
        fecha_limite_respuesta_persist__gte=ahora,
        fecha_limite_respuesta_persist__lte=ahora + datetime.timedelta(hours=48),
    ).count()

    context = {
        'titulo_pagina': 'Dashboard Ventanilla',
        'kpi_pendientes_radicacion': kpi_pendientes_radicacion,
        'kpi_radicados_hoy': kpi_radicados_hoy,
        'kpi_sla_vencido': kpi_sla_vencido,
        'kpi_sla_por_vencer': kpi_sla_por_vencer,
        # Formularios para modales
        'form_contacto': contacto_form,
        'form_entidad': entidad_form,
        'form_radicacion': radicacion_form,
        # Formularios de radicación rápida
        'form_rapida_entrante': rapida_entrante_form,
        'form_rapida_saliente': rapida_saliente_form,
        # Flags para reabrir modales tras POST inválido
        'open_contacto_modal': open_contacto_modal,
        'open_radicacion_modal': open_radicacion_modal,
        'open_entidad_modal': open_entidad_modal,
        'open_rapida_entrante_modal': open_rapida_entrante_modal,
        'open_rapida_saliente_modal': open_rapida_saliente_modal,
    }

    return render(request, 'correspondencia/admin/dashboard_ventanilla.html', context)
# ------------- ENDPOINT AJAX PARA SUBSERIES -------------
@require_GET
def api_subseries(request):
    """
    Endpoint AJAX que devuelve {id, nombre} de SubserieDocumental
    filtradas por serie_id.
    """
    serie_id = request.GET.get("serie_id")
    if serie_id:
        try:
            # Validar que serie_id sea un número entero
            serie_id = int(serie_id)
            qs = SubserieDoc.objects.filter(serie_id=serie_id).order_by('nombre')
            data = list(qs.values("id", "nombre"))
        except (ValueError, TypeError):
            # Si serie_id no es válido, devolver lista vacía
            data = []
    else:
        data = []
    return JsonResponse(data, safe=False)
# ------------- ENDPOINT AJAX PARA SUBSERIES -------------
# ------------- ENDPOINT PARA CONTACTOS EN FORMULARIO DE RADICACIÓN MANUAL -------------
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def buscar_contactos(request):
    q = request.GET.get('q', '').strip()
    contactos = Contacto.objects.filter(nombres__icontains=q).order_by('nombres')[:10]
    data = {
        'contactos': [
            {
                'id': c.id,
                'nombre': f'{c.nombres} {c.apellidos}',  # Usa los campos reales
                'correo': c.correo_electronico
            }
            for c in contactos
        ]
    }

    return JsonResponse(data)

# ------------- ENDPOINT AJAX PARA CONCTACTOS  -------------


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(),
                 login_url='correspondencia:welcome')
def visor_correo_completo_view(request, correo_id):
    """Vista dedicada para visualizar el contenido completo de un correo: cuerpo HTML, imágenes inline y adjuntos."""
    correo = get_object_or_404(
        CorreoEntrante.objects.prefetch_related('adjuntos'),
        pk=correo_id
    )
    cuerpo_html_renderizado = correo.obtener_cuerpo_html_renderizado()
    adjuntos = correo.adjuntos.all()
    adjuntos_imagen = [a for a in adjuntos if (a.tipo_mime or '').lower().startswith('image/')]
    adjuntos_otros = [a for a in adjuntos if not (a.tipo_mime or '').lower().startswith('image/')]

    return render(request, 'correspondencia/admin/visor_correo_completo.html', {
        'correo': correo,
        'cuerpo_html_renderizado': cuerpo_html_renderizado,
        'adjuntos_imagen': adjuntos_imagen,
        'adjuntos_otros': adjuntos_otros,
        'titulo_pagina': f'Visor de Correo: {correo.asunto[:50]}...',
    })


@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(),
                 login_url='correspondencia:welcome')
def detalle_correo_entrante_view(request, correo_id):
    correo = get_object_or_404(
        CorreoEntrante.objects.prefetch_related('adjuntos', 'problemas_origen'),
        pk=correo_id
    )
    cuerpo_html_renderizado = correo.obtener_cuerpo_html_renderizado()
    titulo_pagina = f'Detalle Correo: {correo.asunto[:50]}...'
    problema_origen = correo.problemas_origen.order_by('-fecha_recibida_gmail', '-fecha_lectura_imap').first()

    # Flags de control de modales (deben existir tanto en GET como en POST)
    open_radicacion_modal = False
    open_rapida_entrante_modal = False

    # 1) Formulario de contacto
    # --- DOMINIO y entidad sugerida para el formulario de contacto ---
    from correspondencia.utils.blocked_recipients import normalizar_email_destinatario

    email_remitente_normalizado = normalizar_email_destinatario(correo.remitente)
    entidad_sugerida = None
    if email_remitente_normalizado and '@' in email_remitente_normalizado:
        dominio = email_remitente_normalizado.split('@', 1)[1]
        if dominio not in DOMINIOS_GENERICOS:
            entidad_sugerida = EntidadExterna.buscar_por_dominio(dominio)

    # 1) Formulario de contacto con entidad sugerida (si existe)
    if entidad_sugerida:
        form_contacto = ContactoForm(
            request.POST or None,
            prefix="contacto",
            initial={'entidad_externa': entidad_sugerida}
        )
    else:
        form_contacto = ContactoForm(request.POST or None, prefix="contacto")
    
    # 1.1) Formulario de entidad
    form_entidad = EntidadExternaForm(request.POST or None, prefix="entidad")
    
    # 1.2) Formulario de urgencia (solo si NO está radicado)
    from .forms import UrgenciaRadicacionForm
    form_urgencia = None
    if not correo.radicado_asociado:
        form_urgencia = UrgenciaRadicacionForm()
    
    # 2) Formulario de radicación (solo si NO está radicado)
    form_radicacion = None
    contacto_sugerido = None
    if not correo.radicado_asociado:
        if email_remitente_normalizado:
            contacto_sugerido = Contacto.objects.filter(
                correo_electronico__iexact=email_remitente_normalizado
            ).select_related('entidad_externa').first()

        initial_radicacion = {
            'asunto': correo.asunto,
            'medio_recepcion': 'ELECTRONICO',  # 🔥 AÑADIDO: Valor por defecto para correos
            'remitente': contacto_sugerido,
        }
        form_radicacion = ManualRadicacionCorreoForm(
            data=request.POST if request.method == 'POST' and request.POST.get('form_prefix') == 'radicar' else None,
            initial=initial_radicacion if request.method == 'GET' else None,
            prefix="radicar"
        )

        # Lógica de subseries (igual que antes)…
        serie_id = None
        if request.method == 'POST' and request.POST.get('form_prefix') == 'radicar':
            serie_id = request.POST.get('radicar-serie')
        elif request.method == 'GET' and form_radicacion.initial.get('serie'):
            serie_id = form_radicacion.initial.get('serie').pk
        elif request.method == 'GET' and 'serie' in initial_radicacion and initial_radicacion['serie']:
            serie_id = initial_radicacion['serie'].pk

        if serie_id:
            try:
                form_radicacion.fields['subserie'].queryset = (
                    SubserieDoc.objects.filter(serie_id=int(serie_id)).order_by('nombre')
                )
                form_radicacion.fields['subserie'].disabled = False
            except (ValueError, TypeError):
                form_radicacion.fields['subserie'].queryset = SubserieDoc.objects.none()
                form_radicacion.fields['subserie'].disabled = True
        else:
            form_radicacion.fields['subserie'].queryset = SubserieDoc.objects.none()
            form_radicacion.fields['subserie'].disabled = True

        # Formulario de radicación rápida (pre-llenado desde correo)
        initial_rapida = {
            'asunto': correo.asunto,
            'remitente_texto': correo.remitente or '',
            'medio_recepcion': 'ELECTRONICO',
            'remitente': contacto_sugerido,
        }
        form_rapida_entrante = RadicacionRapidaEntranteForm(
            data=request.POST if request.method == 'POST' and request.POST.get('form_prefix') == 'rapida_entrante' else None,
            initial=initial_rapida if request.method == 'GET' else None,
            prefix='rapida_ent'
        )
    else:
        form_rapida_entrante = None

    # === Procesamiento POST ===
    if request.method == 'POST':
        form_prefix = request.POST.get('form_prefix')

        # 1) Crear Contacto
        if form_prefix == 'contacto':
            form_contacto = ContactoForm(request.POST, prefix="contacto")
            if form_contacto.is_valid():
                try:
                    nuevo_contacto = form_contacto.save()
                    messages.success(request, f"Contacto '{nuevo_contacto}' creado exitosamente.")
                    return redirect('correspondencia:detalle_correo_entrante', correo_id=correo.id)
                except IntegrityError as e:
                    error_msg = f"Error al crear contacto: Ya existe un contacto similar o con ese correo. {e}"
                    if 'contacto_unico_por_entidad' in str(e):
                        error_msg = "Error: Ya existe un contacto con esos nombres y correo para esa entidad."
                    elif 'correo_electronico' in str(e):
                        error_msg = "Error: Ya existe un contacto con ese correo electrónico."
                    messages.error(request, error_msg)
                except Exception as e:
                    messages.error(request, f"Error inesperado al crear contacto: {e}")
            else:
                messages.error(request, "Por favor corrija los errores en el formulario de contacto.")

        # 1.1) Crear Entidad
        elif form_prefix == 'entidad':
            form_entidad = EntidadExternaForm(request.POST, prefix="entidad")
            if form_entidad.is_valid():
                entidad_nueva = form_entidad.save()
                messages.success(request, f"Entidad '{entidad_nueva}' creada exitosamente.")
                return redirect('correspondencia:detalle_correo_entrante', correo_id=correo.id)
            else:
                messages.error(request, "Corrige los errores del formulario de entidad.")

        # 2.1) Radicación Rápida (desde detalle de correo - asocia el correo al radicado)
        elif form_prefix == 'rapida_entrante' and not correo.radicado_asociado and not correo.urgencia_asociada:
            rapida_form = RadicacionRapidaEntranteForm(request.POST, prefix='rapida_ent')
            if rapida_form.is_valid():
                try:
                    with transaction.atomic():
                        correspondencia = rapida_form.save(commit=False)
                        correspondencia.tipo_radicado = 'ENTRANTE'
                        correspondencia.origen_radicacion = 'RAPIDA'
                        correspondencia.usuario_radicador = request.user
                        correspondencia.save()

                        HistorialCorrespondencia.objects.create(
                            correspondencia=correspondencia,
                            evento='RADICADA',
                            usuario=request.user,
                            descripcion=f"Radicada rápidamente desde CorreoEntrante ID {correo.id}"
                        )
                        correo.radicado_asociado = correspondencia
                        correo.save(update_fields=['radicado_asociado'])

                        # Copiar adjuntos del correo a la correspondencia
                        for adj_origen in correo.adjuntos.all():
                            adj_destino = AdjuntoCorreo(
                                correspondencia=correspondencia,
                                nombre_original=adj_origen.nombre_original,
                                tipo_mime=adj_origen.tipo_mime
                            )
                            if adj_origen.archivo:
                                try:
                                    file_content = ContentFile(adj_origen.archivo.read())
                                    file_name = adj_origen.nombre_original or os.path.basename(adj_origen.archivo.name)
                                    adj_destino.archivo.save(file_name, file_content, save=False)
                                except Exception as e:
                                    messages.warning(request, f"No se pudo copiar el adjunto: {adj_origen.nombre_original}")
                            adj_destino.save()

                    # === ENVÍO DE CORREO AL FUNCIONARIO RESPONSABLE (fuera del atomic) ===
                    email_funcionario = correspondencia.email_funcionario_responsable
                    if email_funcionario:
                        try:
                            from correspondencia.utils.radicacion_rapida_email import (
                                adjuntos_desde_queryset,
                                enviar_notificacion_radicacion_rapida_entrante,
                            )

                            contexto_email = {
                                'nombre_funcionario': correspondencia.funcionario_responsable_tramite or 'Funcionario',
                                'numero_radicado': correspondencia.numero_radicado,
                                'fecha_radicacion': timezone.localtime(correspondencia.fecha_radicacion).strftime('%d/%m/%Y %H:%M'),
                                'remitente': correspondencia.entidad_persona_remitente or correo.remitente or 'No especificado',
                                'direccion_correo_remitente': correspondencia.direccion_correo_remitente or correo.email_remitente or '',
                                'oficina_destino': str(correspondencia.oficina_destino) if correspondencia.oficina_destino else 'No especificada',
                                'medio_recepcion': correspondencia.get_medio_recepcion_display() if correspondencia.medio_recepcion else '',
                                'tipo_tramite': correspondencia.tipo_tramite or '',
                                'fecha_limite': correspondencia.fecha_limite_respuesta_manual.strftime('%d/%m/%Y') if correspondencia.fecha_limite_respuesta_manual else '',
                                'asunto': correspondencia.asunto,
                                'cuerpo_correo': correo.cuerpo_html or correo.cuerpo_texto or '',
                                'usuario_radicador': request.user.get_full_name() or request.user.username,
                            }

                            adjuntos_notificacion = (
                                adjuntos_desde_queryset(correo.adjuntos.all())
                                + adjuntos_desde_queryset(correspondencia.adjuntos_rapidos.all())
                            )

                            enviar_notificacion_radicacion_rapida_entrante(
                                email_funcionario=email_funcionario,
                                contexto_email=contexto_email,
                                asunto=f"Correspondencia asignada - {correspondencia.numero_radicado}",
                                adjuntos=adjuntos_notificacion,
                            )

                            HistorialCorrespondencia.objects.create(
                                correspondencia=correspondencia,
                                evento='NOTIFICACION',
                                usuario=request.user,
                                descripcion=(
                                    f"Notificación enviada a {correspondencia.funcionario_responsable_tramite} "
                                    f"({email_funcionario}) — radicación rápida desde correo entrante"
                                )
                            )
                            messages.info(request, f"📧 Notificación enviada a {email_funcionario}")

                        except Exception as e_email:
                            HistorialCorrespondencia.objects.create(
                                correspondencia=correspondencia,
                                evento='ERROR',
                                usuario=request.user,
                                descripcion=f"Error al enviar notificación a {email_funcionario}: {e_email}"
                            )
                            messages.warning(
                                request,
                                f"⚠️ Radicación exitosa, pero falló el envío de notificación a {email_funcionario}: {e_email}"
                            )

                    messages.success(
                        request,
                        f"Correo radicado exitosamente como {correspondencia.numero_radicado} (rápida)."
                    )
                    return redirect('correspondencia:detalle_correo_entrante', correo_id=correo.id)
                except Exception as e:
                    messages.error(request, f"Error al radicar: {e}")
                    traceback.print_exc()
            else:
                form_rapida_entrante = rapida_form
                open_rapida_entrante_modal = True

        # 2) Radicar Correo
        elif form_prefix == 'radicar' and form_radicacion:
            form_radicacion = ManualRadicacionCorreoForm(request.POST, request.FILES, prefix="radicar")
            if form_radicacion.is_valid():
                try:
                    with transaction.atomic():
                        correspondencia = form_radicacion.save(commit=False)
                        # 🔥 ACTUALIZADO: Usar el medio_recepcion del formulario en lugar de forzar ELECTRONICO
                        correspondencia.usuario_radicador = request.user
                        correspondencia.origen_radicacion = 'CORREO'  # Desde correo electrónico
                        correspondencia.save()

                        adjuntos_guardados = _guardar_adjuntos_radicacion_fisica(correspondencia, request)

                        HistorialCorrespondencia.objects.create(
                            correspondencia=correspondencia,
                            evento='RADICADA',
                            usuario=request.user,
                            descripcion=(
                                f"Radicada manualmente por {request.user.username} "
                                f"desde CorreoEntrante ID {correo.id}" + 
                                (f" con {adjuntos_guardados} adjuntos" if adjuntos_guardados > 0 else "")
                            )
                        )
                        correo.radicado_asociado = correspondencia
                        correo.save(update_fields=['radicado_asociado'])

                        # Copiar adjuntos…
                        for adj_origen in correo.adjuntos.all():
                            adj_destino = AdjuntoCorreo(
                                correspondencia=correspondencia,
                                nombre_original=adj_origen.nombre_original,
                                tipo_mime=adj_origen.tipo_mime
                            )
                            if adj_origen.archivo:
                                try:
                                    file_content = ContentFile(adj_origen.archivo.read())
                                    file_name = adj_origen.nombre_original or os.path.basename(adj_origen.archivo.name)
                                    adj_destino.archivo.save(file_name, file_content, save=False)
                                except Exception as e:
                                    print(f"Error copiando archivo adjunto {adj_origen.id}: {e}")
                                    messages.warning(request, f"No se pudo copiar el adjunto: {adj_origen.nombre_original}")
                            adj_destino.save()

                        suffix = _aplicar_distribucion_rapida_desde_form(
                            correspondencia,
                            request,
                            form_radicacion.cleaned_data,
                        )

                        messages.success(
                            request,
                            f"Correo radicado exitosamente como {correspondencia.numero_radicado}{suffix}."
                        )
                        return redirect('correspondencia:bandeja_correos_pendientes')
                except IntegrityError as e:
                    open_radicacion_modal = True
                    messages.error(request, f"Error de integridad al radicar: {e}")
                    traceback.print_exc()
                except Exception as e:
                    open_radicacion_modal = True
                    messages.error(request, f"Error interno inesperado al radicar: {e}")
                    traceback.print_exc()
            else:
                open_radicacion_modal = True
                messages.error(request, "Por favor corrija los errores en el formulario de radicación.")

    # === Preparación del Contexto para GET o si hubo error en POST ===
    context = {
        'correo': correo,
        'form_contacto': form_contacto,
        'form_radicacion': form_radicacion,
        'form_entidad': form_entidad,
        'form': form_urgencia,  # Formulario de urgencia
        'titulo_pagina': titulo_pagina,
        'open_contacto_modal': (
            request.method == 'POST'
            and request.POST.get('form_prefix') == 'contacto'
            and not form_contacto.is_valid()
        ),
        'open_radicacion_modal': open_radicacion_modal,
        'open_entidad_modal': (
            request.method == 'POST'
            and request.POST.get('form_prefix') == 'entidad'
            and not form_entidad.is_valid()
        ),
        'contacto_sugerido': contacto_sugerido,
        'problema_origen': problema_origen,
        'avisar_contacto_asociado': bool(contacto_sugerido),
        'sla_enabled': True,  # Habilitar funcionalidad SLA
        'motivo_papelera_choices': MOTIVO_PAPELERA_CHOICES,
        'form_rapida_entrante': form_rapida_entrante,
        'open_rapida_entrante_modal': open_rapida_entrante_modal,
    }

    return render(
        request,
        'correspondencia/admin/detalle_correo_entrante.html',
        context
    )

# === VISTAS PARA RADICADO DE SALIDA RÁPIDA VINCULADO ===

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def generar_radicado_salida_rapida(request, correo_id):
    """Genera un radicado de salida (CorrespondenciaSalida) vinculado a la correspondencia
    entrante de un correo. Solo para radicación rápida."""
    if request.method != 'POST':
        return redirect('correspondencia:detalle_correo_entrante', correo_id=correo_id)

    correo = get_object_or_404(CorreoEntrante, pk=correo_id)
    correspondencia = correo.radicado_asociado

    if not correspondencia:
        messages.error(request, "Este correo aún no ha sido radicado.")
        return redirect('correspondencia:detalle_correo_entrante', correo_id=correo_id)

    if correspondencia.origen_radicacion != 'RAPIDA':
        messages.error(request, "Esta función solo aplica para radicación rápida.")
        return redirect('correspondencia:detalle_correo_entrante', correo_id=correo_id)

    # Verificar que no exista ya una salida vinculada
    salida_existente = correspondencia.respuestas_salientes.first()
    if salida_existente:
        messages.warning(
            request,
            f"Ya existe un radicado de salida vinculado: {salida_existente.numero_radicado_salida}"
        )
        return redirect('correspondencia:detalle_correo_entrante', correo_id=correo_id)

    try:
        with transaction.atomic():
            salida = CorrespondenciaSalida(
                respuesta_a=correspondencia,
                usuario_redactor=request.user,
                asunto=asunto_respuesta_desde_entrada(correspondencia.asunto, prefijo='Re: '),
                cuerpo=f"Respuesta al radicado de entrada {correspondencia.numero_radicado}",
            )
            salida.save()  # Auto-genera radicado, oficina_emisora, destinatario_contacto

            HistorialCorrespondencia.objects.create(
                correspondencia=correspondencia,
                evento='RADICADA',
                usuario=request.user,
                descripcion=(
                    f"Radicado de salida {salida.numero_radicado_salida} generado "
                    f"y vinculado al radicado de entrada {correspondencia.numero_radicado}"
                )
            )

        messages.success(
            request,
            f"Radicado de salida generado: {salida.numero_radicado_salida} "
            f"(vinculado a {correspondencia.numero_radicado})"
        )
    except Exception as e:
        messages.error(request, f"Error al generar radicado de salida: {e}")
        import traceback
        traceback.print_exc()

    return redirect('correspondencia:detalle_correo_entrante', correo_id=correo_id)


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def enviar_radicado_salida_email(request, salida_id):
    """Envía el número de radicado de salida por email al funcionario responsable."""
    if request.method != 'POST':
        return redirect('correspondencia:bandeja_correos_pendientes')

    salida = get_object_or_404(CorrespondenciaSalida, pk=salida_id)
    correspondencia = salida.respuesta_a

    if not correspondencia:
        messages.error(request, "Este radicado de salida no está vinculado a una correspondencia de entrada.")
        return redirect('correspondencia:bandeja_correos_pendientes')

    email_destino = correspondencia.email_funcionario_responsable
    if not email_destino:
        messages.error(request, "No hay email de funcionario responsable registrado.")
        correo = correspondencia.correo_origen.first()
        if correo:
            return redirect('correspondencia:detalle_correo_entrante', correo_id=correo.id)
        return redirect('correspondencia:bandeja_correos_pendientes')

    try:
        from django.core.mail import EmailMessage as DjangoEmailMessage

        from correspondencia.utils.radicacion_rapida_email import (
            HEADER_NOTIFICACION_CORRESPONDENCIA,
            VALOR_NOTIFICACION_RADICACION_RAPIDA,
            get_radicacion_rapida_entrante_mail_connection,
            preparar_destinatarios_notificacion_radicacion_rapida,
        )
        from correspondencia.aprobacion_envio import _direccion_remitente_visible

        funcionario = correspondencia.funcionario_responsable_tramite or 'Funcionario'
        radicado_entrada = correspondencia.numero_radicado
        radicado_salida = salida.numero_radicado_salida
        asunto_original = correspondencia.asunto

        cuerpo_html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background-color: #0d6efd; color: white; padding: 15px 20px; border-radius: 8px 8px 0 0;">
                <h3 style="margin: 0;">📋 Número de Radicado de Salida</h3>
            </div>
            <div style="border: 1px solid #dee2e6; border-top: none; padding: 20px; border-radius: 0 0 8px 8px;">
                <p>Estimado/a <strong>{funcionario}</strong>,</p>
                <p>Se ha generado el siguiente número de radicado de salida para que lo incluya en su respuesta:</p>
                <div style="background-color: #f8f9fa; border: 2px solid #0d6efd; border-radius: 8px; padding: 20px; text-align: center; margin: 15px 0;">
                    <small style="color: #6c757d;">Radicado de Salida</small><br>
                    <span style="font-size: 24px; font-weight: bold; color: #0d6efd;">{radicado_salida}</span>
                </div>
                <table style="width: 100%; border-collapse: collapse; margin: 15px 0;">
                    <tr style="border-bottom: 1px solid #dee2e6;">
                        <td style="padding: 8px; color: #6c757d;">Radicado de entrada:</td>
                        <td style="padding: 8px; font-weight: bold;">{radicado_entrada}</td>
                    </tr>
                    <tr style="border-bottom: 1px solid #dee2e6;">
                        <td style="padding: 8px; color: #6c757d;">Asunto:</td>
                        <td style="padding: 8px;">{asunto_original}</td>
                    </tr>
                </table>
                <p style="color: #6c757d; font-size: 13px;">Por favor, incluya el número de radicado de salida <strong>{radicado_salida}</strong> en el correo de respuesta que envíe.</p>
                <hr style="border: none; border-top: 1px solid #dee2e6;">
                <p style="color: #999; font-size: 11px; margin-bottom: 0;">Hospital del Sarare - Sistema de Gestión Documental</p>
            </div>
        </div>
        """

        to_recipients, bcc_recipients = preparar_destinatarios_notificacion_radicacion_rapida(
            email_destino
        )
        remitente_visible = _direccion_remitente_visible()
        if not remitente_visible:
            raise ValueError(
                'Configure OUTBOUND_EMAIL_ADDRESS con el buzón Gmail API autorizado.'
            )

        connection = get_radicacion_rapida_entrante_mail_connection()
        connection.open()
        try:
            email_msg = DjangoEmailMessage(
                subject=f"Radicado de Salida {radicado_salida} - {asunto_original}",
                body=cuerpo_html,
                from_email=remitente_visible,
                to=to_recipients,
                bcc=bcc_recipients,
                connection=connection,
                headers={
                    HEADER_NOTIFICACION_CORRESPONDENCIA: VALOR_NOTIFICACION_RADICACION_RAPIDA,
                },
            )
            email_msg.content_subtype = 'html'
            email_msg.send(fail_silently=False)

            HistorialCorrespondencia.objects.create(
                correspondencia=correspondencia,
                evento='NOTIFICACION',
                usuario=request.user,
                descripcion=(
                    f"Número de radicado de salida {radicado_salida} enviado por email "
                    f"a {funcionario} ({email_destino})"
                )
            )
            messages.success(request, f"📧 Radicado de salida {radicado_salida} enviado a {email_destino}")
        finally:
            try:
                connection.close()
            except Exception:
                pass

    except Exception as e:
        HistorialCorrespondencia.objects.create(
            correspondencia=correspondencia,
            evento='ERROR',
            usuario=request.user,
            descripcion=f"Error al enviar radicado de salida por email a {email_destino}: {e}"
        )
        messages.warning(request, f"⚠️ Error al enviar email: {e}")

    # Redirigir al detalle del correo original
    correo = correspondencia.correo_origen.first()
    if correo:
        return redirect('correspondencia:detalle_correo_entrante', correo_id=correo.id)
    return redirect('correspondencia:bandeja_correos_pendientes')

# === FIN VISTAS VENTANILLA (segunda sección) ===
# Nota: Las vistas de CRUD Entidades/Contactos están definidas más arriba (línea ~3500).
# Se eliminó el bloque duplicado que existía aquí.

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def listar_contactos(request):
    """Muestra una lista paginada de todos los contactos externos, incluyendo su entidad."""
    # Optimizar consulta incluyendo la entidad externa relacionada
    contactos_list = Contacto.objects.select_related('entidad_externa').all()
    
    # Búsqueda simple por nombre, apellido, correo o entidad
    query = request.GET.get('q')
    if query:
        contactos_list = contactos_list.filter(
            Q(nombres__icontains=query) | 
            Q(apellidos__icontains=query) | 
            Q(correo_electronico__icontains=query) |
            Q(entidad_externa__nombre__icontains=query) # Buscar por nombre de entidad
        )
    
    paginator = Paginator(contactos_list, 25) # 25 contactos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'contactos': page_obj,
        'titulo_pagina': 'Gestionar Contactos Externos',
        'search_query': query or "" # Pasar query a la plantilla para mostrarlo en el input
    }
    return render(request, 'correspondencia/admin/lista_contactos.html', context)

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def crear_contacto(request):
    """Crea un nuevo Contacto Externo (ahora requiere seleccionar EntidadExterna)."""
    if request.method == 'POST':
        # Usar ContactoForm actualizado
        form = ContactoForm(request.POST)
        if form.is_valid():
            try:
                contacto = form.save()
                messages.success(request, f"Contacto '{contacto.nombre_completo}' de la entidad '{contacto.entidad_externa.nombre}' creado exitosamente.")
                return redirect('correspondencia:listar_contactos') # Redirigir a la lista
            except Exception as e:
                 messages.error(request, f"Error al crear el contacto: {e}")
        else:
             messages.error(request, "Por favor corrija los errores en el formulario.")
    else:
        # GET: mostrar formulario vacío
        form = ContactoForm()
    
    context = {
        'form': form,
        'titulo_pagina': 'Crear Nuevo Contacto Externo'
    }
    # Reutilizar plantilla de formulario genérica o crear una específica
    return render(request, 'correspondencia/admin/contacto_form.html', context)

# Aquí podríamos añadir vistas para editar_contacto y eliminar_contacto en el futuro.

# === FIN CRUD Contactos ===

# === VISTAS PARA VENTANILLA - RADICACIÓN MANUAL DE CORREOS ===

@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def procesar_emails_manual(request):
    """Vista para encolar el procesamiento de correos en Celery."""
    try:
        procesar_emails_periodico.delay()
        messages.success(request, 'Sincronización de correos encolada. Se ejecutará en segundo plano.')
    except Exception as e:
        messages.error(request, f'Error al encolar la sincronización: {str(e)}')
    
    from correspondencia.utils.safe_redirect import safe_redirect_back

    return safe_redirect_back(
        request,
        fallback_name='correspondencia:bandeja_correos_pendientes',
    )

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def procesar_emails_imap_manual_view(request):
    """Encola una sincronización puntual por IMAP, independiente de Gmail API."""
    if request.method != 'POST':
        return redirect('correspondencia:bandeja_correos_pendientes')

    try:
        async_result = procesar_emails_imap_manual.delay()
        messages.success(
            request,
            f'Procesamiento IMAP encolado. Task ID: {async_result.id}. Recargue la bandeja en unos minutos.'
        )
    except Exception as e:
        messages.error(request, f'Error al encolar el procesamiento IMAP: {str(e)}')

    from correspondencia.utils.safe_redirect import safe_redirect_back

    return safe_redirect_back(
        request,
        fallback_name='correspondencia:bandeja_correos_pendientes',
    )


# === VISTAS PARA RESPUESTA DE CORRESPONDENCIA (USUARIO REGULAR) ===

@login_required
def crear_o_editar_respuesta(request, correspondencia_entrada_id):
    """Crea una nueva respuesta o edita una existente en estado Borrador o Rechazada."""
    correspondencia_entrada = get_object_or_404(Correspondencia, pk=correspondencia_entrada_id)
    respuesta_existente = CorrespondenciaSalida.objects.filter(respuesta_a=correspondencia_entrada).first()

    puede_responder = usuario_puede_responder_correspondencia(correspondencia_entrada, request.user)
    es_respuesta_discrecional = bool(
        respuesta_existente and respuesta_existente.tipo_respuesta == 'DISCRECIONAL'
    )
    if not es_respuesta_discrecional:
        es_respuesta_discrecional = (
            not correspondencia_entrada.requiere_respuesta and
            usuario_tiene_permiso_respuesta_discrecional(request.user)
        )

    if not puede_responder:
        messages.error(request, "No tienes permiso para responder a esta correspondencia.")
        return redirect('correspondencia:detalle_correspondencia', pk=correspondencia_entrada_id)

    if not correspondencia_entrada.requiere_respuesta and not es_respuesta_discrecional:
         messages.error(request, "Esta correspondencia no requiere respuesta.")
         return redirect('correspondencia:detalle_correspondencia', pk=correspondencia_entrada_id)
        
    # Determinar si se está creando o editando y si el estado permite edición
    if respuesta_existente:
        if respuesta_existente.estado not in ['BORRADOR', 'RECHAZADA']:
            messages.warning(request, f"La respuesta ya está en estado '{respuesta_existente.get_estado_display()}' y no se puede editar.")
            # Redirigir a la vista de detalle de la *respuesta* si existe, o a la entrada
            # return redirect('correspondencia:detalle_respuesta', pk=respuesta_existente.pk) # Crear esta vista luego
            return redirect('correspondencia:detalle_correspondencia', pk=correspondencia_entrada_id)
        instance = respuesta_existente
        titulo_pagina = f"Editando Respuesta a {correspondencia_entrada.numero_radicado}"
    else:
        instance = None
        titulo_pagina = f"Creando Respuesta a {correspondencia_entrada.numero_radicado}"

    if request.method == 'POST':
        form = RespuestaCorrespondenciaForm(
            request.POST,
            request.FILES,
            instance=instance,
            es_respuesta_discrecional=es_respuesta_discrecional,
        )
        if form.is_valid():
            try:
                with transaction.atomic():
                    respuesta = form.save(commit=False)
                    respuesta.tipo_respuesta = 'DISCRECIONAL' if es_respuesta_discrecional else 'OBLIGATORIA'
                    if not es_respuesta_discrecional:
                        respuesta.motivo_respuesta_discrecional = ''
                    if not instance: # Si es nueva
                        respuesta.respuesta_a = correspondencia_entrada
                        respuesta.usuario_redactor = request.user
                        # Asignar explícitamente el destinatario basado en el remitente original
                        respuesta.destinatario_contacto = correspondencia_entrada.remitente
                        # El estado por defecto es BORRADOR (definido en el modelo)
                        respuesta.estado = 'BORRADOR'
                    respuesta.save()

                    # Guardar adjuntos múltiples
                    adjuntos = request.FILES.getlist('adjuntos_respuesta')
                    for f in adjuntos:
                        AdjuntoSalida.objects.create(
                            correspondencia_salida=respuesta,
                            archivo=f,
                            nombre_original=f.name
                        )
                    
                    # Registrar historial
                    evento_historial = 'RESPUESTA_DISCRECIONAL' if (es_respuesta_discrecional and not instance) else ('CREACION' if not instance else 'MODIFICACION')
                    HistorialSalida.objects.create(
                        correspondencia_salida=respuesta,
                        tipo_evento=evento_historial,
                        usuario=request.user,
                        descripcion=(
                            f"Respuesta discrecional guardada. Motivo: {respuesta.motivo_respuesta_discrecional}"
                            if es_respuesta_discrecional else
                            "Borrador guardado."
                        )
                    )
                    
                    # Determinar a dónde ir según el botón presionado
                    if 'enviar_aprobacion' in request.POST:
                        # Cambiar estado y redirigir a aprobación (o a la vista de envío)
                        respuesta.estado = 'PENDIENTE_APROBACION'
                        respuesta.save(update_fields=['estado'])
                        HistorialSalida.objects.create(
                            correspondencia_salida=respuesta, tipo_evento='ENVIO_APROBACION', usuario=request.user
                        )
                        messages.success(request, f"Respuesta enviada a aprobación.")
                        return redirect('correspondencia:detalle_correspondencia', pk=correspondencia_entrada_id)
                    else: # Guardar borrador
                        messages.success(request, f"Borrador de respuesta guardado exitosamente.")
                        # Permanecer en la misma página para seguir editando
                        return redirect('correspondencia:crear_respuesta', correspondencia_entrada_id=correspondencia_entrada_id)
                        
            except Exception as e:
                messages.error(request, f"Error al guardar la respuesta: {e}")
        else:
             messages.error(request, "Por favor corrija los errores en el formulario.")
    else:
        form = RespuestaCorrespondenciaForm(instance=instance, es_respuesta_discrecional=es_respuesta_discrecional)

    context = {
        'form': form,
        'correspondencia_entrada': correspondencia_entrada,
        'respuesta_existente': respuesta_existente,
        'titulo_pagina': titulo_pagina,
        'adjuntos_actuales': instance.adjuntos.all() if instance else None,
        'es_respuesta_discrecional': es_respuesta_discrecional,
    }
    return render(request, 'correspondencia/usuario/respuesta_form.html', context)

# --- VISTAS PARA APROBACIÓN DE RESPUESTAS (VENTANILLA) ---

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def bandeja_respuestas_pendientes(request):
    """Muestra respuestas de salida pendientes de aprobación y aprobadas en pestañas separadas."""
    # Determinar qué pestaña mostrar (pendientes o aprobadas)
    tab = request.GET.get('tab', 'pendientes')

    respuestas_base = CorrespondenciaSalida.objects.annotate(
        destinatarios_total=Count('destinatarios', distinct=True),
        destinatarios_ok=Count('destinatarios', filter=Q(destinatarios__estado='ENVIADO'), distinct=True),
        destinatarios_fallo=Count('destinatarios', filter=Q(destinatarios__estado='FALLO'), distinct=True),
        destinatarios_rebote=Count('destinatarios', filter=Q(destinatarios__estado='REBOTE'), distinct=True),
        destinatarios_pendiente=Count('destinatarios', filter=Q(destinatarios__estado='PENDIENTE'), distinct=True),
    )
    
    if tab == 'aprobadas':
        # Respuestas aprobadas y enviadas
        respuestas = respuestas_base.filter(
            estado__in=['APROBADA', 'ENVIADA']
        ).select_related(
            'respuesta_a', 'usuario_redactor', 'destinatario_contacto', 'usuario_aprobador'
        ).order_by('-fecha_aprobacion', '-fecha_creacion')
        titulo = 'Correspondencia de Salida Aprobada'
    else:
        # Respuestas pendientes de aprobación
        respuestas = respuestas_base.filter(
            estado__in=['PENDIENTE_APROBACION', 'ERROR_ENVIO']
        ).select_related(
            'respuesta_a', 'usuario_redactor', 'destinatario_contacto'
        ).order_by('-fecha_creacion')
        titulo = 'Respuestas Pendientes de Aprobación'
    
    # Contar totales para las pestañas
    total_pendientes = CorrespondenciaSalida.objects.filter(
        estado__in=['PENDIENTE_APROBACION', 'ERROR_ENVIO']
    ).count()
    
    total_aprobadas = CorrespondenciaSalida.objects.filter(
        estado__in=['APROBADA', 'ENVIADA']
    ).count()
    
    context = {
        'respuestas': respuestas,
        'titulo_pagina': titulo,
        'tab': tab,
        'total_pendientes': total_pendientes,
        'total_aprobadas': total_aprobadas
    }
    return render(request, 'correspondencia/admin/bandeja_respuestas.html', context)
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def revisar_respuesta(request, respuesta_id):
    
    respuesta = get_object_or_404(
        CorrespondenciaSalida.objects.select_related(
            'respuesta_a__remitente',
            'respuesta_a__oficina_destino',
            'usuario_redactor__perfil',
            'destinatario_contacto'
        ).prefetch_related('adjuntos', 'historial__usuario'),
        pk=respuesta_id
    )

    if respuesta.estado not in ['PENDIENTE_APROBACION', 'ERROR_ENVIO']:
        messages.warning(request, f"Esta respuesta ya fue procesada (Estado: {respuesta.get_estado_display()}). Puedes revisarla en modo consulta.")
        return redirect('correspondencia:bandeja_respuestas_pendientes')


    # Ajustar si es rechazo
    if request.method == 'POST' and 'rechazar' in request.POST:
        form_rechazo = AprobarRechazarRespuestaForm(request.POST)
    else:
        form_rechazo = AprobarRechazarRespuestaForm()
        form_rechazo.fields['motivo_rechazo'].required = False

    if request.method == 'POST':
        if 'aprobar_enviar' in request.POST:
            from .aprobacion_envio import aprobar_y_enviar_una_respuesta

            try:
                enviados, total = aprobar_y_enviar_una_respuesta(respuesta, request.user)
                if total > 0 and enviados == total:
                    messages.success(request, f"Respuesta enviada a {enviados} destinatarios.")
                elif enviados > 0:
                    messages.warning(
                        request,
                        f"Envío parcial en {respuesta.numero_radicado_salida}: {enviados}/{total} destinatarios OK.",
                    )
                else:
                    messages.error(request, "No se pudo enviar la respuesta a ningún destinatario.")
                return redirect('correspondencia:bandeja_respuestas_pendientes')
            except Exception as e:
                respuesta.estado = 'ERROR_ENVIO'
                respuesta.save(update_fields=['estado'])
                HistorialSalida.objects.create(
                    correspondencia_salida=respuesta,
                    tipo_evento='ENVIO_FALLIDO',
                    usuario=request.user,
                    descripcion=f"Error: {e}\n{traceback.format_exc()}",
                )
                messages.error(request, f"Error al enviar la respuesta: {e}")
                return redirect('correspondencia:revisar_respuesta', respuesta_id=respuesta.id)

        elif 'rechazar' in request.POST:
            if form_rechazo.is_valid():
                try:
                    with transaction.atomic():
                        respuesta.estado = 'RECHAZADA'
                        respuesta.usuario_aprobador = request.user
                        respuesta.fecha_aprobacion = timezone.now()
                        respuesta.motivo_rechazo = form_rechazo.cleaned_data['motivo_rechazo']
                        respuesta.save(update_fields=['estado', 'usuario_aprobador', 'fecha_aprobacion', 'motivo_rechazo'])
                        
                        HistorialSalida.objects.create(
                            correspondencia_salida=respuesta,
                            tipo_evento='RECHAZO',
                            usuario=request.user,
                            descripcion=respuesta.motivo_rechazo
                        )
                    messages.warning(request, f"Respuesta {respuesta.numero_radicado_salida} rechazada.")
                    return redirect('correspondencia:bandeja_respuestas_pendientes')
                except Exception as e:
                    messages.error(request, f"Error al rechazar la respuesta: {e}")
            else:
                messages.error(request, "Debe indicar un motivo para el rechazo.")

    return render(request, 'correspondencia/admin/revisar_respuesta.html', {
        'respuesta': respuesta,
        'correspondencia_original': respuesta.respuesta_a,
        'adjuntos_respuesta': respuesta.adjuntos.all(),
        'historial_respuesta': respuesta.historial.all(),
        'form_rechazo': form_rechazo,
        'titulo_pagina': f"Revisar Respuesta {respuesta.numero_radicado_salida}"
    })
# --- FIN VISTAS RESPUESTA --- 
@login_required
def detalle_respuesta_salida(request, respuesta_id):
    respuesta = get_object_or_404(
        CorrespondenciaSalida.objects.select_related(
            'respuesta_a__remitente',
            'respuesta_a__oficina_destino',
            'usuario_redactor',
            'usuario_aprobador',
            'destinatario_contacto'
        ).prefetch_related('adjuntos', 'historial__usuario', 'destinatarios__contacto'),
        pk=respuesta_id
    )

    if not _usuario_puede_ver_salida(request.user, respuesta):
        messages.error(request, "No tienes permiso para ver esta respuesta.")
        return redirect('correspondencia:home')

    from correspondencia.utils.evidencia_envio import (
        destinatario_entrega_confirmada_servidor,
        destinatario_tiene_problema_entrega,
        resumir_confirmacion_entrega_destinatario,
    )

    destinatarios_salientes = []
    for destinatario in respuesta.destinatarios.all():
        destinatario.mostrar_problema_entrega = destinatario_tiene_problema_entrega(destinatario)
        destinatario.mostrar_entrega_servidor = destinatario_entrega_confirmada_servidor(destinatario)
        if destinatario.mostrar_entrega_servidor and (destinatario.detalle_error or '').strip():
            resumir_confirmacion_entrega_destinatario(destinatario)
        destinatarios_salientes.append(destinatario)

    return render(request, 'correspondencia/admin/detalle_respuesta_salida.html', {
        'respuesta': respuesta,
        'correspondencia_original': respuesta.respuesta_a,
        'adjuntos_respuesta': respuesta.adjuntos.all(),
        'destinatarios_salientes': destinatarios_salientes,
        'historial_respuesta': respuesta.historial.all(),
        'titulo_pagina': f"Detalle Respuesta {respuesta.numero_radicado_salida}",
    })


def _resumir_error_destinatario(destinatario):
    """Compacta rebotes/errores repetidos para una vista más legible."""
    from correspondencia.utils.evidencia_envio import detalle_error_es_solo_confirmacion_entrega

    detalle_error = (getattr(destinatario, 'detalle_error', None) or '').strip()
    if not detalle_error:
        destinatario.detalle_error_resumen = None
        destinatario.detalle_error_compacto = None
        destinatario.detalle_error_repeticiones = 0
        return destinatario

    if detalle_error_es_solo_confirmacion_entrega(detalle_error):
        destinatario.detalle_error_resumen = (
            'El servidor del destinatario aceptó el mensaje (confirmación SMTP 250).'
        )
        destinatario.detalle_error_compacto = detalle_error
        destinatario.detalle_error_repeticiones = 1
        return destinatario

    bloques = []
    actual = []
    for linea in detalle_error.splitlines():
        texto = linea.rstrip()
        if texto.startswith('DSN:') and actual:
            bloques.append('\n'.join(actual).strip())
            actual = [texto]
        else:
            actual.append(texto)

    if actual:
        bloques.append('\n'.join(actual).strip())

    unicos = []
    vistos = set()
    for bloque in bloques or [detalle_error]:
        normalizado = ' '.join(bloque.split())
        if normalizado and normalizado not in vistos:
            vistos.add(normalizado)
            unicos.append(bloque)

    primer_detalle = unicos[0] if unicos else detalle_error
    texto_base = ' '.join(primer_detalle.split())
    texto_base_lower = texto_base.lower()
    smtp_code = getattr(destinatario, 'smtp_code', None)
    dsn_status = getattr(destinatario, 'dsn_status', None)

    if dsn_status == '5.1.1' or 'does not exist' in texto_base_lower or 'nosuchuser' in texto_base_lower:
        resumen = 'La cuenta de correo no existe o la dirección está mal escrita.'
    elif dsn_status == '5.2.2' or 'mailbox full' in texto_base_lower:
        resumen = 'El buzón del destinatario está lleno.'
    elif dsn_status and str(dsn_status).startswith('4.'):
        resumen = 'El servidor reportó un problema temporal de entrega.'
    elif smtp_code and str(smtp_code).startswith('5'):
        resumen = 'El servidor del destinatario rechazó el correo de forma permanente.'
    else:
        resumen = 'Se presentó un problema de entrega con este destinatario.'

    destinatario.detalle_error_resumen = resumen
    destinatario.detalle_error_compacto = primer_detalle
    destinatario.detalle_error_repeticiones = len(bloques) if bloques else 1
    return destinatario


# --- VISTA: Mis rebotes (todos los rebotes del usuario) ---
@login_required
def mis_rebotes(request):
    """Lista todos los rebotes de correspondencia saliente del usuario actual."""
    rebotes_qs = SalidaDestinatario.objects.filter(
        estado='REBOTE',
        correspondencia_salida__usuario_redactor=request.user,
    ).select_related(
        'correspondencia_salida__respuesta_a',
        'contacto',
    ).order_by('-ultimo_evento_at', '-fecha_envio')

    total_rebotes = rebotes_qs.count()

    # Agrupar por salida para resumen
    salidas_con_rebote = (
        CorrespondenciaSalida.objects.filter(
            usuario_redactor=request.user,
            destinatarios__estado='REBOTE',
        )
        .distinct()
        .select_related('respuesta_a')
        .order_by('-fecha_creacion')
    )
    total_salidas_afectadas = salidas_con_rebote.count()

    # Paginar rebotes
    paginator = Paginator(rebotes_qs, 25)
    page = request.GET.get('page')
    try:
        rebotes_page = paginator.page(page)
    except PageNotAnInteger:
        rebotes_page = paginator.page(1)
    except EmptyPage:
        rebotes_page = paginator.page(paginator.num_pages)

    context = {
        'titulo_pagina': 'Mis rebotes de envío',
        'rebotes': rebotes_page,
        'total_rebotes': total_rebotes,
        'total_salidas_afectadas': total_salidas_afectadas,
    }
    return render(request, 'correspondencia/usuario/mis_rebotes.html', context)


# --- NUEVA VISTA: Detalle de Errores/Logs de una salida ---
@login_required
def detalle_respuesta_salida_errores(request, respuesta_id):
    """Muestra un resumen centrado en errores y rebotes de una salida.

    No altera la lógica de la vista principal; se usa como página separada
    para inspección de fallos, rebotes y metadatos de envío.
    """
    respuesta = get_object_or_404(
        CorrespondenciaSalida.objects.select_related(
            'respuesta_a__remitente',
            'respuesta_a__oficina_destino',
            'usuario_redactor',
            'usuario_aprobador',
            'destinatario_contacto'
        ).prefetch_related('adjuntos', 'historial__usuario', 'destinatarios'),
        pk=respuesta_id
    )

    if not _usuario_puede_ver_salida(request.user, respuesta):
        messages.error(request, "No tienes permiso para ver esta respuesta.")
        return redirect('correspondencia:home')

    from correspondencia.utils.evidencia_envio import (
        destinatario_entrega_confirmada_servidor,
        destinatario_tiene_problema_entrega,
        resumir_confirmacion_entrega_destinatario,
    )

    destinatarios_qs = respuesta.destinatarios.all()
    errores_destinatarios = []
    entregas_confirmadas = []
    enviados_sin_confirmacion = []
    for d in destinatarios_qs.order_by('id'):
        if destinatario_tiene_problema_entrega(d):
            errores_destinatarios.append(_resumir_error_destinatario(d))
        elif destinatario_entrega_confirmada_servidor(d):
            entregas_confirmadas.append(resumir_confirmacion_entrega_destinatario(d))
        elif d.estado == 'ENVIADO':
            enviados_sin_confirmacion.append(d)

    total = destinatarios_qs.count()
    total_enviados = destinatarios_qs.filter(estado='ENVIADO').count()
    total_fallos = destinatarios_qs.filter(estado='FALLO').count()
    total_rebotes = destinatarios_qs.filter(estado='REBOTE').count()
    total_entregas_confirmadas = len(entregas_confirmadas)

    historial_qs = respuesta.historial.filter(
        tipo_evento__in=['INTENTO_ENVIO', 'ENVIO_FALLIDO', 'ENVIO_EXITOSO', 'ENTREGA_CONFIRMADA']
    ).order_by('-fecha_hora')

    # Si no hay destinatarios con error actual pero sí hubo rebotes en el historial,
    # mostrar todos los destinatarios con nota informativa.
    hubo_rebotes_historicos = not errores_destinatarios and historial_qs.filter(tipo_evento='ENVIO_FALLIDO').exists()
    todos_destinatarios = []
    if hubo_rebotes_historicos:
        for d in destinatarios_qs.order_by('id'):
            todos_destinatarios.append(d)

    contexto = {
        'respuesta': respuesta,
        'titulo_pagina': f"Errores y Rebotes · {respuesta.numero_radicado_salida}",
        'errores_destinatarios': errores_destinatarios,
        'entregas_confirmadas': entregas_confirmadas,
        'enviados_sin_confirmacion': enviados_sin_confirmacion,
        'todos_destinatarios': todos_destinatarios,
        'hubo_rebotes_historicos': hubo_rebotes_historicos,
        'total_destinatarios': total,
        'total_enviados': total_enviados,
        'total_fallos': total_fallos,
        'total_rebotes': total_rebotes,
        'total_entregas_confirmadas': total_entregas_confirmadas,
        'historial_respuesta': historial_qs,
    }

    return render(request, 'correspondencia/admin/detalle_respuesta_salida_errores.html', contexto)

# --- Vista para Historial Consolidado (MEJORADA) --- 
class HistorialCorrespondenciaView(LoginRequiredMixin, View):
    template_name = 'correspondencia/admin/historial_correspondencia.html'
    paginate_by = 25 # Elementos por página

    def get(self, request, *args, **kwargs):
        form = HistorialFilterForm(request.GET)

        oficina_seleccionada = None
        search_term = None
        serie_seleccionada = None
        subserie_seleccionada = None
        tipo_seleccionado = None
        estado_entrada_seleccionado = None
        estado_salida_seleccionado = None
        fecha_inicio = None
        fecha_fin = None
        remitente_filtro = None
        destinatario_filtro = None
        estado_vencimiento_seleccionado = None

        if form.is_valid():
            oficina_seleccionada = form.cleaned_data.get('oficina')
            search_term = form.cleaned_data.get('search_term')
            serie_seleccionada = form.cleaned_data.get('serie')
            subserie_seleccionada = form.cleaned_data.get('subserie')
            tipo_seleccionado = form.cleaned_data.get('tipo')
            estado_entrada_seleccionado = form.cleaned_data.get('estado_entrada')
            estado_salida_seleccionado = form.cleaned_data.get('estado_salida')
            fecha_inicio = form.cleaned_data.get('fecha_inicio')
            fecha_fin = form.cleaned_data.get('fecha_fin')
            remitente_filtro = form.cleaned_data.get('remitente')
            destinatario_filtro = form.cleaned_data.get('destinatario')
            estado_vencimiento_seleccionado = form.cleaned_data.get('estado_vencimiento')

        historial_combinado = fetch_historial_combinado({
            'oficina': oficina_seleccionada,
            'serie': serie_seleccionada,
            'subserie': subserie_seleccionada,
            'tipo': tipo_seleccionado,
            'estado_entrada': estado_entrada_seleccionado,
            'estado_salida': estado_salida_seleccionado,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'search_term': search_term,
            'remitente': remitente_filtro,
            'destinatario': destinatario_filtro,
            'estado_vencimiento': estado_vencimiento_seleccionado,
        })

        paginator = Paginator(historial_combinado, self.paginate_by)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        hydrate_historial_urls(page_obj.object_list)

        from documentos.models import SerieDocumental, SubserieDocumental

        oficinas_list = cache.get('historial:oficinas_list:v1')
        if oficinas_list is None:
            oficinas_list = list(
                OficinaProductora.objects.values_list('nombre', flat=True).order_by('nombre')
            )
            cache.set('historial:oficinas_list:v1', oficinas_list, 300)
        series_list = cache.get('historial:series_list:v1')
        if series_list is None:
            series_list = list(
                SerieDocumental.objects.values_list('nombre', flat=True).order_by('nombre')
            )
            cache.set('historial:series_list:v1', series_list, 300)
        subseries_list = cache.get('historial:subseries_list:v1')
        if subseries_list is None:
            subseries_list = list(
                SubserieDocumental.objects.values_list('nombre', flat=True).order_by('nombre')
            )
            cache.set('historial:subseries_list:v1', subseries_list, 300)

        context = {
            'titulo_pagina': 'Historial de Correspondencia',
            'form': form,
            'page_obj': page_obj,
            'oficinas_list': oficinas_list,
            'series_list': series_list,
            'subseries_list': subseries_list,
            'filtros_activos': {},
            'today_date': timezone.now().date(),
        }

        if form.is_valid():
            for key, value in form.cleaned_data.items():
                if value:
                    field_label = form.fields[key].label
                    display_value = value
                    if isinstance(value, models.Model):
                        display_value = str(value)
                    elif key in form.fields and hasattr(form.fields[key], 'choices'):
                        display_value = dict(form.fields[key].choices).get(value, value)

                    context['filtros_activos'][key] = {
                        'label': field_label,
                        'value': display_value,
                    }

        return render(request, self.template_name, context)

@login_required
def generar_informe_correspondencia_excel(request):
    """
    Genera un informe Excel de correspondencias entrantes filtradas por fecha.
    
    Campos incluidos:
    - Radicado
    - Fecha de Radicado
    - Fecha de Entrega (primera distribución)
    - Nombre de la Persona o Entidad Remitente
    - Asunto
    - Nombre del Funcionario Responsable del Trámite
    - Firma de Recibido (campo en blanco)
    - Amerita Respuesta (SI/NO)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Obtener fecha del filtro (por defecto hoy)
    fecha_str = request.GET.get('fecha')
    if fecha_str:
        try:
            fecha_filtro = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_filtro = timezone.now().date()
    else:
        fecha_filtro = timezone.now().date()
    
    # Obtener rango de horas (opcional)
    hora_inicio_str = request.GET.get('hora_inicio', '00:00')
    hora_fin_str = request.GET.get('hora_fin', '23:59')
    
    # Parsear horas
    try:
        hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time()
        hora_fin = datetime.strptime(hora_fin_str, '%H:%M').time()
    except ValueError:
        # Si hay error, usar todo el día
        hora_inicio = datetime.strptime('00:00', '%H:%M').time()
        hora_fin = datetime.strptime('23:59', '%H:%M').time()
    
    # Crear datetime inicio y fin para el filtro
    datetime_inicio = timezone.make_aware(datetime.combine(fecha_filtro, hora_inicio))
    datetime_fin = timezone.make_aware(datetime.combine(fecha_filtro, hora_fin))
    
    # Filtrar correspondencias entrantes del día seleccionado en el rango de horas
    correspondencias = Correspondencia.objects.filter(
        tipo_radicado='ENTRANTE',
        fecha_radicacion__gte=datetime_inicio,
        fecha_radicacion__lte=datetime_fin
    ).select_related(
        'remitente',
        'remitente__entidad_externa',
        'usuario_destino_inicial',
        'usuario_destino_inicial__perfil'
    ).prefetch_related(
        'distribuciones_internas'
    ).order_by('fecha_radicacion')
    
    # Marcar las correspondencias como descargadas en planilla
    correspondencias.update(en_planilla=True)
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    # Título con rango de horas si se especificó (sin : porque Excel no lo permite)
    if hora_inicio_str != '00:00' or hora_fin_str != '23:59':
        hora_inicio_clean = hora_inicio_str.replace(':', '')
        hora_fin_clean = hora_fin_str.replace(':', '')
        ws.title = f"Planilla {fecha_filtro.strftime('%d-%m-%Y')} {hora_inicio_clean}-{hora_fin_clean}"
    else:
        ws.title = f"Informe {fecha_filtro.strftime('%d-%m-%Y')}"
    
    # Estilos
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Headers
    headers = [
        'RADICADO',
        'FECHA DE RADICADO',
        'FECHA DE ENTREGA',
        'NOMBRE DE LA PERSONA O ENTIDAD REMITENTE',
        'ASUNTO',
        'NOMBRE DEL FUNCIONARIO RESPONSABLE DEL TRAMITE',
        'FIRMA DE RECIBIDO',
        'AMERITA RESPUESTA'
    ]
    
    # Escribir headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # Escribir datos
    for row_num, corr in enumerate(correspondencias, 2):
        # Radicado
        ws.cell(row=row_num, column=1, value=corr.numero_radicado).border = thin_border
        
        # Fecha de Radicado
        fecha_rad = timezone.localtime(corr.fecha_radicacion).strftime('%d/%m/%Y %H:%M') if corr.fecha_radicacion else ''
        ws.cell(row=row_num, column=2, value=fecha_rad).border = thin_border
        
        # Fecha de Entrega (primera distribución interna)
        primera_distribucion = corr.distribuciones_internas.order_by('fecha_asignacion').first()
        fecha_entrega = ''
        if primera_distribucion:
            fecha_entrega = timezone.localtime(primera_distribucion.fecha_asignacion).strftime('%d/%m/%Y %H:%M')
        ws.cell(row=row_num, column=3, value=fecha_entrega).border = thin_border
        
        # Nombre de la Persona o Entidad Remitente
        remitente_nombre = ''
        if corr.remitente:
            remitente_nombre = corr.remitente.nombre_completo or ''
            if corr.remitente.entidad_externa:
                remitente_nombre = f"{remitente_nombre} - {corr.remitente.entidad_externa.nombre}"
        ws.cell(row=row_num, column=4, value=remitente_nombre).border = thin_border
        
        # Asunto
        ws.cell(row=row_num, column=5, value=corr.asunto or '').border = thin_border
        
        # Nombre del Funcionario Responsable del Trámite (con oficina)
        funcionario = ''
        if corr.usuario_destino_inicial:
            nombre = corr.usuario_destino_inicial.get_full_name() or corr.usuario_destino_inicial.username
            # Agregar la oficina del funcionario
            perfil = getattr(corr.usuario_destino_inicial, 'perfil', None)
            if perfil and perfil.oficina:
                funcionario = f"{nombre} - {perfil.oficina.nombre}"
            else:
                funcionario = nombre
        ws.cell(row=row_num, column=6, value=funcionario).border = thin_border
        
        # Firma de Recibido (campo en blanco)
        ws.cell(row=row_num, column=7, value='').border = thin_border
        
        # Amerita Respuesta
        amerita = 'SI' if corr.requiere_respuesta else 'NO'
        cell_amerita = ws.cell(row=row_num, column=8, value=amerita)
        cell_amerita.border = thin_border
        cell_amerita.alignment = center_alignment
    
    # Ajustar anchos de columna
    column_widths = [20, 18, 18, 40, 50, 35, 20, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Ajustar altura de fila de encabezado
    ws.row_dimensions[1].height = 40
    
    # Registrar la descarga en el historial
    from .models import InformeDiarioCorrespondencia, HistorialDescargaInforme
    
    # Obtener o crear el informe del día
    informe, created = InformeDiarioCorrespondencia.objects.get_or_create(
        fecha=fecha_filtro,
        defaults={'total_correspondencias': correspondencias.count()}
    )
    if not created:
        informe.total_correspondencias = correspondencias.count()
        informe.save(update_fields=['total_correspondencias'])
    
    # Registrar la descarga
    def get_client_ip(request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    
    HistorialDescargaInforme.objects.create(
        informe=informe,
        usuario=request.user,
        ip_address=get_client_ip(request)
    )
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Nombre del archivo con rango de horas si se especificó
    if hora_inicio_str != '00:00' or hora_fin_str != '23:59':
        filename = f"Planilla_{fecha_filtro.strftime('%Y-%m-%d')}_{hora_inicio_str.replace(':', '')}-{hora_fin_str.replace(':', '')}.xlsx"
    else:
        filename = f"Informe_Correspondencia_{fecha_filtro.strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def generar_informe_correspondencia_excel_mensual(request):
    """
    Genera un Excel mensual con todas las correspondencias entrantes del mes.
    """
    import calendar as py_calendar
    from datetime import date, datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    year_str = request.GET.get('year')
    month_str = request.GET.get('month')

    today = timezone.now().date()

    try:
        year = int(year_str) if year_str else today.year
        month = int(month_str) if month_str else today.month
        first_day = date(year, month, 1)
        last_day = date(year, month, py_calendar.monthrange(year, month)[1])
    except (TypeError, ValueError):
        return HttpResponse('Parámetros inválidos. Use year=YYYY&month=MM', status=400)

    correspondencias = Correspondencia.objects.filter(
        tipo_radicado='ENTRANTE',
        fecha_radicacion__date__gte=first_day,
        fecha_radicacion__date__lte=last_day
    ).select_related(
        'remitente',
        'remitente__entidad_externa',
        'usuario_destino_inicial',
        'usuario_destino_inicial__perfil',
        'oficina_destino'
    ).prefetch_related(
        'distribuciones_internas'
    ).order_by('fecha_radicacion')

    correspondencias.update(en_planilla=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Comunicaciones {year}-{month:02d}"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def style_header(worksheet, headers):
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment
        worksheet.row_dimensions[1].height = 28

    def style_cell(cell, alignment=None):
        cell.border = thin_border
        cell.alignment = alignment or left_alignment

    def get_remitente_nombre(corr):
        remitente_nombre = ''
        if corr.remitente:
            remitente_nombre = corr.remitente.nombre_completo or ''
            if corr.remitente.entidad_externa:
                entidad_nombre = corr.remitente.entidad_externa.nombre or ''
                remitente_nombre = f"{remitente_nombre} - {entidad_nombre}" if remitente_nombre else entidad_nombre
        return remitente_nombre

    def get_funcionario(corr):
        if not corr.usuario_destino_inicial:
            return ''
        nombre = corr.usuario_destino_inicial.get_full_name() or corr.usuario_destino_inicial.username
        perfil = getattr(corr.usuario_destino_inicial, 'perfil', None)
        if perfil and perfil.oficina:
            return f"{nombre} - {perfil.oficina.nombre}"
        return nombre

    def get_fecha_entrega(corr):
        primera_distribucion = corr.distribuciones_internas.all()
        primera_distribucion = min(
            primera_distribucion,
            key=lambda item: item.fecha_asignacion,
            default=None
        )
        if not primera_distribucion or not primera_distribucion.fecha_asignacion:
            return ''
        return timezone.localtime(primera_distribucion.fecha_asignacion).strftime('%d/%m/%Y %H:%M')

    headers_principales = [
        'RADICADO',
        'FECHA DE RADICADO',
        'FECHA DE ENTREGA',
        'REMITENTE',
        'ASUNTO',
        'FUNCIONARIO RESPONSABLE',
        'OFICINA',
        'AMERITA RESPUESTA'
    ]
    style_header(ws, headers_principales)

    for row_num, corr in enumerate(correspondencias, 2):
        fecha_rad = timezone.localtime(corr.fecha_radicacion).strftime('%d/%m/%Y %H:%M') if corr.fecha_radicacion else ''
        oficina = corr.oficina_destino.nombre if getattr(corr, 'oficina_destino', None) else ''

        values = [
            corr.numero_radicado,
            fecha_rad,
            get_fecha_entrega(corr),
            get_remitente_nombre(corr),
            corr.asunto or '',
            get_funcionario(corr),
            oficina,
            'SI' if corr.requiere_respuesta else 'NO'
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            alignment = center_alignment if col_num in [2, 3, 8] else left_alignment
            style_cell(cell, alignment)

    column_widths_principales = [20, 20, 20, 35, 45, 35, 24, 16]
    for col_num, width in enumerate(column_widths_principales, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Informe_Correspondencia_Mensual_{year}-{month:02d}.xlsx"'
    )

    wb.save(response)
    return response


@login_required
@user_passes_test(usuario_puede_gestion_operativa, login_url='correspondencia:dashboard_ventanilla')
def calendario_informes_view(request):
    """
    Vista del calendario de informes diarios de correspondencia.
    Muestra un calendario con los días que tienen correspondencias y su estado de firma.
    """
    import calendar
    from datetime import date, timedelta
    from .models import InformeDiarioCorrespondencia
    
    # Obtener mes y año de los parámetros GET (por defecto mes actual)
    today = timezone.now().date()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    
    # Calcular el primer y último día del mes
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    # Obtener correspondencias del mes agrupadas por día
    # order_by por la misma expresión del GROUP BY evita error en SQL Server
    # (el modelo tiene ordering=['-fecha_radicacion'], incompatible con GROUP BY por fecha_radicacion__date)
    correspondencias_por_dia = {}
    correspondencias_mes = Correspondencia.objects.filter(
        tipo_radicado='ENTRANTE',
        fecha_radicacion__date__gte=first_day,
        fecha_radicacion__date__lte=last_day
    ).values('fecha_radicacion__date').annotate(
        total=models.Count('id')
    ).order_by('fecha_radicacion__date')
    
    for item in correspondencias_mes:
        correspondencias_por_dia[item['fecha_radicacion__date']] = item['total']
    
    # Obtener informes del mes
    informes_mes = {
        informe.fecha: informe
        for informe in InformeDiarioCorrespondencia.objects.filter(
            fecha__gte=first_day,
            fecha__lte=last_day
        )
    }
    
    # Construir datos del calendario
    cal = calendar.Calendar(firstweekday=0)  # Lunes como primer día
    semanas = []
    
    for semana in cal.monthdatescalendar(year, month):
        dias_semana = []
        for dia in semana:
            dia_data = {
                'fecha': dia,
                'es_mes_actual': dia.month == month,
                'es_hoy': dia == today,
                'es_futuro': dia > today,
                'total_correspondencias': correspondencias_por_dia.get(dia, 0),
                'informe': informes_mes.get(dia),
                'tiene_correspondencias': dia in correspondencias_por_dia,
            }
            dias_semana.append(dia_data)
        semanas.append(dias_semana)
    
    # Calcular mes anterior y siguiente
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
    
    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year
    
    context = {
        'titulo_pagina': 'Calendario de Planillas',
        'semanas': semanas,
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'today': today,
    }
    
    return render(request, 'correspondencia/admin/calendario_informes.html', context)


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def subir_informe_firmado(request):
    """
    Vista para subir el archivo firmado de un informe diario.
    """
    from .models import InformeDiarioCorrespondencia
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    fecha_str = request.POST.get('fecha')
    archivo = request.FILES.get('archivo_firmado')
    
    if not fecha_str:
        return JsonResponse({'success': False, 'error': 'Fecha no proporcionada'})
    
    if not archivo:
        return JsonResponse({'success': False, 'error': 'Archivo no proporcionado'})
    
    # Validar tipo de archivo (PDF, imágenes)
    allowed_types = ['application/pdf', 'image/jpeg', 'image/png', 'image/jpg']
    if archivo.content_type not in allowed_types:
        return JsonResponse({'success': False, 'error': 'Tipo de archivo no permitido. Use PDF o imagen.'})
    
    # Validar tamaño (máximo 10MB)
    if archivo.size > 10 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'El archivo excede el tamaño máximo de 10MB'})
    
    try:
        from datetime import datetime
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Obtener o crear el informe
        informe, created = InformeDiarioCorrespondencia.objects.get_or_create(
            fecha=fecha,
            defaults={'total_correspondencias': 0}
        )
        
        # Si ya tiene un archivo, eliminarlo
        if informe.archivo_firmado:
            informe.archivo_firmado.delete(save=False)
        
        # Guardar el nuevo archivo
        informe.archivo_firmado = archivo
        informe.estado = 'FIRMADO'
        informe.fecha_subida_firma = timezone.now()
        informe.subido_por = request.user
        informe.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'Archivo firmado subido correctamente para {fecha.strftime("%d/%m/%Y")}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:welcome')
def guardar_firma_correspondencia(request):
    """
    Vista AJAX para guardar la firma digital de una correspondencia.
    Recibe la imagen de firma en base64 desde el canvas.
    """
    from .models import FirmaCorrespondencia
    from django.core.files.base import ContentFile
    import base64
    import uuid
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    correspondencia_id = request.POST.get('correspondencia_id')
    firma_base64 = request.POST.get('firma_imagen')
    nombre_firmante = request.POST.get('nombre_firmante', '').strip()
    cargo_firmante = request.POST.get('cargo_firmante', '').strip()
    observaciones = request.POST.get('observaciones', '').strip()
    
    if not correspondencia_id:
        return JsonResponse({'success': False, 'error': 'ID de correspondencia no proporcionado'})
    
    if not firma_base64:
        return JsonResponse({'success': False, 'error': 'Firma no proporcionada'})
    
    if not nombre_firmante:
        return JsonResponse({'success': False, 'error': 'El nombre del firmante es obligatorio'})
    
    try:
        correspondencia = Correspondencia.objects.get(pk=correspondencia_id)
    except Correspondencia.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Correspondencia no encontrada'})
    
    # Verificar si ya tiene firma
    if hasattr(correspondencia, 'firma_recibida') and correspondencia.firma_recibida:
        return JsonResponse({'success': False, 'error': 'Esta correspondencia ya tiene una firma registrada'})
    
    try:
        # Decodificar la imagen base64
        # El formato viene como: data:image/png;base64,XXXXXX
        if ',' in firma_base64:
            header, firma_data = firma_base64.split(',', 1)
        else:
            firma_data = firma_base64
        
        imagen_bytes = base64.b64decode(firma_data)
        
        # Generar nombre único para el archivo
        filename = f"firma_{correspondencia.numero_radicado}_{uuid.uuid4().hex[:8]}.png"
        
        # Obtener oficina del funcionario destino si existe
        oficina_firmante = None
        if correspondencia.usuario_destino_inicial:
            if hasattr(correspondencia.usuario_destino_inicial, 'perfil'):
                oficina_firmante = correspondencia.usuario_destino_inicial.perfil.oficina
        
        # Obtener IP del cliente
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip_address = x_forwarded_for.split(',')[0]
        else:
            ip_address = request.META.get('REMOTE_ADDR')
        
        # Crear el registro de firma
        firma = FirmaCorrespondencia(
            correspondencia=correspondencia,
            nombre_firmante=nombre_firmante,
            cargo_firmante=cargo_firmante or None,
            oficina_firmante=oficina_firmante,
            firmado_por=correspondencia.usuario_destino_inicial,
            recolector=request.user,
            ip_address=ip_address,
            observaciones=observaciones or None
        )
        
        # Guardar la imagen
        firma.firma_imagen.save(filename, ContentFile(imagen_bytes), save=True)
        
        return JsonResponse({
            'success': True, 
            'message': f'Firma guardada correctamente para {correspondencia.numero_radicado}',
            'firma_id': firma.id,
            'fecha_firma': firma.fecha_firma.strftime('%d/%m/%Y %H:%M')
        })
        
    except Exception as e:
        import traceback
        logger.error(f"Error guardando firma: {e}")
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'error': f'Error al guardar la firma: {str(e)}'})


@login_required
@user_passes_test(usuario_puede_gestion_operativa, login_url='correspondencia:dashboard_ventanilla')
def detalle_dia_informe(request, fecha_str):
    """
    Vista de detalle de un día específico del informe.
    Muestra las correspondencias del día, historial de descargas y permite subir archivo firmado.
    Incluye funcionalidad de recolección de firmas digitales.
    """
    from datetime import datetime
    from .models import InformeDiarioCorrespondencia, HistorialDescargaInforme
    
    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Fecha inválida')
        return redirect('correspondencia:calendario_informes')
    
    # Obtener correspondencias del día con firmas incluidas
    correspondencias = Correspondencia.objects.filter(
        tipo_radicado='ENTRANTE',
        fecha_radicacion__date=fecha
    ).select_related(
        'remitente',
        'remitente__entidad_externa',
        'usuario_destino_inicial',
        'usuario_destino_inicial__perfil',
        'usuario_destino_inicial__perfil__oficina',
        'oficina_destino',
        'firma_recibida'  # Incluir la firma para mostrar estado
    ).order_by('fecha_radicacion')
    
    # Obtener o crear el informe
    informe, created = InformeDiarioCorrespondencia.objects.get_or_create(
        fecha=fecha,
        defaults={'total_correspondencias': correspondencias.count()}
    )
    
    if not created and informe.total_correspondencias != correspondencias.count():
        informe.total_correspondencias = correspondencias.count()
        informe.save(update_fields=['total_correspondencias'])
    
    # Historial de descargas
    historial_descargas = HistorialDescargaInforme.objects.filter(
        informe=informe
    ).select_related('usuario').order_by('-fecha_descarga')[:10]
    
    # Estadísticas de firmas
    total_correspondencias = correspondencias.count()
    total_firmadas = sum(1 for c in correspondencias if hasattr(c, 'firma_recibida') and c.firma_recibida)
    total_pendientes = total_correspondencias - total_firmadas
    
    context = {
        'titulo_pagina': f'Informe del {fecha.strftime("%d/%m/%Y")}',
        'fecha': fecha,
        'fecha_str': fecha_str,
        'correspondencias': correspondencias,
        'informe': informe,
        'historial_descargas': historial_descargas,
        'stats_firmas': {
            'total': total_correspondencias,
            'firmadas': total_firmadas,
            'pendientes': total_pendientes,
            'porcentaje': round((total_firmadas / total_correspondencias * 100) if total_correspondencias > 0 else 0, 1)
        }
    }
    
    return render(request, 'correspondencia/admin/detalle_dia_informe.html', context)


@login_required
def calcular_plazo_sla(request):
    """
    API endpoint para calcular plazo SLA basado en tipo de trámite seleccionado.
    
    El cálculo se basa directamente en el campo dias_respuesta del modelo
    TipoTramite seleccionado por el usuario. Ya no se usa SubserieTramite/TRD.
    
    Parámetros POST:
        tipo_tramite_codigo (str): Código del tipo de trámite seleccionado
        requiere_respuesta (bool): Si la correspondencia requiere respuesta
        check_tipo_tramite (bool): Solo consultar info del tipo de trámite
        
    Returns:
        JsonResponse con plazo_dias, fecha_limite, plazo_origen, etc.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        tipo_tramite_codigo = request.POST.get('tipo_tramite_codigo')
        requiere_respuesta = request.POST.get('requiere_respuesta') == 'true'
        check_tipo_tramite = request.POST.get('check_tipo_tramite') == 'true'
        
        # Consultar info del tipo de trámite sin cálculo completo
        if check_tipo_tramite and tipo_tramite_codigo:
            try:
                tipo_tramite = TipoTramite.objects.get(
                    codigo=tipo_tramite_codigo, activo=True
                )
                
                tiene_dias = tipo_tramite.dias_respuesta is not None
                tiempo_sugerido = None
                if tiene_dias:
                    if tipo_tramite.dias_respuesta <= 3:
                        tiempo_sugerido = 'MUY_URGENTE'
                    elif tipo_tramite.dias_respuesta <= 5:
                        tiempo_sugerido = 'URGENTE'
                    else:
                        tiempo_sugerido = 'NORMAL'
                
                return JsonResponse({
                    'tiene_plazo': tiene_dias,
                    'tramite_codigo': tipo_tramite.codigo,
                    'tramite_nombre': tipo_tramite.nombre,
                    'plazo_dias': tipo_tramite.dias_respuesta,
                    'tiempo_respuesta_sugerido': tiempo_sugerido,
                })
            except TipoTramite.DoesNotExist:
                return JsonResponse({
                    'tiene_plazo': False,
                    'mensaje': 'Tipo de trámite no encontrado o inactivo'
                })
        
        if not requiere_respuesta:
            return JsonResponse({
                'plazo_dias': None,
                'fecha_limite': None,
                'plazo_origen': 'No requiere respuesta',
                'corte_horario': get_cutoff_time().strftime('%H:%M')
            })
        
        if not tipo_tramite_codigo:
            return JsonResponse({'error': 'Debe seleccionar un tipo de trámite.'}, status=400)

        plazo_dias = None
        plazo_origen = "No configurado"
        
        # Obtener plazo desde el tipo de trámite seleccionado
        try:
            tipo_tramite = TipoTramite.objects.get(
                codigo=tipo_tramite_codigo, activo=True
            )
            if tipo_tramite.dias_respuesta is not None:
                plazo_dias = tipo_tramite.dias_respuesta
                plazo_origen = f"Tipo de trámite: {tipo_tramite.codigo} ({tipo_tramite.nombre})"
        except TipoTramite.DoesNotExist:
            pass
        
        if plazo_dias is None:
            return JsonResponse({
                'plazo_dias': None,
                'fecha_limite': None,
                'plazo_origen': 'Sin plazo configurado para este tipo de trámite',
                'corte_horario': get_cutoff_time().strftime('%H:%M'),
            })
        
        # Calcular fecha límite
        ahora = timezone.now()
        inicio = aplicar_corte(ahora)
        fecha_limite = sumar_habiles(inicio, plazo_dias)
        
        return JsonResponse({
            'plazo_dias': plazo_dias,
            'fecha_limite': fecha_limite.strftime('%d/%m/%Y %H:%M'),
            'plazo_origen': plazo_origen,
            'corte_horario': get_cutoff_time().strftime('%H:%M'),
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def buscar_categorias_ajax(request):
    """Alias de buscar_grupos_ajax para compatibilidad con el frontend."""
    response = buscar_grupos_ajax(request)
    if hasattr(response, 'content'):
        import json
        data = json.loads(response.content.decode('utf-8'))
        if data.get('success') and 'grupos' in data:
            # Cambiar 'grupos' por 'categorias' para compatibilidad con el frontend
            data['categorias'] = data.pop('grupos')
            # También cambiar 'total' por 'contactos_count' para compatibilidad
            for categoria in data['categorias']:
                if 'total' in categoria:
                    categoria['contactos_count'] = categoria.pop('total')
        response.content = json.dumps(data).encode('utf-8')
    return response
 
@login_required
def categoria_detalle_ajax(request, pk: int):
    """Alias de grupo_agenda_detalle_ajax para compatibilidad con el frontend."""
    response = grupo_agenda_detalle_ajax(request, pk)
    if hasattr(response, 'content'):
        import json
        data = json.loads(response.content.decode('utf-8'))
        if data.get('success') and 'grupo' in data:
            # Cambiar 'grupo' por 'categoria' para compatibilidad con el frontend
            data['categoria'] = data.pop('grupo')
        response.content = json.dumps(data).encode('utf-8')
    return response

@login_required
def categoria_detalle_view(request, pk):
    """
    Vista para mostrar el detalle de una categoría de contactos
    """
    try:
        # Obtener el perfil del usuario y su oficina
        perfil_usuario = getattr(request.user, 'perfil', None)
        if not perfil_usuario or not perfil_usuario.oficina:
            messages.error(request, 'Tu usuario no tiene oficina asignada. Contacta al administrador.')
            return redirect('correspondencia:dashboard_usuario')
        
        # Obtener la categoría y verificar que pertenece a la oficina del usuario
        categoria = get_object_or_404(
            GrupoAgenda, 
            id=pk, 
            oficina_propietaria=perfil_usuario.oficina,
            activo=True
        )
        
        # Obtener todos los contactos de esta categoría usando la relación many-to-many
        contactos = categoria.contactos.select_related('entidad_externa').order_by('nombres')
        
        # --- Filtros GET (según estándares) ---
        search_term = (request.GET.get('search_term') or '').strip()
        entidad = (request.GET.get('entidad') or '').strip()
        tiene_email = (request.GET.get('tiene_email') or '').strip()
        tiene_telefono = (request.GET.get('tiene_telefono') or '').strip()

        # Aplicar filtros
        if search_term:
            contactos = contactos.filter(
                Q(nombres__icontains=search_term) |
                Q(apellidos__icontains=search_term) |
                Q(correo_electronico__icontains=search_term) |
                Q(entidad_externa__nombre__icontains=search_term)
            )

        if entidad and entidad.isdigit():
            contactos = contactos.filter(entidad_externa_id=int(entidad))

        if tiene_email:
            if tiene_email == 'true':
                contactos = contactos.filter(correo_electronico__isnull=False).exclude(correo_electronico='')
            elif tiene_email == 'false':
                contactos = contactos.filter(Q(correo_electronico__isnull=True) | Q(correo_electronico=''))

        if tiene_telefono:
            if tiene_telefono == 'true':
                contactos = contactos.filter(telefono_contacto__isnull=False).exclude(telefono_contacto='')
            elif tiene_telefono == 'false':
                contactos = contactos.filter(Q(telefono_contacto__isnull=True) | Q(telefono_contacto=''))
        
        # Estadísticas básicas (después de aplicar filtros)
        total_contactos = contactos.count()
        contactos_con_email = contactos.filter(correo_electronico__isnull=False).exclude(correo_electronico='').count()
        contactos_con_telefono = contactos.filter(telefono_contacto__isnull=False).exclude(telefono_contacto='').count()
        
        # Obtener entidades externas únicas
        entidades_unicas = contactos.values('entidad_externa__nombre').distinct().count()
        
        # Obtener entidades disponibles para el filtro
        entidades_disponibles = EntidadExterna.objects.filter(
            contactos__in=categoria.contactos.all()
        ).distinct().order_by('nombre')
        
        context = {
            'titulo_pagina': f'Detalle de Categoría: {categoria.nombre}',
            'categoria': categoria,
            'contactos': contactos,
            'total_contactos': total_contactos,
            'contactos_con_email': contactos_con_email,
            'contactos_con_telefono': contactos_con_telefono,
            'entidades_unicas': entidades_unicas,
            'entidades_disponibles': entidades_disponibles,
            # Valores de filtros para rehidratación en la UI
            'filtro_search_term': search_term,
            'filtro_entidad': entidad,
            'filtro_tiene_email': tiene_email,
            'filtro_tiene_telefono': tiene_telefono,
        }
        
        return render(request, 'correspondencia/categoria_detalle.html', context)
        
    except Http404:
        messages.error(request, 'La categoría no existe o no tienes permisos para verla.')
        return redirect('correspondencia:dashboard_usuario')
    except Exception as e:
        messages.error(request, f'Error al cargar la categoría: {str(e)}')
        return redirect('correspondencia:dashboard_usuario')

@login_required
@user_passes_test(lambda u: not u.groups.filter(name='Ventanilla').exists(), login_url='correspondencia:pendientes_distribuir')
def bandeja_respuestas_salientes(request):
    """Muestra la bandeja de respuestas salientes del usuario (correspondencias de salida que ha creado)."""
    usuario_actual = request.user
    perfil_usuario = getattr(usuario_actual, 'perfil', None)
    respuestas_list = []
    load_success = False

    # Parámetros de filtro
    search_term = request.GET.get('search_term', '').strip()
    estado_salida = request.GET.get('estado_salida', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    alcance = (request.GET.get('alcance') or 'mias').strip().lower()
    if alcance not in {'mias', 'oficina'}:
        alcance = 'mias'

    if perfil_usuario and perfil_usuario.oficina:
        try:
            respuestas_qs = _base_bandeja_salientes_queryset(
                usuario_actual,
                alcance=alcance,
                oficina=perfil_usuario.oficina,
            )

            if search_term:
                respuestas_qs = respuestas_qs.filter(
                    Q(asunto__icontains=search_term) |
                    Q(numero_radicado_salida__icontains=search_term) |
                    Q(destinatario_contacto__nombres__icontains=search_term) |
                    Q(destinatario_contacto__apellidos__icontains=search_term) |
                    Q(destinatario_contacto__entidad_externa__nombre__icontains=search_term)
                )

            if estado_salida:
                respuestas_qs = respuestas_qs.filter(estado=estado_salida)

            if fecha_inicio:
                try:
                    respuestas_qs = respuestas_qs.filter(fecha_creacion__date__gte=fecha_inicio)
                except Exception:
                    pass

            if fecha_fin:
                try:
                    respuestas_qs = respuestas_qs.filter(fecha_creacion__date__lte=fecha_fin)
                except Exception:
                    pass

            stats = respuestas_qs.aggregate(
                total_respuestas=Count('pk'),
                enviadas_total=Count('pk', filter=Q(estado='ENVIADA')),
                pendientes_total=Count('pk', filter=Q(estado='PENDIENTE_APROBACION')),
                compartidas_total=Count('pk', filter=~Q(usuario_redactor=usuario_actual)),
                propias_total=Count('pk', filter=Q(usuario_redactor=usuario_actual)),
            )
            total_respuestas = stats['total_respuestas']
            enviadas_total = stats['enviadas_total']
            pendientes_total = stats['pendientes_total']
            compartidas_total = stats['compartidas_total']
            propias_total = stats['propias_total']

            ordered_qs = respuestas_qs.order_by('-fecha_creacion')
            paginator = Paginator(ordered_qs, 25)
            paginator.__dict__['count'] = total_respuestas
            page_number = request.GET.get('page', 1)
            try:
                respuestas_page = paginator.page(page_number)
            except (EmptyPage, PageNotAnInteger):
                respuestas_page = paginator.page(1)

            page_offset = (respuestas_page.number - 1) * paginator.per_page
            page_pks = list(
                ordered_qs.values_list('pk', flat=True)[
                    page_offset:page_offset + paginator.per_page
                ]
            )
            if page_pks:
                annotated_by_pk = {
                    obj.pk: obj
                    for obj in _annotate_bandeja_salientes_page(
                        CorrespondenciaSalida.objects.filter(pk__in=page_pks),
                        usuario_actual,
                    )
                }
                respuestas_list = [
                    annotated_by_pk[pk] for pk in page_pks if pk in annotated_by_pk
                ]
            else:
                respuestas_list = []

            load_success = True
            
        except Exception as e:
            print(f"Error crítico cargando bandeja de respuestas salientes {usuario_actual.username}: {e}")
            messages.error(request, f"Ocurrió un error inesperado al cargar tu bandeja de salientes: {e}")
            load_success = False
    else:
        messages.warning(request, "No tienes un perfil o una oficina asignada para ver tu bandeja de salientes.")
        load_success = False

    filtro_query_string = ''
    if load_success:
        filtro_params = request.GET.copy()
        filtro_params.pop('page', None)
        filtro_query_string = filtro_params.urlencode()

    context = {
        'page_title': 'Mis Respuestas Salientes' if alcance == 'mias' else 'Respuestas de Mi Oficina',
        'respuestas': respuestas_list,
        'respuestas_page': respuestas_page if 'respuestas_page' in locals() else None,
        'load_success': load_success,
        'tipo_tabla': 'salientes',
        'alcance_actual': alcance if 'alcance' in locals() else 'mias',
        'total_respuestas': total_respuestas if 'total_respuestas' in locals() else 0,
        'enviadas_total': enviadas_total if 'enviadas_total' in locals() else 0,
        'pendientes_total': pendientes_total if 'pendientes_total' in locals() else 0,
        'compartidas_total': compartidas_total if 'compartidas_total' in locals() else 0,
        'propias_total': propias_total if 'propias_total' in locals() else 0,
        # Valores de filtros para rehidratación en la UI
        'filtro_search_term': search_term,
        'filtro_estado_salida': estado_salida,
        'filtro_fecha_inicio': fecha_inicio,
        'filtro_fecha_fin': fecha_fin,
        'filtro_alcance': alcance if 'alcance' in locals() else 'mias',
        'filtro_query_string': filtro_query_string,
        # Opciones para los filtros
        'estados_salida': [
            ('BORRADOR', 'Borrador'),
            ('PENDIENTE_APROBACION', 'Pendiente Aprobación'),
            ('RECHAZADA', 'Rechazada'),
            ('ENVIADA', 'Enviada'),
            ('ERROR_ENVIO', 'Error de Envío'),
        ],
    }
    return render(request, 'correspondencia/usuario/bandeja_respuestas_salientes.html', context)

# ===============================
# === Validación de Email AJAX ===
# ===============================

EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


@login_required
@require_GET
def validar_email_mx_ajax(request):
    """Valida rápidamente formato y existencia de registros MX del dominio.

    Parámetros:
      - email: dirección a validar
    Respuesta JSON:
      { success: bool, formato_valido: bool, tiene_mx: bool, dominio: str, mx_hosts: [str], message: str }
    """
    email = (request.GET.get('email') or '').strip()
    response = {
        'success': False,
        'formato_valido': False,
        'tiene_mx': False,
        'dominio': '',
        'mx_hosts': [],
        'message': ''
    }
    try:
        if not email:
            response['message'] = 'Falta el parámetro email'
            return JsonResponse(response, status=400)

        formato_valido = bool(EMAIL_REGEX.match(email))
        response['formato_valido'] = formato_valido
        if not formato_valido:
            response['message'] = 'Formato de email inválido'
            return JsonResponse(response, status=200)

        dominio = email.split('@')[-1].lower()
        # Normalizar IDN a punycode si aplica
        try:
            dominio = dominio.encode('idna').decode('ascii')
        except Exception:
            pass
        response['dominio'] = dominio

        policy = _build_domain_policy(dominio, entidad_id=request.GET.get('entidad_id'))
        response.update(policy)

        if response['known_in_system']:
            response['success'] = True
            response['tiene_mx'] = True
            response['mx_hosts'] = []
            if response['known_for_entity']:
                response['message'] = 'Dominio registrado para la entidad seleccionada'
            elif response['matching_entity_name']:
                response['message'] = f"Dominio conocido en sistema. Está asociado a '{response['matching_entity_name']}'."
            else:
                response['message'] = 'Dominio conocido en sistema'
            return JsonResponse(response)

        if dns is None:
            response['message'] = 'No disponible: instale dnspython para verificar MX (pip install dnspython)'
            return JsonResponse(response, status=200)

        try:
            # Usar resolutores públicos para evitar problemas de red local
            resolver = dns.resolver.Resolver()  # type: ignore
            resolver.timeout = 3.0
            resolver.lifetime = 3.0
            # Intentar con múltiples servidores públicos
            resolver.nameservers = ['8.8.8.8', '1.1.1.1', '9.9.9.9']  # Google, Cloudflare, Quad9

            answers = resolver.resolve(dominio, 'MX')  # type: ignore
            mx_hosts = sorted([(r.preference, str(r.exchange).rstrip('.')) for r in answers])
            response['tiene_mx'] = len(mx_hosts) > 0
            response['mx_hosts'] = [host for _, host in mx_hosts]
            response['success'] = True
            if response['tiene_mx'] and response['selected_entity_id'] and not response['known_for_entity']:
                response['message'] = 'Dominio válido con MX. No está registrado para la entidad seleccionada.'
            else:
                response['message'] = 'Dominio con MX válido' if response['tiene_mx'] else 'Dominio sin MX'
            # Si no hay MX, comprobar si existe A/AAAA como fallback informativo
            if not response['tiene_mx']:
                try:
                    a_answers = resolver.resolve(dominio, 'A')  # type: ignore
                    if a_answers:
                        response['message'] = 'Sin MX pero el dominio tiene A (algunos servidores aceptan SMTP).'
                except Exception:
                    pass
            return JsonResponse(response)
        except Exception as e:
            # Último intento: consultar A para indicar si el dominio existe
            try:
                resolver = dns.resolver.Resolver()  # type: ignore
                resolver.timeout = 3.0
                resolver.lifetime = 3.0
                resolver.nameservers = ['8.8.8.8', '1.1.1.1', '9.9.9.9']
                a_answers = resolver.resolve(dominio, 'A')  # type: ignore
                if a_answers:
                    response['message'] = f'No se pudieron resolver MX ({e}), pero el dominio responde A.'
                    # Tratar como inconcluso-permitido: no bloquear la verificación posterior
                    response['success'] = True
                    return JsonResponse(response)
                else:
                    response['message'] = f'No se pudieron resolver MX ({e})'
            except Exception:
                response['message'] = f'No se pudieron resolver MX ({e})'
            response['tiene_mx'] = False
            return JsonResponse(response)

    except Exception as e:  # salvaguarda
        response['message'] = f'Error inesperado: {e}'
        return JsonResponse(response, status=500)
def validar_email_smtp_ajax(request):
    """Opción avanzada y opcional: intenta RCPT TO contra el MX principal.

    Importante: muchos servidores rechazan o devuelven 450/550 por políticas anti-spam.
    No bloquea el flujo de creación de contacto.

    Parámetros:
      - email: dirección a validar
    Respuesta JSON:
      { success: bool, conecto: bool, codigo: int|None, mensaje: str, host_usado: str|None }
    """
    email = (request.GET.get('email') or '').strip()
    timeout_seconds = 6
    result = {
        'success': False,
        'conecto': False,
        'codigo': None,
        'mensaje': '',
        'host_usado': None,
    }
    try:
        if not email or not EMAIL_REGEX.match(email):
            result['mensaje'] = 'Email vacío o con formato inválido'
            return JsonResponse(result, status=200)

        dominio = email.split('@')[-1].lower()

        # Resolver MX (si no hay dnspython, devolvemos mensaje informativo)
        mx_hosts = []
        if dns is not None:
            try:
                resolver = dns.resolver.Resolver()  # type: ignore
                resolver.timeout = 3.0
                resolver.lifetime = 3.0
                resolver.nameservers = ['8.8.8.8', '1.1.1.1', '9.9.9.9']
                answers = resolver.resolve(dominio, 'MX')  # type: ignore
                mx_hosts = sorted([(r.preference, str(r.exchange).rstrip('.')) for r in answers])
            except Exception as e:
                result['mensaje'] = f'No se pudieron resolver MX: {e}'
        else:
            result['mensaje'] = 'dnspython no instalado; no se puede resolver MX (pip install dnspython)'

        if not mx_hosts:
            return JsonResponse(result, status=200)

        host = mx_hosts[0][1]
        result['host_usado'] = host

        # Intento SMTP simple sin TLS (inbound MX normalmente acepta plano para handshake)
        try:
            with smtplib.SMTP(host=host, port=25, timeout=timeout_seconds) as server:
                server.ehlo_or_helo_if_needed()
                # MAIL FROM vacío está permitido para consultas
                server.mail('')
                code, msg = server.rcpt(email)
                result['conecto'] = True
                result['codigo'] = int(code)
                # Interpretación básica de códigos
                if code in (250, 251, 252):
                    result['success'] = True
                    result['mensaje'] = f'Aparentemente aceptado (código {code})'
                elif code in (450, 451, 452):
                    result['success'] = False
                    result['mensaje'] = f'Respuesta temporal (código {code})'
                else:
                    result['success'] = False
                    result['mensaje'] = f'Rechazado o desconocido (código {code})'
        except (socket.timeout, smtplib.SMTPConnectError):
            result['mensaje'] = 'Timeout o no fue posible conectar al MX'
        except Exception as e:
            result['mensaje'] = f'Error SMTP: {e}'

        return JsonResponse(result)

    except Exception as e:  # salvaguarda
        result['mensaje'] = f'Error inesperado: {e}'
        return JsonResponse(result, status=500)


@login_required
@require_GET
def validar_email_api_ajax(request):
    """Valida un email usando un proveedor externo configurable.

    Configuración esperada en settings:
      EMAIL_VERIFIER_API_URL (str): Endpoint base
      EMAIL_VERIFIER_API_KEY (str): API key
      EMAIL_VERIFICATION_OFFLINE_MODE (bool): Si True, solo usa validación MX local

    La implementación asume un esquema simple ?email=&api_key= y response con campos
    { success: bool, result: 'deliverable'|'undeliverable'|'risky'|'unknown', reason: str }
    """
    import logging
    logger = logging.getLogger(__name__)
    
    email = (request.GET.get('email') or '').strip()
    if not email:
        return JsonResponse({'success': False, 'message': 'Falta email'}, status=400)

    # Validación mínima local de formato para evitar llamadas innecesarias
    if not EMAIL_REGEX.match(email):
        return JsonResponse({'success': False, 'message': 'Formato inválido'}, status=200)

    # Modo offline: usar solo validación MX local
    if getattr(settings, 'EMAIL_VERIFICATION_OFFLINE_MODE', False):
        logger.info(f"Modo offline activado para {email}")
        # Reutilizar lógica de validar_email_mx_ajax
        dominio = email.split('@')[-1].lower()
        try:
            dominio = dominio.encode('idna').decode('ascii')
        except Exception:
            pass
        policy = _build_domain_policy(dominio, entidad_id=request.GET.get('entidad_id'))
            
        if policy['known_in_system']:
            return JsonResponse({
                'success': True,
                'result': 'deliverable',
                'message': 'Dominio conocido (modo offline)',
                **policy,
            })
            
        # Si no está en lista blanca, usar validación MX
        try:
            import dns.resolver
            resolver = dns.resolver.Resolver()
            resolver.timeout = 3.0
            resolver.lifetime = 3.0
            resolver.nameservers = ['8.8.8.8', '1.1.1.1', '9.9.9.9']
            answers = resolver.resolve(dominio, 'MX')
            if answers:
                return JsonResponse({
                    'success': True, 
                    'result': 'deliverable', 
                    'message': 'MX encontrado (modo offline)',
                    **policy,
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'result': 'undeliverable', 
                    'message': 'No MX encontrado (modo offline)',
                    **policy,
                })
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'result': 'unknown', 
                'message': f'Error DNS: {str(e)} (modo offline)',
                **policy,
            })

    api_url = getattr(settings, 'EMAIL_VERIFIER_API_URL', '')
    api_key = getattr(settings, 'EMAIL_VERIFIER_API_KEY', '')
    api_key_param = getattr(settings, 'EMAIL_VERIFIER_API_KEY_PARAM', 'api_key')
    extra_params = getattr(settings, 'EMAIL_VERIFIER_API_EXTRA_PARAMS', {})
    # Si el endpoint de Abstract no tiene query path, asegurarlo
    if api_url.endswith('/'):
        api_url = api_url.rstrip('/')
    
    # Debug logging
    logger.info(f"=== DEBUG API EMAIL VALIDATION ===")
    logger.info(f"Email: {email}")
    logger.info(f"API URL: {api_url}")
    logger.info(f"API Key: {api_key[:10]}..." if api_key else "No API Key")
    logger.info(f"API Key Param: {api_key_param}")
    logger.info(f"Extra params: {extra_params}")
    
    if not api_url or not api_key:
        logger.error("Servicio no configurado - falta URL o API key")
        return JsonResponse({'success': False, 'message': 'Servicio no configurado'}, status=200)

    try:
        params = {**extra_params}
        params['email'] = email
        params[api_key_param] = api_key
        
        logger.info(f"Parámetros de la petición: {params}")
        logger.info(f"Enviando petición a: {api_url}")
        
        # Configurar timeout específico para conexión y lectura
        resp = requests.get(api_url, params=params, timeout=(10, 30))  # (conexión, lectura)
        
        logger.info(f"Respuesta HTTP: {resp.status_code}")
        logger.info(f"Headers respuesta: {dict(resp.headers)}")
        
        if resp.status_code != 200:
            logger.error(f"Error HTTP: {resp.status_code}")
            logger.error(f"Respuesta texto: {resp.text}")
            return JsonResponse({'success': False, 'message': f'Error proveedor: {resp.status_code}'}, status=200)
        
        data = resp.json()
        logger.info(f"Respuesta API: {data}")
        # Normalizar salida para MailboxLayer
        success = False
        result_label = 'unknown'
        reason = ''

        # MailboxLayer estructura
        if 'smtp_check' in data or 'mx_found' in data:
            smtp_ok = bool(data.get('smtp_check'))
            mx_ok = bool(data.get('mx_found'))
            score = data.get('score')
            success = smtp_ok or (mx_ok and (score is None or float(score) >= 0.5))
            result_label = 'deliverable' if success else 'undeliverable'
            reason = 'smtp_check' if smtp_ok else ('mx_found' if mx_ok else 'no_mx')

        # Abstract API estructura (fallback)
        elif 'email_deliverability' in data:
            deliver = data['email_deliverability']
            status = deliver.get('status')  # 'deliverable' | 'undeliverable' | 'risky' | 'unknown'
            result_label = status or 'unknown'
            is_mx_valid = bool(deliver.get('is_mx_valid'))
            is_smtp_valid = bool(deliver.get('is_smtp_valid'))
            is_format_valid = bool(deliver.get('is_format_valid', True))
            success = (status == 'deliverable') or (is_smtp_valid or is_mx_valid) and is_format_valid
            reason = deliver.get('status_detail') or ''

        # Si no hay claves conocidas, marcar unknown

        result = {
            'success': bool(data.get('success', False)) or success,
            'result': result_label,
            'reason': reason,
            'raw': data,
        }
        logger.info(f"Resultado final: {result}")
        return JsonResponse(result)
    except requests.exceptions.ConnectTimeout:
        logger.error("Timeout de conexión a la API")
        return JsonResponse({'success': False, 'result': 'timeout', 'message': 'Timeout de conexión al proveedor'}, status=200)
    except requests.exceptions.ReadTimeout:
        logger.error("Timeout de lectura de la API")
        return JsonResponse({'success': False, 'result': 'timeout', 'message': 'Timeout de lectura del proveedor'}, status=200)
    except requests.exceptions.Timeout:
        logger.error("Timeout general en la petición a la API")
        return JsonResponse({'success': False, 'result': 'timeout', 'message': 'Timeout del proveedor'}, status=200)
    except Exception as e:
        logger.error(f"Error inesperado: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {e}'}, status=200)


@login_required
@require_GET
def test_api_config(request):
    """Endpoint para probar la configuración de la API de validación de emails."""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        api_url = getattr(settings, 'EMAIL_VERIFIER_API_URL', '')
        api_key = getattr(settings, 'EMAIL_VERIFIER_API_KEY', '')
        api_key_param = getattr(settings, 'EMAIL_VERIFIER_API_KEY_PARAM', 'api_key')
        extra_params = getattr(settings, 'EMAIL_VERIFIER_API_EXTRA_PARAMS', {})
        
        config_info = {
            'api_url': api_url,
            'api_key_length': len(api_key) if api_key else 0,
            'api_key_param': api_key_param,
            'extra_params': extra_params,
            'has_api_key': bool(api_key),
            'has_url': bool(api_url)
        }
        
        logger.info(f"Configuración API: {config_info}")
        
        # Test con un email de prueba
        test_email = "test@gmail.com"
        params = {**extra_params}
        params['email'] = test_email
        params[api_key_param] = api_key
        
        logger.info(f"Probando con email: {test_email}")
        logger.info(f"Parámetros: {params}")
        
        response = requests.get(api_url, params=params, timeout=(10, 30))  # (conexión, lectura)
        
        test_result = {
            'config': config_info,
            'test_email': test_email,
            'http_status': response.status_code,
            'response_headers': dict(response.headers),
            'response_text': response.text[:500] if response.text else '',
            'response_json': None
        }
        
        try:
            test_result['response_json'] = response.json()
        except:
            pass
            
        logger.info(f"Resultado del test: {test_result}")
        
        return JsonResponse(test_result)
        
    except Exception as e:
        logger.error(f"Error en test_api_config: {str(e)}")
        return JsonResponse({
            'error': str(e),
            'config': {
                'api_url': getattr(settings, 'EMAIL_VERIFIER_API_URL', ''),
                'api_key_length': len(getattr(settings, 'EMAIL_VERIFIER_API_KEY', '')),
                'api_key_param': getattr(settings, 'EMAIL_VERIFIER_API_KEY_PARAM', ''),
                'extra_params': getattr(settings, 'EMAIL_VERIFIER_API_EXTRA_PARAMS', {})
            }
        })


@login_required
@require_GET
def simulate_timeout_response(request):
    """Simula la respuesta que debería darse cuando hay timeout en la API."""
    email = request.GET.get('email', 'test@example.com')
    
    # Simular respuesta de timeout
    return JsonResponse({
        'success': False,
        'result': 'timeout',
        'message': 'La verificación externa no pudo completarse. ¿Está seguro de que desea guardar este correo de todas formas?'
    })


# =============================================
# === VISTAS DE NOTIFICACIONES ===
# =============================================

@login_required
def obtener_notificaciones(request):
    """
    Endpoint AJAX para obtener notificaciones del usuario actual.
    Devuelve notificaciones no leídas y las últimas 10 leídas.
    """
    usuario = request.user
    
    # Obtener notificaciones no leídas
    notificaciones_no_leidas = Notificacion.objects.filter(
        usuario=usuario,
        leida=False
    ).select_related('correspondencia').order_by('-fecha_creacion')[:20]
    
    # Obtener últimas notificaciones leídas (máximo 10)
    notificaciones_leidas = Notificacion.objects.filter(
        usuario=usuario,
        leida=True
    ).select_related('correspondencia').order_by('-fecha_creacion')[:10]
    
    # Serializar notificaciones
    def serializar_notificacion(notif):
        return {
            'id': notif.id,
            'tipo': notif.tipo,
            'titulo': notif.titulo,
            'mensaje': notif.mensaje,
            'leida': notif.leida,
            'fecha_creacion': notif.fecha_creacion.isoformat(),
            'url': notif.url or '#',
            'correspondencia_id': notif.correspondencia_id if notif.correspondencia else None,
        }
    
    data = {
        'success': True,
        'no_leidas': [serializar_notificacion(n) for n in notificaciones_no_leidas],
        'leidas': [serializar_notificacion(n) for n in notificaciones_leidas],
        'total_no_leidas': notificaciones_no_leidas.count(),
    }
    
    return JsonResponse(data)


@login_required
@require_POST
def marcar_notificacion_leida(request, notificacion_id):
    """Marca una notificación como leída."""
    try:
        notificacion = Notificacion.objects.get(
            id=notificacion_id,
            usuario=request.user
        )
        notificacion.marcar_leida()
        
        return JsonResponse({
            'success': True,
            'message': 'Notificación marcada como leída'
        })
    except Notificacion.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Notificación no encontrada'
        }, status=404)


@login_required
@require_POST
def marcar_todas_notificaciones_leidas(request):
    """Marca todas las notificaciones del usuario como leídas."""
    count = Notificacion.objects.filter(
        usuario=request.user,
        leida=False
    ).update(
        leida=True,
        fecha_lectura=timezone.now()
    )
    
    return JsonResponse({
        'success': True,
        'message': f'{count} notificaciones marcadas como leídas',
        'count': count
    })


def crear_notificacion_compartir_oficina(correspondencia, usuario, usuario_actual, observaciones):
    """
    Función auxiliar para crear notificaciones cuando se comparte correspondencia con oficina.
    """
    from .models import Notificacion
    from django.urls import reverse
    
    titulo = "Correspondencia compartida con tu oficina"
    mensaje = f"Se compartió contigo la correspondencia {correspondencia.numero_radicado}"
    if correspondencia.remitente:
        mensaje += f" de {correspondencia.remitente.nombre_completo}"
    mensaje += f". Asunto: {correspondencia.asunto[:100]}"
    if observaciones:
        mensaje += f"\nObservaciones: {observaciones[:100]}"
    
    url = reverse('correspondencia:detalle_correspondencia', kwargs={'pk': correspondencia.pk})
    
    Notificacion.objects.create(
        usuario=usuario,
        tipo='compartido',
        titulo=titulo,
        mensaje=mensaje,
        correspondencia=correspondencia,
        url=url
    )


def crear_notificaciones_acceso_oficina(correspondencia, oficina, usuario_actual, observaciones):
    """
    Crea notificaciones para todos los usuarios de la oficina que recibe acceso de solo lectura.
    """
    from django.urls import reverse
    from .models import AccesoCorrespondenciaOficina  # Importación necesaria
    logger = logging.getLogger(__name__)

    try:
        # 1. Determinar nivel de acceso (Solo líder o Todos)
        try:
            acceso = AccesoCorrespondenciaOficina.objects.get(correspondencia=correspondencia, oficina=oficina)
            solo_lider = acceso.solo_lider
        except AccesoCorrespondenciaOficina.DoesNotExist:
            solo_lider = True  # Fallback seguro

        # 2. Obtener usuarios base
        usuarios_qs = User.objects.filter(
            perfil__oficina=oficina,
            is_active=True
        ).select_related('perfil')

        # 3. Filtrar si es solo para líderes
        if solo_lider:
            usuarios_qs = usuarios_qs.filter(groups__name='Lider de Oficina')

        usuarios_destino = usuarios_qs

        if not usuarios_destino.exists():
            logger.warning(f"No se encontraron usuarios activos {'(filtro Lider aplicado)' if solo_lider else ''} en la oficina {oficina.nombre} para notificar.")
            return

        url = reverse('correspondencia:detalle_correspondencia', kwargs={'pk': correspondencia.pk})
        oficina_origen = getattr(getattr(usuario_actual, 'perfil', None), 'oficina', None)
        nombre_oficina_origen = oficina_origen.nombre if oficina_origen else (usuario_actual.get_full_name() or usuario_actual.username)

        notificaciones_creadas = 0
        for usuario in usuarios_destino:
            try:
                Notificacion.objects.create(
                    usuario=usuario,
                    tipo='acceso_oficina',
                    titulo="Acceso de solo lectura otorgado",
                    mensaje=(
                        f"La oficina {nombre_oficina_origen} compartió el radicado {correspondencia.numero_radicado} "
                        f"a {oficina.nombre} con acceso de solo lectura. "
                        f"Asunto: {correspondencia.asunto[:100]}" + (f"\nObservaciones: {observaciones[:100]}" if observaciones else "")
                    ),
                    correspondencia=correspondencia,
                    url=url
                )
                notificaciones_creadas += 1
            except Exception as e:
                logger.error(f"Error al crear notificación para usuario {usuario.username}: {str(e)}")
                continue
        
        logger.info(f"Se crearon {notificaciones_creadas} notificaciones de acceso_oficina para la oficina {oficina.nombre}")
        
    except Exception as e:
        logger.error(f"Error al crear notificaciones de acceso_oficina: {str(e)}")
        # No lanzar excepción para no interrumpir el flujo principal

# =======================================================
# === VISTAS PARA COMUNICACIONES INTERNAS (OFICIOS) ===
# =======================================================
import io

class ComunicacionInternaRecibidasView(LoginRequiredMixin, ListView):
    """Comunicaciones recibidas por el usuario actual."""
    model = ComunicacionInterna
    template_name = 'correspondencia/interna/recibidas.html'
    context_object_name = 'comunicaciones'
    paginate_by = 20

    def _resolver_canal_recepcion(self, comunicacion, user, oficina, proceso_oficina=None):
        if comunicacion.destinatario_usuario_id == user.id:
            return 'MI'

        if any(
            destinatario.tipo == 'USUARIO' and destinatario.usuario_id == user.id
            for destinatario in comunicacion.destinatarios_multiples.all()
        ):
            return 'MI'

        if oficina and comunicacion.destinatario_oficina_id == oficina.id:
            return 'OFICINA'

        if oficina and any(
            destinatario.tipo == 'OFICINA' and destinatario.oficina_id == oficina.id
            for destinatario in comunicacion.destinatarios_multiples.all()
        ):
            return 'OFICINA'

        if proceso_oficina and comunicacion.destinatario_proceso_id == proceso_oficina.id:
            return 'PROCESO'

        if comunicacion.tipo_distribucion == 'ENTIDAD' or comunicacion.es_a_toda_entidad:
            return 'ENTIDAD'

        if any(distribucion.usuario_id == user.id for distribucion in comunicacion.distribuciones.all()):
            return 'MI'

        return 'OTRO'

    def get_queryset(self):
        from .models import ComunicacionInternaDestinatario, ComunicacionInternaDistribucion

        user = self.request.user

        try:
            oficina = user.perfil.oficina
        except AttributeError:
            return ComunicacionInterna.objects.none()

        proceso_oficina = getattr(oficina, 'proceso', None) if oficina else None
        estados_visibles = ['ENVIADA', 'DISTRIBUIDA', 'RESPONDIDA']

        usuario_destinatario_ids = list(
            ComunicacionInternaDestinatario.objects.filter(
                tipo='USUARIO',
                usuario=user,
            ).values_list('comunicacion_id', flat=True)
        )
        oficina_destinatario_ids = list(
            ComunicacionInternaDestinatario.objects.filter(
                tipo='OFICINA',
                oficina=oficina,
            ).values_list('comunicacion_id', flat=True)
        )
        distribucion_usuario_ids = list(
            ComunicacionInternaDistribucion.objects.filter(
                usuario=user,
            ).values_list('comunicacion_id', flat=True)
        )

        qs = (
            ComunicacionInterna.objects.select_related(
                'remitente_oficina',
                'destinatario_oficina',
                'destinatario_usuario',
                'destinatario_proceso',
            )
            .prefetch_related('destinatarios_multiples', 'distribuciones')
            .filter(
                Q(estado__in=estados_visibles),
                (
                    Q(tipo_distribucion='USUARIO', destinatario_usuario=user)
                    | Q(tipo_distribucion='USUARIO', id__in=usuario_destinatario_ids)
                    | Q(tipo_distribucion='OFICINA', destinatario_oficina=oficina)
                    | Q(tipo_distribucion='OFICINA', id__in=oficina_destinatario_ids)
                    | Q(tipo_distribucion='PROCESO', destinatario_proceso=proceso_oficina)
                    | Q(tipo_distribucion='ENTIDAD')
                    | Q(es_a_toda_entidad=True, tipo_distribucion__isnull=True)
                    | Q(destinatario_oficina=oficina, tipo_distribucion__isnull=True)
                    | Q(destinatario_usuario=user, tipo_distribucion__isnull=True)
                    | Q(id__in=distribucion_usuario_ids)
                ),
            )
            .exclude(
                Q(remitente_usuario=user)
                & ~(
                    Q(destinatario_usuario=user)
                    | Q(id__in=usuario_destinatario_ids)
                    | Q(id__in=distribucion_usuario_ids)
                )
            )
            .distinct()
        )

        q = (self.request.GET.get('q') or '').strip()
        if q:
            qs = qs.filter(
                Q(radicado__icontains=q)
                | Q(asunto__icontains=q)
                | Q(remitente_nombre__icontains=q)
                | Q(remitente_oficina__nombre__icontains=q)
            )

        # Filtro por rango de fechas
        fecha_inicio = (self.request.GET.get('fecha_inicio') or '').strip()
        fecha_fin = (self.request.GET.get('fecha_fin') or '').strip()
        if fecha_inicio:
            qs = qs.filter(fecha_creacion__date__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha_creacion__date__lte=fecha_fin)

        estado = (self.request.GET.get('estado') or '').strip()
        if estado in estados_visibles:
            qs = qs.filter(estado=estado)

        destino = (self.request.GET.get('destino') or '').strip()
        if destino == 'mi':
            qs = qs.filter(
                Q(destinatario_usuario=user)
                | Q(id__in=usuario_destinatario_ids)
                | Q(id__in=distribucion_usuario_ids)
            )
        elif destino == 'oficina':
            qs = qs.filter(
                Q(destinatario_oficina=oficina)
                | Q(id__in=oficina_destinatario_ids)
            )
        elif destino == 'proceso':
            qs = qs.filter(destinatario_proceso=proceso_oficina)
        elif destino == 'entidad':
            qs = qs.filter(Q(tipo_distribucion='ENTIDAD') | Q(es_a_toda_entidad=True))

        return qs.order_by('-fecha_creacion')

    def get_object(self, queryset=None):
        """Validar que el usuario tenga acceso a la comunicación."""
        from django.http import Http404
        from .models import ComunicacionInternaDistribucion
        
        obj = super().get_object(queryset)
        user = self.request.user
        
        try:
            oficina = user.perfil.oficina
        except AttributeError:
            oficina = None
        
        # === VALIDACIÓN DE ACCESO ===
        tiene_acceso = False
        
        # 1. USUARIO: Solo remitente o destinatario específico
        if obj.tipo_distribucion == 'USUARIO':
            tiene_acceso = (obj.remitente_usuario == user or obj.destinatario_usuario == user)
        
        # 2. OFICINA: Solo si es de su oficina Y está en la distribución
        elif obj.tipo_distribucion == 'OFICINA':
            if oficina and (obj.remitente_oficina == oficina or obj.destinatario_oficina == oficina):
                tiene_acceso = ComunicacionInternaDistribucion.objects.filter(
                    comunicacion=obj, usuario=user
                ).exists()
        
        # 3. PROCESO: Solo si está en la distribución
        elif obj.tipo_distribucion == 'PROCESO':
            tiene_acceso = ComunicacionInternaDistribucion.objects.filter(
                comunicacion=obj, usuario=user
            ).exists()
        
        # 4. ENTIDAD: Solo si está en la distribución
        elif obj.tipo_distribucion == 'ENTIDAD':
            tiene_acceso = ComunicacionInternaDistribucion.objects.filter(
                comunicacion=obj, usuario=user
            ).exists()
        
        # Fallback: Si no tiene tipo_distribucion asignado (datos antiguos)
        elif obj.tipo_distribucion is None:
            if oficina:
                tiene_acceso = (
                    obj.remitente_usuario == user or
                    obj.destinatario_usuario == user or
                    obj.remitente_oficina == oficina or
                    obj.destinatario_oficina == oficina or
                    ComunicacionInternaDistribucion.objects.filter(
                        comunicacion=obj, usuario=user
                    ).exists()
                )
        
        if not tiene_acceso:
            raise Http404("No tienes acceso a esta comunicación.")
        
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['titulo_pagina'] = 'Comunicaciones Recibidas'
        context['tipo_bandeja'] = 'recibidas'
        context['filtro_q'] = (self.request.GET.get('q') or '').strip()
        context['filtro_estado'] = (self.request.GET.get('estado') or '').strip()
        context['filtro_destino'] = (self.request.GET.get('destino') or '').strip()
        context['filtro_fecha_inicio'] = (self.request.GET.get('fecha_inicio') or '').strip()
        context['filtro_fecha_fin'] = (self.request.GET.get('fecha_fin') or '').strip()
        context['filtros_activos'] = sum(
            1
            for value in (context['filtro_q'], context['filtro_estado'], context['filtro_destino'], context['filtro_fecha_inicio'], context['filtro_fecha_fin'])
            if value
        )
        try:
            context['oficina_actual'] = self.request.user.perfil.oficina
        except AttributeError:
            context['oficina_actual'] = None

        proceso_oficina = getattr(context['oficina_actual'], 'proceso', None) if context['oficina_actual'] else None
        for comunicacion in context['comunicaciones']:
            comunicacion.canal_recepcion = self._resolver_canal_recepcion(
                comunicacion,
                self.request.user,
                context['oficina_actual'],
                proceso_oficina=proceso_oficina,
            )
        return context

def _usuario_puede_ver_comunicacion_interna(user, obj):
    """
    Valida si un usuario tiene acceso a una comunicación interna.
    Misma lógica que ComunicacionInternaDetailView.get_object().
    """
    from .models import ComunicacionInternaDestinatario, ComunicacionInternaDistribucion
    
    try:
        perfil = user.perfil
        oficina = perfil.oficina
    except AttributeError:
        perfil = None
        oficina = None
    proceso_oficina = getattr(oficina, 'proceso', None) if oficina else None
    es_lider_oficina = user.groups.filter(name='Lider de Oficina').exists()

    if obj.remitente_usuario == user:
        return True

    if oficina and obj.remitente_oficina == oficina:
        return True

    if ComunicacionInternaDistribucion.objects.filter(
        comunicacion=obj, usuario=user
    ).exists():
        return True
    
    # 1. USUARIO: destinatario directo, destinatario múltiple o remitente
    if obj.tipo_distribucion == 'USUARIO':
        if obj.destinatario_usuario == user:
            return True

        return ComunicacionInternaDestinatario.objects.filter(
            comunicacion=obj,
            tipo='USUARIO',
            usuario=user
        ).exists()
    
    # 2. OFICINA: por oficina legacy o destinatario múltiple de oficina
    elif obj.tipo_distribucion == 'OFICINA':
        if not oficina:
            return False

        if obj.remitente_oficina == oficina or obj.destinatario_oficina == oficina:
            return True

        return ComunicacionInternaDestinatario.objects.filter(
            comunicacion=obj,
            tipo='OFICINA',
            oficina=oficina
        ).exists()
    
    # 3. PROCESO: por proceso de la oficina o distribución explícita
    elif obj.tipo_distribucion == 'PROCESO':
        return bool(proceso_oficina and obj.destinatario_proceso == proceso_oficina)
    
    # 4. ENTIDAD: visible para cualquier usuario autenticado
    elif obj.tipo_distribucion == 'ENTIDAD' or obj.es_a_toda_entidad:
        return user.is_authenticated
    
    # Fallback: Si no tiene tipo_distribucion asignado (datos antiguos)
    elif obj.tipo_distribucion is None:
        if oficina:
            return (
                obj.remitente_usuario == user or
                obj.destinatario_usuario == user or
                obj.remitente_oficina == oficina or
                obj.destinatario_oficina == oficina or
                ComunicacionInternaDestinatario.objects.filter(
                    comunicacion=obj,
                    tipo='USUARIO',
                    usuario=user
                ).exists() or
                ComunicacionInternaDestinatario.objects.filter(
                    comunicacion=obj,
                    tipo='OFICINA',
                    oficina=oficina
                ).exists() or
                ComunicacionInternaDistribucion.objects.filter(
                    comunicacion=obj, usuario=user
                ).exists()
            )
    
    return False


class ComunicacionInternaPDFView(LoginRequiredMixin, View):
    """Vista para previsualizar el documento PDF."""
    def get(self, request, pk):
        obj = get_object_or_404(ComunicacionInterna, pk=pk)
        
        # Verificar que el usuario tenga acceso a esta comunicación
        if not _usuario_puede_ver_comunicacion_interna(request.user, obj):
            return HttpResponseForbidden("No tienes acceso a esta comunicación.")
        
        if obj.archivo_generado:
            # Generar nombre descriptivo basado en la oficina del redactor
            nombre_archivo = "comunicacion_interna"
            
            # Intentar obtener el nombre de la oficina del usuario redactor
            try:
                if obj.remitente_usuario and hasattr(obj.remitente_usuario, 'perfil'):
                    perfil = obj.remitente_usuario.perfil
                    if perfil and perfil.oficina:
                        nombre_oficina = perfil.oficina.nombre
                        # Limpiar el nombre para usarlo en archivo
                        nombre_oficina = nombre_oficina.replace(' ', '_').replace('/', '_').lower()
                        nombre_archivo = f"comunicacion_interna_{nombre_oficina}"
                    
                elif obj.remitente_oficina:
                    nombre_oficina = obj.remitente_oficina.nombre
                    nombre_oficina = nombre_oficina.replace(' ', '_').replace('/', '_').lower()
                    nombre_archivo = f"comunicacion_interna_{nombre_oficina}"
            except Exception as e:
                logger.warning(f"Error al obtener nombre de oficina para descargar PDF: {e}")
            
            # Agregar el radicado si existe
            if obj.radicado:
                nombre_archivo = f"{nombre_archivo}_{obj.radicado.replace('/', '_')}"
            
            response = HttpResponse(obj.archivo_generado, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
            return response
        else:
            messages.error(request, "El documento no ha sido generado.")
            return redirect('correspondencia:interna_detalle', pk=pk)

def usuarios_por_oficina_ajax(request):
    """Retorna lista de usuarios para una oficina dada."""
    oficina_id = request.GET.get('oficina_id')
    if not oficina_id:
        return JsonResponse([], safe=False)
    
    try:
        users = User.objects.filter(
            perfil__oficina_id=oficina_id,
            is_active=True
        ).values('id', 'first_name', 'last_name', 'username').order_by('first_name')
        
        data = []
        for u in users:
            nombre = f"{u['first_name']} {u['last_name']}".strip() or u['username']
            data.append({'id': u['id'], 'nombre': nombre})
            
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def oficinas_todas_interna_ajax(request):
    """Endpoint para obtener todas las oficinas para comunicación interna (sin restricción de grupo)."""
    from .models import OficinaProductora
    oficinas = OficinaProductora.objects.all().order_by('nombre')
    data = [
        {
            'id': o.id,
            'text': o.nombre
        }
        for o in oficinas
    ]
    return JsonResponse(data, safe=False)


@login_required
def procesos_todos_ajax(request):
    """Endpoint para obtener todos los procesos para comunicación interna."""
    from documentos.models import Proceso
    procesos = Proceso.objects.all().order_by('macroproceso__numero', 'numero')
    data = [
        {
            'id': p.id,
            'text': f"{p.numero} - {p.nombre} ({p.sigla})"
        }
        for p in procesos
    ]
    return JsonResponse(data, safe=False)


@login_required
def destinatarios_interna_ajax(request, pk):
    """
    Devuelve los destinatarios de una comunicación interna en formato JSON.
    Usado en el modal de destinatarios de comunicaciones recibidas.
    """
    from .models import ComunicacionInterna
    
    if request.method != 'GET' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'})

    try:
        comunicacion = get_object_or_404(ComunicacionInterna, pk=pk)
        
        # Obtener destinatarios múltiples
        destinatarios_data = []
        
        for dest in comunicacion.destinatarios_multiples.select_related('usuario', 'oficina').all():
            if dest.tipo == 'USUARIO' and dest.usuario:
                destinatarios_data.append({
                    'id': dest.id,
                    'tipo': 'USUARIO',
                    'nombre': dest.usuario.get_full_name() or dest.usuario.username,
                })
            elif dest.tipo == 'OFICINA' and dest.oficina:
                destinatarios_data.append({
                    'id': dest.id,
                    'tipo': 'OFICINA',
                    'nombre': dest.oficina.nombre,
                })
        
        return JsonResponse({
            'success': True,
            'comunicacion': {
                'radicado': comunicacion.radicado,
                'asunto': comunicacion.asunto,
            },
            'destinatarios': destinatarios_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def guardar_firma_interna(request):
    """Vista AJAX para guardar la firma capturada."""
    try:
        import json
        import base64
        from django.core.files.base import ContentFile
        from documentos.models import PerfilUsuario
        
        data = json.loads(request.body)
        firma_data = data.get('firma')  # data:image/png;base64,...
        
        if not firma_data:
            return JsonResponse({'success': False, 'error': 'No se recibió la firma'})
        
        # Obtener o crear perfil
        perfil, created = PerfilUsuario.objects.get_or_create(user=request.user)
        
        # Extraer la parte base64
        format, imgstr = firma_data.split(';base64,')
        ext = format.split('/')[-1]  # png
        
        # Decodificar y crear archivo
        firma_decoded = base64.b64decode(imgstr)
        filename = f'firma_{request.user.id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.{ext}'
        
        # Guardar en el perfil del usuario
        perfil.firma_digital.save(
            filename,
            ContentFile(firma_decoded),
            save=True
        )
        perfil.fecha_firma_creada = timezone.now()
        perfil.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Firma guardada exitosamente',
            'fecha': perfil.fecha_firma_creada.strftime('%d/%m/%Y %H:%M')
        })
    except Exception as e:
        logger.error(f"Error guardando firma: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def crear_comunicacion_interna_ajax(request):
    """Vista AJAX para crear comunicación interna desde modal."""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            from .models import HistorialComunicacionInterna, OficinaProductora
            from django.db import transaction
            
            # Obtener datos del formulario
            ciudad = request.POST.get('ciudad', 'Saravena').strip()
            fecha_documento = request.POST.get('fecha_documento')
            serie_id = request.POST.get('serie') or None
            subserie_id = request.POST.get('subserie') or None
            tipo_distribucion = request.POST.get('tipo_distribucion', 'USUARIO')
            es_a_toda_entidad = request.POST.get('es_a_toda_entidad') == 'on' or tipo_distribucion == 'ENTIDAD'
            destinatario_oficina_id = request.POST.get('destinatario_oficina') or None
            destinatario_usuario_id = request.POST.get('destinatario_usuario') or None
            destinatario_proceso_id = request.POST.get('destinatario_proceso') or None
            asunto = request.POST.get('asunto', '').strip()
            cuerpo = request.POST.get('cuerpo', '').strip()
            enviar_inmediatamente = request.POST.get('enviar_inmediatamente') == 'on'
            
            # Debug: Log de datos recibidos
            import logging
            logger = logging.getLogger(__name__)
            logger.debug(f"Crear comunicación interna - tipo_distribucion: {tipo_distribucion}, destinatario_oficina_id: {destinatario_oficina_id}, destinatario_usuario_id: {destinatario_usuario_id}, POST data: {dict(request.POST)}")
            
            # Validaciones básicas
            if not all([ciudad, fecha_documento, asunto, cuerpo]):
                return JsonResponse({'success': False, 'error': 'Todos los campos obligatorios deben ser completados.'})
            
            # Validar fecha
            try:
                from datetime import datetime
                fecha_doc = datetime.strptime(fecha_documento, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Fecha inválida.'})
            
            # Obtener destinatarios múltiples (nuevo sistema)
            destinatarios_usuarios_ids = request.POST.getlist('destinatarios_usuarios')
            destinatarios_oficinas_ids = request.POST.getlist('destinatarios_oficinas')
            
            # Validar según tipo de distribución
            if tipo_distribucion == 'USUARIO':
                # Para tipo USUARIO, debe haber al menos un usuario seleccionado
                if not destinatarios_usuarios_ids:
                    return JsonResponse({'success': False, 'error': 'Debe seleccionar al menos un usuario.'})
            elif tipo_distribucion == 'OFICINA':
                # Para tipo OFICINA, debe haber al menos una oficina en destinatarios_oficinas
                if not destinatarios_oficinas_ids:
                    return JsonResponse({'success': False, 'error': 'Debe seleccionar al menos una oficina (subproceso).'})
            elif tipo_distribucion == 'PROCESO' and not destinatario_proceso_id:
                return JsonResponse({'success': False, 'error': 'Debe seleccionar un proceso para distribución completa.'})
            
            # Obtener perfil del usuario
            user = request.user
            try:
                perfil = user.perfil
                if not perfil.oficina:
                    return JsonResponse({'success': False, 'error': 'Su usuario no tiene una oficina asignada.'})
            except Exception:
                return JsonResponse({'success': False, 'error': 'Su usuario no tiene perfil configurado.'})

            try:
                serie_documental, subserie_documental, trd = _resolver_clasificacion_comunicacion_interna(
                    perfil.oficina,
                    serie_id=serie_id,
                    subserie_id=subserie_id,
                )
            except ValidationError as exc:
                return JsonResponse({'success': False, 'error': exc.message})
            
            # Verificar si el usuario es líder de oficina
            es_lider = user.groups.filter(name='Lider de Oficina').exists()
            
            # Determinar el estado inicial basado en el tipo de distribución y si es líder
            # Una comunicación es "pre-aprobada" (lista para distribuir) si:
            # - Es enviada por un líder de oficina
            # - No requiere aprobación de otro líder
            # Se verá como APROBADA o DISTRIBUIDA según si requiere firma
            
            # Calcular si requiere firma (igual a la propiedad del modelo)
            requiere_firma = tipo_distribucion in ('OFICINA', 'PROCESO', 'ENTIDAD') or es_a_toda_entidad
            
            # Determinar estado inicial
            # Si es líder, crear directamente como pre-aprobada
            # Si no es líder, crear como pendiente de aprobación
            if es_lider:
                # Líder crea como pre-aprobada
                estado_inicial = 'APROBADA' if requiere_firma else 'DISTRIBUIDA'
            else:
                # No líder: crea como pendiente de aprobación
                estado_inicial = 'PENDIENTE_APROBACION'
            
            with transaction.atomic():
                # Para compatibilidad, mantener el primer destinatario en los campos originales
                # pero ahora también usamos el nuevo modelo de destinatarios múltiples
                primera_oficina_id = destinatarios_oficinas_ids[0] if destinatarios_oficinas_ids else destinatario_oficina_id
                primer_usuario_id = destinatarios_usuarios_ids[0] if destinatarios_usuarios_ids else destinatario_usuario_id
                
                # Log para debugging
                logger.info(f"Creando ComunicacionInterna: tipo={tipo_distribucion}, es_lider={es_lider}, estado_inicial={estado_inicial}")
                
                # Crear la comunicación especificando el estado inicial
                # El signal SIEMPRE cambia BORRADOR a PENDIENTE_APROBACION, por lo que:
                # - Si estado_inicial != 'BORRADOR', el signal no lo modificará
                # - Líderes: APROBADA o DISTRIBUIDA (pre-aprobadas)
                # - No líderes: PENDIENTE_APROBACION
                comunicacion = ComunicacionInterna.objects.create(
                    estado=estado_inicial,
                    ciudad=ciudad,
                    fecha_documento=fecha_doc,
                    serie_documental=serie_documental,
                    subserie_documental=subserie_documental,
                    trd=trd,
                    tipo_distribucion=tipo_distribucion,
                    es_a_toda_entidad=es_a_toda_entidad,  # Mantener para compatibilidad
                    destinatario_oficina_id=primera_oficina_id if tipo_distribucion in ('USUARIO', 'OFICINA') else None,
                    destinatario_usuario_id=primer_usuario_id if tipo_distribucion == 'USUARIO' else None,
                    destinatario_proceso_id=destinatario_proceso_id if tipo_distribucion == 'PROCESO' else None,
                    asunto=asunto,
                    cuerpo=cuerpo,
                    remitente_usuario=user,
                    remitente_nombre=f"{user.first_name} {user.last_name}".strip() or user.username,
                    remitente_cargo=getattr(perfil, 'cargo', None),
                    remitente_oficina=perfil.oficina
                )
                
                logger.info(f"ComunicacionInterna creada: ID={comunicacion.id}, estado={comunicacion.estado}")
                
                # Crear destinatarios múltiples
                from .models import ComunicacionInternaDestinatario
                if tipo_distribucion == 'USUARIO':
                    # Agregar usuarios seleccionados
                    for usuario_id in destinatarios_usuarios_ids:
                        try:
                            usuario_obj = User.objects.get(id=usuario_id, is_active=True)
                            ComunicacionInternaDestinatario.objects.create(
                                comunicacion=comunicacion,
                                tipo='USUARIO',
                                usuario=usuario_obj
                            )
                        except User.DoesNotExist:
                            logger.warning(f"Usuario {usuario_id} no encontrado, omitiendo")
                
                elif tipo_distribucion == 'OFICINA':
                    # Agregar oficinas seleccionadas (oficinas completas)
                    for oficina_id in destinatarios_oficinas_ids:
                        try:
                            oficina_obj = OficinaProductora.objects.get(id=oficina_id)
                            ComunicacionInternaDestinatario.objects.create(
                                comunicacion=comunicacion,
                                tipo='OFICINA',
                                oficina=oficina_obj
                            )
                        except OficinaProductora.DoesNotExist:
                            logger.warning(f"Oficina {oficina_id} no encontrada, omitiendo")
                
                # Refrescar la instancia para asegurar que los destinatarios múltiples estén disponibles
                comunicacion.refresh_from_db()
                
                # Procesar anexos
                from .models import AnexoComunicacionInterna
                anexos_subidos = request.FILES.getlist('anexos')
                
                if anexos_subidos:
                    max_anexos = 10
                    max_size_per_file = 25 * 1024 * 1024  # 25MB
                    max_total_size = 25 * 1024 * 1024  # 25MB
                    allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx']
                    
                    if len(anexos_subidos) > max_anexos:
                        comunicacion.delete()
                        return JsonResponse({
                            'success': False,
                            'error': f'Máximo {max_anexos} anexos permitidos.'
                        }, status=400)
                    
                    total_size = 0
                    for archivo in anexos_subidos:
                        ext = '.' + archivo.name.lower().split('.')[-1]
                        if ext not in allowed_extensions:
                            comunicacion.delete()
                            return JsonResponse({
                                'success': False,
                                'error': f'Archivo "{archivo.name}" tiene formato no permitido.'
                            }, status=400)
                        
                        if archivo.size > max_size_per_file:
                            comunicacion.delete()
                            return JsonResponse({
                                'success': False,
                                'error': f'Archivo "{archivo.name}" excede 25MB.'
                            }, status=400)
                        
                        total_size += archivo.size
                    
                    if total_size > max_total_size:
                        comunicacion.delete()
                        return JsonResponse({
                            'success': False,
                            'error': 'El tamaño total de los anexos excede 25MB.'
                        }, status=400)
                    
                    # Guardar anexos
                    for archivo in anexos_subidos:
                        AnexoComunicacionInterna.objects.create(
                            comunicacion=comunicacion,
                            archivo=archivo,
                            nombre_original=archivo.name,
                            subido_por=user
                        )
                
                # Registrar historial inicial
                HistorialComunicacionInterna.objects.create(
                    comunicacion=comunicacion,
                    evento='CREADA',
                    usuario=user,
                    descripcion="Comunicación creada desde modal"
                )
                
                # Si el usuario es líder, generar radicado y completar aprovación
                # (comunicación pre-aprobada, lista para distribuir o esperando firma)
                if es_lider and estado_inicial in ('APROBADA', 'DISTRIBUIDA'):
                    # El líder crea y aprueba automáticamente
                    comunicacion.revisado_por = user
                    comunicacion.revisado_nombre = f"{user.first_name} {user.last_name}".strip() or user.username
                    comunicacion.revisado_cargo = getattr(perfil, 'cargo', None)
                    comunicacion.fecha_revision = timezone.now()
                    
                    # Si requiere firma (OFICINA, PROCESO, ENTIDAD), queda en APROBADA esperando firma
                    # Si es a usuario específico, se distribuye directamente
                    if comunicacion.requiere_firma:
                        comunicacion.estado = 'APROBADA'
                        if not comunicacion.radicado:
                            comunicacion.radicado = comunicacion._generar_radicado()
                        
                        # PRIMERO guardar el estado y radicado en la base de datos
                        comunicacion.save()
                        
                        # LUEGO refrescar para asegurar que los destinatarios múltiples estén disponibles
                        comunicacion.refresh_from_db()
                        
                        # Generar documento
                        generar_pdf_comunicacion_interna(comunicacion, request)
                        comunicacion.save()
                        
                        # Registrar historial
                        HistorialComunicacionInterna.objects.create(
                            comunicacion=comunicacion,
                            evento='APROBADA',
                            usuario=user,
                            descripcion="Aprobada automáticamente por líder. Pendiente de firma digital."
                        )
                        
                        tipo_texto = 'a toda la entidad' if comunicacion.tipo_distribucion == 'ENTIDAD' else 'al proceso completo' if comunicacion.tipo_distribucion == 'PROCESO' else 'a la oficina completa'
                        message = f"Comunicación {comunicacion.radicado} aprobada. Debe subir el documento firmado para distribuirla {tipo_texto}."
                    else:
                        # A usuario específico: distribuir directamente
                        comunicacion.estado = 'DISTRIBUIDA'
                        comunicacion.fecha_distribucion = timezone.now()
                        if not comunicacion.radicado:
                            comunicacion.radicado = comunicacion._generar_radicado()
                        
                        # PRIMERO guardar el estado y radicado en la base de datos
                        comunicacion.save()
                        
                        # LUEGO refrescar para asegurar que los destinatarios múltiples estén disponibles
                        comunicacion.refresh_from_db()
                        
                        # Generar documento
                        generar_pdf_comunicacion_interna(comunicacion, request)
                        comunicacion.save()
                        
                        # Registrar historial
                        HistorialComunicacionInterna.objects.create(
                            comunicacion=comunicacion,
                            evento='APROBADA',
                            usuario=user,
                            descripcion="Aprobada automáticamente por líder"
                        )
                        
                        # Distribuir a usuarios destino
                        _distribuir_comunicacion_interna(comunicacion, user)
                        
                        message = f"Comunicación {comunicacion.radicado} distribuida exitosamente."
                    
                # Si se envía inmediatamente (para compatibilidad con código anterior)
                elif enviar_inmediatamente:
                    if es_lider:
                        # El líder crea y aprueba automáticamente
                        comunicacion.revisado_por = user
                        comunicacion.revisado_nombre = f"{user.first_name} {user.last_name}".strip() or user.username
                        comunicacion.revisado_cargo = getattr(perfil, 'cargo', None)
                        comunicacion.fecha_revision = timezone.now()
                        
                        # Si requiere firma (OFICINA, PROCESO, ENTIDAD), queda en APROBADA esperando firma
                        # Si es a usuario específico, se distribuye directamente
                        if comunicacion.requiere_firma:
                            comunicacion.estado = 'APROBADA'
                            if not comunicacion.radicado:
                                comunicacion.radicado = comunicacion._generar_radicado()
                            
                            # PRIMERO guardar el estado y radicado en la base de datos
                            comunicacion.save()
                            
                            # LUEGO refrescar para asegurar que los destinatarios múltiples estén disponibles
                            comunicacion.refresh_from_db()
                            
                            # Generar documento
                            generar_pdf_comunicacion_interna(comunicacion, request)
                            comunicacion.save()
                            
                            # Registrar historial
                            HistorialComunicacionInterna.objects.create(
                                comunicacion=comunicacion,
                                evento='APROBADA',
                                usuario=user,
                                descripcion="Aprobada automáticamente por líder. Pendiente de firma digital."
                            )
                            
                            tipo_texto = 'a toda la entidad' if comunicacion.tipo_distribucion == 'ENTIDAD' else 'al proceso completo' if comunicacion.tipo_distribucion == 'PROCESO' else 'a la oficina completa'
                            message = f"Comunicación {comunicacion.radicado} aprobada. Debe subir el documento firmado para distribuirla {tipo_texto}."
                        else:
                            # A usuario específico: distribuir directamente
                            comunicacion.estado = 'DISTRIBUIDA'
                            comunicacion.fecha_distribucion = timezone.now()
                            if not comunicacion.radicado:
                                comunicacion.radicado = comunicacion._generar_radicado()
                            
                            # PRIMERO guardar el estado y radicado en la base de datos
                            comunicacion.save()
                            
                            # LUEGO refrescar para asegurar que los destinatarios múltiples estén disponibles
                            comunicacion.refresh_from_db()
                            
                            # Generar documento
                            generar_pdf_comunicacion_interna(comunicacion, request)
                            comunicacion.save()
                            
                            # Registrar historial
                            HistorialComunicacionInterna.objects.create(
                                comunicacion=comunicacion,
                                evento='APROBADA',
                                usuario=user,
                                descripcion="Aprobada automáticamente por líder"
                            )
                            
                            # Distribuir a usuarios destino
                            _distribuir_comunicacion_interna(comunicacion, user)
                            
                            message = f"Comunicación {comunicacion.radicado} distribuida exitosamente."
                    else:
                        # Usuario normal: pasa a pendiente de aprobación
                        comunicacion.estado = 'PENDIENTE_APROBACION'
                        comunicacion.save()
                        
                        # Registrar historial
                        HistorialComunicacionInterna.objects.create(
                            comunicacion=comunicacion,
                            evento='ENVIADA_APROBACION',
                            usuario=user,
                            descripcion="Enviada para aprobación del líder de oficina"
                        )
                        
                        message = "Comunicación enviada para aprobación del líder de oficina."
                else:
                    # Guardar como borrador
                    comunicacion.estado = 'BORRADOR'
                    comunicacion.save()
                    message = "Borrador guardado."
            
            from django.urls import reverse
            return JsonResponse({
                'success': True,
                'message': message,
                'radicado': comunicacion.radicado,
                'id': comunicacion.id,
                'redirect_url': reverse('correspondencia:interna_detalle', kwargs={'pk': comunicacion.id})
            })
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al crear comunicación interna: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': f'Ocurrió un error inesperado: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def obtener_comunicacion_interna_ajax(request, pk):
    """
    Vista AJAX para obtener datos de una comunicación interna para edición.
    Solo permite obtener comunicaciones en estado BORRADOR o PENDIENTE_APROBACION.
    """
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Solicitud inválida.'}, status=400)
    
    try:
        comunicacion = get_object_or_404(
            ComunicacionInterna.objects.prefetch_related('anexos'),
            pk=pk
        )
        user = request.user
        
        # Verificar permisos: solo creador o líder de la oficina remitente
        puede_editar = False
        es_lider = user.groups.filter(name='Lider de Oficina').exists()
        
        if comunicacion.remitente_usuario == user:
            puede_editar = True
        elif es_lider:
            try:
                if user.perfil.oficina == comunicacion.remitente_oficina:
                    puede_editar = True
            except AttributeError:
                pass
        
        if not puede_editar:
            return JsonResponse({
                'success': False, 
                'error': 'No tiene permisos para editar esta comunicación.'
            }, status=403)
        
        # Verificar estado editable
        if comunicacion.estado not in ('BORRADOR', 'PENDIENTE_APROBACION'):
            return JsonResponse({
                'success': False, 
                'error': 'Solo se pueden editar comunicaciones en estado Borrador o Pendiente de Aprobación.'
            }, status=400)
        
        # Obtener destinatarios múltiples
        from .models import ComunicacionInternaDestinatario
        destinatarios = ComunicacionInternaDestinatario.objects.filter(
            comunicacion=comunicacion
        ).select_related('usuario', 'oficina', 'usuario__perfil', 'usuario__perfil__oficina')
        
        destinatarios_usuarios = []
        destinatarios_oficinas = []
        
        for dest in destinatarios:
            if dest.tipo == 'USUARIO' and dest.usuario:
                destinatarios_usuarios.append({
                    'id': dest.usuario.id,
                    'nombre': dest.usuario.get_full_name() or dest.usuario.username,
                    'oficina': dest.usuario.perfil.oficina.nombre if hasattr(dest.usuario, 'perfil') and dest.usuario.perfil.oficina else None,
                    'oficina_id': dest.usuario.perfil.oficina.id if hasattr(dest.usuario, 'perfil') and dest.usuario.perfil.oficina else None
                })
            elif dest.tipo == 'OFICINA' and dest.oficina:
                destinatarios_oficinas.append({
                    'id': dest.oficina.id,
                    'nombre': dest.oficina.nombre
                })
        
        # Obtener anexos
        anexos_data = []
        for anexo in comunicacion.anexos.all():
            anexos_data.append({
                'id': anexo.id,
                'nombre': anexo.nombre_original,
                'tipo': anexo.get_tipo_archivo(),
                'url': anexo.archivo.url if anexo.archivo else '',
                'es_pdf': anexo.es_pdf()
            })
        
        # Construir respuesta
        data = {
            'success': True,
            'comunicacion': {
                'id': comunicacion.id,
                'ciudad': comunicacion.ciudad,
                'fecha_documento': comunicacion.fecha_documento.strftime('%Y-%m-%d') if comunicacion.fecha_documento else '',
                'trd': comunicacion.trd or '',
                'serie_id': comunicacion.serie_documental_id,
                'subserie_id': comunicacion.subserie_documental_id,
                'tipo_distribucion': comunicacion.tipo_distribucion or 'USUARIO',
                'asunto': comunicacion.asunto,
                'cuerpo': comunicacion.cuerpo,
                'estado': comunicacion.estado,
                'radicado': comunicacion.radicado,
                # Destinatarios legacy (para compatibilidad)
                'destinatario_oficina_id': comunicacion.destinatario_oficina_id,
                'destinatario_usuario_id': comunicacion.destinatario_usuario_id,
                'destinatario_proceso_id': comunicacion.destinatario_proceso_id,
                # Destinatarios múltiples
                'destinatarios_usuarios': destinatarios_usuarios,
                'destinatarios_oficinas': destinatarios_oficinas,
                # Anexos
                'anexos': anexos_data,
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al obtener comunicación interna {pk}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)


@login_required
def editar_comunicacion_interna_ajax(request, pk):
    """
    Vista AJAX para editar una comunicación interna existente.
    Solo permite editar:
    - Asunto y cuerpo
    - Destinatarios (agregar/quitar)
    No permite cambiar tipo de distribución.
    Solo en estados BORRADOR o PENDIENTE_APROBACION.
    """
    if request.method != 'POST' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    try:
        from .models import HistorialComunicacionInterna, ComunicacionInternaDestinatario, OficinaProductora
        from django.db import transaction
        
        comunicacion = get_object_or_404(ComunicacionInterna, pk=pk)
        user = request.user
        
        # Verificar permisos: solo creador o líder de la oficina remitente
        puede_editar = False
        es_lider = user.groups.filter(name='Lider de Oficina').exists()
        
        if comunicacion.remitente_usuario == user:
            puede_editar = True
        elif es_lider:
            try:
                if user.perfil.oficina == comunicacion.remitente_oficina:
                    puede_editar = True
            except AttributeError:
                pass
        
        if not puede_editar:
            return JsonResponse({
                'success': False, 
                'error': 'No tiene permisos para editar esta comunicación.'
            }, status=403)
        
        # Verificar estado editable
        if comunicacion.estado not in ('BORRADOR', 'PENDIENTE_APROBACION'):
            return JsonResponse({
                'success': False, 
                'error': 'Solo se pueden editar comunicaciones en estado Borrador o Pendiente de Aprobación.'
            }, status=400)
        
        # Obtener datos del formulario
        ciudad = request.POST.get('ciudad', '').strip()
        fecha_documento = request.POST.get('fecha_documento', '').strip()
        asunto = request.POST.get('asunto', '').strip()
        cuerpo = request.POST.get('cuerpo', '').strip()
        trd = request.POST.get('trd', '').strip()
        serie_id = request.POST.get('serie') or None
        subserie_id = request.POST.get('subserie') or None
        
        # Validaciones básicas
        if not ciudad or not fecha_documento or not asunto or not cuerpo:
            return JsonResponse({'success': False, 'error': 'La ciudad, fecha del documento, asunto y contenido son obligatorios.'})
        
        # Obtener destinatarios del formulario
        destinatarios_usuarios_ids = request.POST.getlist('destinatarios_usuarios')
        destinatarios_oficinas_ids = request.POST.getlist('destinatarios_oficinas')
        destinatario_proceso_id = request.POST.get('destinatario_proceso') or None
        
        tipo_distribucion = comunicacion.tipo_distribucion
        
        # Validar destinatarios según tipo de distribución
        if tipo_distribucion == 'USUARIO':
            if not destinatarios_usuarios_ids:
                return JsonResponse({'success': False, 'error': 'Debe seleccionar al menos un usuario.'})
        elif tipo_distribucion == 'OFICINA':
            if not destinatarios_oficinas_ids:
                return JsonResponse({'success': False, 'error': 'Debe seleccionar al menos una oficina.'})
        elif tipo_distribucion == 'PROCESO':
            if not destinatario_proceso_id:
                return JsonResponse({'success': False, 'error': 'Debe seleccionar un proceso.'})
        
        import logging
        logger = logging.getLogger(__name__)
        
        with transaction.atomic():
            try:
                serie_documental, subserie_documental, _trd = _resolver_clasificacion_comunicacion_interna(
                    comunicacion.remitente_oficina,
                    serie_id=serie_id,
                    subserie_id=subserie_id,
                )
            except ValidationError as exc:
                return JsonResponse({'success': False, 'error': exc.message}, status=400)

            # Actualizar campos editables
            comunicacion.ciudad = ciudad
            comunicacion.fecha_documento = fecha_documento
            comunicacion.asunto = asunto
            comunicacion.cuerpo = cuerpo
            comunicacion.serie_documental = serie_documental
            comunicacion.subserie_documental = subserie_documental
            if trd:
                comunicacion.trd = trd
            
            # Si es tipo PROCESO, actualizar el proceso destino
            if tipo_distribucion == 'PROCESO' and destinatario_proceso_id:
                comunicacion.destinatario_proceso_id = destinatario_proceso_id
            
            # Para tipos USUARIO y OFICINA, actualizar destinatarios múltiples
            if tipo_distribucion in ('USUARIO', 'OFICINA'):
                # Eliminar destinatarios actuales
                ComunicacionInternaDestinatario.objects.filter(comunicacion=comunicacion).delete()
                
                if tipo_distribucion == 'USUARIO':
                    # Agregar nuevos usuarios
                    primera_oficina_id = None
                    primer_usuario_id = None
                    
                    for idx, usuario_id in enumerate(destinatarios_usuarios_ids):
                        try:
                            usuario_obj = User.objects.get(id=usuario_id, is_active=True)
                            ComunicacionInternaDestinatario.objects.create(
                                comunicacion=comunicacion,
                                tipo='USUARIO',
                                usuario=usuario_obj
                            )
                            if idx == 0:
                                primer_usuario_id = usuario_id
                                if hasattr(usuario_obj, 'perfil') and usuario_obj.perfil.oficina:
                                    primera_oficina_id = usuario_obj.perfil.oficina.id
                        except User.DoesNotExist:
                            logger.warning(f"Usuario {usuario_id} no encontrado al editar")
                    
                    # Actualizar campos legacy
                    comunicacion.destinatario_usuario_id = primer_usuario_id
                    comunicacion.destinatario_oficina_id = primera_oficina_id
                
                elif tipo_distribucion == 'OFICINA':
                    # Agregar nuevas oficinas
                    primera_oficina_id = None
                    
                    for idx, oficina_id in enumerate(destinatarios_oficinas_ids):
                        try:
                            oficina_obj = OficinaProductora.objects.get(id=oficina_id)
                            ComunicacionInternaDestinatario.objects.create(
                                comunicacion=comunicacion,
                                tipo='OFICINA',
                                oficina=oficina_obj
                            )
                            if idx == 0:
                                primera_oficina_id = oficina_id
                        except OficinaProductora.DoesNotExist:
                            logger.warning(f"Oficina {oficina_id} no encontrada al editar")
                    
                    # Actualizar campo legacy
                    comunicacion.destinatario_oficina_id = primera_oficina_id
            
            comunicacion.save()
            
            # Procesar anexos nuevos
            from .models import AnexoComunicacionInterna
            anexos_subidos = request.FILES.getlist('anexos')
            anexos_count_client = int(request.POST.get('anexos_count_client', 0) or 0)

            # Log para depurar recepción de anexos
            logger.info(
                "Editar CI %s: anexos_count_client=%s, recibidos=%s, nombres=%s",
                comunicacion.id,
                anexos_count_client,
                len(anexos_subidos),
                [getattr(f, 'name', '') for f in anexos_subidos]
            )

            if anexos_count_client > 0 and not anexos_subidos:
                logger.warning("Editar CI %s: cliente envió %s anexos pero request.FILES llegó vacío", comunicacion.id, anexos_count_client)
                return JsonResponse({
                    'success': False,
                    'error': 'No se recibieron los anexos en el servidor. Verifique conexión o tamaño del archivo.'
                }, status=400)
            
            if anexos_subidos:
                # Verificar límites
                anexos_actuales = comunicacion.anexos.count()
                max_anexos = 10
                max_size_per_file = 25 * 1024 * 1024  # 25MB
                max_total_size = 25 * 1024 * 1024  # 25MB
                allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx']
                
                if anexos_actuales + len(anexos_subidos) > max_anexos:
                    return JsonResponse({
                        'success': False,
                        'error': f'Máximo {max_anexos} anexos permitidos. Ya tiene {anexos_actuales}.'
                    }, status=400)
                
                total_size = 0
                for archivo in anexos_subidos:
                    ext = '.' + archivo.name.lower().split('.')[-1]
                    if ext not in allowed_extensions:
                        return JsonResponse({
                            'success': False,
                            'error': f'Archivo "{archivo.name}" tiene formato no permitido.'
                        }, status=400)
                    
                    if archivo.size > max_size_per_file:
                        return JsonResponse({
                            'success': False,
                            'error': f'Archivo "{archivo.name}" excede 25MB.'
                        }, status=400)
                    
                    total_size += archivo.size
                
                if total_size > max_total_size:
                    return JsonResponse({
                        'success': False,
                        'error': 'El tamaño total de los anexos excede 25MB.'
                    }, status=400)
                
                # Guardar anexos
                for archivo in anexos_subidos:
                    AnexoComunicacionInterna.objects.create(
                        comunicacion=comunicacion,
                        archivo=archivo,
                        nombre_original=archivo.name,
                        subido_por=user
                    )
            
            # Registrar en historial
            HistorialComunicacionInterna.objects.create(
                comunicacion=comunicacion,
                evento='EDITADA',
                usuario=user,
                descripcion=f"Comunicación editada por {user.get_full_name() or user.username}"
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Comunicación actualizada exitosamente.',
            'id': comunicacion.id,
            'radicado': comunicacion.radicado
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al editar comunicación interna {pk}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)
def _crear_notificaciones_aceptacion_borrador(comunicacion, actor, nuevo_estado):
    detalle_url = reverse('correspondencia:interna_detalle', kwargs={'pk': comunicacion.pk})

    if nuevo_estado == 'PENDIENTE_APROBACION':
        lideres = User.objects.filter(
            is_active=True,
            groups__name='Lider de Oficina',
            perfil__oficina=comunicacion.remitente_oficina,
        ).exclude(pk=actor.pk).distinct()

        notificaciones = [
            Notificacion(
                usuario=lider,
                tipo='aprobacion_pendiente',
                titulo=f"Aprobación pendiente: {comunicacion.asunto[:60]}",
                mensaje=(
                    f"{comunicacion.remitente_nombre} envió la comunicación interna "
                    f"{comunicacion.radicado or 'sin radicado'} para revisión."
                ),
                comunicacion_interna=comunicacion,
                url=detalle_url,
            )
            for lider in lideres
        ]
        if notificaciones:
            Notificacion.objects.bulk_create(notificaciones)
        return

    if nuevo_estado == 'APROBADA' and comunicacion.remitente_usuario_id and comunicacion.remitente_usuario_id != actor.pk:
        Notificacion.objects.create(
            usuario=comunicacion.remitente_usuario,
            tipo='comunicacion_interna',
            titulo=f"Comunicación aceptada: {comunicacion.asunto[:60]}",
            mensaje=(
                f"{actor.get_full_name() or actor.username} aceptó la comunicación interna "
                f"{comunicacion.radicado or 'sin radicado'} y quedó preaprobada."
            ),
            comunicacion_interna=comunicacion,
            url=detalle_url,
        )


@login_required
def aceptar_borrador_comunicacion_ajax(request, pk):
    """
    Vista AJAX para aceptar un borrador y convertirlo en comunicación pre-aprobada.
    - Si el usuario es líder: convierte a APROBADA (si requiere firma) o DISTRIBUIDA (si es a usuario)
    - Si no es líder: convierte a PENDIENTE_APROBACION
    
    Solo permitido si la comunicación está en estado BORRADOR.
    """
    if request.method != 'POST' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    try:
        from .models import HistorialComunicacionInterna
        from django.db import transaction
        
        comunicacion = get_object_or_404(ComunicacionInterna, pk=pk)
        user = request.user
        
        # Verificar que sea borrador
        if comunicacion.estado != 'BORRADOR':
            return JsonResponse({
                'success': False,
                'error': f'Solo se pueden aceptar comunicaciones en estado Borrador. Estado actual: {comunicacion.get_estado_display()}'
            }, status=400)
        
        # Verificar permisos: solo creador o líder de la oficina remitente
        puede_aceptar = False
        es_lider = user.groups.filter(name='Lider de Oficina').exists()
        
        if comunicacion.remitente_usuario == user:
            puede_aceptar = True
        elif es_lider:
            try:
                if user.perfil.oficina == comunicacion.remitente_oficina:
                    puede_aceptar = True
            except AttributeError:
                pass
        
        if not puede_aceptar:
            return JsonResponse({
                'success': False, 
                'error': 'No tiene permisos para aceptar esta comunicación.'
            }, status=403)
        
        import logging
        logger = logging.getLogger(__name__)
        
        with transaction.atomic():
            # Calcular estado final
            requiere_firma = comunicacion.tipo_distribucion in ('OFICINA', 'PROCESO', 'ENTIDAD') or comunicacion.es_a_toda_entidad
            
            if es_lider:
                # Líder: crear como pre-aprobada - SIEMPRE va a APROBADA primero
                # El distribuir es un paso posterior
                nuevo_estado = 'APROBADA'
                
                # Completar información de revisión
                perfil = user.perfil
                comunicacion.revisado_por = user
                comunicacion.revisado_nombre = f"{user.first_name} {user.last_name}".strip() or user.username
                comunicacion.revisado_cargo = getattr(perfil, 'cargo', None)
                comunicacion.fecha_revision = timezone.now()
            else:
                # No líder: cambiar a pendiente de aprobación
                nuevo_estado = 'PENDIENTE_APROBACION'
            
            # Generar radicado si es necesario
            if nuevo_estado in ('APROBADA', 'DISTRIBUIDA') and not comunicacion.radicado:
                comunicacion.radicado = comunicacion._generar_radicado()
            
            # Establecer estado y guardar
            comunicacion.estado = nuevo_estado
            comunicacion.save()
            
            # Si es líder, generar PDF
            if es_lider:
                generar_pdf_comunicacion_interna(comunicacion, request)
                comunicacion.save()
                
                # SIEMPRE va a APROBADA - el siguiente paso es revisar/firmar y distribuir
                if requiere_firma:
                    mensaje = f"Comunicación {comunicacion.radicado} aceptada. Requiere firma digital para distribuir."
                    evento = 'APROBADA'
                    descripcion = "Aceptada como comunicación interna pre-aprobada. Pendiente de firma digital"
                else:
                    mensaje = f"Comunicación {comunicacion.radicado} aceptada. Requiere revisión del líder para distribuir."
                    evento = 'APROBADA'
                    descripcion = "Aceptada como comunicación interna pre-aprobada. Pendiente de revisión y distribución"
            else:
                mensaje = "Comunicación enviada para aprobación del líder de oficina."
                evento = 'ENVIADA_APROBACION'
                descripcion = "Aceptada como comunicación interna y enviada para aprobación"
            
            # Registrar en historial
            HistorialComunicacionInterna.objects.create(
                comunicacion=comunicacion,
                evento=evento,
                usuario=user,
                descripcion=descripcion
            )

            _crear_notificaciones_aceptacion_borrador(comunicacion, user, nuevo_estado)
            
            logger.info(f"Borrador {comunicacion.id} aceptado como {nuevo_estado} por {user.username}")
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'id': comunicacion.id,
            'radicado': comunicacion.radicado,
            'estado': comunicacion.get_estado_display()
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al aceptar borrador {pk}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)


@login_required
def revertir_borrador_comunicacion_ajax(request, pk):
    """
    Vista AJAX para revertir una comunicación de PENDIENTE_APROBACION a BORRADOR.
    Solo permite revertir comunicaciones en estado PENDIENTE_APROBACION.
    Solo el creador o líder de la oficina pueden hacer esto.
    """
    if request.method != 'POST' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    try:
        from .models import HistorialComunicacionInterna
        from django.db import transaction
        
        comunicacion = get_object_or_404(ComunicacionInterna, pk=pk)
        user = request.user
        
        # Verificar que esté en PENDIENTE_APROBACION
        if comunicacion.estado != 'PENDIENTE_APROBACION':
            return JsonResponse({
                'success': False,
                'error': f'Solo se pueden revertir comunicaciones en estado Pendiente de Aprobación. Estado actual: {comunicacion.get_estado_display()}'
            }, status=400)
        
        # Verificar permisos: solo creador o líder de la oficina remitente
        puede_revertir = False
        es_lider = user.groups.filter(name='Lider de Oficina').exists()
        
        if comunicacion.remitente_usuario == user:
            puede_revertir = True
        elif es_lider:
            try:
                if user.perfil.oficina == comunicacion.remitente_oficina:
                    puede_revertir = True
            except AttributeError:
                pass
        
        if not puede_revertir:
            return JsonResponse({
                'success': False, 
                'error': 'No tiene permisos para revertir esta comunicación.'
            }, status=403)
        
        import logging
        logger = logging.getLogger(__name__)
        
        with transaction.atomic():
            # Cambiar a BORRADOR
            comunicacion.estado = 'BORRADOR'
            # Limpiar información de revisión si existe
            comunicacion.revisado_por = None
            comunicacion.revisado_nombre = None
            comunicacion.revisado_cargo = None
            comunicacion.fecha_revision = None
            comunicacion.save()
            
            # Registrar en historial
            HistorialComunicacionInterna.objects.create(
                comunicacion=comunicacion,
                evento='REVERTIDA',
                usuario=user,
                descripcion=f"Revertida a borrador por {user.get_full_name() or user.username}"
            )
            
            logger.info(f"Comunicación {comunicacion.id} revertida a BORRADOR por {user.username}")
        
        return JsonResponse({
            'success': True,
            'message': 'Comunicación revertida a borrador exitosamente.',
            'id': comunicacion.id,
            'estado': comunicacion.get_estado_display()
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al revertir comunicación {pk}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)


@login_required
def eliminar_anexo_comunicacion_interna(request, pk):
    """
    Vista AJAX para eliminar un anexo de una comunicación interna.
    Solo permitido si la comunicación está en estado BORRADOR o PENDIENTE_APROBACION.
    """
    if request.method != 'POST' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    try:
        from .models import AnexoComunicacionInterna
        
        anexo = get_object_or_404(AnexoComunicacionInterna, pk=pk)
        comunicacion = anexo.comunicacion
        user = request.user
        
        # Verificar estado de la comunicación
        if comunicacion.estado not in ('BORRADOR', 'PENDIENTE_APROBACION'):
            return JsonResponse({
                'success': False,
                'error': 'No se pueden eliminar anexos de comunicaciones ya distribuidas.'
            }, status=400)
        
        # Verificar permisos
        puede_editar = False
        es_lider = user.groups.filter(name='Lider de Oficina').exists()
        
        if comunicacion.remitente_usuario == user:
            puede_editar = True
        elif es_lider:
            try:
                if user.perfil.oficina == comunicacion.remitente_oficina:
                    puede_editar = True
            except AttributeError:
                pass
        
        if not puede_editar:
            return JsonResponse({
                'success': False,
                'error': 'No tiene permisos para eliminar anexos de esta comunicación.'
            }, status=403)
        
        # Eliminar el anexo (el signal pre_delete borrará el archivo físico)
        nombre_archivo = anexo.nombre_original
        anexo.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Anexo "{nombre_archivo}" eliminado exitosamente.'
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al eliminar anexo {pk}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)


# =======================================================
# === BANDEJA DE APROBACIÓN PARA LÍDERES DE OFICINA ===
# =======================================================

class ComunicacionInternaPendientesView(LoginRequiredMixin, ListView):
    """Bandeja para líderes: muestra comunicaciones pendientes de aprobación de su oficina."""
    model = ComunicacionInterna
    template_name = 'correspondencia/interna/pendientes_aprobacion.html'
    context_object_name = 'comunicaciones'
    paginate_by = 20

    def dispatch(self, request, *args, **kwargs):
        # Solo líderes de oficina pueden acceder
        if not request.user.groups.filter(name='Lider de Oficina').exists():
            messages.error(request, "Solo los líderes de oficina pueden acceder a esta sección.")
            return redirect('correspondencia:interna_bienvenida')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user
        try:
            oficina = user.perfil.oficina
        except AttributeError:
            return ComunicacionInterna.objects.none()
        
        # Comunicaciones de MI oficina pendientes de aprobación
        qs = ComunicacionInterna.objects.filter(
            remitente_oficina=oficina,
            estado='PENDIENTE_APROBACION'
        ).order_by('-fecha_creacion')
        
        return qs

@login_required
def aprobar_comunicacion_interna(request, pk):
    """Vista para aprobar una comunicación interna pendiente."""
    from .models import HistorialComunicacionInterna, ComunicacionInternaDistribucion, Notificacion

    comunicacion_qs = ComunicacionInterna.objects.select_related(
        'remitente_oficina',
        'destinatario_oficina',
        'destinatario_usuario',
        'remitente_usuario',
    )
    comunicacion = get_object_or_404(comunicacion_qs, pk=pk)
    user = request.user
    
    # Verificar que el usuario es líder de oficina
    if not user.groups.filter(name='Lider de Oficina').exists():
        messages.error(request, "Solo los líderes de oficina pueden aprobar comunicaciones.")
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    # Verificar que la comunicación es de su oficina
    try:
        if comunicacion.remitente_oficina != user.perfil.oficina:
            messages.error(request, "Solo puede aprobar comunicaciones de su propia oficina.")
            return redirect('correspondencia:interna_detalle', pk=pk)
    except AttributeError:
        messages.error(request, "Su usuario no tiene oficina asignada.")
        return redirect('correspondencia:interna_bienvenida')
    
    # Verificar estado
    if comunicacion.estado != 'PENDIENTE_APROBACION':
        messages.warning(request, "Esta comunicación no está pendiente de aprobación.")
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    if request.method == 'POST':
        accion = request.POST.get('accion')

        with transaction.atomic():
            comunicacion = get_object_or_404(comunicacion_qs.select_for_update(), pk=pk)

            try:
                if comunicacion.remitente_oficina != user.perfil.oficina:
                    messages.error(request, "Solo puede aprobar comunicaciones de su propia oficina.")
                    return redirect('correspondencia:interna_detalle', pk=pk)
            except AttributeError:
                messages.error(request, "Su usuario no tiene oficina asignada.")
                return redirect('correspondencia:interna_bienvenida')

            if comunicacion.estado != 'PENDIENTE_APROBACION':
                messages.warning(request, "Esta comunicación ya no está pendiente de aprobación.")
                return redirect('correspondencia:interna_detalle', pk=pk)

            if accion == 'aprobar':
                comunicacion.revisado_por = user
                comunicacion.revisado_nombre = f"{user.first_name} {user.last_name}".strip() or user.username
                perfil = getattr(user, 'perfil', None)
                comunicacion.revisado_cargo = getattr(perfil, 'cargo', None)
                comunicacion.fecha_revision = timezone.now()

                if not comunicacion.radicado:
                    comunicacion.radicado = comunicacion._generar_radicado()

                if comunicacion.requiere_firma:
                    comunicacion.estado = 'APROBADA'
                else:
                    comunicacion.estado = 'DISTRIBUIDA'
                    comunicacion.fecha_distribucion = timezone.now()

                comunicacion.save()

                generar_pdf_comunicacion_interna(comunicacion, request)
                comunicacion.save()

                if comunicacion.requiere_firma:
                    HistorialComunicacionInterna.objects.create(
                        comunicacion=comunicacion,
                        evento='APROBADA',
                        usuario=user,
                        descripcion="Aprobada por líder. Pendiente de firma digital para distribución."
                    )

                    tipo_texto = 'a toda la entidad' if comunicacion.tipo_distribucion == 'ENTIDAD' else 'al proceso completo' if comunicacion.tipo_distribucion == 'PROCESO' else 'a la oficina completa'
                    messages.warning(request, f"Comunicación {comunicacion.radicado} aprobada. Debe subir el documento firmado para distribuirla {tipo_texto}.")
                else:
                    HistorialComunicacionInterna.objects.create(
                        comunicacion=comunicacion,
                        evento='APROBADA',
                        usuario=user,
                        descripcion="Aprobada por líder"
                    )

                    _distribuir_comunicacion_interna(comunicacion, user)

                    messages.success(request, f"Comunicación {comunicacion.radicado} aprobada y distribuida exitosamente.")

                return redirect('correspondencia:interna_detalle', pk=pk)

            elif accion == 'rechazar':
                motivo = request.POST.get('motivo', '').strip()
                if not motivo:
                    messages.error(request, "Debe indicar el motivo del rechazo.")
                    return redirect('correspondencia:interna_aprobar', pk=pk)

                comunicacion.estado = 'RECHAZADA'
                comunicacion.motivo_rechazo = motivo
                comunicacion.revisado_por = user
                comunicacion.revisado_nombre = f"{user.first_name} {user.last_name}".strip() or user.username
                perfil = getattr(user, 'perfil', None)
                comunicacion.revisado_cargo = getattr(perfil, 'cargo', None)
                comunicacion.fecha_revision = timezone.now()
                comunicacion.save()

                HistorialComunicacionInterna.objects.create(
                    comunicacion=comunicacion,
                    evento='RECHAZADA',
                    usuario=user,
                    descripcion=f"Rechazada por líder. Motivo: {motivo}"
                )

                Notificacion.objects.create(
                    usuario=comunicacion.remitente_usuario,
                    tipo='comunicacion_interna',
                    titulo=f"Comunicación rechazada: {comunicacion.asunto[:50]}",
                    mensaje=f"Tu comunicación fue rechazada por {user.get_full_name() or user.username}. Motivo: {motivo}",
                    comunicacion_interna=comunicacion,
                    url=reverse('correspondencia:interna_detalle', kwargs={'pk': comunicacion.pk})
                )

                messages.error(request, "Comunicación rechazada. El creador ha sido notificado.")
                return redirect('correspondencia:interna_pendientes')
    
    return render(request, 'correspondencia/interna/aprobar.html', {
        'comunicacion': comunicacion,
        'titulo_pagina': 'Aprobar Comunicación Interna'
    })


def _distribuir_comunicacion_interna(comunicacion, usuario_distribuidor):
    """Función auxiliar para distribuir la comunicación a los usuarios destino."""
    from .models import ComunicacionInternaDistribucion, HistorialComunicacionInterna, Notificacion
    from documentos.models import OficinaProductora
    from django.urls import reverse
    
    usuarios_destino = []
    tipo_dist = comunicacion.tipo_distribucion or ('ENTIDAD' if comunicacion.es_a_toda_entidad else 'USUARIO')
    
    if tipo_dist == 'ENTIDAD' or comunicacion.es_a_toda_entidad:
        # Todos los usuarios activos del sistema
        usuarios_destino = User.objects.filter(is_active=True).exclude(pk=comunicacion.remitente_usuario_id)
    
    elif tipo_dist == 'PROCESO':
        # Todos los usuarios de todas las oficinas del proceso
        if comunicacion.destinatario_proceso:
            oficinas_proceso = OficinaProductora.objects.filter(proceso=comunicacion.destinatario_proceso)
            usuarios_destino = User.objects.filter(
                is_active=True,
                perfil__oficina__in=oficinas_proceso
            ).exclude(pk=comunicacion.remitente_usuario_id)
    
    elif tipo_dist == 'OFICINA':
        # Todos los usuarios de la oficina destino
        if comunicacion.destinatario_oficina:
            usuarios_destino = User.objects.filter(
                is_active=True,
                perfil__oficina=comunicacion.destinatario_oficina
            ).exclude(pk=comunicacion.remitente_usuario_id)
    
    else:  # USUARIO
        # Múltiples destinatarios (usuarios y oficinas)
        from .models import ComunicacionInternaDestinatario
        usuarios_destino = []
        
        # Obtener destinatarios del nuevo modelo
        destinatarios = ComunicacionInternaDestinatario.objects.filter(comunicacion=comunicacion)
        
        for destinatario in destinatarios:
            if destinatario.tipo == 'USUARIO' and destinatario.usuario:
                if destinatario.usuario not in usuarios_destino:
                    usuarios_destino.append(destinatario.usuario)
            elif destinatario.tipo == 'OFICINA' and destinatario.oficina:
                # Para oficinas completas, agregar todos los usuarios de esa oficina
                usuarios_oficina = User.objects.filter(
                    is_active=True,
                    perfil__oficina=destinatario.oficina
                ).exclude(pk=comunicacion.remitente_usuario_id)
                for usuario in usuarios_oficina:
                    if usuario not in usuarios_destino:
                        usuarios_destino.append(usuario)
        
        # Si no hay destinatarios múltiples, usar el sistema antiguo (compatibilidad)
        if not usuarios_destino and comunicacion.destinatario_usuario:
            usuarios_destino = [comunicacion.destinatario_usuario]
    
    count = 0
    for usuario in usuarios_destino:
        # Crear distribución
        dist, created = ComunicacionInternaDistribucion.objects.get_or_create(
            comunicacion=comunicacion,
            usuario=usuario
        )
        if created:
            count += 1
            # Crear notificación
            Notificacion.objects.create(
                usuario=usuario,
                tipo='comunicacion_interna',
                titulo=f"Nueva comunicación interna: {comunicacion.asunto[:50]}",
                mensaje=f"Has recibido una comunicación interna de {comunicacion.remitente_nombre} ({comunicacion.remitente_oficina.nombre}).",
                comunicacion_interna=comunicacion,
                url=reverse('correspondencia:interna_detalle', kwargs={'pk': comunicacion.pk})
            )
    
    # Registrar en historial
    HistorialComunicacionInterna.objects.create(
        comunicacion=comunicacion,
        evento='DISTRIBUIDA',
        usuario=usuario_distribuidor,
        descripcion=f"Distribuida a {count} usuarios"
    )


@login_required
def subir_firma_interna(request, pk):
    """Vista para firmar digitalmente y distribuir una comunicación interna que requiere firma (ENTIDAD, PROCESO, OFICINA)."""
    from .models import ComunicacionInterna, HistorialComunicacionInterna
    
    comunicacion = get_object_or_404(ComunicacionInterna, pk=pk)
    user = request.user
    
    # Verificar que el usuario es líder de oficina
    if not user.groups.filter(name='Lider de Oficina').exists():
        messages.error(request, "Solo los líderes de oficina pueden firmar documentos.")
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    # Verificar que la comunicación es de su oficina
    try:
        if comunicacion.remitente_oficina != user.perfil.oficina:
            messages.error(request, "Solo puede gestionar comunicaciones de su propia oficina.")
            return redirect('correspondencia:interna_detalle', pk=pk)
    except AttributeError:
        messages.error(request, "Su usuario no tiene oficina asignada.")
        return redirect('correspondencia:interna_bienvenida')
    
    # Verificar estado y que requiere firma
    if comunicacion.estado != 'APROBADA':
        messages.warning(request, "Esta comunicación no está en estado aprobada pendiente de firma.")
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    if not comunicacion.requiere_firma:
        messages.warning(request, "Esta comunicación no requiere firma digital.")
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    if request.method == 'POST':
        # Verificar que el líder tiene su firma digital configurada
        try:
            perfil = user.perfil
            if not perfil or not perfil.firma_digital:
                messages.error(request, "Debe configurar su firma digital antes de poder firmar documentos. Vaya a Comunicaciones Internas > Bienvenida para configurarla.")
                return redirect('correspondencia:interna_detalle', pk=pk)
        except AttributeError:
            messages.error(request, "Debe configurar su firma digital antes de poder firmar documentos.")
            return redirect('correspondencia:interna_detalle', pk=pk)
        
        # Actualizar datos de revisión del líder
        comunicacion.revisado_por = user
        comunicacion.revisado_nombre = user.get_full_name() or user.username
        try:
            comunicacion.revisado_cargo = user.perfil.cargo
        except AttributeError:
            pass
        comunicacion.fecha_revision = timezone.now()
        
        # Regenerar PDF con las firmas digitales incluidas
        comunicacion.save()
        generar_pdf_comunicacion_interna(comunicacion, request)
        
        # Marcar como distribuida
        comunicacion.fecha_firma = timezone.now()
        comunicacion.estado = 'DISTRIBUIDA'
        comunicacion.fecha_distribucion = timezone.now()
        comunicacion.save()
        
        # Registrar historial
        HistorialComunicacionInterna.objects.create(
            comunicacion=comunicacion,
            evento='FIRMADO_DIGITAL',
            usuario=user,
            descripcion=f"Documento firmado digitalmente por {user.get_full_name() or user.username}"
        )
        
        # Distribuir a todos los usuarios
        _distribuir_comunicacion_interna(comunicacion, user)
        
        tipo_texto = 'a toda la entidad' if comunicacion.tipo_distribucion == 'ENTIDAD' else 'al proceso completo' if comunicacion.tipo_distribucion == 'PROCESO' else 'a la oficina completa'
        messages.success(request, f"Comunicación {comunicacion.radicado} firmada digitalmente y distribuida {tipo_texto}.")
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    # GET: Mostrar confirmación
    return render(request, 'correspondencia/interna/confirmar_firma.html', {
        'comunicacion': comunicacion,
        'titulo_pagina': 'Firmar y Distribuir'
    })


@login_required
def responder_comunicacion_interna(request, pk):
    """
    Vista para crear una respuesta a una comunicación interna.
    
    Flujo de respuestas (rediseñado):
    - Múltiples destinatarios pueden responder (no hay límite de 1 respuesta)
    - Las respuestas NO se pueden responder (solo ida y vuelta)
    - Líderes y usuarios asignados inicialmente tienen respuestas destacadas (estrellita)
    - El estado de la comunicación origen NO cambia a RESPONDIDA
    """
    from .models import ComunicacionInterna, HistorialComunicacionInterna
    
    comunicacion_origen = get_object_or_404(ComunicacionInterna, pk=pk)
    user = request.user
    
    # Verificar que la comunicación está distribuida
    if comunicacion_origen.estado not in ('DISTRIBUIDA',):
        messages.warning(request, "Solo se puede responder a comunicaciones distribuidas.")
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    # IMPORTANTE: Las respuestas NO se pueden responder (solo ida y vuelta)
    if comunicacion_origen.es_respuesta():
        messages.warning(request, "Las respuestas no se pueden responder. El flujo de comunicación es solo ida y vuelta.")
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    # Verificar que el usuario tiene perfil
    try:
        perfil = user.perfil
        oficina_usuario = perfil.oficina
    except AttributeError:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'Su usuario no tiene perfil configurado.'
            })
        messages.error(request, "Su usuario no tiene perfil configurado.")
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    # Usar el método del modelo para validar si puede responder
    if not comunicacion_origen.puede_responder(user):
        tipo_dist = getattr(comunicacion_origen, 'tipo_distribucion', None) or ('ENTIDAD' if comunicacion_origen.es_a_toda_entidad else 'USUARIO')
        
        if tipo_dist == 'PROCESO':
            mensaje_error = "No se pueden crear respuestas a comunicaciones dirigidas a procesos completos (normativas)."
        else:
            mensaje_error = "No tiene permiso para responder esta comunicación. Solo los destinatarios pueden responder."
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': mensaje_error
            })
        messages.error(request, mensaje_error)
        return redirect('correspondencia:interna_detalle', pk=pk)
    
    if request.method == 'POST':
        from .forms import ComunicacionInternaRespuestaForm
        
        form = ComunicacionInternaRespuestaForm(request.POST)
        if form.is_valid():
            respuesta = form.save(commit=False)
            try:
                serie_documental, subserie_documental, trd = _resolver_clasificacion_comunicacion_interna(
                    oficina_usuario,
                    serie_id=request.POST.get('serie') or None,
                    subserie_id=request.POST.get('subserie') or None,
                )
                respuesta.serie_documental = serie_documental
                respuesta.subserie_documental = subserie_documental
                respuesta.trd = trd
            except ValidationError as exc:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': exc.message})
                form.add_error(None, exc.message)
                return render(request, 'correspondencia/interna/responder.html', {
                    'comunicacion_origen': comunicacion_origen,
                    'form': form,
                    'titulo_pagina': f'Responder a: {comunicacion_origen.asunto}',
                    'num_respuestas_existentes': comunicacion_origen.contar_respuestas(),
                })
            
            # Configurar datos de la respuesta
            respuesta.remitente_usuario = user
            respuesta.remitente_nombre = f"{user.first_name} {user.last_name}".strip() or user.username
            respuesta.remitente_cargo = getattr(perfil, 'cargo', None)
            respuesta.remitente_oficina = oficina_usuario
            respuesta.comunicacion_origen = comunicacion_origen
            
            # El destinatario de la respuesta es el remitente original
            respuesta.destinatario_oficina = comunicacion_origen.remitente_oficina
            respuesta.destinatario_usuario = comunicacion_origen.remitente_usuario
            respuesta.es_a_toda_entidad = False  # Respuestas siempre van a una oficina específica
            respuesta.tipo_distribucion = 'USUARIO'  # Respuestas siempre son a usuario específico
            
            # Determinar si la respuesta debe ser destacada (estrellita)
            # Se marca como destacada si el usuario es líder de oficina destinataria
            # o si estaba como destinatario directo (asignado inicialmente)
            es_lider = user.groups.filter(name='Lider de Oficina').exists()
            
            # Marcar respuesta como destacada según criterios
            respuesta.es_respuesta_destacada = _es_respuesta_destacada(comunicacion_origen, user, oficina_usuario, es_lider)
            
            if 'save_send' in request.POST:
                if es_lider:
                    # El líder aprueba y distribuye directamente
                    respuesta.estado = 'DISTRIBUIDA'
                    respuesta.revisado_por = user
                    respuesta.revisado_nombre = f"{user.first_name} {user.last_name}".strip() or user.username
                    respuesta.revisado_cargo = getattr(perfil, 'cargo', None)
                    respuesta.fecha_revision = timezone.now()
                    respuesta.fecha_distribucion = timezone.now()
                    respuesta.radicado = respuesta._generar_radicado()
                else:
                    # Usuario normal: envía a aprobación
                    respuesta.estado = 'PENDIENTE_APROBACION'
            else:
                # Guardar como borrador
                respuesta.estado = 'BORRADOR'
            
            respuesta.save()
            
            # Si es AJAX, retornar JSON antes de continuar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.urls import reverse
                return JsonResponse({
                    'success': True,
                    'message': 'Respuesta creada exitosamente.',
                    'redirect_url': reverse('correspondencia:interna_detalle', kwargs={'pk': comunicacion_origen.pk})
                })
            
            # Registrar historial en la comunicación origen
            estrellita_txt = " ⭐" if respuesta.es_respuesta_destacada else ""
            HistorialComunicacionInterna.objects.create(
                comunicacion=comunicacion_origen,
                evento='RESPUESTA_CREADA',
                usuario=user,
                descripcion=f"Respuesta creada por {respuesta.remitente_nombre}{estrellita_txt}"
            )
            
            # NOTA: Ya NO se cambia el estado de la comunicación origen a RESPONDIDA
            # Esto permite que múltiples destinatarios puedan responder
            
            # Registrar historial en la respuesta
            if respuesta.estado == 'BORRADOR':
                HistorialComunicacionInterna.objects.create(
                    comunicacion=respuesta,
                    evento='CREADA',
                    usuario=user,
                    descripcion=f"Respuesta a comunicación {comunicacion_origen.radicado} (Borrador)"
                )
                messages.info(request, "Respuesta guardada como borrador.")
            elif respuesta.estado == 'PENDIENTE_APROBACION':
                HistorialComunicacionInterna.objects.create(
                    comunicacion=respuesta,
                    evento='CREADA',
                    usuario=user,
                    descripcion=f"Respuesta a comunicación {comunicacion_origen.radicado}"
                )
                HistorialComunicacionInterna.objects.create(
                    comunicacion=respuesta,
                    evento='ENVIADA_APROBACION',
                    usuario=user,
                    descripcion="Enviada para aprobación del líder"
                )
                messages.info(request, "Respuesta enviada para aprobación del líder de oficina.")
            elif respuesta.estado == 'DISTRIBUIDA':
                # Generar el PDF con el texto de respuesta
                respuesta.refresh_from_db()
                generar_pdf_comunicacion_interna(respuesta, request)
                respuesta.save()
                
                HistorialComunicacionInterna.objects.create(
                    comunicacion=respuesta,
                    evento='CREADA',
                    usuario=user,
                    descripcion=f"Respuesta a comunicación {comunicacion_origen.radicado}"
                )
                HistorialComunicacionInterna.objects.create(
                    comunicacion=respuesta,
                    evento='APROBADA',
                    usuario=user,
                    descripcion="Aprobada automáticamente por líder"
                )
                # Distribuir la respuesta
                _distribuir_comunicacion_interna(respuesta, user)
                destacada_msg = " (Respuesta destacada ⭐)" if respuesta.es_respuesta_destacada else ""
                messages.success(request, f"Respuesta {respuesta.radicado} aprobada y distribuida.{destacada_msg}")
            
            return redirect('correspondencia:interna_detalle', pk=respuesta.pk)
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        from .forms import ComunicacionInternaRespuestaForm
        
        # Pre-llenar el formulario con datos relacionados
        initial_data = {
            'fecha_documento': timezone.now().date(),
            'ciudad': comunicacion_origen.ciudad,
            'asunto': f"RE: {comunicacion_origen.asunto}",
        }
        form = ComunicacionInternaRespuestaForm(initial=initial_data)
    
    # Información adicional para el template
    num_respuestas = comunicacion_origen.contar_respuestas()
    
    return render(request, 'correspondencia/interna/responder.html', {
        'comunicacion_origen': comunicacion_origen,
        'form': form,
        'titulo_pagina': f'Responder a: {comunicacion_origen.asunto}',
        'num_respuestas_existentes': num_respuestas,
    })


def _es_respuesta_destacada(comunicacion_origen, usuario, oficina_usuario, es_lider):
    """
    Determina si una respuesta de este usuario debe ser destacada (estrellita).
    
    Criterios para respuesta destacada:
    - El usuario es líder de oficina Y su oficina es destinataria
    - El usuario está en la lista de destinatarios directos (asignados inicialmente)
    """
    from .models import ComunicacionInternaDestinatario
    
    # Si es líder y su oficina fue destinataria original, es destacada
    if es_lider and oficina_usuario:
        oficina_era_destinataria = ComunicacionInternaDestinatario.objects.filter(
            comunicacion=comunicacion_origen,
            tipo='OFICINA',
            oficina=oficina_usuario
        ).exists()
        
        if oficina_era_destinataria:
            return True
        
        # Fallback legacy
        if comunicacion_origen.destinatario_oficina == oficina_usuario:
            return True
    
    # Si el usuario estaba como destinatario directo (asignado inicialmente), es destacada
    usuario_era_destinatario_directo = ComunicacionInternaDestinatario.objects.filter(
        comunicacion=comunicacion_origen,
        tipo='USUARIO',
        usuario=usuario
    ).exists()
    
    if usuario_era_destinatario_directo:
        return True
    
    # Fallback legacy
    if comunicacion_origen.destinatario_usuario == usuario:
        return True
    
    return False


def formatear_fecha_espanol(fecha):
    """Formatea una fecha en español: 'día de mes de año' (ej: '27 de enero de 2026')"""
    meses_es = {
        'January': 'enero',
        'February': 'febrero',
        'March': 'marzo',
        'April': 'abril',
        'May': 'mayo',
        'June': 'junio',
        'July': 'julio',
        'August': 'agosto',
        'September': 'septiembre',
        'October': 'octubre',
        'November': 'noviembre',
        'December': 'diciembre'
    }
    
    fecha_en = fecha.strftime('%d de %B de %Y')
    for mes_en, mes_es in meses_es.items():
        fecha_en = fecha_en.replace(mes_en, mes_es)
    
    return fecha_en


COMUNICACION_INTERNA_FOOTER_CONTACT_LINES = [
    'Calle 30 No. 19A-82 Barrio los Libertadores / Línea Correspondencia y Administrativa: 313 409 4023 / E-mail: correspondencia@esehospitaldelsarare.gov.co',
    '/ Atención al Usuario: (607) 8859868 Opción 6 / Talento Humano: 316 878 9501',
    '/ Urgencias: 321 471 4096 / Referencia y Contrareferencia: 317 643 1580 Página Web: www.hospitaldelsarare.gov.co',
    '/ Facebook: Hospital del Sarare / Instagram: @hospitaldelsarare / Saravena - Arauca',
]


def generar_pdf_comunicacion_interna(instance, request=None):
    """Función auxiliar para generar el PDF de una comunicación interna."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether, Flowable as RLFlowable
        from reportlab.lib.colors import HexColor
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from django.contrib.staticfiles import finders
        from django.core.files.base import ContentFile
        
        # Registrar fuente Segoe UI si está disponible (Windows)
        segoe_ui_paths = [
            r'C:\Windows\Fonts\segoeui.ttf',  # Segoe UI Regular
            r'C:\Windows\Fonts\segoeuib.ttf',  # Segoe UI Bold
            r'C:\WINDOWS\Fonts\segoeui.ttf',   # Variante con mayúsculas
            r'C:\WINDOWS\Fonts\segoeuib.ttf',
        ]
        
        segoe_ui_registered = False
        segoe_ui_bold_registered = False
        
        # Intentar registrar Segoe UI Regular
        for path in [segoe_ui_paths[0], segoe_ui_paths[2]]:
            if os.path.exists(path) and not segoe_ui_registered:
                try:
                    pdfmetrics.registerFont(TTFont('SegoeUI', path))
                    segoe_ui_registered = True
                    break
                except Exception as e:
                    logger.warning(f"No se pudo registrar Segoe UI desde {path}: {e}")
        
        # Intentar registrar Segoe UI Bold
        for path in [segoe_ui_paths[1], segoe_ui_paths[3]]:
            if os.path.exists(path) and not segoe_ui_bold_registered:
                try:
                    pdfmetrics.registerFont(TTFont('SegoeUI-Bold', path))
                    segoe_ui_bold_registered = True
                    break
                except Exception as e:
                    logger.warning(f"No se pudo registrar Segoe UI Bold desde {path}: {e}")
        
        # Si no se pudo registrar Segoe UI, usar Helvetica como fallback
        font_name = 'SegoeUI' if segoe_ui_registered else 'Helvetica'
        font_name_bold = 'SegoeUI-Bold' if segoe_ui_bold_registered else 'Helvetica-Bold'
        
        # Registrar familia de fuentes para que las etiquetas <b> funcionen
        if segoe_ui_registered and segoe_ui_bold_registered:
            from reportlab.pdfbase.pdfmetrics import registerFontFamily
            registerFontFamily('SegoeUI', normal='SegoeUI', bold='SegoeUI-Bold')
        
        buffer = io.BytesIO()

        compact_gap = 0.28 * cm
        line_gap = 0.14 * cm

        def build_signature_image(user, width=3.6 * cm, height=1.15 * cm):
            from reportlab.platypus import Image as RLImage

            try:
                if user and hasattr(user, 'perfil'):
                    perfil = user.perfil
                    if perfil and perfil.firma_digital:
                        firma_path = perfil.firma_digital.path
                        if os.path.exists(firma_path):
                            firma = RLImage(firma_path, width=width, height=height)
                            firma.hAlign = 'LEFT'
                            return firma
            except Exception as e:
                logger.warning(f"Error obteniendo firma digital: {e}")

            return None

        def build_role_description(nombre, cargo=None, oficina=None):
            if not nombre:
                return ''

            description_lines = [f"<b>{nombre}</b>"]
            if cargo and oficina:
                description_lines.append(f"{cargo} - {oficina}")
            elif cargo:
                description_lines.append(cargo)
            elif oficina:
                description_lines.append(oficina)

            return '<br/>'.join(description_lines)

        class PushToBottom(RLFlowable):
            """Spacer that expands to push subsequent content to the bottom of the frame."""
            def __init__(self, subsequent_height):
                RLFlowable.__init__(self)
                self._subsequent_height = subsequent_height

            def wrap(self, availWidth, availHeight):
                push = max(0, availHeight - self._subsequent_height)
                return (0, push)

            def draw(self):
                pass

        footer_contact_lines = COMUNICACION_INTERNA_FOOTER_CONTACT_LINES

        def _draw_supersalud_logo(canvas, font_name_regular, font_name_strong):
            logo_gray = HexColor('#5f6165')
            logo_teal = HexColor('#28b7b0')
            logo_yellow = HexColor('#f6c21a')
            logo_blue = HexColor('#0d47a1')
            logo_red = HexColor('#d71920')

            vigilado_text = 'Vigilado'
            supersalud_text = 'Supersalud'
            vigilado_font_size = 7.0
            supersalud_font_size = 7.4
            word_gap = 0.12 * cm

            vigilado_width = canvas.stringWidth(vigilado_text, font_name_strong, vigilado_font_size)
            supersalud_width = canvas.stringWidth(supersalud_text, font_name_strong, supersalud_font_size)
            text_width = vigilado_width + word_gap + supersalud_width

            emblem_width = 0.48 * cm
            emblem_height = 0.62 * cm
            emblem_gap = 0.05 * cm
            emblem_x = vigilado_width + word_gap + (supersalud_width / 2) - (emblem_width / 2)
            emblem_y = 0.31 * cm

            bar_gap = 0.02 * cm
            bar_height = 0.10 * cm
            bar_y = -0.17 * cm
            bar_widths = [0.40 * cm, 0.18 * cm, 0.18 * cm]
            bars_total_width = sum(bar_widths) + (2 * bar_gap)
            bars_x = vigilado_width + (word_gap / 2) + (supersalud_width / 2) - (bars_total_width / 2)

            logo_width = text_width

            canvas.saveState()
            canvas.setFillColor(logo_gray)
            canvas.setFont(font_name_strong, vigilado_font_size)
            canvas.drawString(0, 0, vigilado_text)
            canvas.setFillColor(logo_teal)
            canvas.setFont(font_name_strong, supersalud_font_size)
            canvas.drawString(vigilado_width + word_gap, 0, supersalud_text)

            # Escudo lineal simplificado, ubicado sobre la palabra "Supersalud".
            canvas.setStrokeColor(logo_teal)
            canvas.setLineWidth(0.7)
            emblem_mid_x = emblem_x + (emblem_width / 2)
            shield_path = canvas.beginPath()
            shield_path.moveTo(emblem_x + 0.07 * cm, emblem_y + emblem_height)
            shield_path.curveTo(
                emblem_x + 0.03 * cm, emblem_y + emblem_height - 0.08 * cm,
                emblem_x + 0.02 * cm, emblem_y + 0.24 * cm,
                emblem_mid_x, emblem_y,
            )
            shield_path.curveTo(
                emblem_x + emblem_width - 0.02 * cm, emblem_y + 0.24 * cm,
                emblem_x + emblem_width - 0.03 * cm, emblem_y + emblem_height - 0.08 * cm,
                emblem_x + emblem_width - 0.07 * cm, emblem_y + emblem_height,
            )
            canvas.drawPath(shield_path, stroke=1, fill=0)

            canvas.line(emblem_mid_x, emblem_y + 0.06 * cm, emblem_mid_x, emblem_y + emblem_height - 0.10 * cm)
            canvas.line(emblem_x + 0.13 * cm, emblem_y + 0.31 * cm, emblem_x + emblem_width - 0.13 * cm, emblem_y + 0.31 * cm)

            left_wing = canvas.beginPath()
            left_wing.moveTo(emblem_mid_x - 0.03 * cm, emblem_y + emblem_height - 0.03 * cm)
            left_wing.curveTo(
                emblem_x + 0.18 * cm, emblem_y + emblem_height + 0.05 * cm,
                emblem_x + 0.08 * cm, emblem_y + emblem_height + 0.02 * cm,
                emblem_x + 0.05 * cm, emblem_y + emblem_height - 0.10 * cm,
            )
            canvas.drawPath(left_wing, stroke=1, fill=0)

            right_wing = canvas.beginPath()
            right_wing.moveTo(emblem_mid_x + 0.03 * cm, emblem_y + emblem_height - 0.03 * cm)
            right_wing.curveTo(
                emblem_x + emblem_width - 0.18 * cm, emblem_y + emblem_height + 0.05 * cm,
                emblem_x + emblem_width - 0.08 * cm, emblem_y + emblem_height + 0.02 * cm,
                emblem_x + emblem_width - 0.05 * cm, emblem_y + emblem_height - 0.10 * cm,
            )
            canvas.drawPath(right_wing, stroke=1, fill=0)

            current_x = bars_x
            for color, width in zip((logo_yellow, logo_blue, logo_red), bar_widths):
                canvas.setFillColor(color)
                canvas.rect(current_x, bar_y, width, bar_height, stroke=0, fill=1)
                current_x += width + bar_gap
            canvas.restoreState()

            return logo_width
        
        # Buscar imágenes
        encabezado_path = None
        for ext in ['png', 'jpg', 'jpeg']:
            if not encabezado_path:
                path = finders.find(f'correspondencia/img/encabezado.{ext}')
                if path and os.path.exists(path):
                    encabezado_path = path

        supersalud_logo_path = None
        for relative_path in [
            'correspondencia/img/vigilado_supersalud.png',
            'correspondencia/img/vigilado_supersalud.jpg',
            'correspondencia/img/vigilado_supersalud.jpeg',
        ]:
            path = finders.find(relative_path)
            if path and os.path.exists(path):
                supersalud_logo_path = path
                break
        
        # Funciones para dibujar encabezado y pie 
        
        def drawHeader(canvas, doc):
            if encabezado_path:
                try:
                    width, height = A4
                    img_width = width - 4*cm
                    img_height = 3*cm
                    canvas.saveState()
                    canvas.drawImage(encabezado_path, 2*cm, height - 3.5*cm,
                                   width=img_width, height=img_height,
                                   preserveAspectRatio=True, mask='auto')
                    canvas.restoreState()
                except Exception as e:
                    logger.warning(f"Error dibujando encabezado: {e}")
        
        def drawFooter(canvas, doc):
            try:
                width, height = A4

                canvas.saveState()

                # Línea institucional curva sobre el pie de página.
                wave_start_x = 1.55 * cm
                wave_end_x = width - (0.7 * cm)
                # Altura base global de la curva. Este es el valor principal para subir o bajar toda la línea.
                wave_base_y = 2.00 * cm

                def _draw_footer_wave(offset_y=0, color='#3a3a3f', stroke_width=1.4):
                    path = canvas.beginPath()
                    # Punto inicial de la curva. Cambiar 1.35 solo mueve la "punta" izquierda respecto a wave_base_y.
                    path.moveTo(wave_start_x, wave_base_y + 1.35 * cm + offset_y)
                    # Primer tramo Bezier:
                    # - 2.45 / 0.18 = primer punto de control
                    # - 4.35 / 0.05 = segundo punto de control
                    # - 7.1 / 0.28 = punto final del primer tramo
                    # Estos valores cambian la forma local del lado izquierdo, no la altura global completa.
                    path.curveTo(
                        2.45 * cm, wave_base_y + 0.18 * cm + offset_y,
                        4.35 * cm, wave_base_y + 0.05 * cm + offset_y,
                        7.1 * cm, wave_base_y + 0.28 * cm + offset_y,
                    )
                    # Segundo tramo Bezier:
                    # - 12.0 / 0.62 = primer punto de control central
                    # - 17.2 / 0.82 = segundo punto de control hacia la derecha
                    # - wave_end_x / 0.72 = punto final del trazo
                    # Estos valores afinan la panza y la salida derecha de la curva.
                    path.curveTo(
                        12.0 * cm, wave_base_y + 0.62 * cm + offset_y,
                        17.2 * cm, wave_base_y + 0.82 * cm + offset_y,
                        wave_end_x, wave_base_y + 0.72 * cm + offset_y,
                    )
                    canvas.saveState()
                    canvas.setLineCap(1)
                    canvas.setStrokeColor(HexColor(color))
                    canvas.setLineWidth(stroke_width)
                    canvas.drawPath(path, stroke=1, fill=0)
                    canvas.restoreState()

                # offset_y separa un trazo del otro. No mueve el bloque completo, solo la distancia entre ambas líneas.
                _draw_footer_wave(offset_y=0.06 * cm, color='#9ca1aa', stroke_width=0.95)
                _draw_footer_wave(offset_y=0, color='#3b3b40', stroke_width=1.85)

                # Sello vertical "Vigilado Supersalud" usando imagen estática para respetar mejor el logo oficial.
                supersalud_center_y = height * 0.25
                if supersalud_logo_path:
                    supersalud_img = ImageReader(supersalud_logo_path)
                    img_width, img_height = supersalud_img.getSize()
                    supersalud_logo_width = 3.55 * cm
                    supersalud_logo_height = supersalud_logo_width * (img_height / img_width)
                    supersalud_anchor_y = supersalud_center_y - (supersalud_logo_width / 2)
                    canvas.saveState()
                    canvas.translate(supersalud_logo_height + (0.18 * cm), supersalud_anchor_y)
                    canvas.rotate(90)
                    canvas.drawImage(
                        supersalud_img,
                        0,
                        0,
                        width=supersalud_logo_width,
                        height=supersalud_logo_height,
                        preserveAspectRatio=True,
                        mask='auto',
                    )
                    canvas.restoreState()
                else:
                    supersalud_logo_width = (
                        canvas.stringWidth('Vigilado', font_name_bold, 7.0)
                        + (0.12 * cm)
                        + canvas.stringWidth('Supersalud', font_name_bold, 7.4)
                    )
                    supersalud_anchor_y = supersalud_center_y - (supersalud_logo_width / 2)
                    canvas.saveState()
                    canvas.translate(0.42 * cm, supersalud_anchor_y)
                    canvas.rotate(90)
                    _draw_supersalud_logo(canvas, font_name, font_name_bold)
                    canvas.restoreState()

                # Bloque institucional del pie como texto constante y editable.
                canvas.setFillColor(HexColor('#111111'))
                canvas.setFont(font_name_bold, 7.15)
                footer_y = 0.92 * cm
                for index, line in enumerate(footer_contact_lines):
                    y = footer_y + ((len(footer_contact_lines) - 1 - index) * 0.24 * cm)
                    canvas.drawCentredString((width / 2) + (0.42 * cm), y, line)

                canvas.restoreState()
            except Exception as e:
                logger.warning(f"Error dibujando pie de página: {e}")
        
        def drawHeaderFooter(canvas, doc):
            drawHeader(canvas, doc)
            drawFooter(canvas, doc)
        
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                               rightMargin=1.8*cm, leftMargin=1.8*cm,
                               topMargin=4.0*cm, bottomMargin=3.8*cm)
        
        story = []
        styles = getSampleStyleSheet()
        
        encabezado_style = ParagraphStyle(
            'Encabezado', parent=styles['Heading2'],
            fontSize=11, textColor=HexColor('#2c3e50'),
            spaceAfter=12, spaceBefore=12, fontName=font_name_bold
        )
        
        normal_style = ParagraphStyle(
            'Normal', parent=styles['Normal'],
            fontSize=10.5, leading=13, alignment=TA_JUSTIFY, spaceAfter=4, fontName=font_name
        )
        
        # Estilo para negritas (necesario cuando se usan fuentes TTF personalizadas)
        bold_style = ParagraphStyle(
            'Bold', parent=normal_style,
            fontSize=11, fontName=font_name_bold
        )

        # Escala única del cuadro de firmas/revisión.
        # 0.70 = reducción aproximada del 30%.
        review_box_scale = 0.70

        compact_style = ParagraphStyle(
            'Compact', parent=normal_style,
            fontSize=7 * review_box_scale,
            leading=10.5 * review_box_scale,
            spaceAfter=0,
            fontName=font_name
        )

        compact_bold_style = ParagraphStyle(
            'CompactBold', parent=compact_style,
            fontName=font_name_bold
        )

        body_style = normal_style
        if any(len(token) > 45 for token in instance.cuerpo.split()):
            body_style = ParagraphStyle(
                'BodyWrap', parent=normal_style,
                wordWrap='CJK'
            )

        story.append(Spacer(1, line_gap))
        
        # Formato según especificación: Ciudad, día de mes de año
        fecha_str = formatear_fecha_espanol(instance.fecha_documento)
        story.append(Paragraph(f"{instance.ciudad}, {fecha_str}", normal_style))
        story.append(Spacer(1, line_gap))
        
        # Radicado
        story.append(Paragraph(f"{instance.radicado or 'BORRADOR'}", normal_style))
        
        # Si es una respuesta, agregar texto formal
        if instance.comunicacion_origen and instance.comunicacion_origen.radicado:
            texto_respuesta = f"En respuesta a la comunicación con número de radicado {instance.comunicacion_origen.radicado}"
            story.append(Paragraph(texto_respuesta, normal_style))
        
        # TRD
        if instance.trd:
            story.append(Paragraph(f"<b>TRD:</b> {instance.trd}", normal_style))
        
        story.append(Spacer(1, compact_gap))
        
        # Señor(es) y destinatario
        story.append(Paragraph("Señor(es)", normal_style))
        
        # Determinar tipo de distribución (compatibilidad con es_a_toda_entidad)
        tipo_dist = getattr(instance, 'tipo_distribucion', None)
        if not tipo_dist:
            tipo_dist = 'ENTIDAD' if instance.es_a_toda_entidad else 'USUARIO'
        
        if tipo_dist == 'ENTIDAD' or instance.es_a_toda_entidad:
            # Para toda la entidad
            story.append(Paragraph("<b>Funcionarios de la entidad</b>", normal_style))
        
        elif tipo_dist == 'PROCESO':
            # Para proceso completo
            if instance.destinatario_proceso:
                proceso_nombre = instance.destinatario_proceso.nombre
                story.append(Paragraph(f"<b>Funcionarios del Proceso {proceso_nombre}</b>", normal_style))
            else:
                story.append(Paragraph("Funcionarios del Proceso", normal_style))
        
        elif tipo_dist == 'OFICINA':
            # Para oficinas completas (subprocesos) - múltiples
            from .models import ComunicacionInternaDestinatario
            
            # Obtener todas las oficinas destinatarias
            oficinas_destinatarias = ComunicacionInternaDestinatario.objects.filter(
                comunicacion=instance,
                tipo='OFICINA'
            ).select_related('oficina', 'oficina__unidad_administrativa', 'oficina__unidad_administrativa__entidad_productora')
            
            if oficinas_destinatarias.exists():
                # Si hay múltiples oficinas, listarlas todas
                if oficinas_destinatarias.count() > 1:
                    story.append(Paragraph("<b>Funcionarios de los siguientes Subprocesos:</b>", normal_style))
                    for dest in oficinas_destinatarias:
                        oficina = dest.oficina
                        story.append(Paragraph(f"• <b>{oficina.nombre}</b>", normal_style))
                else:
                    # Una sola oficina
                    oficina = oficinas_destinatarias.first().oficina
                    story.append(Paragraph(f"<b>Funcionarios del Subproceso {oficina.nombre}</b>", normal_style))
            elif instance.destinatario_oficina:
                # Fallback para compatibilidad
                oficina_nombre = instance.destinatario_oficina.nombre
                story.append(Paragraph(f"<b>Funcionarios del Subproceso {oficina_nombre}</b>", normal_style))
            else:
                story.append(Paragraph("Funcionarios del Subproceso", normal_style))
        
        else:  # USUARIO - Múltiples destinatarios
            from .models import ComunicacionInternaDestinatario
            
            # Obtener destinatarios del nuevo modelo
            destinatarios = ComunicacionInternaDestinatario.objects.filter(
                comunicacion=instance
            ).select_related('usuario', 'usuario__perfil', 'usuario__perfil__oficina', 'oficina').order_by('tipo', 'oficina', 'usuario')
            
            if destinatarios.exists():
                # Hay destinatarios múltiples
                usuarios_destinatarios = []
                oficinas_destinatarios = []
                
                for dest in destinatarios:
                    if dest.tipo == 'USUARIO' and dest.usuario:
                        usuarios_destinatarios.append(dest)
                    elif dest.tipo == 'OFICINA' and dest.oficina:
                        oficinas_destinatarios.append(dest)
                
                # Mostrar usuarios primero
                if usuarios_destinatarios:
                    for dest in usuarios_destinatarios:
                        usuario = dest.usuario
                        nombre_completo = f"{usuario.first_name} {usuario.last_name}".strip() or usuario.username
                        story.append(Paragraph(f"<b>{nombre_completo}</b>", normal_style))
                        
                        # Cargo y oficina
                        try:
                            perfil_destino = usuario.perfil
                            cargo_destino = getattr(perfil_destino, 'cargo', None)
                            if cargo_destino:
                                oficina_destino_nombre = None
                                try:
                                    if perfil_destino.oficina:
                                        oficina_destino_nombre = perfil_destino.oficina.nombre
                                except AttributeError:
                                    pass
                                
                                if oficina_destino_nombre:
                                    story.append(Paragraph(f"{cargo_destino} - {oficina_destino_nombre}", normal_style))
                                else:
                                    story.append(Paragraph(cargo_destino, normal_style))
                        except AttributeError:
                            pass
                        
                        story.append(Spacer(1, line_gap))
                
                # Mostrar oficinas completas
                if oficinas_destinatarios:
                    for dest in oficinas_destinatarios:
                        oficina = dest.oficina
                        story.append(Paragraph(f"<b>Funcionarios del Subproceso {oficina.nombre}</b>", normal_style))

                        story.append(Spacer(1, line_gap))
            
            else:
                # Sistema antiguo (compatibilidad) - un solo destinatario
                if instance.destinatario_usuario:
                    destinatario_nombre = f"{instance.destinatario_usuario.first_name} {instance.destinatario_usuario.last_name}".strip()
                    story.append(Paragraph(f"<b>{destinatario_nombre}</b>", normal_style))
                    
                    # Cargo del usuario destino
                    try:
                        perfil_destino = instance.destinatario_usuario.perfil
                        cargo_destino = getattr(perfil_destino, 'cargo', None)
                        if cargo_destino:
                            # Obtener oficina del destinatario
                            oficina_destino = None
                            try:
                                if instance.destinatario_oficina:
                                    oficina_destino = instance.destinatario_oficina.nombre
                                elif perfil_destino.oficina:
                                    oficina_destino = perfil_destino.oficina.nombre
                            except AttributeError:
                                pass
                            
                            if oficina_destino:
                                story.append(Paragraph(f"{cargo_destino} - {oficina_destino}", normal_style))
                            else:
                                story.append(Paragraph(cargo_destino, normal_style))
                    except AttributeError:
                        pass
                else:
                    story.append(Paragraph("Encargado", normal_style))
        
        story.append(Spacer(1, compact_gap))
        
        # Asunto
        story.append(Paragraph(f"<b>Asunto:</b> {instance.asunto}", normal_style))
        story.append(Spacer(1, compact_gap))
        
        # Cuerpo
        cuerpo_html = instance.cuerpo.replace('\n', '<br/>')
        story.append(Paragraph(cuerpo_html, body_style))
        story.append(Spacer(1, compact_gap))
        
        # Anexo (si hay archivo generado)
        if instance.archivo_generado:
            story.append(Paragraph("Anexo", normal_style))
            story.append(Spacer(1, line_gap))
        
        story.append(Spacer(1, compact_gap))
        
        # Firma: Cordialmente
        story.append(Paragraph("Cordialmente,", normal_style))
        story.append(Spacer(1, compact_gap))
        
        remitente_user = instance.remitente_usuario
        remitente_signature = build_signature_image(remitente_user)
        if remitente_signature:
            story.append(remitente_signature)
            story.append(Spacer(1, line_gap))
        
        # Nombre del remitente
        story.append(Paragraph(f"<b>{instance.remitente_nombre}</b>", normal_style))
        oficina_remitente = None
        if instance.remitente_cargo:
            # Obtener oficina del remitente
            try:
                if instance.remitente_oficina:
                    oficina_remitente = instance.remitente_oficina.nombre
            except AttributeError:
                pass
            
            if oficina_remitente:
                story.append(Paragraph(f"{instance.remitente_cargo} - {oficina_remitente}", normal_style))
            else:
                story.append(Paragraph(instance.remitente_cargo, normal_style))
        
        review_rows = []

        revisor_user = instance.revisado_por
        revisor_nombre = instance.revisado_nombre
        if not revisor_nombre and revisor_user:
            revisor_nombre = revisor_user.get_full_name() or revisor_user.username

        if revisor_nombre or instance.revisado_cargo or revisor_user:
            revisor_signature = build_signature_image(
                revisor_user,
                width=3.2 * cm * review_box_scale,
                height=0.95 * cm * review_box_scale,
            )
            review_rows.append([
                Paragraph("REVISO", compact_bold_style),
                Paragraph(build_role_description(revisor_nombre, instance.revisado_cargo, oficina_remitente), compact_style),
                revisor_signature or '',
            ])

        proyecto_signature = build_signature_image(
            remitente_user,
            width=3.2 * cm * review_box_scale,
            height=0.95 * cm * review_box_scale,
        )
        review_rows.append([
            Paragraph("PROYECTO", compact_bold_style),
            Paragraph(build_role_description(instance.remitente_nombre, instance.remitente_cargo, oficina_remitente), compact_style),
            proyecto_signature or '',
        ])

        if review_rows:
            review_table = Table(
                review_rows,
                colWidths=[2.2 * cm * review_box_scale, 5.0 * cm * review_box_scale, 5.0 * cm * review_box_scale],
                hAlign='LEFT'
            )
            review_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.6 * review_box_scale, HexColor('#1f2937')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 5 * review_box_scale),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5 * review_box_scale),
                ('TOPPADDING', (0, 0), (-1, -1), 4 * review_box_scale),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4 * review_box_scale),
                ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ]))

            review_gap = Spacer(1, 0.35 * cm * review_box_scale)
            review_block_height = 0
            for flowable in [review_gap, review_table]:
                _, fh = flowable.wrap(doc.width, doc.height)
                review_block_height += fh

            pusher = PushToBottom(review_block_height)
            story.append(pusher)
            story.append(review_gap)
            story.append(review_table)
        
        doc.build(story, onFirstPage=drawHeaderFooter, onLaterPages=drawHeaderFooter)
        buffer.seek(0)
        
        filename = f"OFICIO_{instance.radicado or 'BORRADOR'}.pdf"
        instance.archivo_generado.save(filename, ContentFile(buffer.read()), save=True)
        
    except Exception as e:
        logger.error(f"Error generando documento PDF: {e}")
        import traceback
        logger.error(traceback.format_exc())
        if request:
            messages.warning(request, f"Se guardó el registro pero falló la generación del documento: {e}")


# =============================================
# === TRAZABILIDAD DE COMUNICACIONES INTERNAS ===
# =============================================

class ComunicacionInternaTrazabilidadView(LoginRequiredMixin, DetailView):
    """
    Vista para mostrar la trazabilidad de lectura de una comunicación interna.
    Muestra qué usuarios han visto la comunicación, agrupados por oficina,
    con la fecha y hora exacta de lectura.
    """
    model = ComunicacionInterna
    template_name = 'correspondencia/interna/trazabilidad.html'
    context_object_name = 'comunicacion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import ComunicacionInternaDestinatario, ComunicacionInternaDistribucion
        from collections import defaultdict
        
        comunicacion = self.object
        
        # Verificar que no es una respuesta (las respuestas no tienen respuestas)
        if comunicacion.es_respuesta():
            context['es_respuesta'] = True
            context['comunicacion_origen'] = comunicacion.comunicacion_origen
            context['respuestas'] = []
            context['estadisticas'] = {
                'total_respuestas': 0,
                'respuestas_destacadas': 0,
                'respuestas_normales': 0,
                'por_estado': {},
                'total_usuarios': 0,
                'total_leidos': 0,
                'total_pendientes': 0,
                'porcentaje': 0,
                'total_oficinas': 0,
            }
            context['oficinas_trazabilidad'] = []
            context['titulo_pagina'] = f"Trazabilidad: {comunicacion.radicado or 'BORRADOR'}"
            return context
        
        # ── Trazabilidad de lectura ──
        distribuciones = ComunicacionInternaDistribucion.objects.filter(
            comunicacion=comunicacion,
        ).select_related(
            'usuario',
            'usuario__perfil',
            'usuario__perfil__oficina',
        ).order_by('usuario__perfil__oficina__nombre', 'usuario__last_name', 'usuario__first_name')

        total_usuarios = distribuciones.count()
        total_leidos = distribuciones.filter(leido=True).count()
        total_pendientes = total_usuarios - total_leidos
        porcentaje = round((total_leidos / total_usuarios) * 100) if total_usuarios > 0 else 0

        # Agrupar por oficina
        oficinas_dict = defaultdict(lambda: {'usuarios': [], 'leidos': 0, 'pendientes': 0})
        for dist in distribuciones:
            try:
                oficina = dist.usuario.perfil.oficina
                oficina_nombre = oficina.nombre if oficina else 'Sin oficina asignada'
            except AttributeError:
                oficina_nombre = 'Sin oficina asignada'

            usuario_data = {
                'usuario': dist.usuario,
                'nombre_completo': dist.usuario.get_full_name() or dist.usuario.username,
                'leido': dist.leido,
                'fecha_lectura': dist.fecha_lectura,
                'fecha_distribucion': dist.fecha_distribucion,
            }
            oficinas_dict[oficina_nombre]['usuarios'].append(usuario_data)
            if dist.leido:
                oficinas_dict[oficina_nombre]['leidos'] += 1
            else:
                oficinas_dict[oficina_nombre]['pendientes'] += 1

        oficinas_trazabilidad = []
        for nombre, data in sorted(oficinas_dict.items()):
            total_ofi = data['leidos'] + data['pendientes']
            oficinas_trazabilidad.append({
                'nombre': nombre,
                'leidos': data['leidos'],
                'pendientes': data['pendientes'],
                'porcentaje': round((data['leidos'] / total_ofi) * 100) if total_ofi > 0 else 0,
                'usuarios': data['usuarios'],
            })

        # ── Respuestas ──
        respuestas = comunicacion.get_respuestas_ordenadas().select_related(
            'remitente_usuario',
            'remitente_oficina',
            'revisado_por',
        ).prefetch_related('historial')
        
        respuestas_por_estado = defaultdict(list)
        respuestas_destacadas = 0
        respuestas_normales = 0
        
        respuestas_lista = []
        for resp in respuestas:
            estado_display = resp.get_estado_display()
            respuestas_por_estado[estado_display].append(resp)
            
            if resp.es_respuesta_destacada:
                respuestas_destacadas += 1
            else:
                respuestas_normales += 1
            
            respuestas_lista.append({
                'obj': resp,
                'remitente_nombre': resp.remitente_nombre or (resp.remitente_usuario.get_full_name() if resp.remitente_usuario else 'Desconocido'),
                'oficina_nombre': resp.remitente_oficina.nombre if resp.remitente_oficina else 'Sin oficina',
                'es_destacada': resp.es_respuesta_destacada,
                'estado': resp.estado,
                'estado_display': estado_display,
                'fecha_creacion': resp.fecha_creacion,
                'radicado': resp.radicado,
            })
        
        total_respuestas = len(respuestas_lista)
        
        context['respuestas'] = respuestas_lista
        context['es_respuesta'] = False
        context['oficinas_trazabilidad'] = oficinas_trazabilidad
        context['estadisticas'] = {
            'total_respuestas': total_respuestas,
            'respuestas_destacadas': respuestas_destacadas,
            'respuestas_normales': respuestas_normales,
            'por_estado': dict(respuestas_por_estado),
            'total_usuarios': total_usuarios,
            'total_leidos': total_leidos,
            'total_pendientes': total_pendientes,
            'porcentaje': porcentaje,
            'total_oficinas': len(oficinas_trazabilidad),
        }
        context['titulo_pagina'] = f"Trazabilidad: {comunicacion.radicado or 'BORRADOR'}"
        
        return context


class BandejaRadicacionRapidaView(LoginRequiredMixin, View):
    """Bandeja exclusiva para correspondencia registrada mediante radicación rápida."""
    template_name = 'correspondencia/admin/bandeja_radicacion_rapida.html'
    paginate_by = 25

    def get(self, request, *args, **kwargs):
        # Limpiar filtros si se hace clic en "Limpiar"
        if request.GET.get('reset') == 'Limpiar Filtros':
            return redirect('correspondencia:bandeja_radicacion_rapida')

        # Paginación configurable
        registros_por_pagina = request.GET.get('registros_por_pagina', '25')
        try:
            registros_por_pagina = int(registros_por_pagina)
            if registros_por_pagina not in [10, 25, 50, 100, 150]:
                registros_por_pagina = 25
        except (ValueError, TypeError):
            registros_por_pagina = 25

        # Obtener parámetros de filtro (alineados con el formulario de radicación rápida entrante)
        search_term = request.GET.get('search_term', '').strip()
        oficina_destino = request.GET.get('oficina_destino', '').strip()
        entidad_remitente = request.GET.get('entidad_remitente', '').strip()
        remitente_texto = request.GET.get('remitente_texto', '').strip()
        funcionario_responsable = request.GET.get('funcionario_responsable', '').strip()
        clasificacion = request.GET.get('clasificacion', '').strip()
        medio_recepcion = request.GET.get('medio_recepcion', '').strip()
        medio_recibido = request.GET.get('medio_recibido', '').strip()
        empresa_transportadora = request.GET.get('empresa_transportadora', '').strip()
        estado_respuesta = request.GET.get('estado_respuesta', '').strip()
        anexos = request.GET.get('anexos', '').strip()
        numero_guia = request.GET.get('numero_guia', '').strip()
        fecha_inicio = request.GET.get('fecha_inicio', '').strip()
        fecha_fin = request.GET.get('fecha_fin', '').strip()

        # Queryset base: solo correspondencia con origen_radicacion='RAPIDA'
        qs = Correspondencia.objects.filter(
            origen_radicacion='RAPIDA'
        ).select_related(
            'oficina_destino', 'remitente', 'remitente__entidad_externa', 'usuario_radicador'
        ).prefetch_related(
            'correo_origen'
        ).order_by('-fecha_radicacion')

        # Aplicar filtros (todos los campos del form de radicación rápida entrante)
        if search_term:
            qs = qs.filter(
                Q(numero_radicado__icontains=search_term) |
                Q(asunto__icontains=search_term) |
                Q(entidad_persona_remitente__icontains=search_term) |
                Q(funcionario_responsable_tramite__icontains=search_term) |
                Q(radicado_enviado_respuesta__icontains=search_term) |
                Q(anexos__icontains=search_term) |
                Q(numero_guia__icontains=search_term)
            )
        if oficina_destino:
            qs = qs.filter(oficina_destino__nombre__icontains=oficina_destino)
        if entidad_remitente:
            qs = qs.filter(entidad_persona_remitente__icontains=entidad_remitente)
        if remitente_texto:
            qs = qs.filter(entidad_persona_remitente__icontains=remitente_texto)
        if funcionario_responsable:
            qs = qs.filter(funcionario_responsable_tramite__icontains=funcionario_responsable)
        if clasificacion:
            qs = qs.filter(clasificacion_comunicacion__icontains=clasificacion)
        if medio_recepcion:
            qs = qs.filter(medio_recepcion=medio_recepcion)
        if medio_recibido:
            qs = qs.filter(medio_recibido=medio_recibido)
        if empresa_transportadora:
            qs = qs.filter(empresa_transportadora__icontains=empresa_transportadora)
        if estado_respuesta:
            qs = qs.filter(estado_respuesta=estado_respuesta)
        if anexos:
            qs = qs.filter(anexos__icontains=anexos)
        if numero_guia:
            qs = qs.filter(numero_guia__icontains=numero_guia)
        if fecha_inicio:
            qs = qs.filter(fecha_radicacion__date__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha_radicacion__date__lte=fecha_fin)

        # Paginación
        paginator = Paginator(qs, registros_por_pagina)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Listas para autocompletado
        oficinas_list = OficinaProductora.objects.values_list('nombre', flat=True).order_by('nombre')

        # Filtros activos para mostrar badges (todos los del form de radicación rápida)
        filtros_activos = {}
        if search_term:
            filtros_activos['search_term'] = {'label': 'Búsqueda', 'value': search_term}
        if oficina_destino:
            filtros_activos['oficina_destino'] = {'label': 'Oficina Destino', 'value': oficina_destino}
        if entidad_remitente:
            filtros_activos['entidad_remitente'] = {'label': 'Entidad', 'value': entidad_remitente}
        if remitente_texto:
            filtros_activos['remitente_texto'] = {'label': 'Remitente texto libre', 'value': remitente_texto}
        if funcionario_responsable:
            filtros_activos['funcionario_responsable'] = {'label': 'Funcionario Responsable', 'value': funcionario_responsable}
        if clasificacion:
            filtros_activos['clasificacion'] = {'label': 'Clasificación', 'value': clasificacion}
        if medio_recepcion:
            medio_recepcion_display = dict(MEDIO_RECEPCION_CHOICES).get(medio_recepcion, medio_recepcion)
            filtros_activos['medio_recepcion'] = {'label': 'Medio de Recepción', 'value': medio_recepcion_display}
        if medio_recibido:
            medio_recibido_display = dict(MEDIO_RECIBIDO_CHOICES).get(medio_recibido, medio_recibido)
            filtros_activos['medio_recibido'] = {'label': 'Medio de Recibido', 'value': medio_recibido_display or medio_recibido}
        if empresa_transportadora:
            filtros_activos['empresa_transportadora'] = {'label': 'Empresa Transportadora', 'value': empresa_transportadora}
        if estado_respuesta:
            estado_display = dict(ESTADO_RESPUESTA_RAPIDA_CHOICES).get(estado_respuesta, estado_respuesta)
            filtros_activos['estado_respuesta'] = {'label': 'Estado Respuesta', 'value': estado_display or estado_respuesta}
        if anexos:
            filtros_activos['anexos'] = {'label': 'Anexos', 'value': anexos}
        if numero_guia:
            filtros_activos['numero_guia'] = {'label': 'Nº Guía', 'value': numero_guia}
        if fecha_inicio:
            filtros_activos['fecha_inicio'] = {'label': 'Desde', 'value': fecha_inicio}
        if fecha_fin:
            filtros_activos['fecha_fin'] = {'label': 'Hasta', 'value': fecha_fin}

        # Form vacío para el modal de creación/edición
        form_rapida_entrante = RadicacionRapidaEntranteForm(prefix='rapida_ent')

        context = {
            'titulo_pagina': 'Bandeja de Radicación Rápida (Entrantes)',
            'bandeja_activa': 'entrantes',
            'page_obj': page_obj,
            'filtros_activos': filtros_activos,
            'oficinas_list': oficinas_list,
            'form_rapida_entrante': form_rapida_entrante,
            # Valores actuales para mantener en el formulario (todos los del form)
            'search_term': search_term,
            'oficina_destino': oficina_destino,
            'entidad_remitente': entidad_remitente,
            'remitente_texto': remitente_texto,
            'funcionario_responsable': funcionario_responsable,
            'clasificacion': clasificacion,
            'medio_recepcion': medio_recepcion,
            'medio_recibido': medio_recibido,
            'empresa_transportadora': empresa_transportadora,
            'estado_respuesta': estado_respuesta,
            'anexos': anexos,
            'numero_guia': numero_guia,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'registros_por_pagina': registros_por_pagina,
            # Choices para selects (form radicación rápida entrante)
            'medio_recepcion_choices': MEDIO_RECEPCION_CHOICES,
            'medio_recibido_choices': MEDIO_RECIBIDO_CHOICES,
            'estado_respuesta_choices': [c for c in ESTADO_RESPUESTA_RAPIDA_CHOICES if c[0]],
        }

        return render(request, self.template_name, context)


class BandejaRadicacionRapidaSalienteView(LoginRequiredMixin, View):
    """Bandeja de correspondencia saliente registrada por radicación rápida (independientes y vinculadas a entrantes)."""
    template_name = 'correspondencia/admin/bandeja_radicacion_rapida_saliente.html'
    paginate_by = 25

    def get(self, request, *args, **kwargs):
        if request.GET.get('reset') == 'Limpiar Filtros':
            return redirect('correspondencia:bandeja_radicacion_rapida_saliente')

        # Paginación configurable
        registros_por_pagina = request.GET.get('registros_por_pagina', '25')
        try:
            registros_por_pagina = int(registros_por_pagina)
            if registros_por_pagina not in [10, 25, 50, 100, 150]:
                registros_por_pagina = 25
        except (ValueError, TypeError):
            registros_por_pagina = 25

        search_term = request.GET.get('search_term', '').strip()
        oficina_emisora = request.GET.get('oficina_emisora', '').strip()
        asunto = request.GET.get('asunto', '').strip()
        destinatario = request.GET.get('destinatario', '').strip()
        estado = request.GET.get('estado', '').strip()
        fecha_inicio = request.GET.get('fecha_inicio', '').strip()
        fecha_fin = request.GET.get('fecha_fin', '').strip()

        # Incluir salidas independientes (sin respuesta_a) Y salidas generadas
        # desde radicación rápida entrante (con respuesta_a de origen RAPIDA)
        qs = CorrespondenciaSalida.objects.filter(
            Q(respuesta_a__isnull=True) |
            Q(respuesta_a__origen_radicacion='RAPIDA')
        ).select_related(
            'oficina_emisora', 'destinatario_contacto', 'destinatario_contacto__entidad_externa',
            'usuario_redactor', 'respuesta_a'
        ).order_by('-fecha_creacion')

        if search_term:
            qs = qs.filter(
                Q(numero_radicado_salida__icontains=search_term) |
                Q(asunto__icontains=search_term) |
                Q(cuerpo__icontains=search_term)
            )
        if oficina_emisora:
            qs = qs.filter(oficina_emisora__nombre__icontains=oficina_emisora)
        if asunto:
            qs = qs.filter(asunto__icontains=asunto)
        if destinatario:
            qs = qs.filter(
                Q(destinatario_contacto__nombres__icontains=destinatario) |
                Q(destinatario_contacto__apellidos__icontains=destinatario) |
                Q(destinatario_contacto__entidad_externa__nombre__icontains=destinatario)
            )
        if estado:
            qs = qs.filter(estado=estado)
        if fecha_inicio:
            qs = qs.filter(fecha_creacion__date__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha_creacion__date__lte=fecha_fin)

        paginator = Paginator(qs, registros_por_pagina)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        oficinas_list = OficinaProductora.objects.values_list('nombre', flat=True).order_by('nombre')
        from .models import ESTADOS_SALIDA
        estado_choices = list(ESTADOS_SALIDA)

        filtros_activos = {}
        if search_term:
            filtros_activos['search_term'] = {'label': 'Búsqueda', 'value': search_term}
        if oficina_emisora:
            filtros_activos['oficina_emisora'] = {'label': 'Oficina Emisora', 'value': oficina_emisora}
        if asunto:
            filtros_activos['asunto'] = {'label': 'Asunto', 'value': asunto}
        if destinatario:
            filtros_activos['destinatario'] = {'label': 'Destinatario', 'value': destinatario}
        if estado:
            estado_display = dict(ESTADOS_SALIDA).get(estado, estado)
            filtros_activos['estado'] = {'label': 'Estado', 'value': estado_display}
        if fecha_inicio:
            filtros_activos['fecha_inicio'] = {'label': 'Desde', 'value': fecha_inicio}
        if fecha_fin:
            filtros_activos['fecha_fin'] = {'label': 'Hasta', 'value': fecha_fin}

        # Form vacío para el modal de creación/edición
        form_rapida_saliente = RadicacionRapidaSalienteForm(prefix='rapida_sal', user=request.user)

        context = {
            'titulo_pagina': 'Bandeja de Radicación Rápida (Salientes)',
            'bandeja_activa': 'salientes',
            'page_obj': page_obj,
            'filtros_activos': filtros_activos,
            'oficinas_list': oficinas_list,
            'form_rapida_saliente': form_rapida_saliente,
            'search_term': search_term,
            'oficina_emisora': oficina_emisora,
            'asunto': asunto,
            'destinatario': destinatario,
            'estado': estado,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'estado_choices': estado_choices,
            'registros_por_pagina': registros_por_pagina,
        }
        return render(request, self.template_name, context)

# ===== Clases restauradas (eliminadas por error en limpieza) =====

class BandejaRevisionManualView(LoginRequiredMixin, EsVentanillaMixin, ListView):
    model = CorreoEntrante
    template_name = 'correspondencia/admin/bandeja_clasificados.html' # Reutilizar plantilla
    context_object_name = 'correos_clasificados' # Mantener nombre de contexto para la plantilla
    # paginate_by = 15 # Quitar paginación del servidor

    def get_queryset(self):
        """Filtra correos procesados, no radicados y que SÍ requieren revisión manual."""
        queryset = CorreoEntrante.objects.filter(
            procesado=True,
            radicado_asociado__isnull=True,
            urgencia_asociada__isnull=True,
            requiere_revision_manual=True # <--- La diferencia clave
        ).select_related(
            'oficina_clasificada',
            'serie_clasificada',
            'subserie_clasificada'
        ).prefetch_related(
            'adjuntos'
        ).order_by('-fecha_clasificacion')
        return queryset


# ==============================================
# === GRUPOS DE AGENDA Y COMUNICACIÓN MASIVA ===
# ==============================================



class DetalleCorreoClasificadoView(LoginRequiredMixin, EsVentanillaMixin, DetailView):
    model = CorreoEntrante
    template_name = 'correspondencia/admin/detalle_correo_clasificado.html'
    context_object_name = 'correo'

    def get_queryset(self):
        """Asegurar que solo se puedan ver correos procesados y no radicados."""
        return super().get_queryset().filter(
            procesado=True,
            radicado_asociado__isnull=True,
            urgencia_asociada__isnull=True
        ).prefetch_related('adjuntos') # Cargar adjuntos

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Detalle Correo: {self.object.asunto[:50]}..."
        return context
# --- FIN VISTA DETALLE CORREO ---


class ComunicacionInternaEnviadasView(LoginRequiredMixin, ListView):
    """Bandeja de comunicaciones enviadas por la oficina del usuario logueado."""
    model = ComunicacionInterna
    template_name = 'correspondencia/interna/enviadas.html'
    context_object_name = 'comunicaciones'
    paginate_by = 20

    def get_queryset(self):
        user = self.request.user

        try:
            oficina = user.perfil.oficina
        except AttributeError:
            return ComunicacionInterna.objects.none()

        qs = ComunicacionInterna.objects.select_related(
            'remitente_usuario',
            'remitente_oficina',
            'destinatario_oficina',
            'destinatario_usuario',
            'destinatario_proceso',
        ).prefetch_related('destinatarios_multiples').filter(
            remitente_oficina=oficina
        )
        
        # Filtrar por estado si se especifica
        estado = self.request.GET.get('estado')
        if estado:
            qs = qs.filter(estado=estado)
        
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(radicado__icontains=q) |
                Q(asunto__icontains=q)
            )

        # Filtro por rango de fechas
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        if fecha_inicio:
            qs = qs.filter(fecha_creacion__date__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha_creacion__date__lte=fecha_fin)

        # Filtro por destinatario (nombre o nombre de oficina destino)
        destinatario = (self.request.GET.get('destinatario') or '').strip()
        if destinatario:
            qs = qs.filter(
                Q(destinatario_oficina__nombre__icontains=destinatario)
                | Q(destinatario_usuario__first_name__icontains=destinatario)
                | Q(destinatario_usuario__last_name__icontains=destinatario)
            )

        # Filtro por tipo de distribución
        tipo_dist = self.request.GET.get('tipo_distribucion')
        if tipo_dist:
            qs = qs.filter(tipo_distribucion=tipo_dist)

        return qs.order_by('-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = 'Comunicaciones Enviadas por Mi Oficina'
        context['tipo_bandeja'] = 'enviadas'
        try:
            context['oficina_actual'] = self.request.user.perfil.oficina
        except AttributeError:
            context['oficina_actual'] = None
        return context




class InternaBienvenidaView(LoginRequiredMixin, TemplateView):
    """Página de bienvenida para comunicaciones internas."""
    template_name = 'correspondencia/interna/bienvenida.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            perfil = self.request.user.perfil
            context['tiene_firma'] = bool(perfil.firma_digital)
            if context['tiene_firma']:
                context['fecha_firma'] = perfil.fecha_firma_creada
        except AttributeError:
            context['tiene_firma'] = False
        return context




class InternaConfigurarFirmaView(LoginRequiredMixin, TemplateView):
    """Página para configurar la firma digital con términos y condiciones."""
    template_name = 'correspondencia/interna/configurar_firma.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            perfil = self.request.user.perfil
            context['tiene_firma'] = bool(perfil.firma_digital)
            if context['tiene_firma']:
                context['fecha_firma'] = perfil.fecha_firma_creada
        except AttributeError:
            context['tiene_firma'] = False
        return context




class ComunicacionInternaListView(LoginRequiredMixin, RedirectView):
    """Ruta legacy de internas; redirige a la entrada vigente del módulo."""
    permanent = False
    pattern_name = 'correspondencia:interna_bienvenida'



class ComunicacionInternaDetailView(LoginRequiredMixin, DetailView):
    model = ComunicacionInterna
    template_name = 'correspondencia/interna/detalle.html'
    context_object_name = 'comunicacion'

    def get_object(self, queryset=None):
        """Validar que el usuario tenga acceso a la comunicación."""
        from django.http import Http404
        
        obj = super().get_object(queryset)
        if not _usuario_puede_ver_comunicacion_interna(self.request.user, obj):
            raise Http404("No tienes acceso a esta comunicación.")
        
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import HistorialComunicacionInterna, ComunicacionInternaDistribucion
        
        user = self.request.user
        comunicacion = self.object
        
        # Obtener historial de la comunicación
        context['historial'] = HistorialComunicacionInterna.objects.filter(
            comunicacion=comunicacion
        ).order_by('-fecha')
        
        # Marcar como leída si el usuario es destinatario
        try:
            distribucion = ComunicacionInternaDistribucion.objects.get(
                comunicacion=comunicacion,
                usuario=user
            )
            if not distribucion.leido:
                distribucion.marcar_leido()
                # Registrar en historial
                HistorialComunicacionInterna.objects.create(
                    comunicacion=comunicacion,
                    evento='LEIDA',
                    usuario=user,
                    descripcion=f"Leída por {user.get_full_name() or user.username}"
                )
        except ComunicacionInternaDistribucion.DoesNotExist:
            pass
        
        # Calcular si el usuario puede responder (para pasar al template)
        context['puede_responder'] = comunicacion.puede_responder(user)
        
        return context




class ComunicacionInternaDescargarView(LoginRequiredMixin, View):
    """Vista para descargar el documento PDF."""
    def get(self, request, pk):
        obj = get_object_or_404(ComunicacionInterna, pk=pk)
        
        # Verificar que el usuario tenga acceso a esta comunicación
        if not _usuario_puede_ver_comunicacion_interna(request.user, obj):
            return HttpResponseForbidden("No tienes acceso a esta comunicación.")
        
        if obj.archivo_generado:
            # Generar nombre descriptivo basado en la oficina del redactor
            nombre_archivo = "comunicacion_interna"
            
            # Intentar obtener el nombre de la oficina del usuario redactor
            try:
                if obj.remitente_usuario and hasattr(obj.remitente_usuario, 'perfil'):
                    perfil = obj.remitente_usuario.perfil
                    if perfil and perfil.oficina:
                        nombre_oficina = perfil.oficina.nombre
                        # Limpiar el nombre para usarlo en archivo
                        nombre_oficina = nombre_oficina.replace(' ', '_').replace('/', '_').lower()
                        nombre_archivo = f"comunicacion_interna_{nombre_oficina}"
                    
                elif obj.remitente_oficina:
                    nombre_oficina = obj.remitente_oficina.nombre
                    nombre_oficina = nombre_oficina.replace(' ', '_').replace('/', '_').lower()
                    nombre_archivo = f"comunicacion_interna_{nombre_oficina}"
            except Exception as e:
                logger.warning(f"Error al obtener nombre de oficina para descargar PDF: {e}")
            
            # Agregar el radicado si existe
            if obj.radicado:
                nombre_archivo = f"{nombre_archivo}_{obj.radicado.replace('/', '_')}"
            
            response = HttpResponse(obj.archivo_generado, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
            return response
        else:
            messages.error(request, "El documento no ha sido generado.")
            return redirect('correspondencia:interna_detalle', pk=pk)



class ComunicacionInternaHistorialRespuestasView(LoginRequiredMixin, DetailView):
    """
    Vista para mostrar el historial de respuestas a una comunicación interna.
    
    Muestra todas las respuestas recibidas, destacando con estrellita las
    respuestas de líderes y usuarios asignados inicialmente.
    Las respuestas se ordenan: destacadas primero, luego por fecha.
    """
    model = ComunicacionInterna
    template_name = 'correspondencia/interna/historial_respuestas.html'
    context_object_name = 'comunicacion'

    def get_object(self, queryset=None):
        from django.http import Http404

        obj = super().get_object(queryset)
        if not _usuario_puede_ver_comunicacion_interna(self.request.user, obj):
            raise Http404("No tienes acceso a esta comunicación.")
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import ComunicacionInternaDestinatario, ComunicacionInternaDistribucion
        from collections import defaultdict
        
        comunicacion = self.object
        
        # Verificar que no es una respuesta (las respuestas no tienen respuestas)
        if comunicacion.es_respuesta():
            context['es_respuesta'] = True
            context['comunicacion_origen'] = comunicacion.comunicacion_origen
            context['respuestas'] = []
            context['estadisticas'] = {
                'total_respuestas': 0,
                'respuestas_destacadas': 0,
                'respuestas_normales': 0,
                'por_estado': {},
            }
            context['titulo_pagina'] = f"Historial de Respuestas: {comunicacion.radicado or 'BORRADOR'}"
            return context
        
        # Obtener todas las respuestas ordenadas: destacadas primero, luego por fecha
        respuestas = comunicacion.get_respuestas_ordenadas().select_related(
            'remitente_usuario',
            'remitente_oficina',
            'revisado_por',
        ).prefetch_related('historial')
        
        # Agrupar respuestas por estado
        respuestas_por_estado = defaultdict(list)
        respuestas_destacadas = 0
        respuestas_normales = 0
        
        respuestas_lista = []
        for resp in respuestas:
            estado_display = resp.get_estado_display()
            respuestas_por_estado[estado_display].append(resp)
            
            if resp.es_respuesta_destacada:
                respuestas_destacadas += 1
            else:
                respuestas_normales += 1
            
            # Información adicional para cada respuesta
            respuestas_lista.append({
                'obj': resp,
                'remitente_nombre': resp.remitente_nombre or (resp.remitente_usuario.get_full_name() if resp.remitente_usuario else 'Desconocido'),
                'oficina_nombre': resp.remitente_oficina.nombre if resp.remitente_oficina else 'Sin oficina',
                'es_destacada': resp.es_respuesta_destacada,
                'estado': resp.estado,
                'estado_display': estado_display,
                'fecha_creacion': resp.fecha_creacion,
                'radicado': resp.radicado,
            })
        
        # Estadísticas
        total_respuestas = len(respuestas_lista)
        
        context['respuestas'] = respuestas_lista
        context['es_respuesta'] = False
        context['estadisticas'] = {
            'total_respuestas': total_respuestas,
            'respuestas_destacadas': respuestas_destacadas,
            'respuestas_normales': respuestas_normales,
            'por_estado': dict(respuestas_por_estado),
        }
        context['titulo_pagina'] = f"Historial de Respuestas: {comunicacion.radicado or 'BORRADOR'}"
        
        return context

# ═══════════════════════════════════════════════════════════════
# DIRECTORIO DE USUARIOS
# ═══════════════════════════════════════════════════════════════

@login_required
def directorio_usuarios(request):
    """
    Directorio de usuarios del sistema agrupado por oficina.
    Cada oficina se presenta como un bloque y el detalle de usuarios
    se consulta en un modal.
    """
    q = request.GET.get('q', '').strip()

    usuarios_qs = User.objects.filter(
        is_active=True,
        perfil__isnull=False,
        perfil__oficina__isnull=False,
    ).select_related('perfil', 'perfil__oficina')

    if q:
        usuarios_qs = usuarios_qs.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(username__icontains=q)
            | Q(perfil__cargo__icontains=q)
            | Q(perfil__oficina__nombre__icontains=q)
        )

    usuarios_qs = usuarios_qs.order_by(
        'perfil__oficina__nombre',
        'first_name',
        'last_name',
        'username',
    )

    usuarios = list(usuarios_qs)
    ahora = timezone.now()
    oficinas_agrupadas = {}

    for usuario in usuarios:
        oficina = usuario.perfil.oficina
        bloque = oficinas_agrupadas.setdefault(
            oficina.pk,
            {
                'oficina': oficina,
                'modal_id': f'usuariosOficinaModal{oficina.pk}',
                'usuarios': [],
                'total': 0,
            }
        )

        nombre_completo = usuario.get_full_name().strip() or usuario.username
        partes_nombre = [parte[0].upper() for parte in nombre_completo.split()[:2] if parte]
        iniciales = ''.join(partes_nombre) or usuario.username[:2].upper()

        if usuario.last_login:
            delta = ahora - usuario.last_login
            if delta <= timedelta(days=2):
                actividad_clase = 'is-recent'
            elif delta <= timedelta(days=15):
                actividad_clase = 'is-moderate'
            else:
                actividad_clase = 'is-old'
        else:
            actividad_clase = 'is-never'

        bloque['usuarios'].append({
            'obj': usuario,
            'nombre_completo': nombre_completo,
            'iniciales': iniciales,
            'cargo': usuario.perfil.cargo or 'Cargo no registrado',
            'actividad_clase': actividad_clase,
        })
        bloque['total'] += 1

    oficinas_lista = list(oficinas_agrupadas.values())
    paginator = Paginator(oficinas_lista, 25)
    oficinas_page = paginator.get_page(request.GET.get('page'))

    mi_oficina = getattr(getattr(request.user, 'perfil', None), 'oficina', None)
    total_mi_oficina = 0
    if mi_oficina:
        total_mi_oficina = User.objects.filter(
            is_active=True, perfil__oficina=mi_oficina
        ).count()

    params = request.GET.copy()
    params.pop('page', None)
    querystring = params.urlencode()

    context = {
        'oficinas_page': oficinas_page,
        'filtro_q': q,
        'total_usuarios': len(usuarios),
        'oficinas_visibles': len(oficinas_lista),
        'mi_oficina': mi_oficina,
        'total_mi_oficina': total_mi_oficina,
        'querystring': querystring,
    }
    return render(request, 'correspondencia/usuario/directorio_usuarios.html', context)


# ─── Chat de Soporte ──────────────────────────────────────────────────────────

@login_required
def chat_soporte_view(request):
    """Página de chat de soporte para usuarios."""
    from correspondencia.utils.safe_redirect import safe_back_url

    return render(request, 'correspondencia/chat_soporte.html', {
        'back_url': safe_back_url(request, fallback='/registros/correspondencia/'),
    })
