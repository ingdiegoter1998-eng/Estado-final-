# Importaciones estándar de Python
from datetime import date, datetime  # Manejo de fechas y horas
import unicodedata

# Importaciones de Django
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden
from guardian.shortcuts import get_perms
from .models import RegistroDeArchivo, SubserieDocumental
from .forms import RegistroDeArchivoForm
from django.utils.decorators import method_decorator
from django.contrib import messages  # Envío de mensajes al contexto (ejemplo: mensajes de éxito o error)
from django.contrib.auth.decorators import login_required  # Decorador para restringir acceso a usuarios autenticados
from django.contrib.auth.mixins import LoginRequiredMixin  # Mixin para vistas basadas en clases que requieren autenticación
from django.contrib.auth.models import User  # Modelo de usuarios de Django
from django.core.paginator import Paginator  # Paginación de listas de objetos
from django.db import IntegrityError  # Manejo de errores de integridad en la base de datos
from django.db.models import Q, Count, Avg, Max  # Operadores para consultas avanzadas a la base de datos
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse  # Respuestas HTTP y JSON
from django.shortcuts import render, redirect, get_object_or_404  # Métodos para renderizar vistas y manejar redirecciones
from django.views.decorators.http import require_POST  # Decorador para requerir método POST
from django.urls import reverse_lazy
from django.utils.timezone import now, timedelta  # Fechas y tiempos con soporte de zona horaria
from django.views.generic.edit import CreateView, UpdateView  # Vistas genéricas para creación y edición de objetos
from django.http import HttpResponseRedirect

# Librerías de terceros

import openpyxl  # Librería para trabajar con archivos Excel
from openpyxl.utils import get_column_letter  # Utilidad para obtener letras de columnas en Excel
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font  # Estilos y formato para celdas en Excel
from openpyxl.drawing.image import Image  # Insertar imágenes en hojas de cálculo Excel

# Framework Django Rest Framework
from rest_framework.response import Response  # Respuestas de APIs
from rest_framework.views import APIView  # Clase base para construir APIs

# Importaciones específicas del proyecto
from .forms import (
    RegistroDeArchivoForm,
    FUIDForm,
    FichaPacienteForm,
    SolicitudPrestamoForm,
    GestionPrestamoForm,
    ConfirmacionPrestamoForm,
    RegistroUsuarioForm,
)  # Formularios personalizados
from .models import (  # Modelos de la base de datos
    RegistroDeArchivo,
    SubserieDocumental,
    SerieDocumental,
    FUID,
    FichaPaciente,
    Documento,
    PrestamoDocumental,
    HistorialPrestamo,
    HistorialDescargaPrestamo,
    DocumentoEscaneadoPrestamo,
    OficinaProductora,
    NotificacionAvisoPrestamo
)
from correspondencia.trd_interna import (
    queryset_series_comunicacion_interna,
    queryset_subseries_comunicacion_interna,
)


@login_required
def cargar_series(request):
    """Cargar series asociadas a la oficina del usuario logueado.
    Para superusers devuelve todas las series.
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Usuario no autenticado'}, status=401)

    try:
        if request.GET.get('scope') == 'comunicacion_interna':
            series = queryset_series_comunicacion_interna().values(
                'id', 'codigo', 'nombre', 'codigo_trd'
            )
            return JsonResponse(list(series), safe=False)

        # Superusuarios ven todas las series
        if request.user.is_superuser:
            series = SerieDocumental.objects.all().values(
                'id', 'codigo', 'nombre', 'codigo_trd'
            ).distinct()
            return JsonResponse(list(series), safe=False)

        perfil = request.user.perfil
        if not perfil.oficina:
            return JsonResponse({'error': 'Usuario sin oficina asignada'}, status=400)

        # Filtrar series por la oficina del usuario
        series = SerieDocumental.objects.filter(
            oficinas_productoras=perfil.oficina
        ).values('id', 'codigo', 'nombre', 'codigo_trd').distinct()
        return JsonResponse(list(series), safe=False)
    except Exception as e:
        return JsonResponse({'error': f'Error al cargar series: {str(e)}'}, status=500)
@login_required
@login_required
def cargar_subseries(request):
    """Vista AJAX para cargar subseries según la serie seleccionada"""
    serie_id = request.GET.get('serie_id')
    scope = request.GET.get('scope')
    
    if not serie_id:
        return JsonResponse({'error': 'serie_id es requerido'}, status=400)
    
    try:
        serie_id = int(serie_id)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'serie_id debe ser un número'}, status=400)
    
    if scope == 'comunicacion_interna':
        subseries = queryset_subseries_comunicacion_interna().filter(serie_id=serie_id).values('id', 'nombre', 'codigo_trd')
    else:
        subseries = SubserieDocumental.objects.filter(serie_id=serie_id).values('id', 'nombre', 'codigo_trd')
    return JsonResponse(list(subseries), safe=False)

from guardian.shortcuts import assign_perm  # <-- Importamos assign_perm

@login_required
def lista_registros(request):
    # Solo superusuarios pueden acceder
    if not request.user.is_superuser:
        from django.http import Http404
        raise Http404("Página no encontrada")
    
    # El superusuario ve todo
    registros = RegistroDeArchivo.objects.all()
    
    # Filtros
    GET = request.GET
    if GET.get('numero_orden'):
        registros = registros.filter(numero_orden__icontains=GET.get('numero_orden').strip())
    if GET.get('codigo'):
        registros = registros.filter(codigo__icontains=GET.get('codigo').strip())
    if GET.get('codigo_serie'):
        registros = registros.filter(codigo_serie__nombre__icontains=GET.get('codigo_serie').strip())
    if GET.get('codigo_subserie'):
        registros = registros.filter(codigo_subserie__nombre__icontains=GET.get('codigo_subserie').strip())
    if GET.get('unidad_documental'):
        registros = registros.filter(unidad_documental__icontains=GET.get('unidad_documental').strip())
    if GET.get('fecha_archivo'):
        registros = registros.filter(fecha_archivo__icontains=GET.get('fecha_archivo').strip())
    if GET.get('documento'):
        registros = registros.filter(documentos__isnull=False)
    if GET.get('soporte_fisico'):
        soporte_fisico = GET.get('soporte_fisico').strip().lower()
        if soporte_fisico in ['true', '1', 'si', 'yes']:
            registros = registros.filter(soporte_fisico=True)
        elif soporte_fisico in ['false', '0', 'no']:
            registros = registros.filter(soporte_fisico=False)
    if GET.get('soporte_electronico'):
        soporte_electronico = GET.get('soporte_electronico').strip().lower()
        if soporte_electronico in ['true', '1', 'si', 'yes']:
            registros = registros.filter(soporte_electronico=True)
        elif soporte_electronico in ['false', '0', 'no']:
            registros = registros.filter(soporte_electronico=False)
    if GET.get('creado_por'):
        registros = registros.filter(creado_por__username__icontains=GET.get('creado_por').strip())
    
    # Paginación
    page = request.GET.get('page', 1)
    paginator = Paginator(registros, 50)  # 50 registros por página
    
    try:
        registros_paginados = paginator.page(page)
    except:
        registros_paginados = paginator.page(1)
    
    context = {
        'registros': registros_paginados,
        'total_registros': registros.count(),
        'filtros_activos': {k: v for k, v in GET.items() if v and k != 'page'}
    }
    
    return render(request, 'registro_list.html', context)



@login_required
def crear_registro(request):
    if not request.user.has_perm('documentos.add_registrodearchivo'):
        return HttpResponseForbidden("No tienes permiso para crear registros.")

    if request.method == 'POST':
        # Incluimos request.FILES para manejar archivos
        form = RegistroDeArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.creado_por = request.user  # Asigna el usuario autenticado
            registro.save()

            # Crear Documento si se subió un archivo
            archivo_subido = form.cleaned_data.get('archivo')
            if archivo_subido:
                Documento.objects.create(
                    registro=registro,
                    archivo=archivo_subido
                )

            # Asignar permisos a nivel de objeto
            assign_perm('documentos.view_own_registro', request.user, registro)
            assign_perm('documentos.edit_own_registro', request.user, registro)
            # assign_perm('documentos.delete_own_registro', request.user, registro)

            # Mensaje de éxito
            messages.success(request, 'Registro de archivo creado exitosamente.')

            # Limpiamos el formulario
            form = RegistroDeArchivoForm()
        else:
            for field, errors in form.errors.items():
                field_name = form.fields[field].label
                for error in errors:
                    messages.error(request, f"{field_name}: {error}")

    else:
        form = RegistroDeArchivoForm()
        # Subseries vacío por defecto (si no se selecciona serie)
        form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.none()

    return render(request, 'registro_form.html', {'form': form})





@login_required
def editar_registro(request, pk):
    registro = get_object_or_404(RegistroDeArchivo, id=pk)

    # Verifica si el usuario tiene permiso de edición a nivel de objeto
    if not request.user.is_superuser and 'edit_own_registro' not in get_perms(request.user, registro):
        return HttpResponseForbidden("No tienes permiso para editar este registro.")

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST, instance=registro)
        # Lógica de subseries
        codigo_serie = request.POST.get('codigo_serie')
        if codigo_serie:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.filter(serie_id=codigo_serie)

        if form.is_valid():
            reg = form.save(commit=False)
            reg.editado_por = request.user
            reg.save()
            return redirect('lista_registros')
    else:
        form = RegistroDeArchivoForm(instance=registro)
        if registro.codigo_serie:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.filter(serie=registro.codigo_serie)
        else:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.none()

    return render(request, 'registro_form.html', {'form': form})




from guardian.utils import get_anonymous_user
from guardian.shortcuts import get_perms


from guardian.shortcuts import get_perms

@login_required
def eliminar_registro(request, pk):
    registro = get_object_or_404(RegistroDeArchivo, pk=pk)

    # Si superuser, ok
    if request.user.is_superuser:
        registro.delete()
        return redirect('lista_registros')

    # Verifica si el user tiene permiso de delete a nivel de objeto
    perms = get_perms(request.user, registro)
    if 'delete_own_registro' in perms:
        registro.delete()
        return redirect('lista_registros')
    else:
        return HttpResponseForbidden("No tienes permiso para eliminar este registro.")


@login_required
def lista_completa_registros(request):
    # Solo superusuarios pueden acceder
    if not request.user.is_superuser:
        from django.http import Http404
        raise Http404("Página no encontrada")

    # El superusuario ve todo
    registros = RegistroDeArchivo.objects.all()
    
    # Filtros
    GET = request.GET
    if GET.get('numero_orden'):
        registros = registros.filter(numero_orden__icontains=GET.get('numero_orden').strip())
    if GET.get('codigo'):
        registros = registros.filter(codigo__icontains=GET.get('codigo').strip())
    if GET.get('codigo_serie'):
        registros = registros.filter(codigo_serie__nombre__icontains=GET.get('codigo_serie').strip())
    if GET.get('codigo_subserie'):
        registros = registros.filter(codigo_subserie__nombre__icontains=GET.get('codigo_subserie').strip())
    if GET.get('unidad_documental'):
        registros = registros.filter(unidad_documental__icontains=GET.get('unidad_documental').strip())
    if GET.get('fecha_archivo'):
        registros = registros.filter(fecha_archivo__icontains=GET.get('fecha_archivo').strip())
    if GET.get('fecha_inicial'):
        registros = registros.filter(fecha_inicial__icontains=GET.get('fecha_inicial').strip())
    if GET.get('fecha_final'):
        registros = registros.filter(fecha_final__icontains=GET.get('fecha_final').strip())
    if GET.get('soporte_fisico'):
        soporte_fisico = GET.get('soporte_fisico').strip().lower()
        if soporte_fisico in ['true', '1', 'si', 'yes']:
            registros = registros.filter(soporte_fisico=True)
        elif soporte_fisico in ['false', '0', 'no']:
            registros = registros.filter(soporte_fisico=False)
    if GET.get('soporte_electronico'):
        soporte_electronico = GET.get('soporte_electronico').strip().lower()
        if soporte_electronico in ['true', '1', 'si', 'yes']:
            registros = registros.filter(soporte_electronico=True)
        elif soporte_electronico in ['false', '0', 'no']:
            registros = registros.filter(soporte_electronico=False)
    if GET.get('caja'):
        registros = registros.filter(caja__icontains=GET.get('caja').strip())
    if GET.get('carpeta'):
        registros = registros.filter(carpeta__icontains=GET.get('carpeta').strip())
    if GET.get('tomo_legajo_libro'):
        registros = registros.filter(tomo_legajo_libro__icontains=GET.get('tomo_legajo_libro').strip())
    if GET.get('numero_folios'):
        registros = registros.filter(numero_folios__icontains=GET.get('numero_folios').strip())
    if GET.get('tipo'):
        registros = registros.filter(tipo__icontains=GET.get('tipo').strip())
    if GET.get('cantidad'):
        registros = registros.filter(cantidad__icontains=GET.get('cantidad').strip())
    if GET.get('ubicacion'):
        registros = registros.filter(ubicacion__icontains=GET.get('ubicacion').strip())
    if GET.get('cantidad_documentos_electronicos'):
        registros = registros.filter(cantidad_documentos_electronicos__icontains=GET.get('cantidad_documentos_electronicos').strip())
    if GET.get('tamano_documentos_electronicos'):
        registros = registros.filter(tamano_documentos_electronicos__icontains=GET.get('tamano_documentos_electronicos').strip())
    if GET.get('notas'):
        registros = registros.filter(notas__icontains=GET.get('notas').strip())
    if GET.get('creado_por'):
        registros = registros.filter(creado_por__username__icontains=GET.get('creado_por').strip())
    if GET.get('fecha_creacion'):
        registros = registros.filter(fecha_creacion__icontains=GET.get('fecha_creacion').strip())
    if GET.get('Estado_archivo'):
        estado = GET.get('Estado_archivo').strip().lower()
        if estado in ['true', '1', 'si', 'yes']:
            registros = registros.filter(Estado_archivo=True)
        elif estado in ['false', '0', 'no']:
            registros = registros.filter(Estado_archivo=False)
    
    # Paginación
    page = request.GET.get('page', 1)
    paginator = Paginator(registros, 50)  # 50 registros por página
    
    try:
        registros_paginados = paginator.page(page)
    except:
        registros_paginados = paginator.page(1)
    
    context = {
        'registros': registros_paginados,
        'total_registros': registros.count(),
        'filtros_activos': {k: v for k, v in GET.items() if v and k != 'page'}
    }
    
    return render(request, 'registro_completo.html', context)



@login_required
def registros_api(request):
    # 1) Filtra por oficina
    if request.user.is_superuser:
        registros = RegistroDeArchivo.objects.all()
    else:
        oficina_user = request.user.perfil.oficina
        usuarios_de_mi_oficina = User.objects.filter(perfil__oficina=oficina_user)
        registros = RegistroDeArchivo.objects.filter(creado_por__in=usuarios_de_mi_oficina)

    # 2) Aplicas la búsqueda por columnas que DataTables envía
    draw = request.GET.get("draw", 1)
    start = int(request.GET.get("start", 0))
    length = int(request.GET.get("length", 10))

    i = 0
    while True:
        col_data = request.GET.get(f'columns[{i}][data]')
        if col_data is None:
            break
        col_search_value = request.GET.get(f'columns[{i}][search][value]', '').strip()

        if col_search_value:
            if col_data == 'numero_orden':
                registros = registros.filter(numero_orden__icontains=col_search_value)
            elif col_data == 'codigo':
                registros = registros.filter(codigo__icontains=col_search_value)
            elif col_data == 'codigo_serie':
                registros = registros.filter(codigo_serie__nombre__icontains=col_search_value)
            # ... repites el resto de tus filtros ...
            elif col_data == 'creado_por':
                registros = registros.filter(creado_por__username__icontains=col_search_value)
        i += 1

    # 3) Total sin filtros (para recordsTotal) — generalmente también queremos
    #    "total sin filtros" solo en la oficina, pero si prefieres
    #    contar todo, deja la línea original. Ojo: usualmente DataTables
    #    quiere "recordsTotal" = total *después* de filtrar por oficina*,
    #    y "recordsFiltered" = total *después* de filtrar además por búsqueda*
    total_registros = registros.count()

    # 4) Paginación
    paginator = Paginator(registros, length)
    page_number = start // length + 1
    page = paginator.get_page(page_number)

    # 5) Construir data
    data = []
    for registro in page:
        data.append({
            "numero_orden": registro.numero_orden,
            "codigo": registro.codigo,
            "codigo_serie": registro.codigo_serie.nombre if registro.codigo_serie else "",
            "codigo_subserie": registro.codigo_subserie.nombre if registro.codigo_subserie else "",
            "unidad_documental": registro.unidad_documental,
            "fecha_archivo": registro.fecha_archivo,
            "documento": [{"url": doc.archivo.url} for doc in registro.documentos.all()],
            "soporte_fisico": registro.soporte_fisico,
            "soporte_electronico": registro.soporte_electronico,
            "creado_por": registro.creado_por.username if registro.creado_por else "N/A",
            "id": registro.id,
        })


    response = {
        "draw": int(draw),
        "recordsTotal": total_registros,
        "recordsFiltered": total_registros,  # Podrías usar `registros.count()` si la semántica lo requiere
        "data": data,
    }
    return JsonResponse(response)


from django.shortcuts import get_object_or_404
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.conf import settings
# from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions
from datetime import datetime, timedelta

@login_required
def registros_api_completo(request):
    # Filtrar registros según permisos
    if request.user.is_superuser:
        registros = RegistroDeArchivo.objects.all()
    else:
        oficina_user = request.user.perfil.oficina
        usuarios_de_mi_oficina = User.objects.filter(perfil__oficina=oficina_user)
        registros = RegistroDeArchivo.objects.filter(creado_por__in=usuarios_de_mi_oficina)

    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))

    # Aquí se pueden aplicar filtros de búsqueda si es necesario
    i = 0
    while True:
        col_data = request.GET.get(f'columns[{i}][data]')
        if col_data is None:
            break
        search_value = request.GET.get(f'columns[{i}][search][value]', '').strip()
        if search_value:
            if col_data == 'numero_orden':
                registros = registros.filter(numero_orden__icontains=search_value)
            elif col_data == 'codigo':
                registros = registros.filter(codigo__icontains=search_value)
            # Agrega otros filtros según necesites
        i += 1

    paginator = Paginator(registros, length)
    page_number = start // length + 1
    page = paginator.get_page(page_number)

    # En almacenamiento local, cada documento se accede vía .url
    data = []
    for registro in page:
        documentos_urls = []
        for documento in registro.documentos.all():
            documentos_urls.append({
                "nombre": documento.archivo.name.split('/')[-1],
                "url": documento.archivo.url
            })
        data.append({
            "numero_orden": registro.numero_orden,
            "codigo": registro.codigo,
            "codigo_serie": registro.codigo_serie.nombre if registro.codigo_serie else "",
            "codigo_subserie": registro.codigo_subserie.nombre if registro.codigo_subserie else "",
            "unidad_documental": registro.unidad_documental,
            "fecha_archivo": registro.fecha_archivo,
            "soporte_fisico": registro.soporte_fisico,
            "soporte_electronico": registro.soporte_electronico,
            "creado_por": registro.creado_por.username if registro.creado_por else "",
            "documentos": documentos_urls,
            "id": registro.id,
        })

    response = {
        "draw": draw,
        "recordsTotal": RegistroDeArchivo.objects.count(),
        "recordsFiltered": registros.count(),
        "data": data,
    }
    return JsonResponse(response)

from django.shortcuts import render, get_object_or_404

# from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions
from datetime import datetime, timedelta
import os




from django.shortcuts import get_object_or_404
from django.http import JsonResponse, HttpResponseForbidden
from .models import Documento
# from .utils import generar_url_sas

from django.shortcuts import get_object_or_404
from django.http import JsonResponse, HttpResponseForbidden
from .models import Documento
# from .utils import generar_url_sas
# views.py (o donde tengas el import)
from django.urls import reverse_lazy



from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from .models import RegistroDeArchivo
# from .utils import generar_url_sas

# documentos/views.py
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from .models import RegistroDeArchivo
# from .utils import generar_url_sas

from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from .models import RegistroDeArchivo
# from .utils import generar_url_sas
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required

from django.shortcuts import get_object_or_404, render
from django.http import HttpResponseForbidden
from .models import RegistroDeArchivo, Documento

def ver_documento(request, registro_id):
    registro = get_object_or_404(RegistroDeArchivo, id=registro_id)
    
    # 1. Recuperar la oficina del usuario
    oficina_usuario = request.user.perfil.oficina
    
    # 2. Recuperar la oficina del registro (desde el primer FUID asociado)
    fuid = registro.fuids.first()
    if not fuid:
        return HttpResponseForbidden("Este registro no tiene FUID asignado. No se puede verificar oficina.")
    oficina_registro = fuid.oficina_productora
    
    # 3. Comparar oficinas
    if oficina_registro != oficina_usuario and not request.user.is_superuser:
        return HttpResponseForbidden("No tienes permiso para ver este documento.")
    
    # Si pasa la verificación, procedemos
    documentos = registro.documentos.all()
    for doc in documentos:
        doc.sas_url = doc.archivo.url  # URL local
    
    return render(request, "documento_detalle.html", {
        "registro": registro,
        "documentos": documentos,
        "fuid": fuid,
    })


####    
@login_required
def registros_api_con_id(request):
    registros = RegistroDeArchivo.objects.all()

    if not request.user.is_superuser:
        oficina_user = request.user.perfil.oficina
        usuarios_de_mi_oficina = User.objects.filter(perfil__oficina=oficina_user)
        registros = registros.filter(creado_por__in=usuarios_de_mi_oficina)

    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))

    # Filtro por columnas
    i = 0
    while True:
        col_data = request.GET.get(f'columns[{i}][data]')
        if col_data is None:
            break
        search_value = request.GET.get(f'columns[{i}][search][value]', '').strip()

        if search_value:
            if col_data == 'numero_orden':
                registros = registros.filter(numero_orden__icontains=search_value)
            elif col_data == 'codigo':
                registros = registros.filter(codigo__icontains=search_value)
            elif col_data == 'codigo_serie':
                registros = registros.filter(codigo_serie__nombre__icontains=search_value)
            elif col_data == 'codigo_subserie':
                registros = registros.filter(codigo_subserie__nombre__icontains=search_value)
            elif col_data == 'unidad_documental':
                registros = registros.filter(unidad_documental__icontains=search_value)
            elif col_data == 'fecha_archivo':
                registros = registros.filter(fecha_archivo__icontains=search_value)
            elif col_data == 'fecha_inicial':
                registros = registros.filter(fecha_inicial__icontains=search_value)
            elif col_data == 'fecha_final':
                registros = registros.filter(fecha_final__icontains=search_value)
            elif col_data == 'soporte_fisico':
                registros = registros.filter(soporte_fisico=search_value.lower() in ['true', '1', '✔'])
            elif col_data == 'soporte_electronico':
                registros = registros.filter(soporte_electronico=search_value.lower() in ['true', '1', '✔'])
            elif col_data == 'caja':
                registros = registros.filter(caja__icontains=search_value)
            elif col_data == 'carpeta':
                registros = registros.filter(carpeta__icontains=search_value)
            elif col_data == 'ubicacion':
                registros = registros.filter(ubicacion__icontains=search_value)
            elif col_data == 'Estado_archivo':  # 🔹 Asegurar que coincida con el modelo
                registros = registros.filter(Estado_archivo=search_value.lower() in ['true', '1', '✔'])

        i += 1

    # Paginación
    paginator = Paginator(registros, length)
    page_number = start // length + 1
    page = paginator.get_page(page_number)

    # Construcción de data con la columna estado_archivo
    data = []
    for registro in page:
        data.append({
            "id": registro.id,
            "numero_orden": registro.numero_orden,
            "codigo": registro.codigo,
            "codigo_serie": registro.codigo_serie.nombre if registro.codigo_serie else "",
            "codigo_subserie": registro.codigo_subserie.nombre if registro.codigo_subserie else "",
            "unidad_documental": registro.unidad_documental,
            "fecha_archivo": registro.fecha_archivo,
            "fecha_inicial": registro.fecha_inicial,
            "fecha_final": registro.fecha_final,
            "soporte_fisico": registro.soporte_fisico,
            "soporte_electronico": registro.soporte_electronico,
            "caja": registro.caja,
            "carpeta": registro.carpeta,
            "tomo_legajo_libro": registro.tomo_legajo_libro,
            "numero_folios": registro.numero_folios,
            "tipo": registro.tipo,
            "cantidad": registro.cantidad,
            "ubicacion": registro.ubicacion,
            "cantidad_documentos_electronicos": registro.cantidad_documentos_electronicos,
            "tamano_documentos_electronicos": registro.tamano_documentos_electronicos,
            "notas": registro.notas,
            "creado_por": registro.creado_por.username if registro.creado_por else "",
            "fecha_creacion": registro.fecha_creacion,
            "Estado_archivo": "✔" if registro.Estado_archivo else "✖",
        })

    response = {
        "draw": draw,
        "recordsTotal": registros.count(),
        "recordsFiltered": registros.count(),
        "data": data,
    }
    return JsonResponse(response)










# Vista para crear un FUID

# vistas.py (parte)

from django.template.loader import render_to_string
from django.http import JsonResponse

@login_required
def form_registro_fuid_ajax(request, fuid_id):
    """
    Retorna el HTML parcial de un formulario para crear un RegistroDeArchivo
    asociado a un FUID, listo para inyectar en un modal.
    """
    fuid = get_object_or_404(FUID, pk=fuid_id)
    form = RegistroDeArchivoForm()  # Form vacío

    # Renderizamos un template parcial con el formulario
    html_form = render_to_string(
        'partials/_form_registro.html',
        {'form': form, 'fuid': fuid},
        request=request
    )
    return JsonResponse({'html_form': html_form})

@login_required
def crear_registro_fuid_ajax(request, fuid_id):
    fuid = get_object_or_404(FUID, pk=fuid_id)

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.creado_por = request.user
            registro.save()

            # Asignar permisos a nivel de objeto (si usas Guardian)
            assign_perm('documentos.view_own_registro', request.user, registro)
            assign_perm('documentos.edit_own_registro', request.user, registro)
            # assign_perm('documentos.delete_own_registro', request.user, registro)

            # Asociar con FUID
            fuid.registros.add(registro)

            # Devolvemos éxito y el registro en JSON, para actualizar la tabla sin recargar
            return JsonResponse({
                'ok': True,
                'message': 'Registro creado exitosamente.',
                'registro': {
                    'id': registro.id,
                    'numero_orden': registro.numero_orden,
                    'codigo': registro.codigo or '',
                    # ... agrega las demás propiedades que necesites en el JS
                }
            })
        else:
            # Devolvemos el HTML del form con errores
            html_form = render_to_string(
                'partials/_form_registro.html',
                {'form': form, 'fuid': fuid},
                request=request
            )
            return JsonResponse({'ok': False, 'html_form': html_form})

    # Si no es POST, retornamos un 405 o algo similar
    return JsonResponse({'ok': False, 'message': 'Método no permitido'}, status=405)


from django.http import HttpResponseForbidden
from guardian.shortcuts import assign_perm
class FUIDCreateView(LoginRequiredMixin, CreateView):
    model = FUID
    form_class = FUIDForm  # sin registros
    template_name = "fuid_form.html"
    success_url = reverse_lazy("lista_fuids")

    def dispatch(self, request, *args, **kwargs):
        if not request.user.has_perm('documentos.add_fuid'):
            return HttpResponseForbidden("No tienes permiso para crear FUIDs.")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        fuid = form.save()
        # permisos con django-guardian...
        for perm in ('view_own_fuid','edit_own_fuid','delete_own_fuid'):
            assign_perm(f'documentos.{perm}', self.request.user, fuid)
        return super().form_valid(form)


class FUIDUpdateView(LoginRequiredMixin, UpdateView):
    model = FUID
    form_class = FUIDForm  # sigue incluyendo registros en el Form, pero lo quitamos abajo
    template_name = "fuid_form.html"
    success_url = reverse_lazy("lista_fuids")

    def dispatch(self, request, *args, **kwargs):
        fuid = self.get_object()
        if not (request.user.is_superuser or fuid.creado_por == request.user):
            return HttpResponseForbidden("Solo el creador de este FUID puede editarlo.")
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)
        # elimina cualquier referencia a registros
        form.fields.pop('registros', None)
        return form

    def form_valid(self, form):
        # guardamos SOLO los campos del formulario, sin tocar la M2M
        fuid = form.save(commit=False)
        fuid.save()
        return HttpResponseRedirect(self.get_success_url())

@login_required
def lista_fuids(request):
    if request.user.is_superuser:
        fuids = FUID.objects.all()
    else:
        fuids = FUID.objects.filter(oficina_productora=request.user.perfil.oficina)
    return render(request, 'fuid_list.html', {'fuids': fuids})


@login_required
def detalle_fuid(request, pk):

    """Vista principal que carga el detalle del FUID sin los registros"""
    fuid = get_object_or_404(FUID, pk=pk)

    # Verificar permisos
    if not request.user.is_superuser and fuid.oficina_productora != request.user.perfil.oficina:
        return HttpResponseForbidden("No tienes permiso para ver este FUID.")

    if not request.user.has_perm('documentos.view_own_fuid', fuid):
        assign_perm('documentos.view_own_fuid', request.user, fuid)

    # Renderizar la plantilla sin cargar los registros todavía
    return render(request, 'fuid_complete_list.html', {'fuid': fuid})



@login_required
def registros_fuid_json(request, fuid_id):
    from django.db.models import Q
    from django.core.paginator import Paginator
    from django.urls import reverse
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from documentos.models import FUID

    fuid = get_object_or_404(FUID, pk=fuid_id)

    # Verificar permisos: mismo criterio que detalle_fuid (evitar IDOR)
    if not request.user.is_superuser and fuid.oficina_productora != request.user.perfil.oficina:
        return JsonResponse({'error': 'No tienes permiso para ver los registros de este FUID.'}, status=403)

    # DataTables puede enviar parámetros por GET o POST.
    params = request.POST if request.method == 'POST' else request.GET

    start = int(params.get('start', 0))
    length = int(params.get('length', 30))

    registros_query = fuid.registros.all()

    # ==================== Lectura de filtros (Añadir Serie/Subserie) ====================
    numero_orden = params.get("filtro_numero_orden", "").strip()
    codigo = params.get("filtro_codigo", "").strip()
    # Añadir lectura de filtros para serie y subserie
    serie_nombre = params.get("filtro_serie", "").strip()
    subserie_nombre = params.get("filtro_subserie", "").strip()
    unidad_documental = params.get("filtro_unidad_documental", "").strip()
    identificador_documento = params.get("filtro_identificador_documento", "").strip()
    fecha_archivo = params.get("filtro_fecha_archivo", "").strip()
    fecha_inicial = params.get("filtro_fecha_inicial", "").strip()
    fecha_final = params.get("filtro_fecha_final", "").strip()
    soporte_fisico = params.get("filtro_soporte_fisico", "").strip()
    soporte_electronico = params.get("filtro_soporte_electronico", "").strip()
    caja = params.get("filtro_caja", "").strip()
    carpeta = params.get("filtro_carpeta", "").strip()
    tomo_legajo = params.get("filtro_tomo_legajo_libro", "").strip()
    numero_folios = params.get("filtro_numero_folios", "").strip()
    tipo = params.get("filtro_tipo", "").strip()
    cantidad = params.get("filtro_cantidad", "").strip()
    ubicacion = params.get("filtro_ubicacion", "").strip()
    cant_elec = params.get("filtro_cant_elec", "").strip()
    tamano_elec = params.get("filtro_tamano_elec", "").strip()
    notas = params.get("filtro_notas", "").strip()
    creado_por = params.get("filtro_creado_por", "").strip()
    fecha_creacion = params.get("filtro_fecha_creacion", "").strip()
    estado_archivo = params.get("filtro_estado_archivo", "").strip()

    # ==================== Filtros Q() (Añadir Serie/Subserie) ====================
    q_filter = Q()
    
    if numero_orden:
        q_filter &= Q(numero_orden__icontains=numero_orden) 
            
    if codigo:
        q_filter &= Q(codigo__icontains=codigo)
        
    # Añadir lógica de filtrado para serie y subserie
    if serie_nombre:
        q_filter &= Q(codigo_serie__nombre__icontains=serie_nombre)
        
    if subserie_nombre:
        q_filter &= Q(codigo_subserie__nombre__icontains=subserie_nombre)

    if unidad_documental:
        q_filter &= Q(unidad_documental__icontains=unidad_documental)
        
    if identificador_documento:
        q_filter &= Q(identificador_documento__icontains=identificador_documento)
        
    if fecha_archivo:
        q_filter &= Q(fecha_archivo=fecha_archivo)
        
    if fecha_inicial:
        q_filter &= Q(fecha_inicial=fecha_inicial)
        
    if fecha_final:
        q_filter &= Q(fecha_final=fecha_final)
        
    if soporte_fisico:
        q_filter &= Q(soporte_fisico=(soporte_fisico == "✔"))
        
    if soporte_electronico:
        q_filter &= Q(soporte_electronico=(soporte_electronico == "✔"))
        
    # Mantener exacto para caja y carpeta
    if caja:
        try:
            q_filter &= Q(caja=int(caja))
        except ValueError:
            pass # Ignorar si no es un número válido para caja
            
    if carpeta:
        try:
            q_filter &= Q(carpeta=int(carpeta))
        except ValueError:
            pass # Ignorar si no es un número válido para carpeta
            
    if tomo_legajo:
        q_filter &= Q(tomo_legajo_libro__icontains=tomo_legajo)
        
    if numero_folios:
        q_filter &= Q(numero_folios__icontains=numero_folios) # Volver a icontains
            
    if tipo:
        q_filter &= Q(tipo__icontains=tipo)
        
    if cantidad:
        q_filter &= Q(cantidad__icontains=cantidad) # Volver a icontains
            
    if ubicacion:
        q_filter &= Q(ubicacion__icontains=ubicacion)
        
    if cant_elec:
        q_filter &= Q(cantidad_documentos_electronicos__icontains=cant_elec) # Volver a icontains
            
    if tamano_elec:
        q_filter &= Q(tamano_documentos_electronicos__icontains=tamano_elec)
        
    if notas:
        q_filter &= Q(notas__icontains=notas)
        
    if creado_por:
        q_filter &= Q(creado_por__username__icontains=creado_por)
        
    if fecha_creacion:
        q_filter &= Q(fecha_creacion__startswith=fecha_creacion)
        
    if estado_archivo:
        q_filter &= Q(Estado_archivo=(estado_archivo == "✔"))

    registros_query = registros_query.filter(q_filter)

    # ==================== ORDENAMIENTO SERVER-SIDE (NUEVO) ====================
    # DataTables envía order[0][column], order[0][dir], columns[x][data], etc.
    order_col_index = params.get('order[0][column]', None)
    order_dir = params.get('order[0][dir]', 'asc')
    column_name = None

    if order_col_index is not None:
        column_name = params.get(f'columns[{order_col_index}][data]', None)

    # Mapeamos las columnas "visibles" en DataTables a campos en la BD:
    FIELD_MAP = {
        "numero_orden": "numero_orden",
        "codigo": "codigo",
        "nombre_serie": "codigo_serie__nombre",
        "nombre_subserie": "codigo_subserie__nombre",
        "unidad_documental": "unidad_documental",
        "identificador_documento": "identificador_documento",
        "fecha_archivo": "fecha_archivo",
        "fecha_inicial": "fecha_inicial",
        "fecha_final": "fecha_final",
        "soporte_fisico": "soporte_fisico",
        "soporte_electronico": "soporte_electronico",
        "caja": "caja",
        "carpeta": "carpeta",
        "tomo_legajo_libro": "tomo_legajo_libro",
        "numero_folios": "numero_folios",
        "tipo": "tipo",
        "cantidad": "cantidad",
        "ubicacion": "ubicacion",
        "cantidad_documentos_electronicos": "cantidad_documentos_electronicos",
        "tamano_documentos_electronicos": "tamano_documentos_electronicos",
        "notas": "notas",
        "creado_por": "creado_por__username",
        "fecha_creacion": "fecha_creacion",
        "fecha_edicion": "fecha_edicion",
        "editado_por": "editado_por__username",
        "Estado_archivo": "Estado_archivo",
        # "documento" y "acciones" NO son campos reales, no se ordenan
    }

    mapped_field = FIELD_MAP.get(column_name, None)
    if mapped_field:
        if order_dir == 'desc':
            registros_query = registros_query.order_by('-' + mapped_field)
        else:
            registros_query = registros_query.order_by(mapped_field)
    else:
        # Orden por defecto (antes era 'id')
        registros_query = registros_query.order_by('id')

    # ==================== PAGINACIÓN ====================
    total_registros = registros_query.count()
    paginator = Paginator(registros_query, length)
    page_obj = paginator.get_page(start // length + 1)

    # ==================== CONSTRUCCIÓN DEL JSON ====================
    data = []
    for r in page_obj:
        data.append({
            "numero_orden": r.numero_orden,
            "codigo": r.codigo or "",
            "nombre_serie": r.codigo_serie.nombre if r.codigo_serie else "",
            "nombre_subserie": r.codigo_subserie.nombre if r.codigo_subserie else "",
            "unidad_documental": r.unidad_documental or "",
            "identificador_documento": r.identificador_documento or "",
            "fecha_archivo": r.fecha_archivo.strftime("%Y-%m-%d") if r.fecha_archivo else "",
            "fecha_inicial": r.fecha_inicial.strftime("%Y-%m-%d") if r.fecha_inicial else "",
            "fecha_final": r.fecha_final.strftime("%Y-%m-%d") if r.fecha_final else "",
            "soporte_fisico": "✔" if r.soporte_fisico else "✖",
            "soporte_electronico": "✔" if r.soporte_electronico else "✖",
            "caja": r.caja or "",
            "carpeta": r.carpeta or "",
            "tomo_legajo_libro": r.tomo_legajo_libro or "",
            "numero_folios": r.numero_folios if r.numero_folios else "",
            "tipo": r.tipo or "",
            "cantidad": r.cantidad if r.cantidad else "",
            "ubicacion": r.ubicacion or "",
            "cantidad_documentos_electronicos": r.cantidad_documentos_electronicos if r.cantidad_documentos_electronicos else "",
            "tamano_documentos_electronicos": r.tamano_documentos_electronicos if r.tamano_documentos_electronicos else "",
            "notas": r.notas or "",
            "creado_por": r.creado_por.username if r.creado_por else "",
            "fecha_creacion": r.fecha_creacion.strftime("%Y-%m-%d %H:%M") if r.fecha_creacion else "",
            "fecha_edicion": r.fecha_edicion.strftime("%Y-%m-%d %H:%M") if r.fecha_edicion else "",
            "editado_por": r.editado_por.username if r.editado_por else "",
            "Estado_archivo": "✔" if r.Estado_archivo else "✖",
            "documento": (
                f'<a href="/registros/documento/{r.id}/" target="_blank">📁 Ver Documento</a>'
                if r.documentos.all()
                else "✖ No hay documento"
            ),
            "acciones": f'<a href="/registros/fuids/{fuid.id}/editar_registro/{r.id}/" class="btn btn-sm btn-warning">Editar</a>'
        })

    return JsonResponse({
        "draw": int(params.get('draw', 1)),
        "recordsTotal": total_registros,
        "recordsFiltered": total_registros,
        "data": data
    })


# @login_required
# def registros_fuid_json(request, fuid_id):
#     from django.db.models import Q
#     from django.core.paginator import Paginator
#     from django.urls import reverse
#     from django.http import JsonResponse
#     from django.shortcuts import get_object_or_404
#     from documentos.models import FUID

#     fuid = get_object_or_404(FUID, pk=fuid_id)

#     start = int(request.GET.get('start', 0))
#     length = int(request.GET.get('length', 30))

#     registros_query = fuid.registros.all()

#     # ==================== Lectura de filtros (SIN CAMBIOS) ====================
#     numero_orden = request.GET.get("filtro_numero_orden", "").strip()
#     codigo = request.GET.get("filtro_codigo", "").strip()
#     unidad_documental = request.GET.get("filtro_unidad_documental", "").strip()
#     identificador_documento = request.GET.get("filtro_identificador_documento", "").strip()
#     fecha_archivo = request.GET.get("filtro_fecha_archivo", "").strip()
#     fecha_inicial = request.GET.get("filtro_fecha_inicial", "").strip()
#     fecha_final = request.GET.get("filtro_fecha_final", "").strip()
#     soporte_fisico = request.GET.get("filtro_soporte_fisico", "").strip()
#     soporte_electronico = request.GET.get("filtro_soporte_electronico", "").strip()
#     caja = request.GET.get("filtro_caja", "").strip()
#     carpeta = request.GET.get("filtro_carpeta", "").strip()
#     tomo_legajo = request.GET.get("filtro_tomo_legajo_libro", "").strip()
#     numero_folios = request.GET.get("filtro_numero_folios", "").strip()
#     tipo = request.GET.get("filtro_tipo", "").strip()
#     cantidad = request.GET.get("filtro_cantidad", "").strip()
#     ubicacion = request.GET.get("filtro_ubicacion", "").strip()
#     cant_elec = request.GET.get("filtro_cant_elec", "").strip()
#     tamano_elec = request.GET.get("filtro_tamano_elec", "").strip()
#     notas = request.GET.get("filtro_notas", "").strip()
#     creado_por = request.GET.get("filtro_creado_por", "").strip()
#     fecha_creacion = request.GET.get("filtro_fecha_creacion", "").strip()
#     estado_archivo = request.GET.get("filtro_estado_archivo", "").strip()

#     # ==================== Filtros Q() (SIN CAMBIOS) ====================
#     q_filter = Q()
#     if numero_orden:
#         q_filter &= Q(numero_orden__icontains=numero_orden)
#     if codigo:
#         q_filter &= Q(codigo__icontains=codigo)
#     if unidad_documental:
#         q_filter &= Q(unidad_documental__icontains=unidad_documental)
#     if identificador_documento:
#         q_filter &= Q(identificador_documento__icontains=identificador_documento)
#     if fecha_archivo:
#         q_filter &= Q(fecha_archivo=fecha_archivo)
#     if fecha_inicial:
#         q_filter &= Q(fecha_inicial=fecha_inicial)
#     if fecha_final:
#         q_filter &= Q(fecha_final=fecha_final)
#     if soporte_fisico:
#         q_filter &= Q(soporte_fisico=(soporte_fisico == "✔"))
#     if soporte_electronico:
#         q_filter &= Q(soporte_electronico=(soporte_electronico == "✔"))
#     if caja:
#         q_filter &= Q(caja__icontains=caja)
#     if carpeta:
#         q_filter &= Q(carpeta__icontains=carpeta)
#     if tomo_legajo:
#         q_filter &= Q(tomo_legajo_libro__icontains=tomo_legajo)
#     if numero_folios:
#         q_filter &= Q(numero_folios__icontains=numero_folios)
#     if tipo:
#         q_filter &= Q(tipo__icontains=tipo)
#     if cantidad:
#         q_filter &= Q(cantidad__icontains=cantidad)
#     if ubicacion:
#         q_filter &= Q(ubicacion__icontains=ubicacion)
#     if cant_elec:
#         q_filter &= Q(cantidad_documentos_electronicos__icontains=cant_elec)
#     if tamano_elec:
#         q_filter &= Q(tamano_documentos_electronicos__icontains=tamano_elec)
#     if notas:
#         q_filter &= Q(notas__icontains=notas)
#     if creado_por:
#         q_filter &= Q(creado_por__username__icontains=creado_por)
#     if fecha_creacion:
#         q_filter &= Q(fecha_creacion__icontains=fecha_creacion)
#     if estado_archivo:
#         q_filter &= Q(Estado_archivo=(estado_archivo == "✔"))

#     registros_query = registros_query.filter(q_filter)

#     total_registros = registros_query.count()
#     paginator = Paginator(registros_query.order_by('id'), length)
#     page_obj = paginator.get_page(start // length + 1)

#     data = []
#     for r in page_obj:
#         data.append({
#             "numero_orden": r.numero_orden,
#             "codigo": r.codigo or "",
#             # ==================== NUEVOS CAMPOS ====================
#             "nombre_serie": r.codigo_serie.nombre if r.codigo_serie else "",
#             "nombre_subserie": r.codigo_subserie.nombre if r.codigo_subserie else "",
#             # =======================================================
#             "unidad_documental": r.unidad_documental or "",
#             "identificador_documento": r.identificador_documento or "",
#             "fecha_archivo": r.fecha_archivo.strftime("%Y-%m-%d") if r.fecha_archivo else "",
#             "fecha_inicial": r.fecha_inicial.strftime("%Y-%m-%d") if r.fecha_inicial else "",
#             "fecha_final": r.fecha_final.strftime("%Y-%m-%d") if r.fecha_final else "",
#             "soporte_fisico": "✔" if r.soporte_fisico else "✖",
#             "soporte_electronico": "✔" if r.soporte_electronico else "✖",
#             "caja": r.caja or "",
#             "carpeta": r.carpeta or "",
#             "tomo_legajo_libro": r.tomo_legajo_libro or "",
#             "numero_folios": r.numero_folios if r.numero_folios else "",
#             "tipo": r.tipo or "",
#             "cantidad": r.cantidad if r.cantidad else "",
#             "ubicacion": r.ubicacion or "",
#             "cantidad_documentos_electronicos": r.cantidad_documentos_electronicos if r.cantidad_documentos_electronicos else "",
#             "tamano_documentos_electronicos": r.tamano_documentos_electronicos if r.tamano_documentos_electronicos else "",
#             "notas": r.notas or "",
#             "creado_por": r.creado_por.username if r.creado_por else "",
#             "fecha_creacion": r.fecha_creacion.strftime("%Y-%m-%d %H:%M") if r.fecha_creacion else "",
#             "Estado_archivo": "✔" if r.Estado_archivo else "✖",
#             "documento": (
#                 f'<a href="/registros/documento/{r.id}/" target="_blank">📁 Ver Documento</a>'
#                 if r.documentos.all()
#                 else "✖ No hay documento"
#             ),
#             "acciones": f'<a href="/registros/fuids/{fuid.id}/editar_registro/{r.id}/" class="btn btn-sm btn-warning">Editar</a>'
#         })

#     return JsonResponse({
#         "draw": int(request.GET.get('draw', 1)),
#         "recordsTotal": total_registros,
#         "recordsFiltered": total_registros,
#         "data": data
#     })





@login_required
def agregar_registro_a_fuid(request, fuid_id):
    fuid = get_object_or_404(FUID, pk=fuid_id)
    # Aquí podrías añadir verificaciones de permisos si son necesarias
    # if not request.user.has_perm('documentos.add_registro_to_fuid', fuid):
    #     return HttpResponseForbidden("No tienes permiso para agregar registros a este FUID.")

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST, request.FILES) # Incluir request.FILES
        if form.is_valid():
            registro = form.save(commit=False)
            registro.creado_por = request.user

            # --- Calcular y asignar el siguiente numero_orden para este FUID ---
            max_orden_data = fuid.registros.aggregate(max_orden=Max('numero_orden'))
            max_orden = max_orden_data.get('max_orden')
            if max_orden is None: # Si no hay registros o el campo es nulo
                next_numero_orden = 1
            else:
                next_numero_orden = max_orden + 1
            registro.numero_orden = next_numero_orden
            # --- Fin del cálculo ---

            registro.save() # Guardar el registro principal

            # --- Crear el objeto Documento si se subió un archivo ---            
            archivo_subido = form.cleaned_data.get('archivo')
            if archivo_subido:
                try:
                    Documento.objects.create(
                        registro=registro,
                        archivo=archivo_subido
                    )
                except Exception as e:
                    # Manejar error de creación de documento si es necesario
                    messages.error(request, f"Error al guardar el documento adjunto: {e}")
                    # Podrías decidir si continuar o no, aquí continuamos pero mostramos error
            # --- Fin manejo de archivo ---

            # Asociar el registro recién guardado al FUID
            fuid.registros.add(registro)

            # Asignar permisos si es necesario (ejemplo)
            # assign_perm('view_registrodearchivo', request.user, registro)
            # assign_perm('change_registrodearchivo', request.user, registro)

            messages.success(request, f'Registro con orden #{next_numero_orden} agregado exitosamente al FUID #{fuid.id}.')
            return redirect('detalle_fuid', pk=fuid.id) # Redirigir a la vista de detalle del FUID
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                try:
                    field_label = form.fields[field].label if field != '__all__' else 'Errores generales'
                except KeyError:
                    field_label = field
                for error in errors:
                    messages.error(request, f"{field_label}: {error}")

    else: # Método GET
        form = RegistroDeArchivoForm()
        # Puedes pre-configurar la serie/subserie si es necesario aquí

    context = {
        'form': form,
        'fuid': fuid
    }
    return render(request, 'agregar_registro_a_fuid.html', context)


@login_required
@require_POST
def agregar_registro_modal_ajax(request, fuid_id):
    """
    Vista AJAX para agregar un registro a un FUID desde el modal.
    Replica la lógica completa de agregar_registro_a_fuid pero devuelve JSON.
    """
    fuid = get_object_or_404(FUID, pk=fuid_id)

    form = RegistroDeArchivoForm(request.POST, request.FILES)
    if form.is_valid():
        registro = form.save(commit=False)
        registro.creado_por = request.user

        # Calcular siguiente numero_orden para este FUID
        max_orden_data = fuid.registros.aggregate(max_orden=Max('numero_orden'))
        max_orden = max_orden_data.get('max_orden')
        next_numero_orden = (max_orden + 1) if max_orden is not None else 1
        registro.numero_orden = next_numero_orden

        registro.save()

        # Crear Documento si se subió un archivo
        archivo_subido = form.cleaned_data.get('archivo')
        if archivo_subido:
            try:
                Documento.objects.create(registro=registro, archivo=archivo_subido)
            except Exception as e:
                pass  # El registro ya se guardó, el archivo es secundario

        # Asociar al FUID
        fuid.registros.add(registro)

        return JsonResponse({
            'ok': True,
            'message': f'Registro #{next_numero_orden} agregado exitosamente al FUID #{fuid.id}.',
        })
    else:
        # Devolver errores del formulario
        errores = {}
        for field, errors in form.errors.items():
            try:
                label = form.fields[field].label if field != '__all__' else 'General'
            except KeyError:
                label = field
            errores[field] = [f"{label}: {e}" for e in errors]

        return JsonResponse({
            'ok': False,
            'errores': errores,
            'message': 'Hay errores en el formulario.',
        }, status=400)


@login_required
def editar_registro_de_fuid(request, fuid_id, registro_id):
    # Verificar permiso de edición (superusuario siempre tiene acceso)
    if not (request.user.is_superuser or request.user.has_perm('documentos.change_registrodearchivo')):
        return HttpResponseForbidden("No tienes permiso para editar registros.")

    # Obtener FUID y Registro
    fuid     = get_object_or_404(FUID, pk=fuid_id)
    registro = get_object_or_404(RegistroDeArchivo, pk=registro_id)

    # Comprobación eficiente de asociación en ManyToMany (sin cargar todo el conjunto)
    if not fuid.registros.filter(pk=registro.id).exists():
        return HttpResponseForbidden("El registro no está asociado a este FUID.")

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST, request.FILES, instance=registro)
        if form.is_valid():
            updated_registro = form.save(commit=False)
            updated_registro.editado_por = request.user
            updated_registro.save()
            archivo_subido = form.cleaned_data.get('archivo')
            if archivo_subido:
                Documento.objects.create(
                    registro=updated_registro,
                    archivo=archivo_subido
                )
            messages.success(request, 'Registro actualizado correctamente.')
            return redirect('detalle_fuid', pk=fuid_id)
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label
                for error in errors:
                    messages.error(request, f"{label}: {error}")
    else:
        form = RegistroDeArchivoForm(instance=registro)
        if registro.codigo_serie:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.filter(
                serie_id=registro.codigo_serie.id
            )
        else:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.none()

    return render(request, 'editar_registro_de_fuid.html', {
        'form':     form,
        'fuid':     fuid,
        'registro': registro,
    })

@login_required
def welcome_view(request):
    puede_gestionar_prestamos_usuario = puede_gestionar_prestamos(request.user)
    return render(request, 'welcome.html', {
        'puede_gestionar_prestamos': puede_gestionar_prestamos_usuario
    })


#panel de control para administradores, solo pueden acceder los usuarios con el grupo "administradores"
@login_required
def panel_view(request):
    if not request.user.is_superuser:
        return mi_error_403(request)  # Llamamos a la función de error si no es superusuario

    return render(request, 'panel_de_control.html')

from django.http import HttpResponseForbidden

@login_required
def crear_ficha_paciente(request):
    # Verificar si el usuario tiene permiso global para agregar fichas
    if not request.user.has_perm('documentos.add_fichapaciente'):
        return HttpResponseForbidden("No tienes permiso para crear fichas de pacientes.")

    if request.method == 'POST':
        form = FichaPacienteForm(request.POST)
        if form.is_valid():
            ficha = form.save()
            messages.success(request, 'Ficha del paciente registrada exitosamente.')
            return redirect('crear_ficha')  # Redirige a la misma página o a otra URL
        else:
            # Manejo de errores en el formulario
            for field, errors in form.errors.items():
                field_name = form.fields[field].label
                for error in errors:
                    messages.error(request, f"{field_name}: {error}")
    else:
        form = FichaPacienteForm()

    return render(request, 'ficha_paciente_form.html', {'form': form})




from django.shortcuts import render

def mi_error_403(request, exception=None):
    return render(request, '403.html', status=403)



from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponseForbidden
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, permission_required

@login_required
@permission_required('documentos.view_fichapaciente', raise_exception=True)
def lista_fichas_paciente(request):
    fichas = FichaPaciente.objects.all()
    return render(request, 'lista_fichas_paciente.html', {'fichas': fichas})



@method_decorator(login_required, name='dispatch')
class EditarFichaPaciente(UpdateView):
    model = FichaPaciente
    fields = '__all__'
    template_name = 'ficha_paciente_form.html'
    success_url = reverse_lazy('lista_fichas')
    pk_url_kwarg = 'consecutivo'  # Usar 'consecutivo' en lugar de 'pk'

    def dispatch(self, request, *args, **kwargs):
        # Verificar si el usuario tiene permiso global para editar fichas
        if not request.user.has_perm('documentos.change_fichapaciente'):
            return HttpResponseForbidden("No tienes permiso para editar fichas de pacientes.")

        # Verificar que la ficha existe
        self.object = get_object_or_404(FichaPaciente, consecutivo=kwargs.get(self.pk_url_kwarg))
        return super().dispatch(request, *args, **kwargs)



@login_required
def detalle_ficha_paciente(request, consecutivo):
    ficha = get_object_or_404(FichaPaciente, consecutivo=consecutivo)
    return render(request, 'detalle_ficha_paciente.html', {'ficha': ficha})




class ListaFichasAPIView(APIView):
    def get(self, request):
        # Parámetros enviados desde el frontend
        fecha_inicio = request.GET.get('fecha_inicio', None)
        fecha_fin = request.GET.get('fecha_fin', None)
        filtro_identificacion = request.GET.get('filtro_identificacion', None)
        filtro_historia = request.GET.get('filtro_historia', None)
        filtro_nombre = request.GET.get('filtro_nombre', None)
        filtro_similar = request.GET.get('filtro_similar', None)
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 250))

        # Ordenamiento
        order_column = int(request.GET.get('order[0][column]', 0))
        order_dir = request.GET.get('order[0][dir]', 'asc')

        # Mapear columnas de DataTables a campos del modelo
        column_mapping = {
            0: 'consecutivo',
            1: 'primer_nombre',  # Ordenar por primer nombre
            2: 'tipo_identificacion',
            3: 'num_identificacion',
            4: 'sexo',
            5: 'activo',  # Ordenar por estado
            6: 'fecha_nacimiento',
            7: 'Numero_historia_clinica',
        }

        # Determinar el campo para ordenar
        order_field = column_mapping.get(order_column, 'consecutivo')  # Campo predeterminado: consecutivo
        if order_dir == 'desc':
            order_field = f"-{order_field}"  # Prefijo "-" para orden descendente

        # Base queryset
        queryset = FichaPaciente.objects.all()

        # Filtros avanzados
        if fecha_inicio and fecha_fin:
            queryset = queryset.filter(fecha_nacimiento__range=[fecha_inicio, fecha_fin])
        if filtro_identificacion:
            queryset = queryset.filter(num_identificacion__icontains=filtro_identificacion)
        if filtro_historia:
            queryset = queryset.filter(Numero_historia_clinica__icontains=filtro_historia)
        if filtro_nombre:
            queryset = queryset.filter(
                primer_nombre__icontains=filtro_nombre
            ) | queryset.filter(
                primer_apellido__icontains=filtro_nombre
            )
        if filtro_similar:
            queryset = queryset.filter(
                primer_nombre__icontains=filtro_similar
            ) | queryset.filter(
                segundo_nombre__icontains=filtro_similar
            ) | queryset.filter(
                primer_apellido__icontains=filtro_similar
            ) | queryset.filter(
                segundo_apellido__icontains=filtro_similar
            )

        # Aplicar ordenamiento dinámico
        queryset = queryset.order_by(order_field)

        # Paginación
        total_records = queryset.count()
        paginator = Paginator(queryset, length)
        fichas = paginator.get_page(start // length + 1).object_list

        # Formato JSON para DataTables
        data = [
            {
                "consecutivo": ficha.consecutivo,
                "nombre_completo": f"{ficha.primer_nombre} {ficha.segundo_nombre or ''} {ficha.primer_apellido} {ficha.segundo_apellido}",
                "tipo_identificacion": ficha.tipo_identificacion,
                "num_identificacion": ficha.num_identificacion,
                "sexo": ficha.sexo,
                "estado": ficha.activo,
                "fecha_nacimiento": ficha.fecha_nacimiento.strftime("%Y-%m-%d"),
                "numero_historia_clinica": ficha.Numero_historia_clinica,
            }
            for ficha in fichas
        ]

        return Response(
            {
                "draw": request.GET.get("draw", 1),
                "recordsTotal": total_records,
                "recordsFiltered": total_records,
                "data": data,
            }
        )


def export_fuid_to_excel(request, pk):
    # Obtener el FUID específico
    fuid = FUID.objects.get(pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"FUID #{fuid.id}"

    # Función para truncar valores largos
    def truncate_value(value, max_length=200):
        if not value:
            return "N/A"
        value = str(value)
        return value if len(value) <= max_length else value[:max_length - 3] + "..."

    # Crear estilos
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    header_fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")

    # Combinar celdas para la imagen
    ws.merge_cells(start_row=1, start_column=1, end_row=6, end_column=22)

    # Insertar la imagen

    # Ruta física al archivo estático
    img_path = settings.STATICFILES_DIRS[0] / 'img' / 'fuid_logo.png'

    if not img_path.exists():
        raise FileNotFoundError(f"Logo no encontrado en {img_path}")
    # img_path = r"D:\descargas d\xtz\pino-d-angio-c92c3fc03f2f716d1835fcf5b169efc11833deab\hospital_document_management\documentos\templates\images\fuid_logo.png"
    img = Image(img_path)
    img.width = 1000
    img.height = 120
    ws.add_image(img, "A1")

    # Mover el cursor de escritura a la fila 7 para continuar con el contenido
    current_row = 7

    # Encabezados de datos generales
    ws.cell(row=current_row, column=1, value="Campo")
    ws.cell(row=current_row, column=2, value="Valor")
    ws.cell(row=current_row, column=17, value="AÑO")
    ws.cell(row=current_row, column=18, value="MES")
    ws.cell(row=current_row, column=19, value="DÍA")
    ws.cell(row=current_row, column=20, value="N.T.")
    current_row += 1

    # Datos generales del FUID
    fuid_data = [
        ("Entidad Productora", fuid.entidad_productora.nombre if fuid.entidad_productora else "N/A", fuid.fecha_creacion.year, fuid.fecha_creacion.month, fuid.fecha_creacion.day, ""),
        ("Unidad Administrativa", fuid.unidad_administrativa.nombre if fuid.unidad_administrativa else "N/A", "", "", "", ""),
        ("Oficina Productora", fuid.oficina_productora.nombre if fuid.oficina_productora else "N/A", "", "", "", ""),
        ("Objeto", fuid.objeto.nombre if fuid.objeto else "N/A", "", "", "", ""),
    ]
    for row_data in fuid_data:
        ws.cell(row=current_row, column=1, value=row_data[0])  # Campo
        ws.cell(row=current_row, column=2, value=row_data[1])  # Valor
        ws.cell(row=current_row, column=17, value=row_data[2])  # AÑO
        ws.cell(row=current_row, column=18, value=row_data[3])  # MES
        ws.cell(row=current_row, column=19, value=row_data[4])  # DÍA
        ws.cell(row=current_row, column=20, value=row_data[5])  # N.T.
        current_row += 1

    # Aplicar bordes solo a celdas con contenido
    for row in ws.iter_rows(min_row=7, max_row=current_row-1):
        for cell in row:
            if cell.value:  # Aplica bordes solo si hay contenido
                cell.border = border

    # Espacio antes de la sección de registros
    current_row += 1
    ws.cell(row=current_row, column=1, value="")
    current_row += 1

    # Encabezados de los registros (sin "Fecha Archivo")
    headers = [
        "N° Orden", "Código", "Código Serie", "Código Subserie", "Unidad Documental",
        "Fecha Inicial", "Fecha Final", "Soporte Físico", "Soporte Electrónico",
        "Caja", "Carpeta", "Tomo/Legajo/Libro", "N° Folios", "Tipo", "Cantidad",
        "Ubicación", "Cantidad Electrónicos", "Tamaño Electrónico", "Notas"
    ]
    start_row = current_row + 1
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=start_row+3, end_column=col_idx)
        cell = ws[f"{col_letter}{start_row}"]
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        if cell.value:  # Aplica bordes solo si hay contenido
            cell.border = border

    # Mover el current_row debajo de las cabeceras
    current_row = start_row + 4

    # Agregar registros asociados (sin "Fecha Archivo")
    registros = fuid.registros.all()
    if registros.exists():
        for registro in registros:
            row_data = [
                registro.numero_orden,
                truncate_value(registro.codigo or "N/A"),
                truncate_value(registro.codigo_serie.nombre if registro.codigo_serie else "N/A"),
                truncate_value(registro.codigo_subserie.nombre if registro.codigo_subserie else "N/A"),
                truncate_value(registro.unidad_documental),
                registro.fecha_inicial.strftime('%Y-%m-%d') if registro.fecha_inicial else "N/A",
                registro.fecha_final.strftime('%Y-%m-%d') if registro.fecha_final else "N/A",
                "Sí" if registro.soporte_fisico else "No",
                "Sí" if registro.soporte_electronico else "No",
                truncate_value(registro.caja or "N/A"),
                truncate_value(registro.carpeta or "N/A"),
                truncate_value(registro.tomo_legajo_libro or "N/A"),
                registro.numero_folios or "N/A",
                truncate_value(registro.tipo or "N/A"),
                registro.cantidad or "N/A",
                truncate_value(registro.ubicacion),
                registro.cantidad_documentos_electronicos or "N/A",
                truncate_value(registro.tamano_documentos_electronicos or "N/A"),
                truncate_value(registro.notas or "N/A"),
                # registro.creado_por.username if registro.creado_por else "N/A",
                # registro.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                c = ws.cell(row=current_row, column=col_idx, value=val)
                if c.value:  # Aplica bordes solo si hay contenido
                    c.border = border
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="Sin registros asociados")
        current_row += 1

    # Espacio antes de la sección de roles
    current_row += 1

    # Datos de roles
    roles_data = [
        ["Elaborado Por (Nombre)", truncate_value(fuid.elaborado_por_nombre or "N/A"),
         "Entregado Por (Nombre)", truncate_value(fuid.entregado_por_nombre or "N/A"),
         "Recibido Por (Nombre)", truncate_value(fuid.recibido_por_nombre or "N/A")],
        ["Elaborado Por (Cargo)", truncate_value(fuid.elaborado_por_cargo or "N/A"),
         "Entregado Por (Cargo)", truncate_value(fuid.entregado_por_cargo or "N/A"),
         "Recibido Por (Cargo)", truncate_value(fuid.recibido_por_cargo or "N/A")],
        ["Elaborado Por (Lugar)", truncate_value(fuid.elaborado_por_lugar or "N/A"),
         "Entregado Por (Lugar)", truncate_value(fuid.entregado_por_lugar or "N/A"),
         "Recibido Por (Lugar)", truncate_value(fuid.recibido_por_lugar or "N/A")],
        ["Firma", "", "Firma", "", "Firma", ""],
        ["Lugar", "", "Lugar", "", "Lugar", ""],
        ["Elaborado Por (Fecha)", fuid.elaborado_por_fecha.strftime('%Y-%m-%d') if fuid.elaborado_por_fecha else "N/A",
         "Entregado Por (Fecha)", fuid.entregado_por_fecha.strftime('%Y-%m-%d') if fuid.entregado_por_fecha else "N/A",
         "Recibido Por (Fecha)", fuid.recibido_por_fecha.strftime('%Y-%m-%d') if fuid.recibido_por_fecha else "N/A"],
    ]

    # Asegurar bordes para todas las celdas de roles (rango expandido)
    start_col = 1  # Columna inicial para los datos de roles
    end_col = 10  # Aumentamos el rango de columnas ocupadas
    for row_idx, row_data in enumerate(roles_data, start=current_row):
        for col_idx, val in enumerate(row_data, start=start_col):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = border  # Aplicar bordes incluso si está vacío
        current_row += 1

    # Ajustar el ancho de las columnas automáticamente
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Configurar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=FUID_{fuid.id}.xlsx'

    wb.save(response)
    return response


#  @login_required
def calcular_edad(fecha_nacimiento):
    """
    Calcula la edad actual basada en la fecha de nacimiento.
    """
    if fecha_nacimiento:
        hoy = date.today()
        return hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
    return None
# @login_required
def estadisticas_pacientes(request):
    """
    API para devolver estadísticas de pacientes considerando varios atributos.
    """
    usuario = request.GET.get('usuario')
    pacientes = FichaPaciente.objects.all()

    if usuario:
        pacientes = pacientes.filter(creado_por__username=usuario)

    # Calcular edades
    edades = [calcular_edad(p.fecha_nacimiento) for p in pacientes if p.fecha_nacimiento]

    # Clasificar por grupos de edad
    grupos_edad = {
        "0-18": sum(1 for e in edades if e <= 18),
        "19-35": sum(1 for e in edades if 19 <= e <= 35),
        "36-60": sum(1 for e in edades if 36 <= e <= 60),
        "60+": sum(1 for e in edades if e > 60)
    }

    datos = {
        'total_pacientes': pacientes.count(),
        'por_genero': list(pacientes.values('sexo').annotate(cantidad=Count('sexo'))),
        'por_tipo_identificacion': list(pacientes.values('tipo_identificacion').annotate(cantidad=Count('tipo_identificacion'))),
        'activos': pacientes.filter(activo=True).count(),
        'promedio_edad': round(sum(edades) / len(edades), 2) if edades else None,
        'grupos_edad': grupos_edad
    }

    return JsonResponse(datos, safe=False)

# @login_required
def estadisticas_registros(request):
    """
    API para devolver estadísticas de registros, organizados por series documentales y tipos.
    """
    try:
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        registros = RegistroDeArchivo.objects.all()

        # Filtrar por rango de fechas si se proporcionan
        if fecha_inicio and fecha_fin:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
            registros = registros.filter(fecha_archivo__range=(fecha_inicio, fecha_fin))

        # Generar estadísticas
        datos = {
            'total_registros': registros.count(),
            'por_serie': list(
                registros.values('codigo_serie__nombre').annotate(cantidad=Count('id'))
            ),
            'por_soporte': list(
                registros.values('soporte_fisico', 'soporte_electronico').annotate(cantidad=Count('id'))
            ),
            'por_tipo': list(
                registros.values('tipo').annotate(cantidad=Count('id'))
            ),
        }

        return JsonResponse(datos, safe=False)
    except Exception as e:
        print("Error en estadisticas_registros:", e)
        return JsonResponse({"error": str(e)}, status=500)



# @login_required
def estadisticas_fuids(request):
    """
    API para devolver estadísticas de FUIDs, organizados por oficinas productoras.
    """
    usuario = request.GET.get('usuario')
    fuids = FUID.objects.all()

    if usuario:
        fuids = fuids.filter(creado_por__username=usuario)

    datos = {
        'total_fuids': fuids.count(),
        'por_oficina': list(fuids.values('oficina_productora__nombre').annotate(cantidad=Count('id'))),
        'por_objeto': list(fuids.values('objeto__nombre').annotate(cantidad=Count('id'))),
        'por_entidad': list(fuids.values('entidad_productora__nombre').annotate(cantidad=Count('id'))),
    }

    return JsonResponse(datos, safe=False)

# @login_required
def pagina_estadisticas(request):
    """
    Página principal para mostrar gráficos de las estadísticas.
    """
    return render(request, 'pagina_estadisticas.html')


@login_required
def obtener_usuarios(request):
    usuarios = User.objects.values('username')
    return JsonResponse(list(usuarios), safe=False)

# mixins.py
from django.http import HttpResponseForbidden

class OficinaFilterMixin:
    """
    Filtra los objetos para que el usuario solo vea y manipule
    aquellos creados por su oficina. También bloquea la edición
    de objetos de otras oficinas.
    """
    def get_queryset(self):
        qs = super().get_queryset()
        # Si deseas que el superusuario vea todo, déjalo pasar:
        if self.request.user.is_superuser:
            return qs
        # Caso contrario, filtra por la oficina del perfil
        return qs.filter(oficina_productora=self.request.user.perfil.oficina)

    def dispatch(self, request, *args, **kwargs):
        # Bloqueo adicional para edición/eliminación
        if hasattr(self, 'get_object'):
            obj = self.get_object()
            if (not request.user.is_superuser) and (obj.oficina_productora != request.user.perfil.oficina):
                return HttpResponseForbidden("No tienes permiso sobre este recurso.")
        return super().dispatch(request, *args, **kwargs)


from django.shortcuts import render

def soporte_view(request):
    return render(request, 'soporte.html')




from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib.auth.decorators import (
    login_required,
    permission_required,
)

from .forms  import UbicacionFisicaForm, BuscaFichaForm, FichaPacienteForm
from .models import FichaPaciente


# ═══════════════════════════════════════════════════════════
# FLUJO "GAVETA"  (Paso 0 y Paso 1)
# ═══════════════════════════════════════════════════════════
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import UbicacionFisicaForm

from django.http import Http404

@login_required
def preparar_fichas1(request):
    """
    Paso 0 – El funcionario selecciona gaveta/caja/carpeta.
    Guarda la elección en la sesión y redirige al buscador o al formulario de creación.
    """
    if not request.user.has_perm('documentos.add_fichapaciente'):
        raise Http404("No tienes permiso para crear fichas.")

    if request.method == "POST":
        form = UbicacionFisicaForm(request.POST)
        if form.is_valid():
            # Guarda los datos en la sesión para prellenar formularios después
            request.session["prefill"] = form.cleaned_data
            messages.success(request, "Ubicación física seleccionada correctamente.")
            return redirect("crear_ficha1")
    else:
        # Si ya hay datos prellenados, los carga
        form = UbicacionFisicaForm(initial=request.session.get("prefill", {}))

    return render(request, "templatesfichas/preparar.html", {"form": form})

@login_required
def buscar_ficha1(request):
    """
    Paso 1 – Busca una ficha por número de identificación o historia clínica.
    Si existe → redirige a edición; si no existe → redirige a creación
    con los campos de búsqueda en la URL para pre‑rellenar.
    """
    if not request.user.has_perm('documentos.add_fichapaciente'):
        raise Http404("No tienes permiso para crear fichas.")

    if "prefill" not in request.session:
        return redirect("preparar_fichas")

    if request.method == "POST":
        form = BuscaFichaForm(request.POST)
        if form.is_valid():
            num_id  = form.cleaned_data.get("num_identificacion")
            hist_id = form.cleaned_data.get("Numero_historia_clinica")

            qs = FichaPaciente.objects.all()
            if num_id:
                qs = qs.filter(num_identificacion=num_id)
            if hist_id:
                qs = qs.filter(Numero_historia_clinica=hist_id)
            if qs.exists():
                ficha = qs.first()
                messages.info(request, "Ficha encontrada; puedes verificar o editar.")
                return redirect("editar_ficha", consecutivo=ficha.consecutivo)

            # No existe → construir querystring y redirigir a creación
            params = []
            if num_id:  params.append(f"num_identificacion={num_id}")
            if hist_id: params.append(f"Numero_historia_clinica={hist_id}")
            qs_str = "&".join(params)
            return redirect(reverse("crear_ficha") + (f"?{qs_str}" if qs_str else ""))
    else:
        form = BuscaFichaForm()

    return render(request, "templatesfichas/buscar.html", {"form": form})

PAGE_SIZE = 50       
                      # 50 consultas y 50 renderizadas
from .models import TipoDocumento
PAGE_SIZE = 30          # o el que uses

@login_required
def lista_fichas1(request):
    """
    Listado paginado y filtrable de FichaPaciente.
    – Carga solo columnas imprescindibles (only)
    – select_related para FK de documento
    – Todos los filtros son AND
    """
    GET = request.GET

    # -------- 1.  Queryset base -----------
    qs = (
        FichaPaciente.objects
        .only(
            "consecutivo", "primer_nombre", "segundo_nombre",
            "primer_apellido", "segundo_apellido",
            "num_identificacion", "num_identificacion_secundario",
            "Numero_historia_clinica",
            "tipo_identificacion_id", "tipo_identificacion_secundario_id",
            "sexo", "activo", "estado_de_migracion",
            "gabeta", "caja", "carpeta",
            "fecha_nacimiento",
            "Fecha_de_visita_de_la_tarjeta",
            "ultimo_registro_de_visita_en_la_base_de_datos"
        )
        .select_related("tipo_identificacion", "tipo_identificacion_secundario")
    )

    # -------- 2.  Filtros dinámicos (todos AND) ----------
    f = Q()

    # Búsqueda rápida en varias columnas
    q = GET.get("q", "").strip()
    if q:
        f &= (
            Q(num_identificacion__icontains=q) |
            Q(Numero_historia_clinica__icontains=q) |
            Q(primer_nombre__icontains=q) |
            Q(segundo_nombre__icontains=q) |
            Q(primer_apellido__icontains=q) |
            Q(segundo_apellido__icontains=q)
        )

    # Por campo exacto / icontains
    mapa_icontains = {
        "num_identificacion":   "num_identificacion__icontains",
        "num_identificacion_secundario": "num_identificacion_secundario__icontains",
        "caja":     "caja__iexact",
        "carpeta":  "carpeta__iexact",
    }
    for param, lookup in mapa_icontains.items():
        val = GET.get(param, "").strip()
        if val:
            f &= Q(**{lookup: val})

    # Filtros para nombres separados
    primer_nombre = GET.get("primer_nombre", "").strip()
    if primer_nombre:
        f &= Q(primer_nombre__icontains=primer_nombre)

    segundo_nombre = GET.get("segundo_nombre", "").strip()
    if segundo_nombre:
        f &= Q(segundo_nombre__icontains=segundo_nombre)

    primer_apellido = GET.get("primer_apellido", "").strip()
    if primer_apellido:
        f &= Q(primer_apellido__icontains=primer_apellido)

    segundo_apellido = GET.get("segundo_apellido", "").strip()
    if segundo_apellido:
        f &= Q(segundo_apellido__icontains=segundo_apellido)

    # Historia clínica (entero)
    hist = GET.get("Numero_historia_clinica", "").strip()
    if hist.isdigit():
        f &= Q(Numero_historia_clinica=int(hist))

    # Tipo documento primario / secundario
    tipo_doc = GET.get("tipo_identificacion", "").strip()
    if tipo_doc:
        f &= Q(tipo_identificacion__nombre__iexact=tipo_doc)

    tipo_doc_sec = GET.get("tipo_identificacion_secundario", "").strip()
    if tipo_doc_sec:
        f &= Q(tipo_identificacion_secundario__nombre__iexact=tipo_doc_sec)

    # Activo / Migrada
    if GET.get("activo") in {"si", "no"}:
        f &= Q(activo=(GET["activo"] == "si"))
    if GET.get("migrada") in {"si", "no"}:
        f &= Q(estado_de_migracion=(GET["migrada"] == "si"))

    # Sexo
    if GET.get("sexo"):
        f &= Q(sexo__iexact=GET["sexo"])

    # Rango fecha nacimiento
    ini, fin = GET.get("f_ini"), GET.get("f_fin")
    if ini and fin:
        f &= Q(fecha_nacimiento__range=[ini, fin])

    # -------- 3.  Aplicar filtros y paginar --------------
    qs = qs.filter(f).order_by("consecutivo")
    page_obj = Paginator(qs, PAGE_SIZE).get_page(GET.get("page"))

    return render(
        request,
        "templatesfichas/fichas_list.html",
        {
            "page_obj": page_obj,
            "tipos_doc": TipoDocumento.objects.all(),
        },
    )
# ----- Paso 2a: Crear ficha nueva ------------------------------------------
@permission_required("documentos.add_fichapaciente")
def crear_ficha1(request):
    """
    Alta de ficha.  
    Si el usuario llegó desde el flujo "gaveta", los valores se
    pre‑rellenan con `request.session["prefill"]`.
    """
    inicial = request.session.get("prefill", {}).copy()
    inicial.update({k: v for k, v in request.GET.items() if v})

    if request.method == "POST":
        form = FichaPacienteForm(request.POST, initial=inicial)
        if form.is_valid():
            ficha = form.save()
            messages.success(request, "Ficha creada con éxito.")
            return redirect("detalle_ficha1", consecutivo=ficha.consecutivo)
    else:
        form = FichaPacienteForm(initial=inicial)

    return render(
        request,
        "templatesfichas/ficha_paciente_form1.html",
        {"form": form, "modo": "crear", "ficha": None},
    )
    

# ----- Paso 2b: Editar ficha existente -------------------------------------
@permission_required("documentos.change_fichapaciente")
def editar_ficha1(request, consecutivo):
    ficha = get_object_or_404(FichaPaciente, consecutivo=consecutivo)

    if request.method == "POST":
        form = FichaPacienteForm(request.POST, instance=ficha)
        if form.is_valid():
            form.save()
            messages.success(request, "Ficha actualizada correctamente.")
            return redirect("detalle_ficha1", consecutivo=ficha.consecutivo)
    else:
        form = FichaPacienteForm(instance=ficha)

    return render(
        request,
        "templatesfichas/ficha_paciente_form1.html",
        {"form": form, "modo": "editar", "ficha": ficha},
    )

from django.http import Http404

@login_required
def detalle_ficha1(request, consecutivo):
    if not request.user.has_perm('documentos.view_fichapaciente'):
        raise Http404("No tienes permiso para ver esta ficha.")
    
    ficha = get_object_or_404(FichaPaciente, consecutivo=consecutivo)
    return render(request, "templatesfichas/ficha_paciente_detail.html", {"ficha": ficha})


from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
def bienvenida_fichas(request):
    if not request.user.has_perm('documentos.view_fichapaciente'):
        return render(request, '403.html', status=403)
    
    return render(request, "templatesfichas/bienvenida.html")



#seccion de importacion de documentos#
# views.py
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import RegistroDeArchivo, SerieDocumental, SubserieDocumental
from django.contrib.auth.decorators import login_required
from datetime import datetime

import tempfile

from django.contrib.auth.decorators import login_required, user_passes_test


@user_passes_test(lambda u: u.is_superuser)
@login_required
def importar_excel_archivo(request):
    """
    Vista para importar registros desde Excel.
    En POST: guarda el archivo en /tmp, lanza la tarea Celery y devuelve JSON con task_id.
    En GET: renderiza el formulario con barra de progreso.
    """
    if request.method == "POST":
        archivo = request.FILES.get("excel")
        fuid_id = request.POST.get("fuid_id")

        if not archivo:
            return JsonResponse(
                {"error": "No se adjuntó ningún archivo."}, status=400
            )

        try:
            fuid = FUID.objects.get(id=fuid_id)
        except FUID.DoesNotExist:
            return JsonResponse(
                {"error": "El FUID seleccionado no existe."}, status=400
            )

        # Guardar archivo en /tmp para que Celery lo lea
        suffix = ".xlsx" if archivo.name.endswith(".xlsx") else ".xls"
        with tempfile.NamedTemporaryFile(
            delete=False, suffix=suffix, prefix="import_excel_"
        ) as tmp:
            for chunk in archivo.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        # Lanzar tarea Celery
        from documentos.tasks import importar_excel_task
        task = importar_excel_task.delay(tmp_path, fuid.id, request.user.id)

        return JsonResponse({"task_id": task.id})

    # GET: cargar formulario con lista de FUIDs
    fuids = FUID.objects.all()
    return render(request, "templatesfichas/importar_excel.html", {"fuids": fuids})


@user_passes_test(lambda u: u.is_superuser)
@login_required
def progreso_importacion_excel(request, task_id):
    """
    Endpoint AJAX que devuelve el progreso de la tarea de importación de Excel.
    """
    from celery.result import AsyncResult

    result = AsyncResult(task_id)

    if result.state == 'PENDING':
        data = {
            'state': 'PENDING',
            'current': 0, 'total': 0,
            'fase': 'En cola, esperando...',
            'exitos': 0, 'errores_count': 0,
        }
    elif result.state == 'PROGRESS':
        info = result.info or {}
        data = {
            'state': 'PROGRESS',
            'current': info.get('current', 0),
            'total': info.get('total', 0),
            'fase': info.get('fase', 'Procesando...'),
            'exitos': info.get('exitos', 0),
            'errores_count': info.get('errores_count', 0),
        }
    elif result.state == 'SUCCESS':
        info = result.result or {}
        data = {
            'state': 'SUCCESS',
            'estado': info.get('estado', 'completado'),
            'mensaje': info.get('mensaje', ''),
            'exitos': info.get('exitos', 0),
            'errores': info.get('errores', []),
            'total': info.get('total', 0),
        }
    elif result.state == 'FAILURE':
        data = {
            'state': 'FAILURE',
            'mensaje': str(result.info),
            'exitos': 0, 'errores': [],
        }
    else:
        data = {
            'state': result.state,
            'fase': 'Procesando...',
            'current': 0, 'total': 0,
            'exitos': 0, 'errores_count': 0,
        }

    return JsonResponse(data)


# ==============================================================================
# VISTAS PARA PRÉSTAMOS DOCUMENTALES
# ==============================================================================

def normalizar_texto(valor):
    if not valor:
        return ''
    valor = unicodedata.normalize('NFKD', str(valor))
    return ''.join(char for char in valor if not unicodedata.combining(char)).lower().strip()


def q_prestamos_historia_clinica():
    return (
        Q(subserie__nombre__icontains='historia clínica') |
        Q(subserie__nombre__icontains='historia clinica') |
        Q(subserie__nombre__icontains='historias clínicas') |
        Q(subserie__nombre__icontains='historias clinicas') |
        Q(serie__nombre__icontains='historia clínica') |
        Q(serie__nombre__icontains='historia clinica') |
        Q(serie__nombre__icontains='historias clínicas') |
        Q(serie__nombre__icontains='historias clinicas')
    )


def obtener_nombre_oficina_usuario(user):
    if user.is_superuser:
        return 'Superusuario'
    try:
        return user.perfil.oficina.nombre
    except AttributeError:
        return ''


def obtener_area_gestion_usuario(user):
    if user.is_superuser:
        return 'todas'

    nombre_oficina = normalizar_texto(obtener_nombre_oficina_usuario(user))
    if 'gestion documental' in nombre_oficina or 'archivo central' in nombre_oficina:
        return 'archivo_central'
    if 'historias clinicas' in nombre_oficina or 'historia clinica' in nombre_oficina:
        return 'historia_clinica'
    return None


def obtener_area_desde_nombre_oficina(nombre_oficina):
    nombre_normalizado = normalizar_texto(nombre_oficina)
    if 'gestion documental' in nombre_normalizado or 'archivo central' in nombre_normalizado:
        return 'archivo_central'
    if 'historias clinicas' in nombre_normalizado or 'historia clinica' in nombre_normalizado:
        return 'historia_clinica'
    return None


def puede_gestionar_prestamos(user):
    return obtener_area_gestion_usuario(user) is not None


def q_oficina_responsable_area(area):
    if area == 'historia_clinica':
        return (
            Q(oficina_responsable__nombre__icontains='historias clínicas') |
            Q(oficina_responsable__nombre__icontains='historias clinicas') |
            Q(oficina_responsable__nombre__icontains='historia clínica') |
            Q(oficina_responsable__nombre__icontains='historia clinica')
        )
    if area == 'archivo_central':
        return (
            Q(oficina_responsable__nombre__icontains='gestión documental') |
            Q(oficina_responsable__nombre__icontains='gestion documental') |
            Q(oficina_responsable__nombre__icontains='archivo central')
        )
    return Q()


def aplicar_filtro_area_prestamos(queryset, area):
    if area == 'historia_clinica':
        return queryset.filter(
            q_oficina_responsable_area(area) |
            (Q(oficina_responsable__isnull=True) & q_prestamos_historia_clinica())
        )
    if area == 'archivo_central':
        return queryset.filter(
            q_oficina_responsable_area(area) |
            (Q(oficina_responsable__isnull=True) & ~q_prestamos_historia_clinica())
        )
    return queryset


def prestamo_es_historia_clinica(prestamo):
    nombres = [
        getattr(getattr(prestamo, 'serie', None), 'nombre', ''),
        getattr(getattr(prestamo, 'subserie', None), 'nombre', ''),
    ]
    criterios = ('historia clinica', 'historias clinicas')
    return any(any(criterio in normalizar_texto(nombre) for criterio in criterios) for nombre in nombres)


def obtener_area_responsable_prestamo(prestamo):
    if prestamo.oficina_responsable:
        area_oficina = obtener_area_desde_nombre_oficina(prestamo.oficina_responsable.nombre)
        if area_oficina:
            return area_oficina
    return 'historia_clinica' if prestamo_es_historia_clinica(prestamo) else 'archivo_central'


def obtener_area_label(area):
    if area == 'historia_clinica':
        return 'Historias Clínicas'
    if area == 'archivo_central':
        return 'Gestión Documental'
    return 'Préstamos Documentales'


def obtener_oficina_responsable_label(prestamo):
    if prestamo.oficina_responsable_id:
        return prestamo.oficina_responsable.nombre
    return obtener_area_label(obtener_area_responsable_prestamo(prestamo))


def usuario_puede_gestionar_prestamo(user, prestamo):
    area_usuario = obtener_area_gestion_usuario(user)
    if area_usuario is None:
        return False
    if area_usuario == 'todas':
        return True
    return area_usuario == obtener_area_responsable_prestamo(prestamo)


def registrar_historial_prestamo(prestamo, evento, descripcion='', usuario=None):
    return HistorialPrestamo.objects.create(
        prestamo=prestamo,
        usuario=usuario,
        evento=evento,
        descripcion=descripcion or None,
    )


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def aplicar_filtros_listado_prestamos(queryset, *, estado='', tipo='', alerta='', oficina_id='', q_search='', filtro_tipo_archivo='', year=None, month=None):
    queryset = aplicar_filtro_area_prestamos(queryset, filtro_tipo_archivo)

    if year:
        queryset = queryset.filter(fecha_solicitud__year=year)

    if month:
        queryset = queryset.filter(fecha_solicitud__month=month)

    if estado:
        queryset = queryset.filter(estado=estado)

    if tipo:
        queryset = queryset.filter(tipo_prestamo=tipo)

    if oficina_id:
        try:
            queryset = queryset.filter(oficina_solicitante_id=int(oficina_id))
        except (ValueError, TypeError):
            pass

    if alerta == 'devolucion_solicitada':
        queryset = queryset.filter(estado='DEVOLUCION_SOLICITADA')
    elif alerta == 'vencidos':
        queryset = queryset.filter(
            tipo_prestamo='FISICO',
            fecha_vencimiento__isnull=False,
            fecha_vencimiento__lt=now().date()
        )
    elif alerta == 'entregados_pendientes_confirmacion':
        queryset = queryset.filter(estado='ENTREGADO')
    elif alerta == 'pendientes_reintegracion':
        queryset = queryset.filter(estado='DEVUELTO')

    if q_search:
        queryset = queryset.filter(
            Q(solicitante__first_name__icontains=q_search) |
            Q(solicitante__last_name__icontains=q_search) |
            Q(solicitante__username__icontains=q_search) |
            Q(codigo_trd__icontains=q_search) |
            Q(descripcion_documento__icontains=q_search) |
            Q(serie__nombre__icontains=q_search) |
            Q(subserie__nombre__icontains=q_search)
        )

    return queryset


def construir_resumen_historial_prestamo(prestamo, limite=6):
    eventos = []
    historial = list(prestamo.historial.all()[:limite])
    for evento in reversed(historial):
        fecha_txt = evento.fecha.strftime('%d/%m/%Y %H:%M') if evento.fecha else ''
        if evento.usuario:
            usuario_txt = evento.usuario.get_full_name() or evento.usuario.username
        else:
            usuario_txt = 'Sistema'
        descripcion = (evento.descripcion or '').strip()
        detalle = f"{fecha_txt} - {evento.evento} ({usuario_txt})"
        if descripcion:
            detalle = f"{detalle}: {descripcion}"
        eventos.append(detalle)
    return ' | '.join(eventos)

@login_required
def solicitar_prestamo(request):
    """
    Vista para que cualquier usuario autenticado solicite un préstamo.
    """
    if request.method == 'POST':
        form = SolicitudPrestamoForm(request.POST)
        if form.is_valid():
            prestamo = form.save(commit=False)
            prestamo.solicitante = request.user
            prestamo.estado = 'SOLICITADO'
            
            # Datos automáticos
            try:
                prestamo.oficina_solicitante = request.user.perfil.oficina
            except AttributeError:
                pass # Usuario sin perfil u oficina
            
            # Vo.Bo. Jefe automático por ahora
            prestamo.vobo_jefe_nombre = "Eliana Gelves" 
            prestamo.vobo_jefe_cargo = "Jefe de Archivo"
            prestamo.vobo_jefe_fecha = now().date()
            
            prestamo.save()
            
            # Crear historial inicial
            registrar_historial_prestamo(
                prestamo,
                "Solicitud creada",
                f"Solicitud de préstamo {prestamo.tipo_prestamo} creada para {obtener_oficina_responsable_label(prestamo)}.",
                request.user,
            )
            
            messages.success(request, 'Solicitud de préstamo enviada correctamente.')
            return redirect('mis_prestamos')
    else:
        form = SolicitudPrestamoForm()
    
    return render(request, 'prestamos/solicitud_form.html', {'form': form})

@login_required
def mis_prestamos(request):
    """
    Lista de préstamos del usuario actual.
    """
    base_queryset = PrestamoDocumental.objects.filter(
        solicitante=request.user
    ).select_related('serie', 'subserie', 'oficina_solicitante', 'oficina_responsable').order_by('-fecha_solicitud')

    estado_filter = request.GET.get('estado', '').strip()
    tipo_filter = request.GET.get('tipo', '').strip()
    search = request.GET.get('q', '').strip()

    prestamos = base_queryset

    if estado_filter:
        prestamos = prestamos.filter(estado=estado_filter)

    if tipo_filter:
        prestamos = prestamos.filter(tipo_prestamo=tipo_filter)

    if search:
        prestamos = prestamos.filter(
            Q(codigo_trd__icontains=search) |
            Q(descripcion_documento__icontains=search) |
            Q(serie__nombre__icontains=search) |
            Q(subserie__nombre__icontains=search)
        )

    # Verificar vencimientos automáticamente al listar
    for p in prestamos:
        if p.verificar_vencimiento():
            registrar_historial_prestamo(
                p,
                "Vencimiento automático",
                "El préstamo superó la fecha de vencimiento y quedó marcado como VENCIDO.",
            )

    resumen_estados = {estado: 0 for estado, _ in PrestamoDocumental.ESTADO_CHOICES}
    for item in base_queryset.values('estado').annotate(total=Count('id')):
        resumen_estados[item['estado']] = item['total']

    tablero_estados = [
        {
            'codigo': codigo,
            'label': label,
            'total': resumen_estados.get(codigo, 0)
        }
        for codigo, label in PrestamoDocumental.ESTADO_CHOICES
    ]

    stats_totales = {
        'total': base_queryset.count(),
        'fisicos': base_queryset.filter(tipo_prestamo='FISICO').count(),
        'virtuales': base_queryset.filter(tipo_prestamo='VIRTUAL').count(),
        'activos': base_queryset.filter(estado='PRESTAMO_ACTIVO').count(),
    }

    paginator = Paginator(prestamos, 8)
    page_number = request.GET.get('page')
    prestamos_paginados = paginator.get_page(page_number)

    context = {
        'prestamos': prestamos_paginados,
        'resumen_estados': resumen_estados,
        'stats_totales': stats_totales,
        'filtros': {
            'estado': estado_filter,
            'tipo': tipo_filter,
            'q': search,
        },
        'estado_board': tablero_estados,
    }

    return render(request, 'prestamos/mis_prestamos_list.html', context)

@login_required
def gestion_prestamos(request):
    """
    Tablero de control para el área responsable del préstamo.
    """
    if not puede_gestionar_prestamos(request.user):
        return HttpResponseForbidden("No tiene permisos para gestionar préstamos.")

    area_usuario = obtener_area_gestion_usuario(request.user)
    
    # Autocuración: corregir préstamos con fecha_reintegracion pero estado DEVUELTO
    # Esto asegura que los préstamos marcados como reintegrados tengan el estado correcto
    PrestamoDocumental.objects.filter(
        estado='DEVUELTO',
        fecha_reintegracion__isnull=False
    ).update(estado='REINTEGRADO')
    
    # Filtros básicos
    estado = request.GET.get('estado', '').strip()
    tipo = request.GET.get('tipo', '').strip()
    alerta = request.GET.get('alerta', '').strip()
    oficina_id = request.GET.get('oficina', '').strip()
    q_search = request.GET.get('q', '').strip()
    year_str = request.GET.get('year', '').strip()
    month_str = request.GET.get('month', '').strip()
    filtro_tipo_archivo = request.GET.get('tipo_archivo', '').strip()  # 'historia_clinica' o 'archivo_central'
    puede_cambiar_area = area_usuario == 'todas'
    if not puede_cambiar_area:
        filtro_tipo_archivo = area_usuario

    today = now().date()
    try:
        filtro_year = int(year_str) if year_str else None
    except (TypeError, ValueError):
        filtro_year = None
    try:
        filtro_month = int(month_str) if month_str else None
        if filtro_month and (filtro_month < 1 or filtro_month > 12):
            filtro_month = None
    except (TypeError, ValueError):
        filtro_month = None
    
    prestamos = PrestamoDocumental.objects.select_related(
        'solicitante', 'oficina_solicitante', 'oficina_responsable', 'serie', 'subserie'
    ).all().order_by('-fecha_solicitud')
    prestamos = aplicar_filtros_listado_prestamos(
        prestamos,
        estado=estado,
        tipo=tipo,
        alerta=alerta,
        oficina_id=oficina_id,
        q_search=q_search,
        filtro_tipo_archivo=filtro_tipo_archivo,
        year=filtro_year,
        month=filtro_month,
    )
        
    # Verificar vencimientos globalmente para préstamos activos físicos
    activos = PrestamoDocumental.objects.filter(estado='PRESTAMO_ACTIVO', tipo_prestamo='FISICO')
    activos = aplicar_filtro_area_prestamos(activos, filtro_tipo_archivo)
    for p in activos:
        p.verificar_vencimiento()

    paginator = Paginator(prestamos, 15)
    page = request.GET.get('page')
    prestamos_paginados = paginator.get_page(page)

    # Base queryset para KPIs y resumen - aplicar el mismo filtro de tipo_archivo si existe
    kpi_queryset = PrestamoDocumental.objects.all()
    kpi_queryset = aplicar_filtros_listado_prestamos(
        kpi_queryset,
        estado=estado,
        tipo=tipo,
        alerta='',
        oficina_id=oficina_id,
        q_search=q_search,
        filtro_tipo_archivo=filtro_tipo_archivo,
        year=filtro_year,
        month=filtro_month,
    )
    # Importante para MSSQL: limpiar el ordering por defecto antes de agrupar
    # Evita ORDER BY en columna no agrupada (fecha_solicitud) durante GROUP BY
    kpi_queryset = kpi_queryset.order_by()
    
    resumen_estados = {estado: 0 for estado, _ in PrestamoDocumental.ESTADO_CHOICES}
    for item in kpi_queryset.values('estado').annotate(total=Count('id')):
        resumen_estados[item['estado']] = item['total']

    tablero_estados_admin = [
        {
            'codigo': codigo,
            'label': label,
            'total': resumen_estados.get(codigo, 0)
        }
        for codigo, label in PrestamoDocumental.ESTADO_CHOICES
    ]

    kpis = {
        'solicitados': kpi_queryset.filter(estado='SOLICITADO').count(),
        'entregados': kpi_queryset.filter(estado='ENTREGADO').count(),
        'activos': kpi_queryset.filter(estado='PRESTAMO_ACTIVO').count(),
        'devolucion_solicitada': kpi_queryset.filter(estado='DEVOLUCION_SOLICITADA').count(),
        'pendientes_reintegracion': kpi_queryset.filter(estado='DEVUELTO').count(),
        'vencidos': kpi_queryset.filter(estado='VENCIDO').count(),
    }

    alertas_operativas = []
    if kpis['devolucion_solicitada']:
        alertas_operativas.append(
            f"{kpis['devolucion_solicitada']} préstamo(s) esperan recepción física por {obtener_area_label(filtro_tipo_archivo if filtro_tipo_archivo else area_usuario)}."
        )
    if kpis['entregados']:
        alertas_operativas.append(
            f"{kpis['entregados']} préstamo(s) ya fueron entregados y siguen pendientes de confirmación del usuario."
        )
    if kpis['pendientes_reintegracion']:
        alertas_operativas.append(
            f"{kpis['pendientes_reintegracion']} préstamo(s) ya fueron devueltos y siguen pendientes de reintegración física."
        )
    
    # Contar préstamos por tipo de archivo para las cards
    total_historia_clinica = PrestamoDocumental.objects.filter(q_prestamos_historia_clinica()).count()
    total_archivo_central = PrestamoDocumental.objects.exclude(q_prestamos_historia_clinica()).count()
    
    # Obtener todas las oficinas para el filtro
    oficinas = OficinaProductora.objects.all().order_by('nombre')
    series = SerieDocumental.objects.all().order_by('nombre')

    filtro_fecha_desde = request.GET.get('fecha_desde', '').strip()
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    filtro_serie = request.GET.get('serie', '').strip()
    
    return render(request, 'prestamos/gestion_list.html', {
        'prestamos': prestamos_paginados,
        'filtro_estado': estado,
        'filtro_tipo': tipo,
        'filtro_alerta': alerta,
        'filtro_oficina': oficina_id,
        'filtro_tipo_archivo': filtro_tipo_archivo,
        'filtro_year': filtro_year,
        'filtro_month': filtro_month,
        'filtro_year_default': today.year,
        'filtro_month_default': today.month,
        'puede_cambiar_area': puede_cambiar_area,
        'area_gestion_label': obtener_area_label(filtro_tipo_archivo if filtro_tipo_archivo else area_usuario),
        'busqueda': q_search,
        'resumen_estados': resumen_estados,
        'estado_board': tablero_estados_admin,
        'kpis': kpis,
        'alertas_operativas': alertas_operativas,
        'total_historia_clinica': total_historia_clinica,
        'total_archivo_central': total_archivo_central,
        'oficinas': oficinas,
        'series': series,
        'filtro_fecha_desde': filtro_fecha_desde,
        'filtro_fecha_hasta': filtro_fecha_hasta,
        'filtro_serie': filtro_serie,
    })


@login_required
def exportar_prestamos_excel_mensual(request):
    if not puede_gestionar_prestamos(request.user):
        return HttpResponseForbidden("No tiene permisos para exportar reportes de préstamos.")

    import calendar as py_calendar

    area_usuario = obtener_area_gestion_usuario(request.user)
    estado = request.GET.get('estado', '').strip()
    tipo = request.GET.get('tipo', '').strip()
    alerta = request.GET.get('alerta', '').strip()
    oficina_id = request.GET.get('oficina', '').strip()
    q_search = request.GET.get('q', '').strip()
    tipo_archivo = request.GET.get('tipo_archivo', '').strip()
    if area_usuario != 'todas':
        tipo_archivo = area_usuario

    today = now().date()
    try:
        year = int(request.GET.get('year', '') or today.year)
        month = int(request.GET.get('month', '') or today.month)
        if month < 1 or month > 12:
            raise ValueError()
    except (TypeError, ValueError):
        return HttpResponse('Parámetros inválidos. Use year=YYYY&month=MM', status=400)

    queryset = PrestamoDocumental.objects.select_related(
        'solicitante',
        'oficina_solicitante',
        'oficina_responsable',
        'serie',
        'subserie',
        'aprobado_por',
        'procesado_por',
    ).prefetch_related(
        'historial__usuario',
        'notificaciones_aviso',
    ).order_by('fecha_solicitud')

    prestamos = aplicar_filtros_listado_prestamos(
        queryset,
        estado=estado,
        tipo=tipo,
        alerta=alerta,
        oficina_id=oficina_id,
        q_search=q_search,
        filtro_tipo_archivo=tipo_archivo,
        year=year,
        month=month,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Prestamos {year}-{month:02d}"

    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    title_font = Font(bold=True, size=12, color="FFFFFF")
    header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    titulo = f"Reporte mensual de préstamos documentales - {py_calendar.month_name[month].capitalize()} {year}"
    ws.merge_cells('A1:R1')
    ws['A1'] = titulo
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    ws['A1'].border = thin_border

    resumen_filtros = [
        f"Área: {obtener_area_label(tipo_archivo if tipo_archivo else area_usuario)}",
        f"Estado: {estado or 'Todos'}",
        f"Tipo: {tipo or 'Todos'}",
        f"Alerta: {alerta or 'Sin filtro'}",
        f"Oficina: {oficina_id or 'Todas'}",
        f"Búsqueda: {q_search or 'Sin criterio'}",
        f"Total registros: {prestamos.count()}",
    ]
    ws.merge_cells('A2:R2')
    ws['A2'] = ' | '.join(resumen_filtros)
    ws['A2'].alignment = left_alignment
    ws['A2'].border = thin_border

    headers = [
        'ID', 'Solicitante', 'Oficina solicitante', 'Oficina responsable', 'Tipo', 'Estado actual',
        'Descripción', 'Fecha solicitud', 'Fecha entrega', 'Fecha confirmación', 'Fecha vencimiento',
        'Fecha devolución', 'Fecha reintegración', 'Aprobado por', 'Procesado por',
        'Notificaciones', 'Devuelto', 'Historial resumido'
    ]

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    def fmt_dt(value):
        return value.strftime('%d/%m/%Y %H:%M') if value else ''

    def fmt_date(value):
        return value.strftime('%d/%m/%Y') if value else ''

    for row_num, prestamo in enumerate(prestamos, start=5):
        values = [
            prestamo.id,
            prestamo.solicitante.get_full_name() or prestamo.solicitante.username,
            prestamo.oficina_solicitante.nombre if prestamo.oficina_solicitante else '',
            obtener_oficina_responsable_label(prestamo),
            prestamo.get_tipo_prestamo_display(),
            prestamo.get_estado_display(),
            prestamo.descripcion_documento or '',
            fmt_dt(prestamo.fecha_solicitud),
            fmt_dt(prestamo.fecha_entrega),
            fmt_dt(prestamo.fecha_confirmacion),
            fmt_date(prestamo.fecha_vencimiento),
            fmt_dt(prestamo.fecha_devolucion),
            fmt_dt(prestamo.fecha_reintegracion),
            prestamo.aprobado_por.get_full_name() if prestamo.aprobado_por else '',
            prestamo.procesado_por.get_full_name() if prestamo.procesado_por else '',
            prestamo.notificaciones_aviso.count(),
            'Sí' if prestamo.documento_devuelto else 'No',
            construir_resumen_historial_prestamo(prestamo),
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = left_alignment if col_num not in [1, 5, 6, 16, 17] else center_alignment

    widths = {
        'A': 8, 'B': 24, 'C': 24, 'D': 28, 'E': 12, 'F': 20, 'G': 42, 'H': 18,
        'I': 18, 'J': 18, 'K': 15, 'L': 18, 'M': 18, 'N': 22, 'O': 22, 'P': 12,
        'Q': 12, 'R': 80,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = 'A5'

    nombre_archivo = f"prestamos_documentales_{year}_{month:02d}.xlsx"
    HistorialDescargaPrestamo.objects.create(
        usuario=request.user,
        anio=year,
        mes=month,
        nombre_archivo=nombre_archivo,
        total_registros=prestamos.count(),
        filtro_estado=estado or None,
        filtro_tipo=tipo or None,
        filtro_alerta=alerta or None,
        filtro_tipo_archivo=tipo_archivo or None,
        filtro_oficina=OficinaProductora.objects.filter(pk=oficina_id).values_list('nombre', flat=True).first() if oficina_id else None,
        busqueda=q_search or None,
        ip_address=get_client_ip(request),
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


@login_required
def detalle_prestamo(request, pk):
    """
    Vista de detalle enriquecida para cualquier usuario relacionado con el préstamo.
    """
    prestamo = get_object_or_404(
        PrestamoDocumental.objects.select_related(
            'solicitante',
            'oficina_solicitante',
            'oficina_responsable',
            'serie',
            'subserie',
            'fuid',
            'registro'
        ),
        pk=pk
    )

    es_propietario = prestamo.solicitante_id == request.user.id
    puede_gestionar = usuario_puede_gestionar_prestamo(request.user, prestamo)

    if not (es_propietario or puede_gestionar):
        return HttpResponseForbidden("No tiene permisos para ver este préstamo.")

    flujo_definido = [
        ('SOLICITADO', 'Solicitud registrada', 'bi-clipboard-plus'),
        ('ENTREGADO', 'Entregado por Archivo', 'bi-box-seam'),
        ('PRESTAMO_ACTIVO', 'Préstamo activo', 'bi-patch-check'),
        ('DEVOLUCION_SOLICITADA', 'Devolución solicitada', 'bi-box-arrow-up'),
        ('DEVUELTO', 'Devuelto a Archivo', 'bi-box-arrow-in-down'),
        ('REINTEGRADO', 'Reintegrado al archivo', 'bi-archive'),
    ]
    
    # Mapeo de estados a índices para comparación fácil
    estados_orden = [e[0] for e in flujo_definido]
    
    try:
        indice_actual = estados_orden.index(prestamo.estado)
    except ValueError:
        indice_actual = -1  # Estado especial (Rechazado, Vencido, etc.)

    flujo_estados = []
    for codigo, titulo, icono in flujo_definido:
        try:
            indice_paso = estados_orden.index(codigo)
        except ValueError:
            indice_paso = 999
            
        completado = False
        activo = False
        
        if indice_actual >= 0:
            if indice_actual > indice_paso:
                completado = True
            elif indice_actual == indice_paso:
                activo = True
                # Para REINTEGRADO, queremos que se vea verde (completado)
                if codigo == 'REINTEGRADO':
                    completado = True
        
        flujo_estados.append({
            'codigo': codigo,
            'titulo': titulo,
            'icono': icono,
            'completado': completado,
            'activo': activo,
        })

    estado_especial = prestamo.estado in ['RECHAZADO', 'RECHAZADO_USUARIO', 'VENCIDO']

    # Calcular información adicional para el resumen
    from django.utils import timezone
    from datetime import timedelta
    
    dias_transcurridos = None
    dias_restantes = None
    duracion_prestamo = None
    
    if prestamo.fecha_solicitud:
        delta = timezone.now() - prestamo.fecha_solicitud
        dias_transcurridos = delta.days
    
    if prestamo.fecha_vencimiento and prestamo.estado in ['PRESTAMO_ACTIVO', 'ENTREGADO']:
        hoy = timezone.now().date()
        if prestamo.fecha_vencimiento > hoy:
            delta = prestamo.fecha_vencimiento - hoy
            dias_restantes = delta.days
        elif prestamo.fecha_vencimiento < hoy:
            dias_restantes = 0  # Ya venció
    
    # Calcular duración del préstamo (desde entrega hasta devolución)
    if prestamo.estado == 'DEVUELTO' and prestamo.fecha_devolucion and prestamo.fecha_entrega:
        if prestamo.fecha_devolucion.tzinfo:
            fecha_devol = prestamo.fecha_devolucion.date()
        else:
            fecha_devol = prestamo.fecha_devolucion
        
        if prestamo.fecha_entrega.tzinfo:
            fecha_ent = prestamo.fecha_entrega.date()
        else:
            fecha_ent = prestamo.fecha_entrega
        
        delta = fecha_devol - fecha_ent
        duracion_prestamo = delta.days
    
    # Calcular tiempo de reintegración (desde devolución hasta reintegración)
    tiempo_reintegracion = None
    if prestamo.fecha_devolucion and prestamo.fecha_reintegracion:
        if prestamo.fecha_reintegracion.tzinfo:
            fecha_reint = prestamo.fecha_reintegracion.date()
        else:
            fecha_reint = prestamo.fecha_reintegracion
        
        if prestamo.fecha_devolucion.tzinfo:
            fecha_devol = prestamo.fecha_devolucion.date()
        else:
            fecha_devol = prestamo.fecha_devolucion
        
        delta = fecha_reint - fecha_devol
        tiempo_reintegracion = delta.days

    # Obtener notificaciones de aviso
    notificaciones_aviso = prestamo.notificaciones_aviso.select_related(
        'notificado_por', 'oficina_notificada'
    ).all()

    contexto = {
        'prestamo': prestamo,
        'area_responsable_label': obtener_oficina_responsable_label(prestamo),
        'flujo_estados': flujo_estados,
        'estado_especial': estado_especial,
        'puede_gestionar': puede_gestionar,
        'puede_confirmar': es_propietario and prestamo.estado == 'ENTREGADO',
        'puede_solicitar_devolucion': es_propietario and prestamo.tipo_prestamo == 'FISICO' and prestamo.estado == 'PRESTAMO_ACTIVO',
        'historial': prestamo.historial.select_related('usuario').all(),
        'notificaciones_aviso': notificaciones_aviso,
        'dias_transcurridos': dias_transcurridos,
        'dias_restantes': dias_restantes,
        'duracion_prestamo': duracion_prestamo,
        'tiempo_reintegracion': tiempo_reintegracion,
    }

    return render(request, 'prestamos/detalle.html', contexto)

@login_required
def procesar_prestamo(request, pk):
    """
    Vista para gestionar un préstamo específico.
    Flujo:
    - Si está SOLICITADO: Archivo entrega o rechaza la solicitud
    - Cuando Archivo entrega, el préstamo queda en ENTREGADO
    - El usuario solicitante es quien confirma la recepción y lo lleva a PRESTAMO_ACTIVO
    """
    prestamo = get_object_or_404(PrestamoDocumental, pk=pk)
    if not usuario_puede_gestionar_prestamo(request.user, prestamo):
        return HttpResponseForbidden("No tiene permisos para gestionar este préstamo.")
    
    if request.method == 'POST':
        form = GestionPrestamoForm(request.POST, request.FILES, instance=prestamo)
        if form.is_valid():
            accion = form.cleaned_data.get('accion')
            documento_escaneado = form.cleaned_data.get('documento_escaneado')
            motivo_rechazo = form.cleaned_data.get('motivo_rechazo')
            
            # Si el préstamo está SOLICITADO
            if prestamo.estado == 'SOLICITADO':
                if accion == 'cargar_evidencia':
                    # Guardar el formulario primero
                    prestamo = form.save(commit=False)
                    
                    # Procesar múltiples archivos - usar getlist directamente de request.FILES
                    documentos_escaneados = request.FILES.getlist('documentos_escaneados')
                    archivos_subidos = False
                    
                    if documentos_escaneados:
                        import os
                        
                        for archivo in documentos_escaneados:
                            DocumentoEscaneadoPrestamo.objects.create(
                                prestamo=prestamo,
                                archivo=archivo,
                                nombre_archivo_original=archivo.name,
                                uploaded_by=request.user,
                                confirmado=True,
                                fecha_confirmacion=now()
                            )
                            archivos_subidos = True
                    
                    # También manejar el campo legacy documento_escaneado si se usa
                    if documento_escaneado:
                        archivos_subidos = True
                    
                    # Si es préstamo físico, no requiere documentos para cambiar a ENTREGADO
                    # Si es virtual, requiere al menos un archivo
                    tiene_archivos = (archivos_subidos or 
                                     prestamo.documento_escaneado or 
                                     prestamo.documentos_escaneados.exists())
                    
                    # Para préstamos físicos, permitir cambiar a ENTREGADO sin documentos
                    if prestamo.tipo_prestamo == 'FISICO' or tiene_archivos:
                        momento_entrega = now()
                        prestamo.estado = 'ENTREGADO'
                        prestamo.fecha_aprobacion = momento_entrega
                        prestamo.fecha_entrega = momento_entrega
                        prestamo.aprobado_por = request.user
                        prestamo.procesado_por = request.user
                        prestamo.confirmado_por_usuario = False
                        prestamo.fecha_confirmacion = None
                        # Calcular vencimiento si es físico
                        if prestamo.tipo_prestamo == 'FISICO':
                            prestamo.fecha_vencimiento = prestamo.calcular_fecha_vencimiento()
                        
                        prestamo.save()
                        
                        # Guardar también el campo legacy si se usó
                        if documento_escaneado:
                            form.save()
                        
                        # Registrar en historial
                        if tiene_archivos:
                            num_archivos = len(documentos_escaneados) if documentos_escaneados else (1 if documento_escaneado else 0)
                            registrar_historial_prestamo(
                                prestamo,
                                "Entrega registrada por Archivo",
                                f"Se cargaron {num_archivos} archivo(s) de soporte. El préstamo queda en ENTREGADO y espera la confirmación del usuario solicitante.",
                                request.user,
                            )
                            messages.success(request, f'Préstamo #{prestamo.id}: Entrega registrada. Estado actualizado a ENTREGADO y pendiente de confirmación del usuario.')
                        else:
                            # Préstamo físico sin documentos
                            registrar_historial_prestamo(
                                prestamo,
                                "Entrega física registrada por Archivo",
                                "Préstamo físico entregado sin adjuntar soportes adicionales. El préstamo queda en ENTREGADO y espera la confirmación del usuario solicitante.",
                                request.user,
                            )
                            messages.success(request, f'Préstamo #{prestamo.id}: Entrega física registrada. Estado actualizado a ENTREGADO y pendiente de confirmación del usuario.')
                    else:
                        form.save()
                        messages.warning(request, 'Debe subir al menos un documento escaneado.')
                        return redirect('procesar_prestamo', pk=prestamo.id)
                
                elif accion == 'rechazar':
                    # Rechazar el préstamo
                    prestamo = form.save(commit=False)
                    prestamo.estado = 'RECHAZADO'
                    prestamo.motivo_rechazo = motivo_rechazo
                    prestamo.procesado_por = request.user
                    
                    # Guardar el documento de rechazo si se adjuntó
                    if 'documento_rechazo' in request.FILES:
                        prestamo.documento_rechazo = request.FILES['documento_rechazo']
                    
                    prestamo.save()
                    
                    # Registrar en historial
                    descripcion_historial = f"Motivo: {motivo_rechazo}"
                    if prestamo.documento_rechazo:
                        descripcion_historial += f" (Documento adjunto: {prestamo.documento_rechazo.name})"
                    
                    registrar_historial_prestamo(prestamo, "Préstamo rechazado", descripcion_historial, request.user)
                    
                    messages.warning(request, f'Préstamo #{prestamo.id} rechazado.')
                
            else:
                # Para otros estados, guardar normalmente
                prestamo = form.save(commit=False)
                
                # Verificar si se marcó finalizar préstamo
                finalizar_prestamo = form.cleaned_data.get('finalizar_prestamo', False)
                
                if finalizar_prestamo and prestamo.tipo_prestamo == 'FISICO':
                    estado_previo = prestamo.estado
                    if estado_previo not in ['PRESTAMO_ACTIVO', 'DEVOLUCION_SOLICITADA', 'VENCIDO']:
                        messages.warning(request, 'Solo puedes registrar la devolución cuando el préstamo está activo, vencido o ya tiene una solicitud de devolución.')
                        return redirect('procesar_prestamo', pk=prestamo.id)

                    # Finalizar el préstamo: cambiar estado a DEVUELTO
                    prestamo.estado = 'DEVUELTO'
                    prestamo.fecha_devolucion = now()
                    prestamo.documento_devuelto = True
                    prestamo.procesado_por = request.user

                    if estado_previo == 'DEVOLUCION_SOLICITADA':
                        descripcion_devolucion = f"{obtener_oficina_responsable_label(prestamo)} recibió físicamente el documento después de la solicitud de devolución del usuario."
                    elif estado_previo == 'VENCIDO':
                        descripcion_devolucion = f"{obtener_oficina_responsable_label(prestamo)} recibió físicamente el documento y cerró un préstamo vencido."
                    else:
                        descripcion_devolucion = f"{obtener_oficina_responsable_label(prestamo)} recibió físicamente el documento y cerró el préstamo sin solicitud previa de devolución."
                    
                    # Registrar en historial
                    registrar_historial_prestamo(prestamo, "Documento recibido por Archivo", descripcion_devolucion, request.user)
                    
                    messages.success(request, f'Préstamo #{prestamo.id} finalizado exitosamente. Estado actualizado a DEVUELTO.')
                else:
                    messages.success(request, f'Préstamo #{prestamo.id} actualizado.')
                
                prestamo.save()
            
            return redirect('gestion_prestamos')
    else:
        form = GestionPrestamoForm(instance=prestamo)
    
    historial = prestamo.historial.select_related('usuario').order_by('-fecha')[:8]

    return render(request, 'prestamos/procesar_form.html', {
        'form': form,
        'prestamo': prestamo,
        'area_responsable_label': obtener_oficina_responsable_label(prestamo),
        'historial': historial,
    })

@login_required
@require_POST
def enviar_notificacion_aviso(request, pk):
    """
    Vista para enviar una notificación de aviso sobre retraso en un préstamo.
    Solo el área responsable puede enviar notificaciones.
    """
    prestamo = get_object_or_404(PrestamoDocumental, pk=pk)
    if not usuario_puede_gestionar_prestamo(request.user, prestamo):
        return HttpResponseForbidden("No tiene permisos para gestionar este préstamo.")
    
    documento_oficio = request.FILES.get('documento_oficio')
    observaciones = request.POST.get('observaciones', '').strip()
    
    if not documento_oficio:
        messages.error(request, 'Debe adjuntar un documento de oficio.')
        return redirect('gestion_prestamos')
    
    # Crear la notificación
    notificacion = NotificacionAvisoPrestamo.objects.create(
        prestamo=prestamo,
        documento_oficio=documento_oficio,
        observaciones=observaciones if observaciones else None,
        notificado_por=request.user,
        oficina_notificada=prestamo.oficina_solicitante
    )
    
    # Registrar en historial
    registrar_historial_prestamo(
        prestamo,
        "Notificación de aviso enviada",
        f"Se envió una notificación de aviso a la oficina {prestamo.oficina_solicitante.nombre if prestamo.oficina_solicitante else 'N/A'}. {f'Observaciones: {observaciones}' if observaciones else ''}",
        request.user,
    )
    
    messages.success(request, f'Notificación de aviso enviada correctamente para el préstamo #{prestamo.id}.')
    return redirect('gestion_prestamos')


@login_required
@require_POST
def reintegrar_prestamo(request, pk):
    """
    Vista para marcar un préstamo como reintegrado al archivo responsable.
    Cambia el estado a REINTEGRADO y establece la fecha_reintegracion cuando el documento es físicamente reintegrado a su caja/ubicación.
    """
    prestamo = get_object_or_404(PrestamoDocumental, pk=pk)
    if not usuario_puede_gestionar_prestamo(request.user, prestamo):
        return HttpResponseForbidden("No tiene permisos para gestionar este préstamo.")

    area_responsable_label = obtener_oficina_responsable_label(prestamo)
    
    # Log para depuración
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Reintegrar préstamo {pk}: Estado actual = {prestamo.estado}")
    
    if prestamo.estado != 'DEVUELTO':
        messages.error(request, f'Solo se pueden reintegrar préstamos que estén en estado DEVUELTO. Estado actual: {prestamo.get_estado_display()}')
        return redirect('procesar_prestamo', pk=prestamo.id)
    
    # Si ya está reintegrado, no hacer nada
    if prestamo.estado == 'REINTEGRADO':
        messages.info(request, 'Este préstamo ya fue marcado como reintegrado anteriormente.')
        return redirect('procesar_prestamo', pk=prestamo.id)
    
    # Cambiar estado a REINTEGRADO y establecer fecha de reintegración
    prestamo.estado = 'REINTEGRADO'
    prestamo.fecha_reintegracion = now()
    prestamo.save(update_fields=['estado', 'fecha_reintegracion'])
    
    logger.info(f"Préstamo {pk} actualizado: estado = {prestamo.estado}, fecha_reintegracion = {prestamo.fecha_reintegracion}")
    
    # Registrar en historial
    registrar_historial_prestamo(
        prestamo,
        f"Reintegrado a {area_responsable_label}",
        f"Documento reintegrado físicamente a {area_responsable_label} el {prestamo.fecha_reintegracion.strftime('%d/%m/%Y %H:%M')}",
        request.user,
    )
    
    messages.success(request, f'Préstamo #{prestamo.id} marcado como reintegrado a {area_responsable_label}.')
    return redirect('procesar_prestamo', pk=prestamo.id)

@login_required
@require_POST
def solicitar_devolucion(request, pk):
    """
    El usuario con préstamo activo solicita la devolución del documento.
    Cambia el estado a DEVOLUCION_SOLICITADA para que Archivo reciba la notificación
    y se dirija al puesto de trabajo a recibir el documento.
    """
    prestamo = get_object_or_404(PrestamoDocumental, pk=pk, solicitante=request.user)
    
    if prestamo.tipo_prestamo != 'FISICO':
        messages.error(request, 'La solicitud de devolución solo aplica para préstamos físicos.')
        return redirect('mis_prestamos')

    if prestamo.estado != 'PRESTAMO_ACTIVO':
        messages.error(request, 'Solo puedes solicitar devolución cuando el préstamo está activo.')
        return redirect('mis_prestamos')
    
    prestamo.estado = 'DEVOLUCION_SOLICITADA'
    prestamo.save(update_fields=['estado'])
    
    registrar_historial_prestamo(
        prestamo,
        "Devolución solicitada por usuario",
        f"El usuario {request.user.get_full_name() or request.user.username} solicita la devolución y entrega del documento a {obtener_oficina_responsable_label(prestamo)}.",
        request.user,
    )
    
    messages.success(request, f'Solicitud de devolución enviada para el préstamo #{prestamo.id}. El área de Archivo será notificada.')
    return redirect('mis_prestamos')

@login_required
@require_POST
def eliminar_documento_prestamo(request, pk):
    """
    Vista AJAX para eliminar un documento escaneado de un préstamo.
    """
    try:
        documento = get_object_or_404(
            DocumentoEscaneadoPrestamo.objects.select_related('prestamo', 'prestamo__serie', 'prestamo__subserie'),
            pk=pk
        )
        if not usuario_puede_gestionar_prestamo(request.user, documento.prestamo):
            return JsonResponse({'success': False, 'error': 'No tiene permisos para gestionar este préstamo.'}, status=403)
        documento.archivo.delete(save=False)  # Eliminar el archivo del sistema de archivos
        documento.delete()  # Eliminar el registro de la base de datos
        return JsonResponse({'success': True, 'message': 'Archivo eliminado exitosamente.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def confirmar_recepcion(request, pk):
    """
    Acción para que el usuario confirme que recibió el documento.
    """
    prestamo = get_object_or_404(PrestamoDocumental, pk=pk, solicitante=request.user)
    
    if prestamo.estado != 'ENTREGADO':
        messages.error(request, "Solo puedes dar respuesta cuando el préstamo aparece como ENTREGADO.")
        return redirect('mis_prestamos')

    if request.method == 'POST':
        form = ConfirmacionPrestamoForm(request.POST)
        if form.is_valid():
            accion = form.cleaned_data['accion']
            motivo = form.cleaned_data.get('motivo_rechazo_usuario')

            if accion == 'confirmar':
                prestamo.confirmado_por_usuario = True
                prestamo.fecha_confirmacion = now()
                prestamo.estado = 'PRESTAMO_ACTIVO'
                prestamo.save(update_fields=['confirmado_por_usuario', 'fecha_confirmacion', 'estado'])

                registrar_historial_prestamo(
                    prestamo,
                    "Recepción confirmada por usuario",
                    f"El usuario aprobó la entrega realizada por {obtener_oficina_responsable_label(prestamo)}.",
                    request.user,
                )
                messages.success(request, "¡Listo! Confirmaste que el documento corresponde.")

            elif accion == 'rechazar':
                prestamo.estado = 'RECHAZADO_USUARIO'
                prestamo.motivo_rechazo_usuario = motivo
                prestamo.save(update_fields=['estado', 'motivo_rechazo_usuario'])

                registrar_historial_prestamo(
                    prestamo,
                    "Usuario rechazó la entrega",
                    f"Motivo registrado: {motivo}",
                    request.user,
                )
                messages.warning(request, "Registramos tu observación. El área responsable revisará la solicitud.")

            return redirect('mis_prestamos')
    else:
        form = ConfirmacionPrestamoForm()

    return render(request, 'prestamos/confirmacion.html', {
        'form': form,
        'prestamo': prestamo,
    })


# ---- Vista de Registro de Usuario ----------------------------------------

def registro_usuario(request):
    """Vista para registro de nuevos usuarios"""
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            try:
                perfil = form.save()
                messages.success(
                    request,
                    f'¡Registro exitoso! Tu cuenta ha sido creada y está pendiente de aprobación. '
                    f'Recibirás una notificación cuando un administrador active tu usuario.'
                )
                return redirect('login')
            except Exception as e:
                messages.error(request, f'Error al crear el usuario: {str(e)}')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = RegistroUsuarioForm()
    
    from datetime import datetime
    context = {
        'form': form,
        'current_year': datetime.now().year,
    }
    return render(request, 'registration/registro.html', context)


@login_required
def buscar_registros_ajax(request):
    """
    AJAX endpoint para buscar registros sin cargar todos en memoria.
    Optimizado para manejar millones de registros.
    """
    query = request.GET.get('q', '').strip()
    limit = int(request.GET.get('limit', 20))
    
    if not query or len(query) < 2:
        return JsonResponse({'results': []})
    
    # Búsqueda rápida en campos indexados que existen en el modelo
    registros = RegistroDeArchivo.objects.filter(
        Q(codigo__icontains=query) |
        Q(unidad_documental__icontains=query) |
        Q(identificador_documento__icontains=query) |
        Q(notas__icontains=query)
    ).values('id', 'codigo', 'unidad_documental', 'numero_orden', 'fecha_archivo')[:limit]
    
    results = [
        {
            'id': r['id'],
            'text': f"#{r['numero_orden']} - {r['unidad_documental'][:50]}",
            'referencia': r['codigo'] or f"#{r['numero_orden']}",
            'asunto': r['unidad_documental']
        }
        for r in registros
    ]
    
    return JsonResponse({'results': results})
